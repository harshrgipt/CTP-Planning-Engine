"""L0 -- CONTINUOUS LEARNING.  Parameters only, never policy.

    python -m planner.cmbc.l0_learn [--as-of 2026-08-01] [--months 8]

Emits `warehouse/params/params_<as_of>.json` plus per-key parquet tables.

THE ONE RULE THIS LAYER MUST NOT BREAK
  L0 emits NUMBERS. It never emits a decision, a preference, a ranking or a
  hard/soft classification.  The old `planner/learn/` + `kb/promoter.py` path
  does exactly that -- it mines candidates and promotes them to hard/soft rules,
  which is POLICY.  Learning the planner's past decisions reproduces the plant's
  current throughput INCLUDING its mistakes: 56 of 93 PCR GTs are locked to one
  machine by habit, and a policy-learner would faithfully re-learn that lock.
  Learn what the plant IS; decide separately what it SHOULD BE.

  Concretely, L0 may emit "GT X was built on machine M 94% of the time".
  It may NOT emit "GT X should be built on machine M".

PROVENANCE IS PART OF THE OUTPUT
  Every parameter carries n, the window, the method, and -- where the quantity
  is measurable per month -- a stability CV across months.  A parameter without
  provenance cannot be audited later, and several numbers in this project were
  wrong for weeks precisely because they arrived without it.

MEASUREMENT RULES APPLIED HERE (each earned by a real error)
  * every rate names its resource and cardinality: tyres . press^-1 . day^-1
  * never a ratio of medians as a rate -- use totals
  * press busy time by INTERVAL UNION, never per-tyre (co-cured tyres do not
    share a timestamp; per-tyre summing gave 304% utilisation)
  * filter on the Hive partition column `date`, never event_ts
"""
from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path

import numpy as np
import polars as pl

from planner.data.warehouse import duck, set_cutoff
from planner import paths

ROOT = paths.ROOT   # depth-independent; this file moved one level deeper
OUT = ROOT / "warehouse" / "params"
PLANTS = ["PCR", "TBR"]
IDLE_CUT_S = 1800.0        # a build gap beyond 30 min is idle, not a slow tyre
DOWN_CUT_H = 4.0           # a press gap beyond 4 h counts as down, not a change


def _months(as_of: str, n: int) -> list[tuple[str, str]]:
    y, m = int(as_of[:4]), int(as_of[5:7])
    out = []
    for _ in range(n):
        m -= 1
        if m == 0:
            m, y = 12, y - 1
        nm, ny = (m % 12) + 1, y + (m == 12)
        out.append((f"{y:04d}-{m:02d}-01", f"{ny:04d}-{nm:02d}-01"))
    return sorted(out)


def _stab(vals: list[float]) -> float | None:
    """CV across months. Reported, never used to filter."""
    v = [x for x in vals if x is not None and x == x]
    if len(v) < 3:
        return None
    mu = float(np.mean(v))
    return float(np.std(v) / mu) if mu else None


# ---------------------------------------------------------------------------
def tau(w0: str, w1: str, mons: list) -> dict:
    """Coupling buffer: per-TYRE build->cure wait via the barcode join."""
    q = duck().execute(f"""
        SELECT b.plant,
               count(*) AS n,
               quantile_cont(date_diff('second', b.event_ts, c.event_ts), 0.05)/3600.0 AS p05,
               median(date_diff('second', b.event_ts, c.event_ts))/3600.0 AS p50,
               quantile_cont(date_diff('second', b.event_ts, c.event_ts), 0.95)/3600.0 AS p95,
               quantile_cont(date_diff('second', b.event_ts, c.event_ts), 0.001)/3600.0 AS p001
        FROM v_build b JOIN v_curing c ON b.productionID = c.gtbarCode
        WHERE b.stage=2 AND b.date >= DATE '{w0}' AND b.date < DATE '{w1}'
        GROUP BY 1""").pl()
    per = {p: [] for p in PLANTS}
    for a, b in mons:
        m = duck().execute(f"""
            SELECT b.plant, median(date_diff('second', b.event_ts, c.event_ts))/3600.0 AS p50
            FROM v_build b JOIN v_curing c ON b.productionID = c.gtbarCode
            WHERE b.stage=2 AND b.date >= DATE '{a}' AND b.date < DATE '{b}'
            GROUP BY 1""").pl()
        for r in m.iter_rows(named=True):
            per[r["plant"]].append(float(r["p50"]))
    out = {}
    for r in q.iter_rows(named=True):
        p = r["plant"]
        out[p] = {"tau_star_h": round(float(r["p50"]), 3),
                  "tau_min_h": round(max(float(r["p001"]), 0.0), 3),
                  "tau_p05_h": round(float(r["p05"]), 3),
                  "tau_warn_h": round(float(r["p95"]), 3),
                  "n": int(r["n"]), "cv_across_months": _stab(per[p]),
                  "method": "per-tyre productionID=gtbarCode"}
    return out


def _gap_cutoff_h() -> dict[str, float]:
    """Campaign-end gap threshold, DERIVED from mould-change duration.

    24 h was an assumption and the bands were sensitive to it -- TBR's cure p50
    moved 265 h -> 218 h when it was applied. The physical question is "could the
    mould have come out?", so the threshold is the mould-change time itself: a
    press idle on a GT for longer than it takes to change a mould is no longer
    in that campaign.

    Falls back to 24 h only if the mould-change master is absent, and says so.
    """
    fp = ROOT / "warehouse" / "derived" / "press_mould_change.parquet"
    if not fp.exists():
        return {p: 24.0 for p in PLANTS} | {"_source": "fallback 24 h (no master)"}
    d = pl.read_parquet(fp)
    out: dict[str, float] = {}
    for p in PLANTS:
        s2 = d.filter(pl.col("plant") == p)
        out[p] = (round(float(s2["mould_change_min"].median()) / 60.0, 2)
                  if s2.height else 24.0)
    out["_source"] = "press_mould_change.parquet median, per plant"
    return out


CAMP_GAP = _gap_cutoff_h()


def campaign_bands(w0: str, w1: str, mons: list | None = None) -> dict:
    """Cure and build campaign length. A BAND, never a target.

    A run must break on a GAP as well as on a GT change. Without the gap
    break, a press returning to the same GT three weeks later reads as ONE
    campaign -- which produced a p90 of 858 h (PCR) / 1,326 h (TBR), longer
    than the month being measured.
    """
    # MEASURE PER MONTH, THEN AGGREGATE ACROSS MONTHS.
    # Measuring over the whole 8-month window without a date restriction lets a
    # press that runs one GT for months read as ONE campaign, inflating the mean:
    # PCR came out 291.5 h against a July-restricted 169.6 h, TBR 461.3 vs 270.0.
    # Lot sizing driven off the inflated figure produced 1,848-tyre lots that
    # could not be packed. The month is the natural unit -- it is also the
    # planning horizon -- so each month is measured separately and the months are
    # then combined.
    out = {p: {} for p in PLANTS}
    windows = mons if mons else [(w0, w1)]
    for side in ["build", "cure"]:
        if side == "build":
            sql = f"""
            WITH s AS (SELECT plant, machineCode res, itemCode gt, event_ts,
              lag(itemCode) OVER (PARTITION BY plant,machineCode ORDER BY event_ts) pg
              FROM v_build WHERE stage=2 AND itemCode IS NOT NULL
                AND machineCode IS NOT NULL
                AND date >= DATE '{w0}' AND date < DATE '{w1}'),
            g AS (SELECT *, date_diff('second', lag(event_ts) OVER
                     (PARTITION BY plant,res ORDER BY event_ts), event_ts)/3600.0 gap
                  FROM s),
            r AS (SELECT *, sum(CASE WHEN pg IS DISTINCT FROM gt
                                       OR gap > CASE plant WHEN 'PCR' THEN {CAMP_GAP['PCR']} ELSE {CAMP_GAP['TBR']} END THEN 1 ELSE 0 END)
                     OVER (PARTITION BY plant,res ORDER BY event_ts) run FROM g)
            SELECT plant, count(*) qty,
                   date_diff('second', min(event_ts), max(event_ts))/3600.0 hrs
            FROM r GROUP BY plant, res, gt, run"""
        else:
            # A cure campaign ends when the MOULD CHANGES on that press.
            # v_curing.mouldNo is 100% populated, so this is the physical event
            # itself -- no gap proxy needed. Using a gap threshold instead was
            # wrong: TBR MTTR is 11.7 h, so any cutoff below that splits a
            # campaign on every breakdown, even though the mould stays mounted
            # throughout. That drove TBR cure p50 from 218 h down to 48 h.
            sql = f"""
            WITH cg AS (SELECT c.plant, CAST(c.wcID AS VARCHAR) res,
              b.itemCode gt, CAST(c.mouldNo AS VARCHAR) mould, c.event_ts
              FROM v_curing c JOIN v_build b ON c.gtbarCode=b.productionID
              WHERE b.stage=2 AND b.itemCode IS NOT NULL AND c.wcID IS NOT NULL
                AND c.mouldNo IS NOT NULL
                AND c.date >= DATE '{w0}' AND c.date < DATE '{w1}'),
            s AS (SELECT *,
              lag(gt) OVER (PARTITION BY plant,res ORDER BY event_ts) pg,
              lag(mould) OVER (PARTITION BY plant,res ORDER BY event_ts) pm
              FROM cg),
            r AS (SELECT *, sum(CASE WHEN pg IS DISTINCT FROM gt
                                       OR pm IS DISTINCT FROM mould
                                     THEN 1 ELSE 0 END)
                     OVER (PARTITION BY plant,res ORDER BY event_ts) run FROM s)
            SELECT plant, count(*) qty,
                   date_diff('second', min(event_ts), max(event_ts))/3600.0 hrs
            FROM r GROUP BY plant, res, gt, mould, run"""
        frames = []
        for _w0, _w1 in windows:
            frames.append(duck().execute(sql.replace(w0, _w0).replace(w1, _w1)).pl())
        d = pl.concat(frames, how="vertical") if frames else pl.DataFrame()
        for p in PLANTS:
            s = d.filter(pl.col("plant") == p)
            if not s.height:
                continue
            h = np.array(s["hrs"], float); q = np.array(s["qty"], float)
            out[p][side] = {
                "hours_p10": round(float(np.percentile(h, 10)), 2),
                "hours_p50": round(float(np.percentile(h, 50)), 2),
                "hours_p90": round(float(np.percentile(h, 90)), 2),
                "hours_mean": round(float(h.mean()), 2),
                # the SHAPE, not just its summary. Sizing every lot at the mean
                # collapses the plan to a point mass (p50 == p90 == max) and
                # uniform blocks tile badly: a 744 h press fits 4 x 175 h and
                # wastes the 44 h tail, 3,784 h across 86 PCR presses.
                "hours_deciles": [round(float(np.percentile(h, x)), 2)
                                  for x in range(5, 100, 10)],
                "qty_p50": int(np.percentile(q, 50)),
                "qty_mean": round(float(q.mean()), 1),
                "qty_cv": round(float(q.std() / max(q.mean(), 1e-9)), 3),
                "n_campaigns": int(s.height)}
    return out


def cure_cycle(w0: str, w1: str) -> pl.DataFrame:
    """Cure cycle seconds by (plant, mould, press). cycleStart is press-OPEN."""
    return duck().execute(f"""
        SELECT plant, CAST(mouldNo AS VARCHAR) AS mould,
               CAST(wcID AS VARCHAR) AS press,
               count(*) AS n,
               median(date_diff('second', event_ts, cycleStart)) AS cycle_s_p50,
               quantile_cont(date_diff('second', event_ts, cycleStart), 0.9) AS cycle_s_p90
        FROM v_curing
        WHERE wcID IS NOT NULL AND cycleStart IS NOT NULL
          AND date >= DATE '{w0}' AND date < DATE '{w1}'
          AND date_diff('second', event_ts, cycleStart) BETWEEN 60 AND 14400
        GROUP BY 1,2,3 HAVING count(*) >= 20""").pl()


def build_cadence(w0: str, w1: str) -> pl.DataFrame:
    """Per-tyre cadence per machine [seconds . tyre^-1 . machine^-1]."""
    return duck().execute(f"""
        WITH s AS (SELECT plant, machineCode m,
          date_diff('second', lag(event_ts) OVER
            (PARTITION BY plant,machineCode ORDER BY event_ts), event_ts) g
          FROM v_build WHERE stage=2 AND machineCode IS NOT NULL
            AND date >= DATE '{w0}' AND date < DATE '{w1}')
        SELECT plant, m AS machine, count(*) AS n,
               median(g) AS cadence_s_p50
        FROM s WHERE g BETWEEN 1 AND {IDLE_CUT_S} GROUP BY 1,2""").pl()


def press_availability(w0: str, w1: str) -> dict:
    """MTBF / MTTR from press idle gaps.  A gap > DOWN_CUT_H is 'down'."""
    d = duck().execute(f"""
        WITH s AS (SELECT plant, CAST(wcID AS VARCHAR) press, event_ts,
          date_diff('second', lag(event_ts) OVER
            (PARTITION BY plant, CAST(wcID AS VARCHAR) ORDER BY event_ts),
            event_ts)/3600.0 AS gap_h
          FROM v_curing WHERE wcID IS NOT NULL
            AND date >= DATE '{w0}' AND date < DATE '{w1}')
        SELECT plant,
               count(*) AS n,
               sum(CASE WHEN gap_h > {DOWN_CUT_H} THEN 1 ELSE 0 END) AS downs,
               sum(CASE WHEN gap_h > {DOWN_CUT_H} THEN gap_h ELSE 0 END) AS down_h,
               sum(CASE WHEN gap_h <= {DOWN_CUT_H} THEN gap_h ELSE 0 END) AS up_h
        FROM s WHERE gap_h IS NOT NULL GROUP BY 1""").pl()
    out = {}
    for r in d.iter_rows(named=True):
        downs = max(int(r["downs"]), 1)
        out[r["plant"]] = {
            "mtbf_h": round(float(r["up_h"]) / downs, 2),
            "mttr_h": round(float(r["down_h"]) / downs, 2),
            "availability": round(float(r["up_h"]) /
                                  max(float(r["up_h"]) + float(r["down_h"]), 1e-9), 4),
            "n_down_events": int(r["downs"]),
            "method": f"press gap > {DOWN_CUT_H} h counted as down"}
    return out


def yields(w0: str, w1: str) -> dict:
    """Cure completion fraction. Build side carries NO signal -- reported null."""
    c = duck().execute(f"""
        SELECT plant, count(*) n,
               sum(CASE WHEN upper(statuscritical)='NORMAL' THEN 1 ELSE 0 END) ok
        FROM v_curing WHERE date >= DATE '{w0}' AND date < DATE '{w1}'
        GROUP BY 1""").pl()
    b = duck().execute(f"""
        SELECT plant, count(DISTINCT QualityStatus) AS distinct_status
        FROM v_build WHERE stage=2 AND date >= DATE '{w0}' AND date < DATE '{w1}'
        GROUP BY 1""").pl()
    bmap = {r["plant"]: int(r["distinct_status"]) for r in b.iter_rows(named=True)}
    out = {}
    for r in c.iter_rows(named=True):
        p = r["plant"]
        out[p] = {"cure_yield": round(float(r["ok"]) / max(int(r["n"]), 1), 5),
                  "n": int(r["n"]),
                  "build_yield": None,
                  "build_yield_note":
                      f"QualityStatus has {bmap.get(p,0)} distinct value(s) -- "
                      "no variation, so no build scrap signal exists. NOT assumed 1.0."}
    return out


def observed_lot_sizes(w0: str, w1: str) -> pl.DataFrame:
    """Run size per (plant, gt) -- the empirical floor to sanity-check L4.5."""
    return duck().execute(f"""
        WITH s AS (SELECT plant, machineCode res, itemCode gt, event_ts,
          lag(itemCode) OVER (PARTITION BY plant,machineCode ORDER BY event_ts) pg
          FROM v_build WHERE stage=2 AND itemCode IS NOT NULL
            AND machineCode IS NOT NULL
            AND date >= DATE '{w0}' AND date < DATE '{w1}'),
        r AS (SELECT *, sum(CASE WHEN pg IS DISTINCT FROM gt THEN 1 ELSE 0 END)
                 OVER (PARTITION BY plant,res ORDER BY event_ts) run FROM s),
        runs AS (SELECT plant, gt, res, run, count(*) sz FROM r GROUP BY 1,2,3,4)
        SELECT plant, gt AS gt_code, count(*) AS n_runs,
               quantile_cont(sz, 0.10) AS sz_p10,
               median(sz) AS sz_p50,
               quantile_cont(sz, 0.90) AS sz_p90
        FROM runs GROUP BY 1,2""").pl()


def changeover_durations(w0: str, w1: str) -> dict:
    """Observed changeover minutes, alongside the plant's declared cost model.

    The plant costs a changeover as same-size vs different-size PER MACHINE.
    The observed inter-campaign gap is a DIFFERENT quantity -- it includes
    waiting, not just setup. Both are reported; they are never averaged.
    """
    obs = duck().execute(f"""
        WITH s AS (SELECT plant, machineCode m, itemCode g, event_ts,
          lag(itemCode) OVER (PARTITION BY plant,machineCode ORDER BY event_ts) pg,
          date_diff('second', lag(event_ts) OVER
            (PARTITION BY plant,machineCode ORDER BY event_ts), event_ts)/60.0 mins
          FROM v_build WHERE stage=2 AND itemCode IS NOT NULL
            AND machineCode IS NOT NULL
            AND date >= DATE '{w0}' AND date < DATE '{w1}')
        SELECT plant, count(*) n, median(mins) p50,
               quantile_cont(mins, 0.90) p90
        FROM s WHERE pg IS NOT NULL AND g <> pg AND mins BETWEEN 1 AND 240
        GROUP BY 1""").pl()
    out = {}
    for r in obs.iter_rows(named=True):
        out[r["plant"]] = {"observed_p50_min": round(float(r["p50"]), 2),
                           "observed_p90_min": round(float(r["p90"]), 2),
                           "n": int(r["n"])}
    for p, fn in [("PCR", "Master_Building_ChangeoverTime_pcr.csv"),
                  ("TBR", "Master_Building_ChangeoverTime_tbr.csv")]:
        fp = ROOT / "masters" / fn
        if fp.exists() and p in out:
            d = pl.read_csv(fp, infer_schema_length=0)
            out[p]["declared_same_min"] = sorted(
                {float(x) for x in d["Same Size(Minutes)"].drop_nulls()})
            out[p]["declared_diff_min"] = sorted(
                {float(x) for x in d["Different Size(Minutes)"].drop_nulls()})
            out[p]["n_machines"] = int(d.height)
    return out


def _rim_lookup() -> dict[str, str]:
    """PCR GT codes carry no rim -- it lives in ALL PCR CTP SKUS.xlsx as a tyre
    size ("145 R 12"). Without this the PCR sister-lift row silently vanished."""
    import re as _re
    out: dict[str, str] = {}
    fp = paths.raw("ALL PCR CTP SKUS.xlsx")
    if not fp.exists():
        return out
    import warnings
    warnings.filterwarnings("ignore")
    from openpyxl import load_workbook
    wb = load_workbook(fp, read_only=True)
    rows = list(wb["pcr skus running in ctp"].iter_rows(values_only=True))
    wb.close()
    for r in rows[1:]:
        if not r[0] or not r[1]:
            continue
        m = _re.search(r"R\s*(\d{2})", str(r[1]).upper())
        if m:
            code = str(r[0]).strip()
            out[code] = m.group(1)
            out[f"GT {code}"] = m.group(1)
    return out


def sister_lift(w0: str, w1: str) -> dict:
    """P(consecutive campaigns share a rim) vs chance. Measurement, not grouping."""
    import re
    LOOK = _rim_lookup()
    d = duck().execute(f"""
        WITH s AS (SELECT plant, machineCode res, itemCode gt, event_ts,
          lag(itemCode) OVER (PARTITION BY plant,machineCode ORDER BY event_ts) pg
          FROM v_build WHERE stage=2 AND itemCode IS NOT NULL
            AND machineCode IS NOT NULL
            AND date >= DATE '{w0}' AND date < DATE '{w1}')
        SELECT plant, pg, gt FROM s WHERE pg IS NOT NULL AND pg <> gt""").pl()
    pat = re.compile(r"[R/](\d{2})")

    def rim(x):
        g = str(x).strip()
        v = LOOK.get(g) or LOOK.get(g.replace("GT ", "", 1))
        if v:
            return v
        m = pat.search(g.upper().replace(" ", ""))
        return m.group(1) if m else None

    out = {}
    for p in PLANTS:
        s = d.filter(pl.col("plant") == p)
        if not s.height:
            continue
        pairs = [(rim(x), rim(y)) for x, y in zip(s["pg"], s["gt"])]
        n_raw = len(pairs)
        pairs = [(x, y) for x, y in pairs if x and y]
        if not pairs:
            out[p] = {"error": "no rim resolved for any GT -- lookup missing",
                      "n_adjacencies_raw": n_raw}
            continue
        same = sum(1 for x, y in pairs if x == y) / len(pairs)
        _v, cnt = np.unique([x for x, _ in pairs], return_counts=True)
        pr = cnt / cnt.sum()
        chance = float(pr @ pr)
        se = float(np.sqrt(max(same * (1 - same), 0) / len(pairs)))
        out[p] = {"observed": round(same, 4), "chance": round(chance, 4),
                  "lift": round(same / max(chance, 1e-9), 2),
                  "ci95_lift": [round((same - 1.96 * se) / max(chance, 1e-9), 2),
                                round((same + 1.96 * se) / max(chance, 1e-9), 2)],
                  "n_adjacencies": len(pairs),
                  "rim_resolved_pct": round(100 * len(pairs) / max(n_raw, 1), 1)}
    return out


def visit_law(w0: str, w1: str, mons: list) -> dict:
    """n_visits ~ Q^b per GT, refit per month so the exponent's CV is visible."""
    out = {}
    for p in PLANTS:
        bs, r2s, a_s = [], [], []
        for a, b in mons:
            d = duck().execute(f"""
                WITH s AS (SELECT machineCode m, itemCode g, event_ts,
                  lag(itemCode) OVER (PARTITION BY machineCode ORDER BY event_ts) pg
                  FROM v_build WHERE stage=2 AND plant='{p}'
                    AND itemCode IS NOT NULL AND machineCode IS NOT NULL
                    AND date >= DATE '{a}' AND date < DATE '{b}')
                SELECT g, count(*) FILTER (WHERE pg IS DISTINCT FROM g) n,
                       count(*) q FROM s GROUP BY g""").pl()
            d = d.filter((pl.col("q") > 200) & (pl.col("n") > 0))
            if d.height < 10:
                continue
            ly = np.log(np.array(d["n"], float))
            lx = np.log(np.array(d["q"], float))
            A = np.column_stack([np.ones(len(lx)), lx])
            c, *_ = np.linalg.lstsq(A, ly, rcond=None)
            bs.append(float(c[1])); a_s.append(float(c[0]))
            r2s.append(1 - float(((ly - A @ c) ** 2).sum())
                       / float(((ly - ly.mean()) ** 2).sum()))
        if bs:
            out[p] = {"exponent_b": round(float(np.median(bs)), 3),
                      "intercept_a": round(float(np.median(a_s)), 4),
                      "cv_b": round(float(np.std(bs) / max(np.mean(bs), 1e-9)), 3),
                      "r2_min": round(min(r2s), 3), "r2_max": round(max(r2s), 3),
                      "n_months": len(bs)}
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--as-of", default=str(date.today()))
    ap.add_argument("--months", type=int, default=8)
    a = ap.parse_args()
    set_cutoff(None)
    OUT.mkdir(parents=True, exist_ok=True)

    mons = _months(a.as_of, a.months)
    w0, w1 = mons[0][0], mons[-1][1]
    print("=" * 88)
    print(f"L0  CONTINUOUS LEARNING   window {w0} -> {w1}   ({a.months} months)")
    print("=" * 88)
    print("  PARAMETERS ONLY -- this layer emits no rules, rankings or preferences.\n")

    p: dict = {"as_of": a.as_of, "window_start": w0, "window_end": w1,
               "months": a.months,
               "generated_utc": datetime.utcnow().isoformat(timespec="seconds")}

    p["tau"] = tau(w0, w1, mons)
    print("  tau -- coupling buffer")
    for k, v in p["tau"].items():
        cv = f"{v['cv_across_months']:.3f}" if v["cv_across_months"] else "n/a"
        print(f"    {k}: tau* {v['tau_star_h']:>5.2f} h   min {v['tau_min_h']:.2f}   "
              f"warn(p95) {v['tau_warn_h']:>5.1f}   CV {cv}   n={v['n']:,}")

    p["campaign_gap_cutoff_h"] = CAMP_GAP
    p["campaign_bands"] = campaign_bands(w0, w1, mons)
    print(f"\n  campaign bands (hours)")
    print(f"    build: run breaks on GT change or gap > "
          f"PCR {CAMP_GAP['PCR']} h / TBR {CAMP_GAP['TBR']} h (mould-change median)")
    print(f"    cure : run breaks on GT change or MOULD CHANGE "
          f"(v_curing.mouldNo -- the physical event, not a proxy)")
    for k, v in p["campaign_bands"].items():
        for side, b in v.items():
            print(f"    {k} {side:<6} p10 {b['hours_p10']:>6.1f}  p50 {b['hours_p50']:>6.1f}"
                  f"  MEAN {b['hours_mean']:>6.1f}  p90 {b['hours_p90']:>7.1f}   "
                  f"qty p50 {b['qty_p50']:>5,} mean {b['qty_mean']:>6,.0f}   "
                  f"n={b['n_campaigns']:,}")

    p["yields"] = yields(w0, w1)
    print("\n  yields")
    for k, v in p["yields"].items():
        print(f"    {k}: cure {v['cure_yield']:.5f}   build {v['build_yield']}"
              f"  <- {v['build_yield_note'][:58]}")

    p["press_availability"] = press_availability(w0, w1)
    print("\n  press availability")
    for k, v in p["press_availability"].items():
        print(f"    {k}: MTBF {v['mtbf_h']:>7.1f} h   MTTR {v['mttr_h']:>5.2f} h   "
              f"avail {v['availability']:.4f}   downs {v['n_down_events']:,}")

    p["changeover"] = changeover_durations(w0, w1)
    print("\n  changeover durations")
    for k, v in p["changeover"].items():
        print(f"    {k}: observed p50 {v['observed_p50_min']:>6.1f} min "
              f"p90 {v['observed_p90_min']:>6.1f}   declared same "
              f"{v.get('declared_same_min')} diff {v.get('declared_diff_min')}")

    p["sister_lift"] = sister_lift(w0, w1)
    print("\n  sister lift (rim adjacency)")
    for k, v in p["sister_lift"].items():
        if "error" in v:
            print(f"    {k}: *** {v['error']} *** (n={v['n_adjacencies_raw']:,})")
            continue
        print(f"    {k}: observed {v['observed']:.3f}  chance {v['chance']:.3f}  "
              f"lift {v['lift']:.2f}x  CI {v['ci95_lift']}  "
              f"n={v['n_adjacencies']:,}  rim resolved {v['rim_resolved_pct']}%")

    p["visit_law"] = visit_law(w0, w1, mons)
    print("\n  visit law  n ~ Q^b")
    for k, v in p["visit_law"].items():
        print(f"    {k}: b {v['exponent_b']:.3f}  CV(b) {v['cv_b']:.3f}  "
              f"R2 {v['r2_min']:.2f}-{v['r2_max']:.2f}  over {v['n_months']} months")

    tables = {"cure_cycle": cure_cycle(w0, w1),
              "build_cadence": build_cadence(w0, w1),
              "observed_lot_sizes": observed_lot_sizes(w0, w1)}
    print("\n  per-key tables")
    for n, df in tables.items():
        f = OUT / f"{n}_{a.as_of}.parquet"
        df.write_parquet(f)
        print(f"    {n:<20}{df.height:>7,} rows -> {f.name}")
    p["tables"] = {n: f"{n}_{a.as_of}.parquet" for n in tables}

    jf = OUT / f"params_{a.as_of}.json"
    jf.write_text(json.dumps(p, indent=2))
    print(f"\n  -> {jf}")
    print("\n  NOTE: no rule, ranking or preference was emitted. Policy is decided")
    print("        downstream (L2 capability, L4.5 lot sizing, L9 optimiser).")


if __name__ == "__main__":
    main()
