"""Plant-supplied master files (as opposed to the data-derived stubs).

Currently ingests:
  masters/Master_Building_ChangeoverTime_{pcr,tbr}.csv
      Per building machine: changeover minutes for a SAME-size vs a
      DIFFERENT-size transition, manning, and cost/manday. This is the real
      setup-time model -- "setup time" and "changeover time" are the same thing
      here, and it is *size-dependent*, which the mined flat median could not
      express.
  masters/Master_Mapping_Mould_SKU.csv
      Mould -> material (SKU) -> description. The description carries the tyre
      size, which is what decides same-vs-different above.

Writes warehouse/masters/*.parquet and registers v_changeover_build /
v_mould_sku. Safe to run when the files are absent -- it just warns.
"""
from __future__ import annotations

import re
from pathlib import Path

import polars as pl

from planner.config import CONFIG
from planner.runs.logger import log

OUT = "masters"


def _out_dir() -> Path:
    d = CONFIG.paths.warehouse / OUT
    d.mkdir(parents=True, exist_ok=True)
    return d


def _machine_name(plant: str, raw_name: str, wc_id: int) -> str:
    """Map the master's machine label onto the MES `machineCode`.

    PCR already ships MES names (TBMPCR1Stage2). TBR ships plant-floor names
    (SAV-1 .. SAV-9) which are positional, so index off wcID.
    """
    name = (raw_name or "").strip()
    if name.upper().startswith("TBM"):
        return name
    return f"TBMTBR{wc_id + 1}Stage2" if plant == "TBR" else f"TBMPCR{wc_id + 1}Stage2"


def load_changeover() -> pl.DataFrame:
    rows = []
    for plant, fname in (("PCR", "Master_Building_ChangeoverTime_pcr.csv"),
                         ("TBR", "Master_Building_ChangeoverTime_tbr.csv")):
        p = CONFIG.paths.masters / fname
        if not p.exists():
            log.warning("plant_masters.changeover_missing", path=str(p))
            continue
        df = pl.read_csv(p)
        for r in df.iter_rows(named=True):
            rows.append({
                "plant": plant,
                "machine": _machine_name(plant, r.get("Machine Name", ""), int(r["wcID"])),
                "wc_id": int(r["wcID"]),
                "machine_code": str(r["MachineCode"]),
                "same_size_min": float(r["Same Size(Minutes)"]),
                "diff_size_min": float(r["Different Size(Minutes)"]),
                "man_same": float(r.get("Man Required Same Size") or 0),
                "man_diff": float(r.get("Man Required Different Size") or 0),
                "cost_per_manday": float(r.get("Changeover cost/mandays") or 0),
            })
    return pl.DataFrame(rows) if rows else pl.DataFrame()


_SIZE_RE = re.compile(r"^\s*([0-9][0-9./]*\s*[A-Z]?\s*R?\s*[0-9.]+)", re.I)


def _size_of(desc: str | None) -> str | None:
    """Leading size token: '185/65 R15_VECTRA_88_T_TL' -> '185/65 R15'."""
    if not desc:
        return None
    head = str(desc).split("_")[0].strip()
    m = _SIZE_RE.match(head)
    return (m.group(1) if m else head).replace(" ", "").upper() or None


def load_mould_sku() -> pl.DataFrame:
    p = CONFIG.paths.masters / "Master_Mapping_Mould_SKU.csv"
    if not p.exists():
        log.warning("plant_masters.mould_missing", path=str(p))
        return pl.DataFrame()
    df = pl.read_csv(p, infer_schema_length=0)
    out = df.select([
        pl.col("Mould").str.strip_chars().alias("mould"),
        pl.col("Matl.Code").str.strip_chars().alias("sku"),
        pl.col("Matl.Description").str.strip_chars().alias("description"),
        pl.col("Active Flag").alias("active_flag"),
    ]).filter(pl.col("mould").is_not_null() & pl.col("sku").is_not_null())
    return out.with_columns(
        pl.col("description").map_elements(_size_of, return_dtype=pl.Utf8).alias("size")
    )


CTP_FILE = "CTP Set up building ,curing and inspection (1) 2.xlsx"


def load_ctp_mould_change() -> pl.DataFrame:
    """Per-press mould changeover minutes from the CTP workbook.

    Sheet1 = PCR (210-430 min, by platen dia), Sheet2 = TBR (flat 361 min).
    Keyed by plant *asset* number (3615, 4501, 5401 ...), which is NOT the MES
    `wcID` -- see build_press_xwalk.
    """
    p = CONFIG.paths.masters / CTP_FILE
    if not p.exists():
        log.warning("plant_masters.ctp_missing", path=str(p))
        return pl.DataFrame()
    from openpyxl import load_workbook
    wb = load_workbook(p, read_only=True, data_only=True)
    rows = []
    for plant, sheet, col_idx in (("PCR", "Sheet1", 3), ("TBR", "Sheet2", 1)):
        if sheet not in wb.sheetnames:
            continue
        for r in wb[sheet].iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            asset = str(r[0]).strip()
            try:
                mins = float(r[col_idx])
            except (TypeError, ValueError, IndexError):
                continue
            if asset and mins > 0:
                rows.append({"plant": plant, "asset_id": asset, "mould_change_min": mins})
    wb.close()
    return pl.DataFrame(rows) if rows else pl.DataFrame()


def build_press_xwalk() -> pl.DataFrame:
    """MES `wcID` -> plant asset number, via the numeric prefix of pressbarCode.

    `pressbarCode` is '<asset><LHS|RHS>' (e.g. '4817RHS'). This is press
    *identity* -- fixed plant topology, not behaviour -- so deriving it from the
    whole history carries no forward-looking information.
    """
    from planner.data.warehouse import duck
    try:
        df = duck().execute("""
            WITH x AS (
                SELECT plant, wcID::VARCHAR AS press,
                       regexp_extract(pressbarCode, '^[0-9]+') AS asset_id,
                       count(*) AS n
                FROM v_curing
                WHERE pressbarCode IS NOT NULL
                GROUP BY 1, 2, 3
            ),
            r AS (SELECT *, row_number() OVER (PARTITION BY plant, press ORDER BY n DESC) AS rk
                  FROM x WHERE asset_id IS NOT NULL AND asset_id <> '' AND asset_id <> '0')
            SELECT plant, press, asset_id FROM r WHERE rk = 1
        """).pl()
    except Exception as e:  # noqa: BLE001
        log.warning("plant_masters.xwalk_failed", err=str(e))
        return pl.DataFrame()
    return df


def run() -> dict[str, int]:
    d = _out_dir()
    counts: dict[str, int] = {}

    chg = load_changeover()
    if chg.height:
        chg.write_parquet(d / "changeover_building.parquet", compression="zstd")
        counts["changeover_building"] = chg.height
        log.info("plant_masters.changeover.written", rows=chg.height,
                 machines=chg["machine"].to_list())

    ms = load_mould_sku()
    if ms.height:
        ms.write_parquet(d / "mould_sku.parquet", compression="zstd")
        counts["mould_sku"] = ms.height
        log.info("plant_masters.mould_sku.written", rows=ms.height,
                 moulds=ms["mould"].n_unique(), skus=ms["sku"].n_unique(),
                 sizes=ms["size"].n_unique())

    ctp = load_ctp_mould_change()
    if ctp.height:
        ctp.write_parquet(d / "ctp_mould_change.parquet", compression="zstd")
        counts["ctp_mould_change"] = ctp.height
        log.info("plant_masters.ctp.written", rows=ctp.height,
                 plants=ctp["plant"].unique().to_list())

    xw = build_press_xwalk()
    if xw.height:
        xw.write_parquet(d / "press_xwalk.parquet", compression="zstd")
        counts["press_xwalk"] = xw.height
        if ctp.height:
            j = xw.join(ctp, on=["plant", "asset_id"], how="inner")
            log.info("plant_masters.xwalk.written", rows=xw.height,
                     matched_to_ctp=j.height)

    from planner.data.warehouse import refresh_views
    refresh_views()
    return counts


if __name__ == "__main__":
    print(run())
