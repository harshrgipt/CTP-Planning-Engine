"""Build the JULY-MONTH allowable matrix, and ingest the construction signature
that the existing ingest dropped.

    python scripts/make_allowable_july.py 2026-07

Writes:
    masters/allowable_<month>.parquet   plant, gt_code, sku, machine, source
    warehouse/derived/sku_construction.parquet
        sku, gt_code, and the 15 construction component codes + a signature

WHY THIS EXISTS
  * `tbr_machine_certified.parquet` carries only 6 columns -- plant, gt_code,
    sku_code, machine_no, machine, mes_gt. The 15 CONSTRUCTION columns in
    `TBR BUILDING ALLOWABLE MATRIX.xlsx` (PRE ASSEMBLY, NYLON-1, NYLON2&3, GUM
    STRIP, STEEL CHIPPER L/R, BODYPLY, SHOULDER PAD, APEXED BEAD, BELT-1..4,
    BELT EDGE FILLER, Tread Code) were never ingested. They are what actually
    determines setup cost: measured on July TBR, 7.2 of 15 components change on
    an average GT changeover and 54.3% of changeovers are "major" (>=8).
  * PCR has NO release matrix at all, so PCR eligibility falls back to a mined
    rim lock. The July matrix records PCR eligibility observed in production so
    the two plants can be treated the same way.
"""
from __future__ import annotations

import sys
from pathlib import Path

import polars as pl
from openpyxl import load_workbook
from planner import paths

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT.parent.parent          # C:/Users/91810/Downloads/send
TBR_MATRIX = paths.raw("TBR BUILDING ALLOWABLE MATRIX.xlsx")
COMPONENTS = ["PRE ASSEMBLY", "NYLON-1", "NYLON2&3", "GUM STRIP",
              "STEEL CHIPPER LEFT", "STEEL CHIPPER RIGHT", "BODYPLY",
              "SHOULDER PAD", "APEXED BEAD", "BELT-1", "BELT-2",
              "BELT EDGE FILLER", "BELT-3", "BELT-4", "Tread Code"]


def read_matrix() -> tuple[pl.DataFrame, pl.DataFrame]:
    wb = load_workbook(TBR_MATRIX, read_only=True)
    rows = list(wb["SKU-MACHINE-CONSTRUCTION"].iter_rows(values_only=True))
    wb.close()
    hdr = [("" if v is None else str(v)).strip() for v in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    tbm = [(h, i) for h, i in ix.items() if h.upper().startswith("TBM")]
    elig, cons = [], []
    for r in rows[1:]:
        if not r or r[ix["GT CODE"]] is None:
            continue
        gt = str(r[ix["GT CODE"]]).strip()
        sku = ("" if r[ix["SKU CODE"]] is None else str(r[ix["SKU CODE"]])).strip()
        desc = ("" if r[ix["DESCRIPTION"]] is None
                else str(r[ix["DESCRIPTION"]])).strip()
        for h, i in tbm:
            if str(r[i]).strip().upper() == "YES":
                no = int("".join(ch for ch in h if ch.isdigit()))
                elig.append({"plant": "TBR", "gt_code": gt, "sku": sku,
                             "machine_no": int(no),
                             "machine": f"TBMTBR{no}Stage2",
                             "source": "TBR allowable matrix"})
        sig = tuple(("" if r[ix[c]] is None else str(r[ix[c]]).strip())
                    if c in ix else "" for c in COMPONENTS)
        cons.append({"plant": "TBR", "gt_code": gt, "sku": sku,
                     "description": desc,
                     **{c: v for c, v in zip(COMPONENTS, sig)},
                     "signature": "|".join(sig)})
    return pl.DataFrame(elig), pl.DataFrame(cons).unique(subset=["sku"])


def pcr_observed(month: str) -> pl.DataFrame:
    """PCR has no release matrix -- record what production actually ran, in the
    months BEFORE the plan month, so this stays usable as a forward rule."""
    from planner.data.warehouse import duck, set_cutoff
    set_cutoff(None)
    q = duck().execute(
        "SELECT DISTINCT plant, itemCode AS gt_code, machineCode AS machine "
        "FROM v_build WHERE stage=2 AND plant='PCR' AND itemCode IS NOT NULL "
        "AND machineCode IS NOT NULL AND event_ts < ?::DATE",
        [f"{month}-01"]).pl()
    return q.with_columns(pl.lit("").alias("sku"),
                          pl.lit(0, dtype=pl.Int64).alias("machine_no"),
                          pl.lit("PCR observed pre-month").alias("source"))


def main(month: str) -> None:
    elig, cons = read_matrix()
    dem = pl.read_parquet(ROOT / "masters" / "demand" / f"demand_{month}.parquet")
    gts = dem.select(["plant", "gt_code"]).unique()

    pcr = pcr_observed(month)
    allow = pl.concat([elig.select(["plant", "gt_code", "sku", "machine",
                                    "machine_no", "source"]),
                       pcr.select(["plant", "gt_code", "sku", "machine",
                                   "machine_no", "source"])], how="vertical")
    # TBR matrix keys on its own GT codes; map to MES names via the ingest
    cp = ROOT / "warehouse" / "derived" / "tbr_machine_certified.parquet"
    if cp.exists():
        m = (pl.read_parquet(cp).select(["gt_code", "mes_gt"]).unique()
             .rename({"gt_code": "matrix_gt"}))
        allow = (allow.join(m, left_on="gt_code", right_on="matrix_gt", how="left")
                 .with_columns(pl.coalesce(["mes_gt", "gt_code"]).alias("gt_code"))
                 .drop("mes_gt"))
        cons = (cons.join(m, left_on="gt_code", right_on="matrix_gt", how="left")
                .with_columns(pl.coalesce(["mes_gt", "gt_code"]).alias("gt_code"))
                .drop("mes_gt"))

    jul = allow.join(gts, on=["plant", "gt_code"], how="inner").unique()
    out1 = ROOT / "masters" / f"allowable_{month}.parquet"
    jul.sort(["plant", "gt_code", "machine"]).write_parquet(out1)
    out2 = ROOT / "warehouse" / "derived" / "sku_construction.parquet"
    cons.sort(["plant", "gt_code", "sku"]).write_parquet(out2)

    print(f"  {out1}  {jul.height} (plant, GT, machine) pairs")
    for p in ["PCR", "TBR"]:
        s = jul.filter(pl.col("plant") == p)
        planned = gts.filter(pl.col("plant") == p).height
        cov = s["gt_code"].n_unique()
        per = (s.group_by("gt_code").agg(pl.len().alias("m")))
        print(f"    {p}: {cov}/{planned} GTs covered, "
              f"machines per GT p50 {float(per['m'].median()) if per.height else 0:.0f}"
              f"  max {int(per['m'].max()) if per.height else 0}")
    print(f"  {out2}  {cons.height} SKUs x {len(COMPONENTS)} construction "
          f"components")
    n_sig = cons["signature"].n_unique()
    print(f"    distinct construction signatures: {n_sig} "
          f"({cons.height} SKUs -> {cons.height / max(n_sig, 1):.1f} SKUs per "
          f"signature, i.e. genuinely interchangeable groups)")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "2026-07")
