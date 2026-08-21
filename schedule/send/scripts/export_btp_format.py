"""Export a cmbc run in the plant's BTP optimizer workbook format.

    python -m scripts.export_btp_format --run ctPLANT --month 2026-07 \
        --out output/2026_07_btp

Produces, per plant:
    optimizer_building_schedule_full_<month-start>_<PLANT>.xlsx
    optimizer_curing_schedule_full_<month-start>_<PLANT>.xlsx

matching `btpformat/optimizer_*_2026-08-01 4.xlsx` sheet for sheet, column for
column, in the reference's own order and value conventions (percent-as-string on
the building Demand sheet, percent-as-float on the curing one, title row above
the header where the reference has one).

===============================================================================
PROBLEM 1 -- THE FORMAT IS SKU-LEVEL AND THE ENGINE IS GT-LEVEL
===============================================================================
A green tyre is one construction; several finished SKUs can come off it (same
carcass, different branding or market). The engine plans GTs because that is
what a building machine and a mould see. The BTP format reports SKUs.

THE SPLIT RULE, and why this one:
  share(gt, sku) = tyres of that SKU cured off that GT in the month
                 / all tyres cured off that GT in the month
  from `warehouse/derived/gt_sku_share_<month>.parquet`
  (scripts/build_gt_sku_share.py), which is the CURING RECIPE chain
  `v_curing.recipeID -> Recipemaster.SAPMaterialCode`. That chain is documented
  at 100 % of cured volume on both plants and is the same source the demand file
  itself was built from -- `make_demand.py` runs the identical query and then
  keeps only `r = 1`, the modal SKU. So the demand file is this table with the
  split thrown away; here it is put back.

  July 2026: 48 PCR GTs -> 68 GT-SKU pairs (10 GTs split, max 6 SKUs),
             56 TBR GTs -> 72 GT-SKU pairs (15 GTs split, max 3 SKUs).

  Using the demand file's own `sku` column instead would give a 1:1 identity
  split -- it carries exactly one SKU per GT -- which is not a split at all and
  would silently attribute a multi-SKU GT's whole volume to its largest SKU.

HOW THE SHARE IS APPLIED -- ONE SKU PER ROW, NOT A SPLIT WITHIN A ROW:
  A press holds one mould set and a building machine holds one drum setup, so a
  machine-shift is ONE SKU; the reference pack agrees (one SKUCode per Machine x
  Date x Shift row). The share therefore decides HOW MANY ROWS each SKU gets,
  not how each row is divided. Per (plant, GT): integerise the row quantities
  once with largest-remainder against the GT's rounded total, split that total
  across SKUs by share the same way to get each SKU's target, then walk rows
  largest-first and give each whole row to the SKU with the largest remaining
  target.

  Sum over SKU rows = the GT quantity EXACTLY, because no row is ever divided
  and the integerisation happens once per GT rather than once per row. Splitting
  inside a row instead cost 43.0 % zero-qty PCR rows and drifted TBR by 489
  tyres -- both caught by the `SPLIT RECONCILIATION` print, which must read 0.

===============================================================================
PROBLEM 2 -- THE MACHINE CODES. NOT RESOLVED, AND DELIBERATELY NOT GUESSED.
===============================================================================
The reference pack uses building codes 6001-6004 / 6802 / 7001-7004 / 8501-8502
(38 of them, grouped Stage-1 / Stage-2 / Unistage) and curing codes 14801-148xx,
9503, 9701-9704, 24824, 4406, 4923 (170 of them).

Checked against the plant's OWN work-centre master, `wcmaster 1.xlsx`
(294 rows: iD, processID, name, description, SAPCode):

  * 10 of the 38 BTP BUILDING codes exist in wcmaster -- and every one of them
    is a DIFFERENT KIND OF EQUIPMENT:
        6001-6004 -> "PCR INSPECTION STATION 1-4"   (processID 9)
        7001-7004 -> "INSPECTION STATION 1-4"       (processID 6)
        7201      -> "X ray"                        (processID 17)
        7301      -> "TBR RunOut Machine"           (processID 16)
    In the BTP pack 6001 is a Unistage BUILDING machine. In this plant's master
    6001 is PCR Inspection Station 1. That is not a missing mapping, it is a
    direct contradiction.
  * 0 of the BTP CURING codes appear in wcmaster at all -- not as `name`, not as
    `SAPCode`, not as `iD`. This plant's presses are 3601-3616, 43xx, 45xx, 46xx,
    48xx, 52xx and 54xx.
  * This plant's real building machines are wcmaster processID 7
    (3401-3411, "2nd STAGE TBM - MODULE n") and processID 4 (3801-3803,
    "UNISTAGE MACHINE n"), which the engine knows as TBMPCR1..11Stage2 /
    TBMTBR1..9Stage2 via the measured 34NN/38NN identity.
  * Corroboration: only 12 of the reference pack's 91 SKUs match ours even on a
    16-character prefix, and its month totals 689,563 units against our PCR
    398,405. It is a different plant or a different master generation.

CONCLUSION: the `Machine` column carries OUR identifiers. A guessed crosswalk
would put a building machine's output on an inspection station. `machine_code_map_
UNRESOLVED.csv` in the output folder lists both code spaces side by side so the
plant can supply the real bridge; when it does, only `_MACH` below changes.

===============================================================================
CURING `Qty` IS THE PRESS-OUTPUT BASIS. THIS IS THE ZERO-QTY FIX.
===============================================================================
`2_cure_schedule_shift.qty` in the existing pack is bucketed on `cure_ts`, the
instant a build SLICE is handed to a press. That is deliberate (it reconciles to
the headline by construction) but it is spiky: measured on the July pack,
**5,365 of 7,796 PCR press-shift rows read 0** and 2,425 read ~151, for a mean of
49 and a MEDIAN OF ZERO. Read off the floor that says "the press made nothing
this shift" on two shifts in three.

The physical plan is not spiky. On the press-output basis it is 50.9 tyres per
PCR press-shift and 14.3 on TBR, against a plant actual of 53-54 and 16. So this
sheet uses the press-output basis, integerised with the same largest-remainder
apportionment. A zero here means the press genuinely has no campaign seated.

===============================================================================
PROBLEM 3 -- THIS WORKBOOK PUBLISHED PRESS CAPACITY AS CURED OUTPUT.
===============================================================================
FOUND AND FIXED 2026-08-21. Measured on runs/SHIP2_jul and runs/SHIP2_aug.

Every quantity on the curing side came from `cure_by_shift.parquet`, which is
the campaign NAMEPLATE discretised to shifts -- press capacity seated, not
tyres. The csv pack says so in its own KPI sheet ("F is PRESS CAPACITY SEATED,
not tyres. Never quote it as output.") and this workbook quoted it as output:

  August PCR   `Daily Cured tyres` TOTAL / `Planned_Units`      408,929
               csv sheet 2 `qty` = sheet 6 = sheet 7 `cured`    401,551
               the numerator behind the graded headline (L11)   396,636

  `Fulfillment_Pct` TOTAL then divided that nameplate by the raw order book:

    plant-month   this workbook   the SAME pack's 9b_l11_invariants
    PCR Jul            98.37 %                     96.1 %  FAIL
    TBR Jul            99.30 %                     96.1 %  FAIL
    PCR Aug            95.29 %                     92.7 %  FAIL
    TBR Aug            99.47 %                     96.1 %  FAIL

  TBR shipped as PASS here and FAIL there, from one run, in one pack.

  `GT Gap Diagnostic` `Closing_Balance` was `GT_Built - GT_Cured` with the
  correct clipped build and the nameplate cure and NO opening-stock column, so
  it said PCR ends July 5,640 green tyres NEGATIVE where the plan hands 4,357
  forward -- 9,997 tyres, opposite sign, on the hand-off to next month.

  `Daily GT & Carcass` `GT_Produced` summed `gt_events` positive deltas, which
  INCLUDE the OPENING_STOCK credit: 389,030 against this workbook's own
  `Shift Schedule` Qty of 386,264.

FIX. Fulfilment and every cure quantity now come from
`scripts/export_shift_schedule.plan_quantities()` and
`press_shift_cure()` -- the same two functions the csv pack uses, imported, not
copied. Where two quantities are genuinely different things all of them are
kept as named columns (`Qty` press-run, `Qty_Delivered`, `Nameplate_Seated`),
because deleting the nameplate would destroy the press-starvation signal.
`9c_quantity_bridge` in the csv pack is the ladder between them.
"""
from __future__ import annotations

import argparse
import calendar
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import polars as pl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from planner import paths

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
# THE CANONICAL BASIS IS IMPORTED, NEVER RE-IMPLEMENTED. One implementation
# that can be wrong beats two that can disagree -- which is exactly how this
# workbook came to publish a different fulfilment number from the csv pack
# built out of the same run directory. See PROBLEM 3 above.
from scripts.export_shift_schedule import (            # noqa: E402
    plan_quantities, press_shift_cure, plant_closures, quantity_bridge)

D = ROOT / "warehouse" / "derived"
PLANTS = ("PCR", "TBR")
SHIFTS = ("A", "B", "C")
SHIFT_START_H = {"A": 7, "B": 15, "C": 23}
HDR_FILL = PatternFill("solid", start_color="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF")


# --------------------------------------------------------------------------
def _split_int(total: int, weights: list[float]) -> list[int]:
    """Largest-remainder apportionment. Sums to `total` EXACTLY, all integers.

    A tyre is an integer and a GT's SKU rows must add back to the GT. Rounding
    each share independently does neither."""
    n = len(weights)
    if n == 0:
        return []
    total = int(round(total))
    w = sum(weights)
    # ALREADY WHOLE TYRES SUMMING TO THE TARGET -> RETURN THEM UNTOUCHED.
    # `press_shift_cure` now hands over integer quantities that already sum
    # exactly, and re-apportioning them was not a no-op: `total * x / w` is a
    # float division, so a weight of 50 can come back as 49.999999996, truncate
    # to 49, and be handed its missing tyre to a DIFFERENT row by the remainder
    # loop. That moved up to 5 tyres per plant-day between the BTP workbook and
    # the csv pack -- two files disagreeing about one day, from pure float
    # noise. Measured 2026-08-21: worst day gap 14.9 -> 5.0 -> 0.
    if abs(w - total) < 1e-9 and all(
            abs(float(x) - round(float(x))) < 1e-9 for x in weights):
        return [int(round(float(x))) for x in weights]
    if w <= 0:
        out = [0] * n
        out[0] = total
        return out
    raw = [total * x / w for x in weights]
    out = [int(v) for v in raw]
    rem = total - sum(out)
    order = sorted(range(n), key=lambda i: (-(raw[i] - out[i]), i))
    k = 0
    while rem > 0:
        out[order[k % n]] += 1
        rem -= 1
        k += 1
    while rem < 0:
        i = max(range(n), key=lambda j: out[j])
        if out[i] <= 0:
            break
        out[i] -= 1
        rem += 1
    return out


def _intify(df: pl.DataFrame, cols: list[str]) -> pl.DataFrame:
    """Integerise each named column per (plant, gt_code), largest-remainder.

    A tyre is an integer. Rounding 14,771 float rows independently once lost
    489 tyres on TBR and gained 43 on PCR (see `_assign_sku`), so every
    quantity this workbook prints is integerised ONCE per (plant, GT) against
    that GT's own rounded total -- the same rule for all three cure bases, so
    the three columns stay comparable row by row.
    """
    if df.height == 0:
        return df
    df = df.with_columns(pl.arange(0, pl.len()).alias("_ix"))
    out = []
    for _k, grp in df.group_by(["plant", "gt_code"], maintain_order=True):
        g = grp
        vals = {c: _split_int(int(round(float(g[c].sum()))),
                              [max(float(x), 0.0) for x in g[c].to_list()])
                for c in cols if c in g.columns}
        rows = list(g.iter_rows(named=True))
        for i, r in enumerate(rows):
            out.append({**r, **{c: vals[c][i] for c in vals}})
    return pl.DataFrame(out, infer_schema_length=None).sort("_ix").drop("_ix")


def _gt_rounded_total(df: pl.DataFrame, qty_col: str = "qty") -> int:
    """The target `_assign_sku` is actually held to: sum of PER-GT rounded totals.

    `_assign_sku` integerises ONCE PER (plant, gt) -- that is the whole point of
    it -- so the exact quantity it must reproduce is the sum of each GT's rounded
    total, NOT the rounded sum of every float row. The two differ by at most a
    fraction of a tyre per GT and on August TBR they differ by exactly 1 across
    37 GTs. Comparing against the wrong one printed `!! MISMATCH` for a residue
    that is arithmetically unavoidable and is not lost volume, which is precisely
    the kind of false alarm that trains a reader to ignore a real one.
    """
    if df.height == 0:
        return 0
    return int(sum(round(float(v)) for v in
                   df.group_by(["plant", "gt_code"])
                   .agg(pl.col(qty_col).sum().alias("q"))["q"].to_list()))


def _assign_sku(df: pl.DataFrame, shares: dict, order_cols: list[str],
                qty_col: str = "qty") -> pl.DataFrame:
    """Label every row with ONE SKU, integerise, reconcile exactly per GT.

    TWO DEFECTS THIS REPLACES, both found by the reconciliation print.

    1. SPLITTING EVERY ROW ACROSS SKUs MANUFACTURES ZEROS. A PCR press-shift
       holds ~51 tyres; a GT with 6 SKUs whose smallest share is 2 % gets a 1-tyre
       row and several 0-tyre rows. Measured: 43.0 % of PCR and 35.7 % of TBR
       curing rows came out at Qty = 0 -- reproducing, by a different route, the
       exact "so many 0s" defect this export exists to fix. It is also wrong
       physically: a press holds one mould set and a machine builds one drum
       setup, so a machine-shift IS one SKU. The reference pack agrees -- one
       SKUCode per (Machine, Date, Shift) row.

    2. ROUNDING EACH FLOAT ROW INDEPENDENTLY DRIFTS. `cure_by_shift.qty` is a
       float (50.937...); rounding 14,771 rows separately lost 489 tyres on TBR
       and gained 43 on PCR. Integerisation is therefore done ONCE per (plant,
       GT) with largest-remainder against that GT's rounded total.

    Method: per (plant, gt_code) -- integerise the row quantities to the GT
    total; compute each SKU's target as a largest-remainder split of that total
    by its cured share; then walk rows largest-first and give each whole row to
    the SKU with the largest remaining target. Sum over SKU rows equals the GT
    quantity EXACTLY because no row is ever divided.
    """
    if df.height == 0:
        return df.with_columns(pl.lit("").alias("SKUCode"))
    out = []
    df = df.with_columns(pl.arange(0, pl.len()).alias("_ix"))
    for (plant, gt), grp in df.group_by(["plant", "gt_code"], maintain_order=True):
        g = grp.sort(order_cols)
        tot = int(round(float(g[qty_col].sum())))
        qs = _split_int(tot, [max(float(x), 0.0) for x in g[qty_col].to_list()])
        sk = shares.get((plant, gt)) or [(gt, 1.0)]
        need = dict(zip([c for c, _ in sk],
                        _split_int(tot, [s for _c, s in sk])))
        rows = list(g.iter_rows(named=True))
        for i, q in enumerate(qs):
            rows[i] = {**rows[i], qty_col: q}
        for r in sorted(rows, key=lambda x: -x[qty_col]):
            code = max(need, key=lambda k: (need[k], k)) if need else gt
            need[code] = need.get(code, 0) - r[qty_col]
            out.append({**r, "SKUCode": code})
    # Some GT groups only introduce a string-valued audit column after the
    # first 100 rows.  Polars' bounded schema inference then guesses a numeric
    # type and the August export fails when that later value is appended.
    return pl.DataFrame(out, infer_schema_length=None).sort("_ix").drop("_ix")


def _write(path: Path, sheets: list[tuple[str, str | None, list[str], list[list]]]) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, title, hdr, data in sheets:
        ws = wb.create_sheet(name[:31])
        if title is not None:
            ws.append([title])
            ws.append([])
        ws.append(hdr)
        hrow = ws.max_row
        for c in range(1, len(hdr) + 1):
            ws.cell(hrow, c).font = HDR_FONT
            ws.cell(hrow, c).fill = HDR_FILL
        for row in data:
            ws.append(row)
        ws.freeze_panes = ws.cell(hrow + 1, 1).coordinate
        for i, h in enumerate(hdr, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(
                max(11, len(str(h)) + 3), 26)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _pct(x: float) -> str:
    return f"{100.0 * x:.1f}%"


# --------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run", required=True)
    ap.add_argument("--month", default="2026-07")
    ap.add_argument("--out", required=True)
    a = ap.parse_args()
    run = ROOT / "runs" / a.run
    out = ROOT / a.out if not Path(a.out).is_absolute() else Path(a.out)
    out.mkdir(parents=True, exist_ok=True)
    y, mth = int(a.month[:4]), int(a.month[5:7])
    ndays = calendar.monthrange(y, mth)[1]
    t0 = datetime(y, mth, 1, 7, 0)
    stamp = f"{y}-{mth:02d}-01"

    bs = pl.read_parquet(run / "build_by_shift.parquet")
    # THE CANONICAL BASIS. `cs` used to be `cure_by_shift.parquet` -- the
    # campaign NAMEPLATE discretised to shifts. It is still here, as the
    # `qty_planned` column, but it is no longer what this workbook calls cured.
    CLOSE = plant_closures(run, a.month)
    QQ = plan_quantities(run, a.month)
    cs, _l10res = press_shift_cure(run, a.month, CLOSE)
    cs = cs.rename({"plant_day": "day"})
    CLOSED_H = {p: 24.0 * len(CLOSE[p]) for p in PLANTS}
    ge = pl.read_parquet(run / "gt_events.parquet")
    mc = pl.read_parquet(run / "mould_changes.parquet")
    dem = pl.read_parquet(ROOT / "masters" / "demand" / f"demand_{a.month}.parquet")
    shr_f = D / f"gt_sku_share_{a.month}.parquet"
    shr = pl.read_parquet(shr_f)
    cap_m = pl.read_parquet(D / f"cap_machine_{a.month}.parquet")
    cap_p = pl.read_parquet(D / f"cap_press_{a.month}.parquet")
    cure_ct = pl.read_parquet(D / "plant_ct_cure_sku.parquet")
    cob = pl.read_parquet(D / "cap_changeover.parquet")
    pmc = pl.read_parquet(D / "press_mould_change.parquet")
    msku = pl.read_parquet(ROOT / "warehouse" / "masters" / "mould_sku.parquet")
    gsz = pl.read_parquet(D / "gt_size.parquet").select(
        ["plant", "gt_code", "rim"]).unique(subset=["plant", "gt_code"])
    # SAME OVERRIDE AS L4/L5/L7 -- report the opening stock the PLAN used, which
    # is not always the MES-derived master (a manual plant count or the previous
    # month's carry-forward is pointed at with PLANNER_OPENING_GT).
    _ogov = os.environ.get("PLANNER_OPENING_GT", "").strip()
    ogf = (Path(_ogov) if (_ogov and Path(_ogov).is_absolute())
           else (ROOT / "masters" / "opening_gt" / _ogov) if _ogov
           else ROOT / "masters" / "opening_gt" / f"opening_gt_{a.month}.parquet")
    og = (pl.read_parquet(ogf).group_by(["plant", "gt_code"]).len()
          .rename({"len": "open_qty"}) if ogf.exists()
          else pl.DataFrame(schema={"plant": pl.Utf8, "gt_code": pl.Utf8,
                                    "open_qty": pl.UInt32}))

    # ==================================================================
    # MACHINE NUMBERS -- wcmaster ONLY, and never invented.
    # ==================================================================
    # The curing `press` the engine carries is the MES `wcID`, which IS the
    # work-centre master's `iD` column (that identity is measured 175/175 in
    # scripts/ingest_wcmaster.py; joining on `name` matched 0 of 175). So the
    # plant's own press number is one direct lookup away and there is nothing to
    # guess: `wcmaster 1.xlsx` -> `name`, restricted to the three curing
    # processIDs the same ingest establishes (5 = TBR 65.5" GRM, 8 + 27 = PCR
    # 36"/45"/46"/48"/52" LTM).
    #
    # BUILDING IS NOT MAPPED, DELIBERATELY. wcmaster has no row named
    # `TBMPCR<n>Stage2`; the 34NN/38NN link is a MINED identity, not a lookup,
    # and the instruction is wcmaster-or-recipe or leave it. So building keeps
    # the engine's own machine name and `machine_number_source.csv` says so.
    # PREFER THE INGESTED TABLE, not a second parse of the same workbook.
    # `warehouse/derived/wc_master.parquet` IS wcmaster (scripts/ingest_wcmaster
    # .py) and is what `export_cmbc_xlsx.py` already resolves presses with, so
    # both packs carry one identity rather than two that can drift apart.
    _wcm_f = paths.raw("wcmaster 1.xlsx")
    press_no: dict[str, str] = {}
    _wcm_rows: list[dict] = []
    _wcp = D / "wc_master.parquet"
    if _wcp.exists():
        for r in pl.read_parquet(_wcp).iter_rows(named=True):
            if r.get("press_no"):
                press_no[str(r["wc_id"])] = str(r["press_no"])
    if _wcm_f.exists():
        import openpyxl as _ox0
        _wb0 = _ox0.load_workbook(_wcm_f, read_only=True, data_only=True)
        _rw0 = list(_wb0["sheet1"].iter_rows(values_only=True))
        _wb0.close()
        _h0 = [str(x).strip() for x in _rw0[0]]
        _wcm_rows = [dict(zip(_h0, r)) for r in _rw0[1:] if r and r[0] is not None]
        for _r in _wcm_rows:
            if str(_r.get("processID")).strip() in ("5", "8", "27"):
                press_no.setdefault(str(_r["iD"]).strip(), str(_r["name"]).strip())

    def _mach(p) -> str:
        """Plant press number from wcmaster; the raw wcID when it does not resolve."""
        return press_no.get(str(p), str(p))

    # ---- SKU DESCRIPTION -- three plant sources, first hit wins ---------
    # Coverage is asserted below, not hoped for: a silent blank in a column the
    # floor reads as the product name is worse than no column at all.
    desc_of: dict[str, str] = {}
    _gsm = pl.read_parquet(D / "gt_sku_master.parquet")
    for r in _gsm.iter_rows(named=True):                       # 1. GT-SKU master
        if r.get("description"):
            desc_of.setdefault(str(r["sku_code"]), str(r["description"]).strip())
    for r in shr.iter_rows(named=True):                         # 2. curing recipe chain
        if r.get("sku_desc"):
            desc_of.setdefault(str(r["sku_code"]), str(r["sku_desc"]).strip())
    _b6 = paths.raw("Book6(Sheet5).csv")              # 3. demand file
    if _b6.exists():
        import csv as _csv
        with open(_b6, newline="", encoding="cp1252") as _fh:
            for r in _csv.DictReader(_fh):
                k = (r.get("SKUCode") or "").strip()
                v = (r.get("SKU Description") or "").strip()
                if k and v:
                    desc_of.setdefault(k, v)

    def _desc(sku: str) -> str:
        return desc_of.get(str(sku), "")

    # ---- SKU shares, per GT -------------------------------------------
    shares: dict[tuple, list[tuple[str, float]]] = {}
    for r in shr.iter_rows(named=True):
        shares.setdefault((r["plant"], r["gt_code"]), []).append(
            (r["sku_code"], float(r["share"])))
    for k in shares:
        shares[k].sort(key=lambda x: (-x[1], x[0]))
    # any GT the share table misses falls back to the demand file's own SKU
    for r in dem.select(["plant", "gt_code", "sku"]).unique().iter_rows(named=True):
        shares.setdefault((r["plant"], r["gt_code"]), [(r["sku"], 1.0)])

    rim = {(r["plant"], r["gt_code"]): r["rim"] for r in gsz.iter_rows(named=True)}
    co_same = {r["machine"]: float(r["same_min"]) for r in cob.iter_rows(named=True)}
    co_diff = {r["machine"]: float(r["diff_min"]) for r in cob.iter_rows(named=True)}
    mchg = {(r["plant"], r["press"]): float(r["mould_change_min"])
            for r in pmc.iter_rows(named=True)}
    ct_sku = {(r["plant"], r["sku"]): float(r["cure_min"])
              for r in cure_ct.iter_rows(named=True) if r["cure_min"]}
    ct_gt = {}
    for r in pl.read_parquet(D / "plant_ct_cure_gt.parquet").iter_rows(named=True):
        ct_gt[(r["plant"], r["gt_code"])] = float(r["cure_min"])

    def _day_date(d: int) -> str:
        return (t0.date() + timedelta(days=int(d) - 1)).isoformat()

    def _win(d: int, sh: str) -> tuple[str, str]:
        s = t0 + timedelta(days=int(d) - 1,
                           hours=SHIFT_START_H[sh] - 7)
        return s.strftime("%Y-%m-%d %H:%M"), (s + timedelta(hours=8)).strftime(
            "%Y-%m-%d %H:%M")

    def _spread_shifts(st: datetime, en: datetime) -> list[tuple[int, str, float]]:
        """Cut [st, en) at the 8-h shift grid -> [(day, shift, minutes), ...].

        A mould change costs 210-430 min (PCR) / 361 (TBR) and 81 % of them
        cross a shift boundary, so charging the whole cost to the shift the
        change STARTS in would overstate one shift by up to 480 min and show the
        next one as free. Same defect the L10 docstring records for campaign
        quantities, same fix: spread, do not bucket."""
        out: list[tuple[int, str, float]] = []
        cur = st
        while cur < en:
            h = (cur - t0).total_seconds() / 3600.0
            day = int(h // 24) + 1
            sh = SHIFTS[min(int((h % 24) // 8), 2)] if h >= 0 else "A"
            nxt = t0 + timedelta(hours=(int(h // 8) + 1) * 8)
            seg = min(nxt, en)
            out.append((day, sh, (seg - cur).total_seconds() / 60.0))
            cur = seg
        return out

    print("=" * 92)
    print(f"BTP-FORMAT EXPORT   run={a.run}   month={a.month}")
    print("=" * 92)

    for plant in PLANTS:
        b = bs.filter(pl.col("plant") == plant)
        c = cs.filter(pl.col("plant") == plant)
        dm = (dem.filter(pl.col("plant") == plant)
              .group_by(["gt_code", "sku"]).agg(pl.col("qty").sum().alias("demand")))
        # SKU-level demand via the same share rule
        dsku: dict[str, float] = {}
        gt_dem: dict[str, float] = {}
        for r in dm.iter_rows(named=True):
            gt_dem[r["gt_code"]] = gt_dem.get(r["gt_code"], 0.0) + float(r["demand"])
        for gt, q in gt_dem.items():
            for code, qq in zip([s for s, _ in shares.get((plant, gt), [(gt, 1.0)])],
                                _split_int(int(round(q)),
                                           [s for _c, s in shares.get(
                                               (plant, gt), [(gt, 1.0)])])):
                dsku[code] = dsku.get(code, 0.0) + qq
        maxd = max(dsku.values()) if dsku else 1.0
        prio = {k: (v / maxd if maxd else 0.0) for k, v in dsku.items()}
        # Runner / Non-Runner by Pareto: SKUs covering the top 80 % of volume.
        # The reference pack's own split is 46/45 SKUs at 78 %/22 % of volume,
        # which is what an 80 % Pareto cut produces. No classification file
        # exists for July, so the rule is stated rather than a number copied.
        srt = sorted(dsku.items(), key=lambda x: -x[1])
        tot_d = sum(dsku.values()) or 1.0
        cum, runner = 0.0, set()
        for k, v in srt:
            if cum / tot_d < 0.80:
                runner.add(k)
            cum += v
        cat = {k: ("Runner-In" if k in runner else "Non-Runner-In") for k in dsku}

        elig_m: dict[str, int] = {}
        for r in cap_m.filter(pl.col("plant") == plant).iter_rows(named=True):
            elig_m[r["gt_code"]] = elig_m.get(r["gt_code"], 0) + 1
        elig_p: dict[str, int] = {}
        for r in cap_p.filter(pl.col("plant") == plant).iter_rows(named=True):
            elig_p[r["gt_code"]] = elig_p.get(r["gt_code"], 0) + 1
        gt_of_sku = {s: gt for (pp, gt), lst in shares.items() if pp == plant
                     for s, _ in lst}
        QP = QQ[plant]

        def _sku_split(by_gt: dict, total: float) -> dict:
            """A per-GT quantity spread to SKUs by cured share, integerised
            ONCE per plant against `total`.

            The integerisation is done at PLANT level, not per GT, so the
            resulting column sums to `total` EXACTLY -- which is the whole
            point here: `Fulfillment_Pct` on the TOTAL row must be the number
            L11 grades, to the tyre, not that number plus a rounding residue
            spread over 50 GTs. The cost is that a single GT's SKU rows can be
            off its own value by at most one tyre.
            """
            keys, w = [], []
            for gt, v in by_gt.items():
                for sk, sh in (shares.get((plant, gt)) or [(gt, 1.0)]):
                    keys.append(sk)
                    w.append(max(float(v), 0.0) * float(sh))
            ints = _split_int(int(round(total)), w)
            out_: dict = {}
            for k, i in zip(keys, ints):
                out_[k] = out_.get(k, 0) + i
            return out_

        # THE GRADED FULFILMENT, PER SKU. Numerator and denominator are L11's
        # own per-GT quantities (see plan_quantities), split to SKUs by the
        # cured-share rule this file already uses for everything else.
        sku_num = _sku_split(QP["num_by_gt"], QP["numerator"])
        sku_den = _sku_split(QP["den_by_gt"], QP["denominator"])

        # ================= BUILDING =====================================
        # changeover: a run boundary on a machine. Sort by TIME, never by the
        # display key -- PARTITION §1h, a window function reads the row order.
        bb = b.sort(["machine", "start_ts"]).with_columns([
            pl.col("gt_code").shift(1).over("machine").alias("prev_gt")])
        bb = bb.with_columns([
            (pl.col("gt_code") != pl.col("prev_gt")).fill_null(False).alias("is_co")])
        co_rows, ss_rows = [], []
        for r in bb.iter_rows(named=True):
            mach = r["machine"]
            same = (rim.get((plant, r["gt_code"])) ==
                    rim.get((plant, r["prev_gt"]))) if r["prev_gt"] else True
            co_m = 0.0
            co_t = None
            if r["is_co"]:
                co_m = (co_same if same else co_diff).get(
                    mach, 22.0 if same else 42.0)
                co_t = "same_size_CO" if same else "diff_size_CO"
                co_rows.append([mach, _day_date(r["day"]), int(r["day"]),
                                r["prev_gt"], r["gt_code"], co_t, round(co_m, 1),
                                int(r["day"]), "DONE"])
            ss_rows.append({**r, "CO_Mins": co_m, "CO_Type": co_t})
        # CO_Type is null on most rows and a string only at transitions; infer
        # over the complete list so a late first transition cannot be typed as
        # numeric and abort the export.
        ssb = (pl.DataFrame(ss_rows, infer_schema_length=None)
               if ss_rows else bb)
        ssx = _assign_sku(ssb, shares, ["day", "shift", "machine", "start_ts"])
        # reconcile
        _tgt = _gt_rounded_total(ssb)
        gap = int(float(ssx["qty"].sum()) - _tgt)
        _res = int(_tgt - round(float(ssb["qty"].sum())))
        print(f"\n  {plant}  SPLIT RECONCILIATION  building shift: "
              f"GT {float(ssb['qty'].sum()):,.0f} -> SKU "
              f"{float(ssx['qty'].sum()):,.0f}   diff {gap}"
              f"   {'OK' if gap == 0 else '!! MISMATCH'}"
              + (f"   (+{_res} per-GT rounding residue, not volume)" if _res else ""))

        b_shift = []
        for r in ssx.sort(["day", "shift", "machine", "start_ts"]).iter_rows(named=True):
            st, en = _win(r["day"], r["shift"])
            b_shift.append([
                r["machine"], _day_date(r["day"]), r["shift"], r["SKUCode"],
                # GT_Code: the engine PLANS this, and the SKU on the row is a
                # share-assigned label on top of it. Carrying both makes every
                # row traceable back to the planning key instead of only to the
                # reporting one -- the reference pack has no such column.
                r["gt_code"],
                _desc(r["SKUCode"]),
                int(r["qty"]), round(float(r["CO_Mins"]), 1),
                r["start_ts"].strftime("%Y-%m-%d %H:%M"),
                r["end_ts"].strftime("%Y-%m-%d %H:%M"),
                "Stage-2", r["CO_Type"]])

        # SKU classification
        cls_rows = []
        for cname in ("Non-Runner-In", "Runner-In"):
            ks = [k for k, v in cat.items() if v == cname]
            if not ks:
                continue
            cls_rows.append([cname, len(ks), int(sum(dsku[k] for k in ks)),
                             round(sum(prio[k] for k in ks) / len(ks), 4)])
        cls_rows += [[None, None, None, None],
                     ["Building COs scheduled", len(co_rows), None, None]]

        # Daily GT & Carcass
        gg = ge.filter((pl.col("plant") == plant) & (pl.col("d") > 0))
        gg = gg.with_columns(
            (((pl.col("ts") - pl.lit(t0)).dt.total_seconds() / 86400.0)
             .floor() + 1).cast(pl.Int64).alias("day"))
        gd = gg.group_by("day").agg(pl.col("d").sum().alias("gt"),
                                    pl.col("gt_code").n_unique().alias("nsku"))
        bal = ge.filter(pl.col("plant") == plant).sort("ts").with_columns(
            pl.col("d").cum_sum().alias("bal"))
        bal = bal.with_columns(
            (((pl.col("ts") - pl.lit(t0)).dt.total_seconds() / 86400.0)
             .floor() + 1).cast(pl.Int64).alias("day"))
        eod = {int(r["day"]): float(r["bal"])
               for r in bal.group_by("day").agg(pl.col("bal").last()).iter_rows(named=True)}
        # `GT_Produced` USED TO BE THE gt_events POSITIVE DELTAS, WHICH INCLUDE
        # THE OPENING-STOCK CREDIT. On July PCR that summed to 389,030 against
        # this same workbook's `Shift Schedule` Qty of 386,264 -- two columns,
        # one file, 2,766 tyres apart, with nothing saying why. `GT_Produced` is
        # now the tyres this month BUILDS, which is the Shift Schedule total to
        # the tyre; the opening credit and the full ledger credit are their own
        # columns beside it.
        _bd = {int(r["day"]): r for r in
               ssx.group_by("day").agg(
                   pl.col("qty").sum().alias("q"),
                   pl.col("gt_code").n_unique().alias("n")).iter_rows(named=True)}
        # OPENING STOCK IS READ, NOT INFERRED. Deriving it as
        # (gt_events credits - built) came out 2,766 against a true 3,951: the
        # two frames bucket the month boundary differently, so the residual
        # absorbed a boundary artefact and called it opening stock. It is taken
        # straight from the OPENING_STOCK pseudo-machine rows of
        # `build_schedule.parquet`, which is where `9a`'s opening figure and
        # csv `7_daily_summary.opening_gt_credited` also come from.
        _bsf = pl.read_parquet(run / "build_schedule.parquet").filter(
            (pl.col("plant") == plant) & (pl.col("machine") == "OPENING_STOCK"))
        _opd: dict = {}
        for r in _bsf.iter_rows(named=True):
            _d = int((r["end_ts"] - t0).total_seconds() // 86400) + 1
            _opd[_d] = _opd.get(_d, 0.0) + float(r["qty"])
        daily, cum = [], 0.0
        last_bal = 0.0
        for dd in range(1, ndays + 1):
            bq_ = float(_bd.get(dd, {}).get("q", 0.0))
            ns = int(_bd.get(dd, {}).get("n", 0))
            op_ = _opd.get(dd, 0.0)
            cum += bq_
            last_bal = eod.get(dd, last_bal)
            # Carcass (Stage-1) is not modelled by this engine -- it plans and
            # schedules Stage-2 only, so the column is 0 rather than invented.
            daily.append([_day_date(dd), int(round(bq_)), 0, int(round(bq_)), ns,
                          int(round(cum)), int(round(last_bal)),
                          int(round(op_)), int(round(bq_ + op_)),
                          dd in CLOSE[plant]])

        # Demand fulfilment, building side
        built = {}
        for r in ssx.group_by("SKUCode").agg(pl.col("qty").sum()).iter_rows(named=True):
            built[r["SKUCode"]] = float(r["qty"])
        opn = {r["gt_code"]: float(r["open_qty"])
               for r in og.filter(pl.col("plant") == plant).iter_rows(named=True)}
        bdf = []
        for k, dq in sorted(dsku.items(), key=lambda x: -x[1]):
            gt = gt_of_sku.get(k, k)
            gi = opn.get(gt, 0.0) * (dq / max(gt_dem.get(gt, dq), 1e-9))
            pu = built.get(k, 0.0)
            tot = pu + gi
            gapq = max(dq - tot, 0.0)
            # `Fulfillment_Pct` HERE IS A BUILD-COVERAGE RATIO -- built plus
            # opening stock over the raw order book -- and it is NOT the number
            # the gate grades. It shipped under the same header as the curing
            # workbook's, so two sheets in one pack disagreed by up to 6 pt on
            # the same run. It keeps its meaning under an honest name, and the
            # graded ratio ships beside it, identical to the curing workbook.
            bdf.append([k, gt, cat.get(k), round(prio.get(k, 0.0), 7), int(dq),
                        int(round(gi)), int(round(pu)), int(round(tot)),
                        int(round(gapq)), _pct(min(tot / dq, 1.0) if dq else 0),
                        int(round(float(sku_den.get(k, 0.0)))),
                        int(round(float(sku_num.get(k, 0.0)))),
                        (float(sku_num.get(k, 0.0))
                         / max(float(sku_den.get(k, 0.0)), 1e-9)),
                        "FULLY MET" if gapq <= 0 else "PARTIAL",
                        round(ct_sku.get((plant, k),
                                         ct_gt.get((plant, gt), 0.0)), 1),
                        elig_m.get(gt, 0), round(elig_p.get(gt, 0) / 1.0, 2), None])
        # Footer widths track the header. Adding GT_Code at index 1 without
        # padding these left the count sitting under GT_Code instead of under
        # Category -- a summary row silently reading against the wrong header.
        _NB = 18
        bdf += [[None] * _NB,
                ["Total Building COs", None, len(co_rows)] + [None] * (_NB - 3),
                ["TOTAL", None, None, None, int(sum(dsku.values())), None,
                 int(round(QP["built"])), None, None, None,
                 int(round(QP["denominator"])), int(round(QP["numerator"])),
                 QP["numerator"] / max(QP["denominator"], 1), None, None, None,
                 None,
                 "Planned_Units is BUILT this month. Fulfillment_Pct is the "
                 "graded cure ratio and is identical in the curing workbook "
                 "and in 9b_l11_invariants."]]

        # machine utilisation, building
        # AVAILABLE MINUTES EXCLUDE THE PLANT CLOSURE (rule G3): a shut day is
        # not idle machine capacity, and counting it understated occupancy by
        # 2.7 pt on August.
        avail = ndays * 1440.0 - CLOSED_H[plant] * 60.0
        mu = []
        for mach in sorted(b["machine"].unique().to_list()):
            z = b.filter(pl.col("machine") == mach)
            prod = float((z["end_ts"] - z["start_ts"]).dt.total_seconds().sum()) / 60.0
            com = sum(x[6] for x in co_rows if x[0] == mach)
            idle = max(avail - prod - com, 0.0)
            mu.append([mach, "Stage-2", int(avail), int(round(float(z["qty"].sum()))),
                       0, int(round(prod)), int(round(com)), int(round(idle)),
                       round(prod / avail, 4), round(com / avail, 4),
                       round(idle / avail, 4), round((prod + com) / avail, 4),
                       int(z["gt_code"].n_unique()),
                       sum(1 for x in co_rows if x[0] == mach)])
        _u = [x[8] for x in mu] or [0]
        mu_title = (f"Avg GT-machine util (prod only): {100*sum(_u)/len(_u):.1f}%"
                    f"  |  High(>=80%): {sum(1 for v in _u if v >= .8)}"
                    f"  |  Low(<40%): {sum(1 for v in _u if v < .4)}"
                    f"  |  {plant} Stage-2 machines: {len(mu)}"
                    f"  |  Available_Mins excludes {CLOSED_H[plant]:.0f} h of "
                    f"plant closure (days {CLOSE[plant] or 'none'})")

        _write(out / f"optimizer_building_schedule_full_{stamp}_{plant}.xlsx", [
            ("Shift Schedule", f"BC Building Schedule (Rolling Pipeline) - {plant}",
             ["Machine", "Date", "Shift", "SKUCode", "GT_Code", "Description",
              "Qty", "CO_Mins", "StartTime", "EndTime", "Machine_Group",
              "CO_Type"],
             b_shift),
            ("Changeover Plan", None,
             ["Machine", "Date", "Day", "From_SKU", "Target_SKU", "CO_Type",
              "CO_Cost_Mins", "CO_Day_Index", "Status"], co_rows),
            ("SKU Classification", None,
             ["Category", "SKU_Count", "Total_Demand", "Avg_Priority"], cls_rows),
            ("Daily GT & Carcass",
             f"GT_Produced is tyres BUILT and sums to the Shift Schedule Qty "
             f"({QP['built']:,.0f}). Opening_GT_Credited is last month's stock "
             f"entering the ledger ({QP['opening']:,.0f}), read from the "
             f"OPENING_STOCK rows of build_schedule. GT_Ledger_Credits is their "
             f"sum -- what `gt_events` records, and what GT_Produced used to "
             f"show on its own.",
             ["Date", "GT_Produced", "Carcass_Produced", "Total_Units",
              "Active_SKUs", "Cumulative_GT", "EndDay_GT_Inventory",
              "Opening_GT_Credited", "GT_Ledger_Credits", "Plant_Closed"], daily),
            ("Demand Fulfillment (B2C)",
             f"Build_Coverage_Pct = (Planned_Units + GT_Inventory) / Demand -- "
             f"a BUILD ratio against the raw order book. Fulfillment_Pct is the "
             f"graded cure ratio, {QP['numerator']:,.0f} / "
             f"{QP['denominator']:,.0f} = {QP['pct']:.1f}%, identical to the "
             f"curing workbook and to 9b_l11_invariants.",
             ["SKUCode", "GT_Code", "Category", "Priority", "Demand",
              "GT_Inventory", "Planned_Units", "Planned+GT", "Gap",
              "Build_Coverage_Pct", "Requirement", "Cured_In_Month",
              "Fulfillment_Pct", "Status", "CycleTime_min", "Eligible_Machines",
              "Presses_Needed", "Skip_Reason"], bdf),
            ("Machine Utilization", mu_title,
             ["Machine", "Machine_Group", "Available_Mins", "GT_Built",
              "Carcass_Built", "Prod_Mins", "CO_Mins", "Idle_Mins", "Util_Pct",
              "CO_Pct", "Idle_Pct", "Occupancy_Pct", "SKUs_Served",
              "COs_Done"], mu),
        ])

        # ================= CURING =======================================
        # THREE QUANTITIES, ALL INTEGERISED, ALL KEPT, ALL NAMED. See PROBLEM 3
        # in the module docstring: this workbook used to carry exactly one, the
        # NAMEPLATE, under the name `Qty`.
        #   qty_run      tyres fed in-month, spread over the press hours the
        #                campaign actually holds -- the press-output basis, and
        #                the SKU driver because a press-shift is one SKU.
        #   qty          tyres DELIVERED to the press in that shift (`cure_ts`
        #                basis). Spiky; this is csv sheet 2's `qty` exactly.
        #   qty_planned  press capacity SEATED. Never output.
        c = _intify(c, ["qty_run", "qty", "qty_planned"])
        cx = _assign_sku(c, shares, ["day", "shift", "press"],
                         qty_col="qty_run")
        _tgt = _gt_rounded_total(c, "qty_run")
        gap = int(float(cx["qty_run"].sum()) - _tgt)
        _res = int(_tgt - round(float(c["qty_run"].sum())))
        print(f"  {plant}  SPLIT RECONCILIATION  curing shift : "
              f"GT {float(c['qty_run'].sum()):,.0f} -> SKU "
              f"{float(cx['qty_run'].sum()):,.0f}   diff {gap}"
              f"   {'OK' if gap == 0 else '!! MISMATCH'}"
              + (f"   (+{_res} per-GT rounding residue, not volume)" if _res else ""))

        mcp = mc.filter(pl.col("plant") == plant)
        # ---- CURING CHANGEOVER, PER PRESS-SHIFT ------------------------
        # A curing changeover IS the mould change: the press stops, the mould set
        # comes off, the next one goes on. `mould_changes.parquet` is the plan's
        # own record of every one of them, with the per-press duration from the
        # plant's CTP setup master (Sheet1 210-430 PCR / Sheet2 361 TBR).
        #
        # `CO_Mins` was previously the literal 0.0 -- the column existed and was
        # never filled -- while the mould-change minutes were being written into
        # `Mould_Clean_Mins`, which is the wrong column and, worse, was keyed on
        # the change's STARTING shift so an event straddling 23:00 charged one
        # shift twice its share and the next one nothing.
        #
        # `Mould_Clean_Mins` now carries the MOULD-HANDLING component of that same
        # changeover, from the plant's own decomposition in `CTP Set up building,
        # curing and inspection.xlsx`:
        #     PCR Curing  'Mould change time/set up '   90 min (2-piece, all makes)
        #     TBR Curing  'Change & Set-up time of Moulds.'   133.67 min
        # the remainder being press warm-up (120/180 PCR, 180 + 47 dummy TBR).
        # It is therefore a COMPONENT OF CO_Mins, not an addition to it, and
        # occupancy below charges CO_Mins once.
        MOULD_SETUP_MIN = {"PCR": 90.0, "TBR": 133.67}
        co_by: dict[tuple, float] = {}
        cl_by: dict[tuple, float] = {}
        for r in mcp.iter_rows(named=True):
            tot = max(float(r["minutes"]), 1e-9)
            hand = min(MOULD_SETUP_MIN.get(plant, 0.0), tot)
            for dd, ss, mins in _spread_shifts(r["start_ts"], r["end_ts"]):
                k = (r["press"], dd, ss)
                co_by[k] = co_by.get(k, 0.0) + mins
                cl_by[k] = cl_by.get(k, 0.0) + mins * hand / tot
        # running GT balance at the end of each plant-day, for the sheet's
        # GT_Inventory column
        inv_day = {}
        for r in bal.group_by("day").agg(pl.col("bal").last()).iter_rows(named=True):
            inv_day[int(r["day"])] = float(r["bal"])

        c_shift, zero_rows = [], 0
        co_shift_rows = 0
        for r in cx.sort(["day", "shift", "press"]).iter_rows(named=True):
            st, en = _win(r["day"], r["shift"])
            k = (r["press"], int(r["day"]), r["shift"])
            # a shift is 480 min; a change cannot consume more of it than exists
            com = min(co_by.get(k, 0.0), 480.0)
            clm = min(cl_by.get(k, 0.0), com)
            co_shift_rows += (com > 0)
            q = int(r["qty_run"])
            zero_rows += (q == 0)
            c_shift.append([
                _day_date(r["day"]), r["shift"], _mach(r["press"]),
                str(r["press"]),                    # the MES wcID -- see below
                r["SKUCode"],
                r["gt_code"],                       # the planning key, see b_shift
                _desc(r["SKUCode"]), st, en,
                q, int(r["qty"]), int(r["qty_planned"]),
                round(float(r["hours"]), 2),
                round(com, 1), round(clm, 1),
                round(ct_sku.get((plant, r["SKUCode"]),
                                 ct_gt.get((plant, r["gt_code"]), 0.0)), 1),
                int(round(inv_day.get(int(r["day"]), 0.0))), None])

        # ---- DEMAND FULFILMENT, CURING SIDE -- the graded number -----------
        # `Planned_Units` and `Fulfillment_Pct` used to be the NAMEPLATE over
        # raw order-book demand and read PCR Aug 95.29 % against the same
        # pack's graded 92.7 %. They are now L11's own numerator and
        # denominator, split to SKUs by the cured-share rule, and the TOTAL row
        # is the graded ratio to the tyre. The other three cure bases are kept
        # as their own columns so nothing is lost and nothing can be confused.
        cured, deliv, seated = {}, {}, {}
        for r in cx.group_by("SKUCode").agg(
                pl.col("qty_run").sum().alias("r"),
                pl.col("qty").sum().alias("d"),
                pl.col("qty_planned").sum().alias("n")).iter_rows(named=True):
            cured[r["SKUCode"]] = float(r["r"])
            deliv[r["SKUCode"]] = float(r["d"])
            seated[r["SKUCode"]] = float(r["n"])
        cdf = []
        for k, dq in sorted(dsku.items(), key=lambda x: -x[1]):
            gt = gt_of_sku.get(k, k)
            gi = opn.get(gt, 0.0) * (dq / max(gt_dem.get(gt, dq), 1e-9))
            req_k = float(sku_den.get(k, 0.0))
            pu = float(sku_num.get(k, 0.0))
            gq = max(req_k - pu, 0.0)
            ct = ct_sku.get((plant, k))
            cdf.append([k, gt, round(prio.get(k, 0.0), 4), int(dq),
                        int(round(req_k)), int(round(gi)),
                        int(round(pu)), int(round(gq)),
                        (pu / req_k if req_k else 0.0),
                        int(round(cured.get(k, 0.0))),
                        int(round(deliv.get(k, 0.0))),
                        int(round(seated.get(k, 0.0))),
                        "FULLY MET" if gq <= 0 else "PARTIAL",
                        round(ct if ct else ct_gt.get((plant, gt), 0.0), 1),
                        round(ct, 1) if ct else "NA",
                        elig_p.get(gt, 0),
                        round(elig_p.get(gt, 0) / 1.0, 2),
                        None if ct else "missing curing CT"])
        _NC = 18
        cdf += [[None] * _NC,
                ["TOTAL", None, None, int(sum(dsku.values())),
                 int(round(QP["denominator"])), None,
                 int(round(QP["numerator"])),
                 int(round(max(QP["denominator"] - QP["numerator"], 0))),
                 QP["numerator"] / max(QP["denominator"], 1),
                 int(round(QP["fed_in_month"])),
                 int(round(QP["delivered_in_month"])),
                 int(round(QP["nameplate"])),
                 None, None, None, None, None,
                 "Fulfillment_Pct = Planned_Units / Requirement = the number "
                 "9b_l11_invariants grades. Nameplate_Seated is PRESS CAPACITY, "
                 "never output -- see 9c_quantity_bridge in the csv pack."]]

        # AVAILABLE MINUTES EXCLUDE THE PLANT CLOSURE (rule G3). A closed day
        # is not idle press capacity: August booked 24 h of "idle" on every
        # one of 165 presses on 2026-08-15 while the plant was shut.
        p_avail = (ndays * 1440.0) - CLOSED_H[plant] * 60.0
        pu_rows = []
        for pr in sorted(c["press"].unique().to_list()):
            z = c.filter(pl.col("press") == pr)
            used = float(z["hours"].sum()) * 60.0
            zc = mcp.filter(pl.col("press") == pr)
            com = float(zc["minutes"].sum())
            clean = MOULD_SETUP_MIN.get(plant, 0.0) * zc.height
            idle = max(p_avail - used - com, 0.0)
            pu_rows.append([_mach(pr), str(pr), int(p_avail), int(round(used)),
                            int(round(com)), int(round(clean)),
                            int(round(idle)), round(used / p_avail, 4),
                            round(com / p_avail, 4),
                            round(clean / p_avail, 4), round(idle / p_avail, 4),
                            round((used + com) / p_avail, 4),
                            int(z["gt_code"].n_unique()),
                            int(round(float(z["qty_run"].sum()) / 2.0)),
                            int(round(float(z["qty_run"].sum()))),
                            int(round(float(z["qty"].sum()))),
                            int(round(float(z["qty_planned"].sum()))),
                            None])
        _o = [x[11] for x in pu_rows] or [0]
        pu_title = (f"Avg util (occupancy = used+CO): {100*sum(_o)/len(_o):.1f}%"
                    f"  |  High(>=90%): {sum(1 for v in _o if v >= .9)}"
                    f"  |  Idle(<5%): {sum(1 for v in _o if v < .05)}"
                    f"  |  Presses: {len(pu_rows)}"
                    f"  |  Mould changes: {mcp.height}"
                    f"  |  CO_Mins total: {int(mcp['minutes'].sum()):,}"
                    f"  |  Mould_Clean_Mins is the mould-handling COMPONENT of "
                    f"CO_Mins (CTP setup master: PCR 90 / TBR 133.7 per change; "
                    f"the rest is press warm-up), not an addition -- occupancy "
                    f"charges CO_Mins once"
                    f"  |  Machine = plant press no. from wcmaster 1.xlsx `name`; "
                    f"wc_id is the MES press id the csv pack uses -- crosswalk "
                    f"in press_crosswalk.csv and csv/10_press_crosswalk.csv"
                    f"  |  Available_Mins excludes {CLOSED_H[plant]:.0f} h of "
                    f"plant closure"
                    f"  |  Total_Units is tyres FED IN-MONTH (press-run basis); "
                    f"Units_Delivered and Nameplate_Seated are the other two "
                    f"bases -- see the csv pack's 9c_quantity_bridge")

        cco = [[_day_date(int(r["day"])), int(r["day"]), r["shift"],
                _mach(r["press"]), str(r["press"]),
                r["from_gt"], r["to_gt"],
                "same_size_CO" if r["same_rim"] else "diff_size_CO",
                round(float(r["minutes"]), 1)] for r in mcp.iter_rows(named=True)]

        msch = []
        for r in (cx.group_by(["press", "SKUCode"])
                  .agg(pl.col("qty_run").sum(), pl.col("hours").sum())
                  .sort(["press", "qty_run"], descending=[False, True])
                  .iter_rows(named=True)):
            gt = gt_of_sku.get(r["SKUCode"], r["SKUCode"])
            ct = ct_sku.get((plant, r["SKUCode"]), ct_gt.get((plant, gt), 0.0))
            msch.append([_mach(r["press"]), str(r["press"]), r["SKUCode"],
                         round(prio.get(r["SKUCode"], 0), 4),
                         round(ct, 1), int(round(float(r["qty_run"]) / 2.0)),
                         int(round(float(r["qty_run"]))),
                         int(round(float(r["hours"]) * 60.0)),
                         round(float(r["hours"]) / 24.0, 2)])

        # ---- DAILY CURED, ON ALL THREE BASES -------------------------------
        # This sheet drew a completely different day-1 from the csv pack's
        # `7_daily_summary` -- August PCR 7,943 here against 11,974 there --
        # because it prorated the NAMEPLATE over press hours while the csv
        # sheet buckets DELIVERED tyres on `cure_ts`. Both curves are real and
        # both now ship, in both files, under the same three names.
        dcur = []
        for dd in range(1, ndays + 1):
            z = cx.filter(pl.col("day") == dd)
            dcur.append([_day_date(dd),
                         int(round(float(z["qty_run"].sum()))),
                         int(round(float(z["qty"].sum()))),
                         int(round(float(z["qty_planned"].sum()))),
                         round(float(z["hours"].sum()), 1),
                         dd in CLOSE[plant]])
        dcur += [[None] * 6,
                 ["TOTAL", int(round(float(cx["qty_run"].sum()))),
                  int(round(float(cx["qty"].sum()))),
                  int(round(float(cx["qty_planned"].sum()))),
                  round(float(cx["hours"].sum()), 1), None]]

        # ---- GT GAP DIAGNOSTIC -- the hand-off to next month ---------------
        # WAS `Closing_Balance = GT_Built - GT_Cured` with the correctly
        # clipped build and the NAMEPLATE cure and no opening-stock column, so
        # it reported PCR ending July 5,640 green tyres NEGATIVE where the plan
        # hands 4,357 forward: 9,997 tyres, opposite sign, on the one number
        # next month's plan starts from. The balance is
        #     opening + built in-month - delivered-and-cured in-month
        # which is L7's own `carry_forward_gt` definition and the csv pack's
        # sheet-7 `gt_stock_at_0700_next_day`. All four now agree.
        gtg = []
        for k in sorted(cured, key=lambda x: -cured.get(x, 0)):
            gt = gt_of_sku.get(k, k)
            oq = opn.get(gt, 0.0) * (dsku.get(k, 0.0)
                                     / max(gt_dem.get(gt, dsku.get(k, 1.0)), 1e-9))
            bq = built.get(k, 0.0)
            dq_ = deliv.get(k, 0.0)
            gtg.append([k, gt, int(round(oq)), int(round(bq)), int(round(dq_)),
                        int(round(cured.get(k, 0.0))),
                        int(round(seated.get(k, 0.0))),
                        int(round(oq + bq - dq_)),
                        "DEMAND_MET" if cured.get(k, 0.0) >= dsku.get(k, 0) else "SHORT"])
        gtg += [[None] * 9,
                ["TOTAL", None, int(round(QP["opening"])),
                 int(round(QP["built"])),
                 int(round(QP["delivered_in_month"])),
                 int(round(QP["numerator"])),
                 int(round(QP["nameplate"])),
                 int(round(QP["closing_gt"])),
                 "Closing = Opening + Built - Delivered. Equals "
                 "carry_forward_gt.parquet and csv 7_daily_summary's last "
                 "`gt_stock_at_0700_next_day`."]]

        # ---- the three MOULD sheets --------------------------------
        # The engine plans a GT to a COUNT of moulds (`cap_mould`), not to named
        # physical mould serials -- moulds are modelled as `<mould>@<press>`
        # capacity, one set per press (MEMORY: forcing one primary mould per GT
        # produced 416 K phantom double-book violations). So `Mould` here is the
        # engine's own mould-set identity for that (GT, press) seating, and the
        # ELIGIBLE POOL beside it is the real named pool from
        # `warehouse/masters/mould_sku.parquet`. Nothing is invented: a serial we
        # did not plan is never printed as though we had.
        cmp_r = pl.read_parquet(run / "cure_campaigns_reconciled.parquet").filter(
            pl.col("plant") == plant)
        pool: dict[str, int] = {}
        for r in msku.filter(pl.col("active_flag") == "1").group_by("sku").agg(
                pl.col("mould").n_unique().alias("n")).iter_rows(named=True):
            pool[r["sku"]] = int(r["n"])
        mo_cnt = {r["gt_code"]: int(r["moulds"]) for r in
                  pl.read_parquet(D / f"cap_mould_{a.month}.parquet")
                  .filter(pl.col("plant") == plant).iter_rows(named=True)}
        press_on_sku: dict[str, int] = {}
        for r in cmp_r.group_by("gt_code").agg(
                pl.col("press").n_unique().alias("n")).iter_rows(named=True):
            press_on_sku[r["gt_code"]] = int(r["n"])
        mt = []
        for r in cmp_r.sort(["press", "start_ts"]).iter_rows(named=True):
            gt = r["gt_code"]
            sk = (shares.get((plant, gt)) or [(gt, 1.0)])[0][0]
            day = max(1, int((r["start_ts"] - t0).total_seconds() // 86400) + 1)
            npool = pool.get(sk, mo_cnt.get(gt, 1))
            npress = press_on_sku.get(gt, 1)
            mt.append([_mach(r["press"]), r["mould_set"], sk, day, "Yes",
                       "Yes" if npool < npress else "No", npool, npress,
                       round(npress / max(npool, 1), 2)])
        mt_title = (f"Mould SETS seated: {len(mt)}  |  presses "
                    f"{cmp_r['press'].n_unique()}  |  mould swaps this month: "
                    f"{mcp.height}  |  Mould identity is the engine's mould-SET "
                    f"per (GT, press); named serials come from the eligible pool")
        mm = [[int(r["day"]), _mach(r["press"]),
               (shares.get((plant, r["to_gt"])) or [(r["to_gt"], 1.0)])[0][0],
               f"G::{plant}::{r['to_gt']}", f"G::{plant}::{r['from_gt']}"]
              for r in mcp.sort(["day", "press"]).iter_rows(named=True)]
        mm_title = (f"Every mould change during the month ({mcp.height} swaps). "
                    f"A press keeps its mould set until it changes over to a GT "
                    f"those moulds cannot cure, then swaps.")
        # MouldInUse: presses concurrently running the SKU's GT that day = mould
        # sets in use, against the SKU's eligible pool.
        miu = []
        cxd = cx.group_by(["day", "SKUCode", "gt_code"]).agg(
            pl.col("press").n_unique().alias("inuse")).sort(["day", "SKUCode"])
        for r in cxd.iter_rows(named=True):
            miu.append([_day_date(r["day"]), r["SKUCode"], _desc(r["SKUCode"]),
                        int(r["inuse"]),
                        pool.get(r["SKUCode"], mo_cnt.get(r["gt_code"], 0))])
        miu_title = (f"Daily mould occupancy per SKU (grid: {ndays} days x "
                     f"{cx['SKUCode'].n_unique()} SKUs = {len(miu)} rows)  |  "
                     f"Mould in USE = presses concurrently seated on that SKU's "
                     f"GT that day; Total Eligible Moulds = the SKU's named pool "
                     f"in masters/mould_sku.parquet")

        _write(out / f"optimizer_curing_schedule_full_{stamp}_{plant}.xlsx", [
            ("Demand Fulfillment",
             f"Fulfillment_Pct = Planned_Units / Requirement = "
             f"{QP['numerator']:,.0f} / {QP['denominator']:,.0f} = "
             f"{QP['pct']:.1f}% -- the number 9b_l11_invariants grades. "
             f"Demand is the raw order book and is NOT the denominator. "
             f"Nameplate_Seated is press capacity, never output.",
             ["SKUCode", "GT_Code", "Priority", "Demand", "Requirement",
              "GT_Inventory", "Planned_Units", "Gap", "Fulfillment_Pct",
              "Fed_In_Month", "Delivered_In_Month", "Nameplate_Seated",
              "Status", "CycleTime_min", "CT_available", "Eligible_Machines",
              "Presses_Needed", "Skip_Reason"], cdf),
            ("Machine Utilization", pu_title,
             ["Machine", "wc_id", "Available_Mins", "Used_Mins", "CO_Mins",
              "Mould_Clean_Mins", "Idle_Mins", "Utilization_Pct", "CO_Pct",
              "Mould_Clean_Utilization_%", "Idle_Pct", "Occupancy_Pct",
              "SKUs_Count", "total_cycle", "Total_Units", "Units_Delivered",
              "Nameplate_Seated", "Remaining_Mould_Life"], pu_rows),
            ("Shift Schedule", None,
             ["Date", "Shift", "Machine", "wc_id", "SKUCode", "GT_Code",
              "Description", "StartTime", "EndTime", "Qty", "Qty_Delivered",
              "Nameplate_Seated", "Press_Hours", "CO_Mins", "Mould_Clean_Mins",
              "CycleTime_min", "GT_Inventory", "Remarks"], c_shift),
            ("Changeover Plan", None,
             ["Date", "Day", "Shift", "Press", "wc_id", "From_SKU",
              "Target_SKU", "CO_Type", "Mins"], cco),
            ("Mould Tracker", mt_title,
             ["Press", "Mould", "SKU_Built", "Day_Mounted",
              "Mould_Eligible_For_SKU", "Shared_Mould",
              "SKU_Eligible_Mould_Pool", "Presses_On_SKU",
              "SKU_Tooling_Pressure"], mt),
            ("Mould Movement", mm_title,
             ["Day", "Press", "New_SKU", "Moulds_Mounted", "Moulds_Removed"], mm),
            ("MouldInUse", miu_title,
             ["Date", "SKU Code", "Description", "Mould in USE",
              "Total Eligible Moulds"], miu),
            ("Machine Schedule", f"Press-SKU pairs: {len(msch)}  |  "
             f"Total Units (tyres FED IN-MONTH): "
             f"{int(round(float(cx['qty_run'].sum()))):,}",
             ["Machine", "wc_id", "SKUCode", "Priority", "CycleTime_min",
              "Cycles", "Units_Planned", "Mins_Used", "Days_Used"], msch),
            ("Daily Cured tyres",
             f"Cured_Qty is tyres FED IN-MONTH spread over the press hours the "
             f"campaign holds (sums to {QP['fed_in_month']:,.0f}). "
             f"Delivered_Qty is csv 7_daily_summary `cured` "
             f"({QP['delivered_in_month']:,.0f}). Nameplate_Seated is press "
             f"capacity, NOT output ({QP['nameplate']:,.0f} whole-plan, "
             f"{float(cx['qty_planned'].sum()):,.0f} in-month).",
             ["Date", "Cured_Qty", "Delivered_Qty", "Nameplate_Seated",
              "Press_Hours", "Plant_Closed"], dcur),
            ("GT Gap Diagnostic",
             f"Closing_Balance = Opening + Built - Delivered = "
             f"{QP['opening']:,.0f} + {QP['built']:,.0f} - "
             f"{QP['delivered_in_month']:,.0f} = {QP['closing_gt']:,.0f} green "
             f"tyres handed to next month.",
             ["SKUCode", "GT_Code", "Opening_GT", "GT_Built", "GT_Delivered",
              "GT_Cured_In_Month", "Nameplate_Seated", "Closing_Balance",
              "Reason"], gtg),
        ])
        print(f"  {plant}  curing shift rows {len(c_shift):,}  zero-qty "
              f"{zero_rows:,} ({100*zero_rows/max(len(c_shift),1):.1f}%)   "
              f"building shift rows {len(b_shift):,}")
        # ---- PROVE THIS WORKBOOK AGREES WITH THE CSV PACK AND WITH L11 ----
        # Every one of these came out DIFFERENT before 2026-08-21; see
        # PROBLEM 3. They are asserted here, per plant, at export time.
        _chk = [
            ("Demand Fulfillment TOTAL Planned_Units", cdf[-1][6], QP["numerator"]),
            ("Demand Fulfillment TOTAL Requirement", cdf[-1][4], QP["denominator"]),
            ("Daily Cured tyres TOTAL Cured_Qty", dcur[-1][1], QP["fed_in_month"]),
            ("Daily Cured TOTAL Delivered_Qty", dcur[-1][2],
             QP["delivered_in_month"]),
            ("GT Gap TOTAL Closing_Balance", gtg[-1][7], QP["closing_gt"]),
            ("Daily GT & Carcass sum GT_Produced", sum(r[1] for r in daily),
             sum(int(r["qty"]) for r in ssx.iter_rows(named=True))),
            ("building Shift Schedule sum Qty", sum(r[6] for r in b_shift),
             sum(int(r["qty"]) for r in ssx.iter_rows(named=True))),
        ]
        for _lbl, _got, _want in _chk:
            _ok = abs(float(_got) - float(_want)) <= 1.0
            print(f"  {plant}  AGREEMENT  {_lbl:<40}{float(_got):>10,.0f} vs "
                  f"{float(_want):>10,.0f}   {'OK' if _ok else '!! MISMATCH'}")
        _mine = f"{QP['pct']:.1f}%"
        _l11s = QQ["_l11"].get(plant, "(absent)")
        print(f"  {plant}  AGREEMENT  Fulfillment_Pct {_mine} vs L11 {_l11s}   "
              f"{'OK' if _mine == _l11s else '!! MISMATCH'}")
        # ---- the three things the user asked to be able to trust ----
        _pn_hit = sum(1 for p in c["press"].unique().to_list()
                      if str(p) in press_no)
        _pn_n = c["press"].n_unique()
        print(f"  {plant}  MACHINE NO.  curing {_pn_hit}/{_pn_n} presses resolved "
              f"via wcmaster 1.xlsx iD->name"
              f"{'' if _pn_hit == _pn_n else '  !! rest keep the raw wcID'}"
              f"   ·  building {len(mu)} machines NOT mapped (no wcmaster row for "
              f"TBMxxxStage2) -- engine names kept")
        _cur_sk = set(cx["SKUCode"].to_list()) | set(ssx["SKUCode"].to_list())
        _nd = sorted(s for s in _cur_sk if not _desc(s))
        print(f"  {plant}  DESCRIPTION  {len(_cur_sk)-len(_nd)}/{len(_cur_sk)} SKUs "
              f"({100*(len(_cur_sk)-len(_nd))/max(len(_cur_sk),1):.1f}%)"
              f"{'' if not _nd else '  !! MISSING: ' + ', '.join(_nd[:8])}")
        print(f"  {plant}  CHANGEOVER   {mcp.height} mould changes -> "
              f"{co_shift_rows:,} press-shift rows carry CO_Mins "
              f"({float(mcp['minutes'].sum()):,.0f} min total, "
              f"{100*mcp.filter(pl.col('spans_boundary')).height/max(mcp.height,1):.0f}% "
              f"span a shift boundary and are split across both)")

    # ---- WHERE EVERY MACHINE NUMBER IN THIS PACK COMES FROM ------------
    # One row per (stage, plant). No prose, no inference: the file and the
    # column, or the word `unresolved`.
    src_rows = []
    for _pl in PLANTS:
        _pr = sorted(cs.filter(pl.col("plant") == _pl)["press"].unique().to_list())
        _hit = [p for p in _pr if str(p) in press_no]
        _ex = ", ".join(press_no[str(p)] for p in _hit[:4])
        src_rows.append({
            "stage": "curing", "plant": _pl, "n_machines": len(_pr),
            "n_resolved": len(_hit),
            "source_file": "wcmaster 1.xlsx (via warehouse/derived/wc_master.parquet)"
            if _hit else "",
            "source_column": "name -> press_no  (keyed by iD = MES wcID)"
            if _hit else "",
            "fallback": "" if len(_hit) == len(_pr) else "raw wcID",
            "example": _ex})
    for _pl in PLANTS:
        _mn = sorted(bs.filter(pl.col("plant") == _pl)["machine"].unique().to_list())
        src_rows.append({
            "stage": "building", "plant": _pl, "n_machines": len(_mn),
            "n_resolved": 0, "source_file": "", "source_column": "",
            "fallback": "unresolved - engine machine name kept as-is "
                        "(wcmaster has no TBM<plant><n>Stage2 row; the 34NN/38NN "
                        "link is mined, not a lookup)",
            "example": ", ".join(_mn[:3])})
    pl.DataFrame(src_rows).write_csv(out / "machine_number_source.csv")

    # ---- THE PRESS CROSSWALK, SHIPPED ------------------------------------
    # These workbooks name presses with the plant's own `press_no` (4806); the
    # csv pack names the same press with the MES `wc_id` (120). 0 of 165
    # identifiers overlap, and until now the crosswalk shipped in NEITHER file,
    # so a supervisor holding both could not match one row. It ships in both
    # now: here, and as `csv/10_press_crosswalk.csv`.
    _xw = sorted({str(p) for p in cs["press"].unique().to_list()})
    pl.DataFrame([{
        "wc_id": k, "press_no": press_no.get(k, ""),
        "resolved": bool(press_no.get(k)),
        "used_in_this_workbook_as": "Machine / wc_id columns",
        "used_in_csv_pack_as": "press (2_cure_schedule_shift, 2b, 3, 6)",
        "source": ("warehouse/derived/wc_master.parquet, name -> press_no, "
                   "keyed by iD = MES wcID"
                   if press_no.get(k) else "UNRESOLVED -- no wcmaster row"),
    } for k in _xw]).write_csv(out / "press_crosswalk.csv")

    # ---- THE QUANTITY LADDER, SHIPPED BESIDE THE WORKBOOKS ---------------
    # Same frame as the csv pack's `9c_quantity_bridge`, from the same
    # function, so the two packs cannot describe the plan differently.
    quantity_bridge(QQ).write_csv(out / "quantity_bridge.csv")
    print(f"  -> press_crosswalk.csv  ({sum(1 for k in _xw if press_no.get(k))}"
          f"/{len(_xw)} presses resolved)")
    print("  -> quantity_bridge.csv  (identical to csv/9c_quantity_bridge.csv)")

    # ---- the unresolved BTP machine-code map, shipped as evidence ------
    wm = _wcm_rows
    BTP_B = ["6001", "6002", "6003", "6004", "6802", "7001", "7002", "7003",
             "7004", "7201", "7301", "7501", "7502", "7503", "8301", "8302",
             "8501", "8502"]
    BTP_C = ["14801", "14802", "14803", "9503", "9701", "9704", "9802", "9804",
             "24824", "4406", "4923"]
    rows = []
    for code in BTP_B + BTP_C:
        hit = [x for x in wm if str(x["name"]) == code
               or str(x.get("SAPCode")) == code]
        rows.append({
            "btp_code": code,
            "btp_role": "building" if code in BTP_B else "curing",
            "found_in_wcmaster": bool(hit),
            "wcmaster_wcID": hit[0]["iD"] if hit else "",
            "wcmaster_processID": hit[0]["processID"] if hit else "",
            "wcmaster_description": (hit[0].get("description") or "") if hit else "",
            "verdict": ("CONTRADICTS - different equipment class" if hit
                        else "ABSENT from this plant's work-centre master")})
    pl.DataFrame(rows).write_csv(out / "machine_code_map_UNRESOLVED.csv")
    print(f"\n  -> {out}")
    print(f"  -> machine_code_map_UNRESOLVED.csv  ({len(rows)} BTP codes checked "
          f"against wcmaster 1.xlsx)")
    print("  -> machine_number_source.csv")
    for r in src_rows:
        print(f"       {r['stage']:<9}{r['plant']:<5}"
              f"{r['n_resolved']}/{r['n_machines']}  "
              f"{(r['source_file'] + ' / ' + r['source_column']) if r['source_file'] else r['fallback']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
