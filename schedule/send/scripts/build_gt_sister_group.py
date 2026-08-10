"""Build ``gt_sister_group.parquet`` — per-plant GT construction signatures and
sister groups, mined from the RAW SKU-construction workbooks.

Why this exists
---------------
The plant's "sister SKU" notion is, physically, *shared semi-finished components*.
Two green tyres that differ in only one component slot need one prep change
between them, not a full rebuild. This script extracts the per-GT component
signature from the raw workbooks and groups GTs at Hamming distance <= 1.

Output
------
``INPUT/derived/gt_sister_group.parquet`` with columns:

    plant        str   PCR | TBR
    gt_code      str   the PLANNER's GT key (v_build stage-2 ``itemCode``)
    signature    str   pipe-joined slot values, "NONE" for absent
    sister_id    str   stable group id, "<plant>-<lexically-smallest gt_code>"
    group_size   int   number of GTs in the group
    n_slots      int   number of component slots compared
    slots        str   comma-joined slot names used
    month_seen   str   comma-joined months the GT was active in (e.g. "2026-07,2026-08")

Key correctness rules (each one is a defect that was found and fixed)
--------------------------------------------------------------------
1. ``gt_code`` is the PLANNER's key, not the workbook's internal code. The
   planner joins on ``v_build`` stage-2 ``itemCode``; the workbook's "GT 5001" /
   "GT 1412 XPC MM" are internal and are reached only through the SKU bridge.
2. NEVER match on the 4-digit numeric core of a PCR GT code. Those digits encode
   SIZE, not product: "GT 1513 XPC1 MSIL" (the largest July GT, 55 583 tyres)
   and workbook "GT1513 NEO" are different tread patterns. Numeric-core matching
   inflates apparent coverage from 73% to 90% and is silently wrong. Exact SKU
   keys only.
3. PCR legitimately produces all-singleton groups. There are ZERO PCR GT pairs at
   Hamming distance 1, 2 or 3 — every PCR component code is size-specific, so all
   six slots move together and the signature is a fingerprint, not a similarity
   metric. We EMIT those singletons rather than emitting nothing, so a consumer
   reads "PCR has no sister structure" as data rather than as a missing file.
4. The coverage report is printed on EVERY run. A silent artifact is how this
   project's worst defects survived.

Nulls are handled by the "literal" rule: an absent / "-" / "0" slot becomes the
token ``NONE`` and participates in the comparison. Grouping was verified to be
invariant to the alternative ("skip absent slots") rule on both plants.

Usage
-----
    .venv/Scripts/python.exe scripts/build_gt_sister_group.py
    .venv/Scripts/python.exe scripts/build_gt_sister_group.py --months 2026-07,2026-08
    .venv/Scripts/python.exe scripts/build_gt_sister_group.py --max-dist 2 --out <path>
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import duckdb
import openpyxl
import polars as pl

# --------------------------------------------------------------------------- paths
ROOT = Path(__file__).resolve().parent.parent           # schedule/send
WORKSPACE = ROOT.parent.parent                          # repo wrapper root
WAREHOUSE = ROOT / "warehouse"
XLSX_DIR = ROOT / "Sku construction mapping"
PCR_XLSX = XLSX_DIR / "SKU wise construction mapping PCR.xlsx"
TBR_XLSX = XLSX_DIR / "SKU wise construction mapping TBR.xlsx"
DEFAULT_OUT = WORKSPACE / "INPUT" / "derived" / "gt_sister_group.parquet"

# --------------------------------------------------------------------------- slots
# PCR: raw workbook sheet "PCR", two-tier header on rows 5+6, data from row 7.
# The component block is cols S:AC under the merged super-header "Component Code".
# 0-based column indices into the sheet grid:
PCR_COL_SKU = 1        # B  "Product Code"
PCR_COL_GTZ = 25       # Z  "GT code"        (workbook-internal, provenance only)
PCR_COL_FIRST = 18     # S  "Inner Liner Code"
PCR_SLOT_COLS = {
    18: "inner_liner",     # S  "Inner Liner Code"
    19: "ply1_code",       # T  "PLY1"
    20: "ply2_code",       # U  "Ply 2 code"
    21: "bead_apex",       # V  "Bead Apex code"
    22: "sidewall",        # W  "Sidewall code"
    23: "belt1",           # X  "Belt I code"
}
# DROPPED, with reason:
#   Y  "Belt II code"  - 100% redundant with Belt I (same suffix, 177/177 rows)
#   AA "Carcass code"  - literally 'CAR ' + GT code in 174/177 rows
#   AB "Tread code"    - literally 'TRD ' + GT code in 174/177 rows
#   AC "Cure Code"     - 176 distinct over 177 rows; near-unique, no group signal
PCR_SLOTS = [PCR_SLOT_COLS[i] for i in sorted(PCR_SLOT_COLS)]

# TBR: sheet "Sheet6" is the per-GT template (136 GTs) and is a strict superset of
# sheet "Sheet4" (102 GTs); the two agree on 2985/2985 shared slot comparisons.
TBR_SLOTS = [
    "PRE ASSEMBLY",
    "NYLON2&3",
    "STEEL CHIPPER LEFT",
    "BODYPLY",
    "SHOULDER PAD",
    "APEXED BEAD",
    "BELT-1",
    "BELT-2",
    "BELT-3",
    "BELT-4",
]
# DROPPED, with reason (measured over the July-active subset):
#   "GUM STRIP"           - 1 distinct value; zero signal
#   "BELT EDGE FILLER"    - 1 distinct value; zero signal
#   "NYLON-1"             - 1 distinct value (+12% absent); zero signal
#   "STEEL CHIPPER RIGHT" - perfect mirror of LEFT (R-STCnnn vs L-STCnnn)
#   "Tread Code"          - 135 distinct over 136 GTs; near-unique, no group signal

BLANKS = {None, "", "-", "0", "NA", "N/A", "#N/A"}


# --------------------------------------------------------------------------- helpers
def _cell(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _slot(v) -> str:
    """Normalise a slot value under the 'literal' null rule."""
    s = _cell(v)
    return "NONE" if s in BLANKS else s


def _grid(path: Path, sheet: str) -> list[list]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    g = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    while g and all(v is None for v in g[-1]):
        g.pop()
    return g


# --------------------------------------------------------------------------- workbooks
def load_pcr_signatures() -> dict[str, tuple[dict[str, str], str | None]]:
    """workbook SKU -> ({slot: value}, workbook_gt_code)."""
    if not PCR_XLSX.exists():
        print(f"  !! PCR workbook missing: {PCR_XLSX}", file=sys.stderr)
        return {}
    g = _grid(PCR_XLSX, "PCR")
    out: dict[str, tuple[dict[str, str], str | None]] = {}
    blank_block = 0
    for r in range(6, len(g)):            # data starts on sheet row 7 (0-based 6)
        row = g[r]

        def c(i):
            return row[i] if i < len(row) else None

        sku = _cell(c(PCR_COL_SKU))
        if not sku:
            continue
        if _cell(c(PCR_COL_FIRST)) is None:   # no component block on this row
            blank_block += 1
            continue
        sig = {name: _slot(c(i)) for i, name in PCR_SLOT_COLS.items()}
        out.setdefault(sku, (sig, _cell(c(PCR_COL_GTZ))))
    print(f"  PCR workbook: {len(out)} SKUs with a component block "
          f"({blank_block} product rows have none)")
    return out


def load_tbr_signatures() -> tuple[dict[str, str], dict[str, dict[str, str]]]:
    """(Material -> workbook GT code, workbook GT code -> {slot: value})."""
    if not TBR_XLSX.exists():
        print(f"  !! TBR workbook missing: {TBR_XLSX}", file=sys.stderr)
        return {}, {}
    # Sheet4: Material -> Component (workbook GT code). Header on row 1.
    g4 = _grid(TBR_XLSX, "Sheet4")
    mat2gt = {}
    for row in g4[1:]:
        mat, gt = _cell(row[0]), _cell(row[2]) if len(row) > 2 else None
        if mat and gt:
            mat2gt[mat] = gt
    # Sheet6: workbook GT code -> slots. Header on row 1.
    g6 = _grid(TBR_XLSX, "Sheet6")
    hdr = [_cell(v) for v in g6[0]]
    idx = {}
    for want in TBR_SLOTS:
        if want not in hdr:
            raise SystemExit(f"TBR Sheet6 is missing expected column {want!r}; got {hdr}")
        idx[want] = hdr.index(want)
    gt2sig = {}
    for row in g6[1:]:
        gt = _cell(row[0])
        if not gt:
            continue
        gt2sig[gt] = {s: _slot(row[i] if i < len(row) else None) for s, i in idx.items()}
    print(f"  TBR workbook: {len(mat2gt)} Materials (Sheet4), {len(gt2sig)} GT templates (Sheet6)")
    return mat2gt, gt2sig


# --------------------------------------------------------------------------- warehouse
def _bpath(*parts: str) -> str:
    return str(WAREHOUSE.joinpath(*parts)).replace("\\", "/")


def load_active_gts(con: duckdb.DuckDBPyConnection, months: list[str]) -> dict:
    """(plant, gt_code) -> {'tyres': int, 'months': {month: tyres}}."""
    glob = _bpath("building_output", "plant=*", "stage=2", "date=*", "**", "*.parquet")
    lo = f"{min(months)}-01"
    hi_y, hi_m = divmod(int(max(months)[:4]) * 12 + int(max(months)[5:7]), 12)
    hi = f"{hi_y:04d}-{hi_m + 1:02d}-01" if hi_m < 12 else f"{hi_y + 1:04d}-01-01"
    rows = con.execute(
        f"""
        SELECT plant, itemCode AS gt_code, strftime(date, '%Y-%m') AS month, count(*) AS tyres
        FROM read_parquet('{glob}', hive_partitioning=1)
        WHERE date >= DATE '{lo}' AND date < DATE '{hi}'
        GROUP BY 1, 2, 3
        """
    ).fetchall()
    out: dict = defaultdict(lambda: {"tyres": 0, "months": {}})
    for plant, gt, month, n in rows:
        if month not in months:
            continue
        rec = out[(plant, gt)]
        rec["tyres"] += n
        rec["months"][month] = rec["months"].get(month, 0) + n
    return dict(out)


def load_gt_to_sku(con: duckdb.DuckDBPyConnection) -> dict[str, set[str]]:
    """Planner GT key -> candidate SKU codes. Exact keys only; no fuzzy matching."""
    queries = [
        ("gt_sku_master", "SELECT gt_code, sku_code FROM read_parquet('%s')"
         % _bpath("derived", "gt_sku_master.parquet")),
        ("gt_sku_from_recipe", "SELECT gt_code, sku_code FROM read_parquet('%s')"
         % _bpath("derived", "gt_sku_from_recipe.parquet")),
        ("gt_size", "SELECT gt_code, sku FROM read_parquet('%s')"
         % _bpath("derived", "gt_size.parquet")),
        ("bom_gt_map", "SELECT gt_code, super_parent_sku FROM read_parquet('%s')"
         % _bpath("bom", "bom_gt_map.parquet")),
        ("recipe_bridge", "SELECT gt_code, sku FROM read_parquet('%s')"
         % _bpath("derived", "recipe_bridge.parquet")),
        # recipe_gt_sku.gt_name carries the planner-style TBR key
        # ('10.00 R 20 JUH5'); both its gt_code and sku_code are material codes.
        ("recipe_gt_sku.gt_code", "SELECT gt_name, gt_code FROM read_parquet('%s')"
         % _bpath("derived", "recipe_gt_sku.parquet")),
        ("recipe_gt_sku.sku_code", "SELECT gt_name, sku_code FROM read_parquet('%s')"
         % _bpath("derived", "recipe_gt_sku.parquet")),
    ]
    bridge: dict[str, set[str]] = defaultdict(set)
    for name, q in queries:
        try:
            rows = con.execute(q).fetchall()
        except Exception as exc:                                  # noqa: BLE001
            print(f"  !! bridge source {name} unavailable: {exc}")
            continue
        n = 0
        for gt, sku in rows:
            if gt and sku:
                bridge[str(gt).strip()].add(str(sku).strip())
                n += 1
        print(f"  bridge {name:24s} {n:5d} (gt, sku) pairs")
    return dict(bridge)


# --------------------------------------------------------------------------- grouping
def hamming(a: dict[str, str], b: dict[str, str], slots: list[str]) -> int:
    return sum(1 for s in slots if a[s] != b[s])


def connected_components(keys: list[str], edges: list[tuple[str, str]]) -> dict[str, list[str]]:
    """Union-find; returns {representative: members}. Representative is the
    lexically-smallest member so ids are stable across runs."""
    parent = {k: k for k in keys}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    for a, b in edges:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[max(ra, rb)] = min(ra, rb)     # keep the smaller label as root
    groups: dict[str, list[str]] = defaultdict(list)
    for k in keys:
        groups[find(k)].append(k)
    return {rep: sorted(members) for rep, members in groups.items()}


# --------------------------------------------------------------------------- report
def coverage_report(plant, active, resolved, ambiguous, unresolved, groups, months, slots):
    all_p = {gt: rec for (p, gt), rec in active.items() if p == plant}
    tot_tyres = sum(r["tyres"] for r in all_p.values())
    print()
    print("=" * 78)
    print(f"  {plant}: coverage")
    print("=" * 78)
    print(f"  slots compared ({len(slots)}): {', '.join(slots)}")
    for m in months:
        act = {gt: r for gt, r in all_p.items() if m in r["months"]}
        vol = sum(r["months"][m] for r in act.values())
        cov = {gt for gt in act if gt in resolved}
        cvol = sum(act[gt]["months"][m] for gt in cov)
        if not act:
            print(f"  {m}: no build activity")
            continue
        print(f"  {m}: {len(cov):3d}/{len(act):3d} GTs covered ({len(cov) / len(act):6.1%})   "
              f"volume {cvol:>8,}/{vol:>8,} ({cvol / vol if vol else 0:6.2%})")
    print(f"  union  : {len(resolved):3d}/{len(all_p):3d} GTs covered "
          f"({len(resolved) / len(all_p) if all_p else 0:6.1%})   "
          f"volume {sum(all_p[g]['tyres'] for g in resolved):>8,}/{tot_tyres:>8,} "
          f"({sum(all_p[g]['tyres'] for g in resolved) / tot_tyres if tot_tyres else 0:6.2%})")
    if ambiguous:
        av = sum(all_p[g]["tyres"] for g in ambiguous)
        print(f"  EXCLUDED - ambiguous (bridge gave >1 conflicting signature): "
              f"{len(ambiguous)} GTs, {av / tot_tyres if tot_tyres else 0:.2%} of volume")
        for gt in sorted(ambiguous, key=lambda g: -all_p[g]["tyres"])[:8]:
            print(f"      {gt!r} ({all_p[gt]['tyres']:,} tyres)")
    if unresolved:
        uv = sum(all_p[g]["tyres"] for g in unresolved)
        print(f"  EXCLUDED - no signature (GT/SKU absent from workbook): "
              f"{len(unresolved)} GTs, {uv / tot_tyres if tot_tyres else 0:.2%} of volume")
        for gt in sorted(unresolved, key=lambda g: -all_p[g]["tyres"])[:12]:
            print(f"      {gt!r} ({all_p[gt]['tyres']:,} tyres)")
    sizes = sorted((len(v) for v in groups.values()), reverse=True)
    singles = sum(1 for s in sizes if s == 1)
    print(f"  GROUPS : {len(groups)} groups over {len(resolved)} GTs; "
          f"{singles} singletons ({singles / len(groups) if groups else 0:.0%} of groups)")
    print(f"  sizes  : {sizes}")
    if sizes and sizes[0] == 1:
        print("  NOTE   : every group is a singleton -> this plant has NO sister structure "
              "at the chosen distance. That is a real finding, not a data gap.")
    for rep, members in sorted(groups.items(), key=lambda kv: (-len(kv[1]), kv[0]))[:6]:
        if len(members) > 1:
            print(f"      group({len(members)}) {rep}: {members}")


# --------------------------------------------------------------------------- main
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--months", default="2026-07,2026-08",
                    help="comma-separated YYYY-MM to build the GT universe from")
    ap.add_argument("--max-dist", type=int, default=1,
                    help="Hamming distance at which two GTs are sisters (default 1)")
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    args = ap.parse_args()

    months = [m.strip() for m in args.months.split(",") if m.strip()]
    out_path = Path(args.out)

    print("=" * 78)
    print("  build_gt_sister_group")
    print("=" * 78)
    print(f"  warehouse : {WAREHOUSE}")
    print(f"  workbooks : {XLSX_DIR}")
    print(f"  months    : {', '.join(months)}")
    print(f"  max_dist  : {args.max_dist}")
    print(f"  out       : {out_path}")
    print()

    pcr_sku_sig = load_pcr_signatures()
    tbr_mat2gt, tbr_gt2sig = load_tbr_signatures()

    con = duckdb.connect()
    print()
    bridge = load_gt_to_sku(con)
    active = load_active_gts(con, months)
    con.close()
    if not active:
        print("  !! no build activity found for the requested months - nothing to do")
        return 1

    def pcr_sig(sku):
        hit = pcr_sku_sig.get(sku)
        return hit[0] if hit else None

    def tbr_sig(sku):
        gt = tbr_mat2gt.get(sku)
        return tbr_gt2sig.get(gt) if gt else None

    plant_cfg = {
        "PCR": (pcr_sig, PCR_SLOTS),
        "TBR": (tbr_sig, TBR_SLOTS),
    }

    records: list[dict] = []
    for plant, (lookup, slots) in plant_cfg.items():
        plant_gts = {gt: rec for (p, gt), rec in active.items() if p == plant}
        if not plant_gts:
            continue
        resolved: dict[str, dict[str, str]] = {}
        ambiguous: list[str] = []
        unresolved: list[str] = []
        for gt in plant_gts:
            cands = {}
            for sku in bridge.get(gt, set()):
                sig = lookup(sku)
                if sig:
                    cands["|".join(sig[s] for s in slots)] = sig
            if len(cands) == 1:
                resolved[gt] = next(iter(cands.values()))
            elif len(cands) > 1:
                ambiguous.append(gt)
            else:
                unresolved.append(gt)

        keys = sorted(resolved)
        edges = [
            (a, b)
            for i, a in enumerate(keys)
            for b in keys[i + 1:]
            if hamming(resolved[a], resolved[b], slots) <= args.max_dist
        ]
        groups = connected_components(keys, edges)
        coverage_report(plant, active, resolved, ambiguous, unresolved, groups, months, slots)

        slot_str = ",".join(slots)
        for rep, members in groups.items():
            for gt in members:
                records.append({
                    "plant": plant,
                    "gt_code": gt,
                    "signature": "|".join(resolved[gt][s] for s in slots),
                    "sister_id": f"{plant}-{rep}",
                    "group_size": len(members),
                    "n_slots": len(slots),
                    "slots": slot_str,
                    "month_seen": ",".join(m for m in months if m in plant_gts[gt]["months"]),
                })

    if not records:
        print("\n  !! no GT resolved to a signature on any plant - refusing to write an empty artifact")
        return 1

    df = (
        pl.DataFrame(records)
        .with_columns(
            pl.col("group_size").cast(pl.Int32),
            pl.col("n_slots").cast(pl.Int32),
        )
        .sort(["plant", "gt_code"])          # deterministic byte-for-byte output
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.write_parquet(out_path, compression="zstd")

    print()
    print("=" * 78)
    print(f"  WROTE {out_path}  ({df.height} rows)")
    print("=" * 78)
    for plant, n, ng in (
        df.group_by("plant")
          .agg(pl.len().alias("n"), pl.col("sister_id").n_unique().alias("ng"))
          .sort("plant")
          .iter_rows()
    ):
        print(f"    {plant}: {n} GTs in {ng} sister groups")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
