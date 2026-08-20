"""DAY-1 PRESS COVERAGE CONTROL TABLE.  Which presses can start at 07:00, and why not.

    PYTHONPATH=. python scripts/day1_coverage_table.py 2026-08 aug_v13

Writes DAY1_COVERAGE_<month>.xlsx: one row per press seated on day 1, with the
GT it is seated for, the opening stock behind it, the cover that GT needs per
press, when the press actually starts, and the reason if it is not at 07:00.

WHAT THIS IS FOR
  It converts "day 1 is low" into a per-press deficit the plant can act on. The
  optimisation objective it supports is MAXIMISE COVERED PRESSES, not maximise
  total GT inventory -- those are different targets and the second one is what
  produces 46 % unused stock.

WHAT IT IS NOT
  It does NOT invent opening stock. Raising the opening figure to a "full
  coverage target" would hand the plan tyres that are not on the floor at 07:00
  -- the same objection as building before the month opens, which the plant has
  already ruled out. The Deficit column states what is missing; filling it is a
  plant decision about real inventory, not a planner input.

WHAT IS ALREADY IMPLEMENTED, AND IS NOT A GAP
  Press-level partial allocation and progressive starts already work. The early
  budget is bounded BY QUANTITY, so a GT starts as many presses at t0 as its
  stock covers and the rest wait:
      GT 1513 XPC1 MSIL  490 stock -> 6 presses at 07:00, remainder at +11.86 h
      GT 1402 XPC TATA   131 stock -> 1 press   at 07:00, remainder at +11.86 h
  There is no all-or-nothing gate to remove. The residual headroom is smaller
  and different: L5 does not always SEAT as many campaigns at t0 as the stock
  would cover. Measured August PCR: 33 presses start at t0, stock covers 46,
  and after R3 mould limits and per-GT press counts the real headroom is
  11 presses = 130 press-h = ~832 tyres.
"""
from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

import openpyxl
import polars as pl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from planner import paths                                          # noqa: E402

HDR = PatternFill("solid", fgColor="1F3864")
OK = PatternFill("solid", fgColor="C6EFCE")
BAD = PatternFill("solid", fgColor="FFC7CE")
MID = PatternFill("solid", fgColor="FFEB9C")


def main() -> None:
    month = sys.argv[1] if len(sys.argv) > 1 else "2026-08"
    run = Path("runs") / (sys.argv[2] if len(sys.argv) > 2 else "aug_v13")
    y, m = int(month[:4]), int(month[5:7])
    t0 = dt.datetime(y, m, 1, 7, 0)

    cc = pl.read_parquet(run / "cure_campaigns.parquet").with_columns(
        ((pl.col("start_ts") - pl.lit(t0)).dt.total_seconds() / 3600).round(2).alias("h"))
    mo = pl.read_parquet(paths.wh_derived(f"cap_mould_{month}.parquet"))
    mx = {(r["plant"], r["gt_code"]): r["max_concurrent_presses"]
          for r in mo.iter_rows(named=True)}

    ogn = (paths.MASTERS / "opening_gt" / f"opening_gt_manual_{month}.parquet")
    if not ogn.exists():
        ogn = paths.MASTERS / "opening_gt" / f"opening_gt_{month}.parquet"
    og = pl.read_parquet(ogn).filter(pl.col("age_h") <= 72.0)
    bud = {(r["plant"], r["gt_code"]): float(r["len"])
           for r in og.group_by(["plant", "gt_code"]).len().iter_rows(named=True)}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cols = ["Press", "GT seated at day 1", "Opening GT (this GT)",
            "Cover needed / press", "Presses this GT can cover",
            "R3 max concurrent", "Presses seated at 07:00", "This press starts",
            "Covered until", "Status", "Deficit (tyres) to start at 07:00"]

    for plant, cover, rate in (("PCR", 76.0, 6.38), ("TBR", 31.0, 1.79)):
        ws = wb.create_sheet(plant)
        ws.append(cols)
        for c in ws[1]:
            c.fill, c.font = HDR, Font(bold=True, color="FFFFFF", size=9)
            c.alignment = Alignment(wrap_text=True, horizontal="center",
                                    vertical="center")

        x = cc.filter(pl.col("plant") == plant)
        first = (x.group_by("press").agg(pl.col("h").min().alias("h"),
                                         pl.col("gt_code").first())
                  .sort("h"))
        seated = {}
        for r in x.filter(pl.col("h") < 0.5).iter_rows(named=True):
            seated[r["gt_code"]] = seated.get(r["gt_code"], 0) + 1

        n_cov = n_wait = 0
        for r in first.sort(["h", "press"]).iter_rows(named=True):
            g = r["gt_code"]
            s = bud.get((plant, g), 0.0)
            can = int(s // cover)
            at0 = seated.get(g, 0)
            covered_until = (t0 + dt.timedelta(hours=s / max(at0, 1) / rate)
                             if at0 else t0)
            if r["h"] < 0.5:
                st, dfc = "COVERED", 0
                n_cov += 1
            else:
                n_wait += 1
                if can > at0:
                    st, dfc = "SEATING GAP (stock exists)", 0
                else:
                    st = "NO STOCK COVER"
                    dfc = max(0, round(cover - (s - at0 * cover)))
            ws.append([r["press"], g, round(s), round(cover), can,
                       mx.get((plant, g), None), at0, round(r["h"], 2),
                       covered_until.strftime("%d %b %H:%M") if r["h"] < 0.5 else "",
                       st, dfc or None])

        for row in ws.iter_rows(min_row=2):
            c = row[9]
            c.fill = OK if c.value == "COVERED" else (
                MID if str(c.value).startswith("SEATING") else BAD)
            c.alignment = Alignment(horizontal="center")
        for i, w in enumerate([9, 30, 15, 14, 14, 12, 14, 13, 15, 24, 16], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        ws.append([])
        ws.append([f"{plant}: {n_cov} presses covered at 07:00 · {n_wait} waiting · "
                   f"opening GT {sum(v for (p, _), v in bud.items() if p == plant):,.0f} "
                   f"tyres · waiting presses cost "
                   f"{n_wait * 11.86 * rate:,.0f} tyres of day-1 output"])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        print(f"  {plant}: covered {n_cov} · waiting {n_wait}")

    out = paths.ROOT.parent.parent / f"DAY1_COVERAGE_{month}.xlsx"
    wb.save(out)
    print(f"  -> {out}")


if __name__ == "__main__":
    main()
