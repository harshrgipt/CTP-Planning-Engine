"""Mine the plant's OWN threshold bands from referance/ctp_*_schedule_*.xlsx.

    PYTHONPATH=. python scripts/mine_reference_thresholds.py

These workbooks are the plant's published CTP schedules for June/July/Aug 2026,
and their `Exceptions` sheet states, per exception type, the THRESHOLD the plant
holds itself to and whether it is Hard or Soft. We have been deriving those
bands from MES; the plant states them outright.

`export_ctp.py` mirrors the sheet LAYOUT of these files but nothing has ever
read their CONTENT. READ-ONLY -- writes nothing.
"""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

REF = Path(__file__).resolve().parent.parent.parent.parent / "referance"

# sheets that carry a reference band or a plant-set target rather than raw rows
INTEREST = ["Daily GT & Inventory", "Small Build Batches", "Small Cure Campaigns",
            "Press Tube-Type Lock", "Size Change Cadence", "Machine Utilization",
            "Build-Cure Linkage", "Feed NGT Summary"]


def exceptions(path: Path) -> list[tuple]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Exceptions" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Exceptions"]
    rows, hdr = [], None
    for r in ws.iter_rows(values_only=True):
        vals = [("" if c is None else str(c).strip()) for c in r]
        if not any(vals):
            continue
        if hdr is None:
            if "Exception_Type" in vals:
                hdr = {v: i for i, v in enumerate(vals)}
            continue
        get = lambda k: vals[hdr[k]] if k in hdr and hdr[k] < len(vals) else ""
        if get("Exception_Type"):
            rows.append((get("Exception_Type"), get("Threshold"), get("Hard/Soft")))
    wb.close()
    return rows


def sheet_head(path: Path, name: str, n: int = 4) -> list:
    wb = load_workbook(path, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[name]
    out = []
    for i, r in enumerate(ws.iter_rows(max_row=n, values_only=True)):
        out.append([("" if c is None else str(c)[:20]) for c in r[:9]])
    wb.close()
    return out


def main() -> None:
    files = sorted(REF.glob("ctp_*_schedule_*.xlsx"))
    if not files:
        raise SystemExit(f"no reference workbooks under {REF}")

    print(f"PLANT-STATED THRESHOLDS  ({len(files)} reference workbooks)\n")
    agg: dict = defaultdict(lambda: defaultdict(set))
    for f in files:
        kind = "curing" if "curing" in f.name else "building"
        plant = "PCR" if f.name.endswith("_pcr.xlsx") else "TBR"
        for etype, thr, hs in exceptions(f):
            if thr:
                agg[(kind, plant)][etype].add((thr, hs))

    for (kind, plant) in sorted(agg):
        print("=" * 96)
        print(f"{kind.upper()}  ·  {plant}")
        print("=" * 96)
        for etype in sorted(agg[(kind, plant)]):
            for thr, hs in sorted(agg[(kind, plant)][etype]):
                print(f"  {etype[:44]:<46}{thr[:26]:<28}{hs}")
        print()

    print("=" * 96)
    print("REFERENCE SHEETS PRESENT (structure only)")
    print("=" * 96)
    for f in files[:2] + files[-2:]:
        wb = load_workbook(f, read_only=True)
        print(f"\n{f.name}\n  {wb.sheetnames}")
        wb.close()
    for name in INTEREST:
        for f in files:
            head = sheet_head(f, name)
            if head:
                print(f"\n--- {name}   ({f.name}) ---")
                for row in head:
                    print("   ", row)
                break


if __name__ == "__main__":
    main()
