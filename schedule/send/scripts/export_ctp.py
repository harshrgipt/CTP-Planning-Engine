"""Export a planned month as CTP-format workbooks (one per plant, per stage).

Mirrors the sheet/column layout of `referance/ctp_*_schedule_<month><year>_<plant>.xlsx`
so our output can be read side by side with the reference implementation.

    python -m scripts.export_ctp 2026-07 [run_dir] [out_dir]

Produces four files:
    ctp_building_schedule_<month>_<plant>.xlsx   PCR, TBR
    ctp_curing_schedule_<month>_<plant>.xlsx     PCR, TBR

Day boundary is 07:00, matching the reference: shift A 07:00-15:00,
B 15:00-23:00, C 23:00-07:00, so "day 1" runs 01-07 07:00 to 02-07 07:00.
"""
from __future__ import annotations

import calendar
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import polars as pl

from planner.config import CONFIG
from planner.data.warehouse import duck, set_cutoff
from planner.runs.logger import log

DAY_START_H = 7          # plant day boundary, per the reference workbooks
SHIFT_H = 8
SHIFTS = ["A", "B", "C"]


def _desc_map() -> dict[str, str]:
    """GT code -> a representative finished SKU, if the BOM has one.

    The column is `super_parent_sku`; querying a non-existent `sku` threw, the
    except swallowed it, and every description came out blank.
    """
    # From the CURING RECIPE (100% of produced volume, both plants), not the
    # BOM (55% of GTs, ~0% of TBR). This is why TBR descriptions were blank.
    try:
        import polars as _pl
        _p = CONFIG.paths.warehouse / "derived" / "gt_sku_from_recipe.parquet"
        if _p.exists():
            _d = _pl.read_parquet(_p)
            out = {r["gt_code"]: f'{r["sku_code"]} - {r["sku_desc"]}'.strip(" -")
                   for r in _d.iter_rows(named=True) if r["gt_code"]}
            log.info("export_ctp.desc_map", gts=len(out), src="curing_recipe")
            return out
        rows = duck().execute("""
            SELECT gt_code, any_value(super_parent_sku), count(DISTINCT super_parent_sku)
            FROM v_bom_gt WHERE gt_code IS NOT NULL GROUP BY 1
        """).fetchall()
        out = {}
        for gt, sku, n in rows:
            if gt and sku:
                out[gt] = f"{sku}" + (f" (+{n - 1} more FG)" if n and n > 1 else "")
        # KNOWN GAP, not a bug: TBR resolves to ~0. The BOM keys TBR on
        # "GT 5001"-style codes while TBR MES `itemCode` is size-led
        # ("10.00 R 20 JDC3") -- the two do not overlap (MEMORY s3 join trap 2).
        # Leave those blank rather than inventing a mapping.
        log.info("export_ctp.desc_map", gts=len(out))
        return out
    except Exception as e:  # noqa: BLE001
        log.warning("export_ctp.desc_map_failed", err=str(e))
        return {}


def _split_shifts(start: datetime, end: datetime, qty: float,
                  month_start: date) -> list[tuple[int, str, float, float, datetime, datetime]]:
    """Cut a build segment at shift boundaries.

    A lot running 2,023 minutes cannot be one shift. The reference reports such
    a run as Shifts="A,B,C", Segments=3; emitting Segments=1 on a single shift
    letter claims a 33-hour shift. Quantity is pro-rated by time in shift.
    """
    origin = datetime.combine(month_start, datetime.min.time()) + timedelta(hours=DAY_START_H)
    total_s = max(1.0, (end - start).total_seconds())
    out, cur = [], start
    while cur < end:
        off = (cur - origin).total_seconds() / 3600.0
        if off < 0:
            # The planner's horizon opens at 00:00 but the plant day opens at
            # 07:00, so the first 7 hours fall in the PREVIOUS plant day. Cut the
            # piece at the origin rather than letting it run to the end of
            # shift A -- otherwise a lot starting at 00:00 is reported as a
            # single 900-minute "shift A".
            day, idx, bnd = 1, 0, origin
        else:
            day = int(off // 24) + 1
            idx = min(2, int((off % 24) // SHIFT_H))
            bnd = origin + timedelta(days=day - 1, hours=(idx + 1) * SHIFT_H)
        seg_end = min(end, bnd)
        secs = (seg_end - cur).total_seconds()
        if secs <= 0:
            break
        out.append((day, SHIFTS[idx], qty * secs / total_s, secs / 60.0, cur, seg_end))
        cur = seg_end
    return out


def _daypart(ts: datetime, month_start: date) -> tuple[int, str]:
    """(plant day number starting at 1, shift letter) for a timestamp."""
    origin = datetime.combine(month_start, datetime.min.time()) + timedelta(hours=DAY_START_H)
    off = (ts - origin).total_seconds() / 3600.0
    if off < 0:
        return 1, "A"
    day = int(off // 24) + 1
    shift = SHIFTS[min(2, int((off % 24) // SHIFT_H))]
    return day, shift


def _shift_bounds(month_start: date, day: int, shift: str) -> tuple[datetime, datetime]:
    origin = datetime.combine(month_start, datetime.min.time()) + timedelta(hours=DAY_START_H)
    s = origin + timedelta(days=day - 1, hours=SHIFTS.index(shift) * SHIFT_H)
    return s, s + timedelta(hours=SHIFT_H)


def _write(path: Path, sheets: list[tuple[str, list[str], list[list]]]) -> None:
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, header, rows in sheets:
        ws = wb.create_sheet(name[:31])
        ws.append(header)
        for r in rows:
            ws.append(r)
        for i, h in enumerate(header, start=1):
            width = max(10, min(30, len(str(h)) + 4))
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = width
        ws.freeze_panes = "A2"
    wb.save(path)


def export(month: str, run_dir: Path, out_dir: Path) -> list[Path]:
    y, m = int(month[:4]), int(month[5:7])
    month_start = date(y, m, 1)
    ndays = calendar.monthrange(y, m)[1]
    out_dir.mkdir(parents=True, exist_ok=True)

    b = pl.read_parquet(run_dir / "build_schedule.parquet")
    c = pl.read_parquet(run_dir / "cure_schedule.parquet")
    ev = pl.read_parquet(run_dir / "gt_events.parquet")
    desc = _desc_map()
    mon = f"{calendar.month_name[m].lower()}{y}"
    written: list[Path] = []

    for plant in ("PCR", "TBR"):
        bp = b.filter(pl.col("plant") == plant)
        cp = c.filter(pl.col("plant") == plant)
        if bp.height == 0 and cp.height == 0:
            continue

        # ---------------- BUILDING ----------------------------------------
        seg, per_shift, co_rows = [], [], []
        last_gt: dict[str, str] = {}
        for r in bp.sort(["machine", "start_ts"]).iter_rows(named=True):
            d0, s0 = _daypart(r["start_ts"], month_start)
            mins = (r["end_ts"] - r["start_ts"]).total_seconds() / 60.0
            gt, mach = r["gt_code"], r["machine"]
            parts = _split_shifts(r["start_ts"], r["end_ts"], float(r["qty"]), month_start)
            shifts = ",".join(dict.fromkeys(p[1] for p in parts)) or s0
            seg.append([mach, d0, gt, int(round(r["qty"])), round(mins, 1),
                        shifts, len(parts) or 1, str(r["start_ts"]), str(r["end_ts"]),
                        gt, desc.get(gt, "")])
            for pd_, ps_, pq_, pm_, pst_, pen_ in parts:
                per_shift.append([mach, pd_, ps_, gt, int(round(pq_)),
                                  round(pm_, 1), ps_, str(pst_), str(pen_),
                                  gt, desc.get(gt, "")])
            prev = last_gt.get(mach)
            if prev is not None and prev != gt and r["setup_s"]:
                co_rows.append([mach, d0, s0, prev, gt,
                                "size_CO" if desc.get(prev, "")[:3] != desc.get(gt, "")[:3]
                                else "same_size_CO",
                                round(r["setup_s"] / 60.0, 1), str(r["start_ts"]),
                                prev, gt])
            last_gt[mach] = gt

        setup_rows = {}
        for r in co_rows:
            k = (r[1], r[0])
            setup_rows[k] = setup_rows.get(k, 0.0) + r[6]
        setup_tbl = [[d_, mc, round(v, 1), round(v, 1), round(100 * v / 1440, 2)]
                     for (d_, mc), v in sorted(setup_rows.items())]

        # NB: Polars group_by yields the key as a TUPLE, not a scalar.
        variety = [[mc[0] if isinstance(mc, tuple) else mc,
                    int(g["gt_code"].n_unique()), int(g["gt_code"].n_unique()),
                    int(g["qty"].sum())]
                   for mc, g in bp.group_by("machine")]

        bdaily = (bp.with_columns(
            pl.col("start_ts").map_elements(lambda t: _daypart(t, month_start)[0],
                                            return_dtype=pl.Int64).alias("Day"))
            .group_by("Day").agg(pl.col("qty").sum().alias("Built"),
                                 pl.col("gt_code").n_unique().alias("GTs"))
            .sort("Day"))

        p = out_dir / f"ctp_building_schedule_{mon}_{plant.lower()}.xlsx"
        _write(p, [
            ("Version",
             ["Parameter", "Value", "Note"],
             [["Version", "jk-planner curing-first", "-"],
              ["Generated", str(datetime.now())[:16], "-"],
              ["Month", month, f"{ndays} days"],
              ["Plant", plant, "-"],
              ["Day boundary", f"{DAY_START_H:02d}:00", "A/B/C x 480 min"],
              ["Source run", run_dir.name, "-"],
              ["Demand mode", "proxy_prev28", "leak-free, trailing 28 days"]]),
            ("Shift Schedule",
             ["Machine", "Day", "SKU", "Qty", "Mins", "Shifts", "Segments",
              "StartTime", "EndTime", "GT_Code", "SKU_Description"], seg),
            ("Shift Schedule (per-shift)",
             ["Machine", "Day", "Shift", "SKU", "Qty", "Mins", "Phase",
              "StartTime", "EndTime", "GT_Code", "SKU_Description"], per_shift),
            ("Changeover Plan",
             ["Machine", "Day", "Shift", "Old_SKU", "New_SKU", "CO_Type",
              "CO_Mins", "StartTime", "Old_SKU_GT_Code", "New_SKU_GT_Code"], co_rows),
            ("Setup Time %",
             ["Day", "Machine", "CO_Mins", "Setup_Mins", "Setup_%_of_1440"], setup_tbl),
            ("Machine SKU Variety",
             ["Machine", "Distinct_SKUs", "Distinct_GTs", "Total_Qty"], variety),
            ("Daily Built",
             ["Day", "Built", "Distinct_GTs"],
             [[r["Day"], int(r["Built"]), r["GTs"]] for r in bdaily.iter_rows(named=True)]),
        ])
        written.append(p)

        # ---------------- CURING ------------------------------------------
        cs = cp.with_columns(
            pl.col("start_ts").map_elements(lambda t: _daypart(t, month_start)[0],
                                            return_dtype=pl.Int64).alias("Day"),
            pl.col("start_ts").map_elements(lambda t: _daypart(t, month_start)[1],
                                            return_dtype=pl.Utf8).alias("Shift"))
        grid = (cs.group_by(["press", "Day", "Shift", "gt_code"])
                  .agg(pl.len().alias("Cured_Qty"),
                       (pl.col("cycle_s").sum() / 60.0).alias("Used_Mins"))
                  .sort(["press", "Day", "Shift"]))
        crows = []
        for r in grid.iter_rows(named=True):
            st, en = _shift_bounds(month_start, r["Day"], r["Shift"])
            crows.append([r["press"], r["Day"], r["Shift"], r["gt_code"], "RUNNING",
                          int(r["Cured_Qty"]), round(r["Used_Mins"], 1), 0, 0, 0,
                          str(st), str(en), r["gt_code"], desc.get(r["gt_code"], "")])

        # changeovers: press changes mounted GT between consecutive shifts
        seq = grid.sort(["press", "Day", "Shift"]).with_columns(
            pl.col("gt_code").shift(1).over("press").alias("prev"))
        cco = []
        for r in seq.filter(pl.col("prev").is_not_null()
                            & (pl.col("prev") != pl.col("gt_code"))).iter_rows(named=True):
            st, en = _shift_bounds(month_start, r["Day"], r["Shift"])
            cco.append([r["press"], r["prev"], r["gt_code"], r["Day"], r["Shift"],
                        str(st), str(en), 480, "regular",
                        r["prev"], r["gt_code"]])

        dem = (bp.group_by("gt_code").agg(pl.col("qty").sum().alias("Demand_Qty"))
                 .join(cs.group_by("gt_code").agg(pl.len().alias("Cured")),
                       on="gt_code", how="left")
                 .with_columns(pl.col("Cured").fill_null(0))
                 .with_columns((pl.col("Demand_Qty") - pl.col("Cured")).alias("Remaining"),
                               (100 * pl.col("Cured") / pl.col("Demand_Qty")).alias("Coverage_%"))
                 .sort("Demand_Qty", descending=True))

        util = (cs.group_by("press").agg(
            pl.len().alias("Cured_Qty"), (pl.col("cycle_s").sum() / 60.0).alias("Used_Mins"))
            .with_columns((100 * pl.col("Used_Mins") / (ndays * 1440)).round(1)
                          .alias("Utilization_%")).sort("press"))

        cdaily = (cs.group_by("Day").agg(pl.len().alias("Cured"))
                    .sort("Day"))
        # GT inventory per day (rule G8 band)
        evp = ev.filter(pl.col("plant") == plant).with_columns(
            pl.col("ts").map_elements(lambda t: _daypart(t, month_start)[0],
                                      return_dtype=pl.Int64).alias("Day"))
        inv = (evp.group_by("Day").agg(pl.col("qty_delta").sum().alias("net")).sort("Day")
                  .with_columns(pl.col("net").cum_sum().alias("GT_Inventory_EOD")))
        lo = CONFIG.thresholds.gt_wip_min.get(plant)
        hi = CONFIG.thresholds.gt_wip_max.get(plant)

        p = out_dir / f"ctp_curing_schedule_{mon}_{plant.lower()}.xlsx"
        _write(p, [
            ("Version",
             ["Parameter", "Value", "Note"],
             [["Version", "jk-planner curing-first", "-"],
              ["Generated", str(datetime.now())[:16], "-"],
              ["Month", month, f"{ndays} days"],
              ["Plant", plant, "-"],
              ["Engine", "window_plan + shift_grid", "campaign = window"],
              ["Source run", run_dir.name, "-"]]),
            ("Demand Fulfillment",
             ["GT_Code", "SKU_Description", "Demand_Qty", "Cured", "Remaining",
              "Coverage_%"],
             [[r["gt_code"], desc.get(r["gt_code"], ""), int(r["Demand_Qty"]),
               int(r["Cured"]), int(r["Remaining"]), round(r["Coverage_%"] or 0, 1)]
              for r in dem.iter_rows(named=True)]),
            ("Shift Schedule",
             ["Press", "Day", "Shift", "SKU", "Status", "Cured_Qty", "Used_Mins",
              "CO_Mins", "Mould_Clean_Mins", "Starved_Qty", "StartTime", "EndTime",
              "GT_Code", "SKU_Description"], crows),
            ("Changeover Plan",
             ["Press", "Old_SKU", "New_SKU", "Day", "Shift", "StartTime", "EndTime",
              "CO_Mins", "Type", "Old_SKU_GT_Code", "New_SKU_GT_Code"], cco),
            ("Machine Utilization",
             ["Press", "Cured_Qty", "Used_Mins", "Utilization_%"],
             [[r["press"], int(r["Cured_Qty"]), round(r["Used_Mins"], 1),
               r["Utilization_%"]] for r in util.iter_rows(named=True)]),
            ("Daily Cured",
             ["Day", "Date", "Cured"],
             [[r["Day"], str(month_start + timedelta(days=r["Day"] - 1)), int(r["Cured"])]
              for r in cdaily.iter_rows(named=True)]),
            ("GT Inventory (rule G8)",
             ["Day", "GT_Inventory_EOD", "Band_Min", "Band_Max", "Within_Band"],
             [[r["Day"], int(r["GT_Inventory_EOD"]), lo, hi,
               bool(lo <= r["GT_Inventory_EOD"] <= hi)]
              for r in inv.iter_rows(named=True) if r["Day"] >= 1]),
        ])
        written.append(p)

    log.info("export_ctp.done", files=[p.name for p in written])
    return written


if __name__ == "__main__":
    mo = sys.argv[1] if len(sys.argv) > 1 else "2026-07"
    rd = Path(sys.argv[2]) if len(sys.argv) > 2 else CONFIG.paths.runs / "walkforward" / f"month={mo}"
    od = Path(sys.argv[3]) if len(sys.argv) > 3 else CONFIG.paths.root / "output" / mo
    set_cutoff(None)
    for f in export(mo, rd, od):
        print("WROTE", f)
