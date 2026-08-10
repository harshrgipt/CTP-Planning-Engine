"""Export every dataset mined from the production history to shareable XLSX.

    python scripts/export_masters_xlsx.py [out_dir]

Writes to `output/masters_xlsx/`:

    01_derived_masters.xlsx      capability, cycle times, capacity, lot sizes
    02_demand_by_month.xlsx      8 months, one sheet each + summary
    03_opening_gt_by_month.xlsx  7 months (Dec excluded) + summary
    04_calibration_8months.xlsx  constant stability across the whole history
    05_plant_model.xlsx          the measured operating model + formulas

Every workbook opens with an INDEX sheet saying what each table is, where it
came from, and -- where it matters -- what is known to be wrong with it. A
master shared without its caveats gets used as if it were exact.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import polars as pl

from planner.config import CONFIG
from planner.runs.logger import log

MAXROW = 1_000_000


def _clean(v):
    if v is None:
        return ""
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


def _sheet(wb, name: str, df: pl.DataFrame, note: str = "") -> None:
    ws = wb.create_sheet(name[:31])
    if note:
        ws.append([note])
        ws.append([])
    ws.append(list(df.columns))
    for row in df.head(MAXROW).iter_rows():
        ws.append([_clean(v) for v in row])
    ws.freeze_panes = f"A{3 if note else 1}"
    for i, c in enumerate(df.columns, start=1):
        col = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col].width = max(11, min(34, len(str(c)) + 6))


def _index(wb, rows: list[tuple[str, str, str]]) -> None:
    ws = wb.create_sheet("INDEX", 0)
    ws.append(["Sheet", "What it is", "Source / caveat"])
    for r in rows:
        ws.append(list(r))
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 88
    ws.freeze_panes = "A2"


def main(out_dir: Path) -> int:
    import openpyxl
    out_dir.mkdir(parents=True, exist_ok=True)
    der = CONFIG.paths.warehouse / "derived"
    root = CONFIG.paths.root
    written = []

    # ---- 01 derived masters --------------------------------------------
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    idx = []
    spec = [
        ("allowed_machine_matrix", "Building machine eligibility per GT",
         "MINED from history. basis=direct (ran there) / size (widened). "
         "CAVEAT: median GT has only 2 machines -- this is history, not physics. "
         "Engine spills beyond it under the rim-size lock."),
        ("allowed_press_matrix", "Curing press eligibility per GT",
         "MINED. CAVEAT: lists 114 PCR presses against ~87 real -- cross-plant "
         "leak. Engine intersects with actual per-plant presses before use."),
        ("cycle_time_building", "Build cadence s/tyre per machine", "MINED"),
        ("cycle_time_curing", "Cure cadence s/tyre per press",
         "MINED. eff_CT = (raw_dwell + 2.3)/0.94; rate = 3*floor(480/eff_CT)"),
        ("capacity_machine_day", "Machine daily capacity", "MINED p95"),
        ("capacity_press_day", "Press daily capacity", "MINED p95"),
        ("lot_size", "Observed run lengths per GT", "MINED gaps-and-islands"),
        ("gt_size", "GT -> rim size", "Parsed from GT code + construction map"),
        ("calendar_shifts", "Shift pattern", "INFERRED -- no plant calendar master"),
        ("opening_gt_inventory", "Opening WIP snapshot", "MINED"),
    ]
    for name, what, src in spec:
        p = der / f"{name}.parquet"
        if p.exists():
            _sheet(wb, name, pl.read_parquet(p), f"{what}  |  {src}")
            idx.append((name[:31], what, src))
    pp = der / "plant_profile.json"
    if pp.exists():
        prof = json.loads(pp.read_text())
        flat = [{"key": k, "value": json.dumps(v, default=str)[:600]}
                for k, v in sorted(prof.items())]
        _sheet(wb, "plant_profile", pl.DataFrame(flat), "Mined operating envelope")
        idx.append(("plant_profile", "Measured operating envelope", "MINED"))
    _index(wb, idx)
    f = out_dir / "01_derived_masters.xlsx"; wb.save(f); written.append(f)

    # ---- 02 demand ------------------------------------------------------
    dd = CONFIG.paths.masters / "demand"
    if dd.exists():
        wb = openpyxl.Workbook(); wb.remove(wb.active); idx = []
        for p in sorted(dd.glob("demand_2*.csv")):
            tag = p.stem.replace("demand_", "")
            d = pl.read_csv(p)
            _sheet(wb, tag, d,
                   f"Demand {tag}: tyres CURED per GT per day. "
                   f"{d.height} rows, {int(d['qty'].sum()):,} tyres")
            idx.append((tag, f"{int(d['qty'].sum()):,} tyres, {d['gt_code'].n_unique()} GTs",
                        "Demand = finished tyres CURED (not built): a green tyre "
                        "is an intermediate, not an order line"))
        s = dd / "demand_summary.csv"
        if s.exists():
            _sheet(wb, "SUMMARY", pl.read_csv(s), "Per month x plant")
            idx.append(("SUMMARY", "Monthly totals and daily spread", ""))
        _index(wb, idx)
        f = out_dir / "02_demand_by_month.xlsx"; wb.save(f); written.append(f)

    # ---- 03 opening GT --------------------------------------------------
    od = CONFIG.paths.masters / "opening_gt"
    if od.exists():
        wb = openpyxl.Workbook(); wb.remove(wb.active); idx = []
        for p in sorted(od.glob("opening_gt_2*.csv")):
            tag = p.stem.replace("opening_gt_", "")
            d = pl.read_csv(p)
            _sheet(wb, tag, d, f"Opening GT {tag}, as of the 1st at 07:00")
            idx.append((tag, f"{int(d['qty'].sum()):,} tyres, {d.height} GT rows",
                        "As-of 1st 07:00 (plant day boundary). Age-bounded at the "
                        "measured build->cure p99 lag: older = scrap, not stock"))
        s = od / "opening_gt_summary.csv"
        if s.exists():
            _sheet(wb, "SUMMARY", pl.read_csv(s), "Per month x plant, with ages")
            idx.append(("SUMMARY", "Levels and ages", "December excluded: no prior "
                                                      "month to carry stock from"))
        _index(wb, idx)
        f = out_dir / "03_opening_gt_by_month.xlsx"; wb.save(f); written.append(f)

    # ---- 04 calibration --------------------------------------------------
    from planner.data.warehouse import duck, set_cutoff
    set_cutoff(None)
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    q = {
        "cure_cadence_s": """WITH s AS (SELECT plant, date_trunc('month',event_ts) mo,
             wcID::VARCHAR p, CAST(event_ts AS DATE) d, count(*) n FROM v_curing
             WHERE statuscritical='Normal' GROUP BY 1,2,3,4)
             SELECT plant, mo::DATE mo, 28800.0/(quantile_cont(n,0.5)/3.0) v FROM s GROUP BY 1,2""",
        "presses_used": """SELECT plant, date_trunc('month',event_ts)::DATE mo,
             count(DISTINCT wcID::VARCHAR) v FROM v_curing
             WHERE statuscritical='Normal' GROUP BY 1,2""",
        "build_machines": """SELECT plant, date_trunc('month',event_ts)::DATE mo,
             count(DISTINCT machineCode) v FROM v_build WHERE stage=2 GROUP BY 1,2""",
        "cured_per_day": """WITH d AS (SELECT plant, date_trunc('month',event_ts) mo,
             CAST(event_ts AS DATE) dd, count(*) n FROM v_curing
             WHERE statuscritical='Normal' GROUP BY 1,2,3)
             SELECT plant, mo::DATE mo, quantile_cont(n,0.5) v FROM d GROUP BY 1,2""",
        "gts_active": """SELECT plant, date_trunc('month',event_ts)::DATE mo,
             count(DISTINCT itemCode) v FROM v_build WHERE stage=2 GROUP BY 1,2""",
    }
    frames = []
    for name, sql in q.items():
        try:
            d = duck().execute(sql).pl().with_columns(pl.lit(name).alias("quantity"))
            # counts come back Int64, quantiles Float64 -- concat needs one type
            frames.append(d.select([
                pl.col("quantity").cast(pl.Utf8), pl.col("plant").cast(pl.Utf8),
                pl.col("mo").cast(pl.Date), pl.col("v").cast(pl.Float64)]))
        except Exception as e:  # noqa: BLE001
            log.warning("xlsx.calib_failed", q=name, err=str(e))
    if frames:
        cal = pl.concat(frames).sort(["quantity", "plant", "mo"])
        _sheet(wb, "measurements", cal,
               "Every planner constant, measured per month across all 8 months")
        piv = (cal.group_by(["quantity", "plant"])
               .agg(pl.col("v").mean().round(2).alias("mean"),
                    pl.col("v").min().round(2).alias("min"),
                    pl.col("v").max().round(2).alias("max"),
                    (pl.col("v").std() / pl.col("v").mean()).round(4).alias("CV"))
               .sort(["quantity", "plant"]))
        piv = piv.with_columns(
            pl.when(pl.col("CV") <= 0.05).then(pl.lit("STABLE -- safe to fix"))
              .otherwise(pl.lit("DRIFTS -- derive per month")).alias("verdict"))
        _sheet(wb, "stability", piv,
               "CV <= 0.05 = stable. Anything else MUST be derived per month: a "
               "constant read off one month cannot be shown wrong by that month.")
        _index(wb, [("measurements", "Per-month values", "MINED, 8 months"),
                    ("stability", "CV test + keep/derive verdict",
                     "This is the protocol: measure 8 months, keep only what is "
                     "stable, validate on held-out months")])
        f = out_dir / "04_calibration_8months.xlsx"; wb.save(f); written.append(f)

    # ---- 05 plant model --------------------------------------------------
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    model = pl.DataFrame([
        {"item": "Press rate", "formula": "3 x floor(480/eff_CT) x slots",
         "value": "PCR 156 / TBR 48 per press-day",
         "evidence": "eff_CT=(raw+2.3)/0.94; reproduces actuals exactly"},
        {"item": "GT inventory", "formula": "I = lambda x W,  W ~ 9 h",
         "value": "PCR ~4,820 / TBR ~1,297",
         "evidence": "8.4-10.9 h cover, both plants, all 8 months (Little's Law)"},
        {"item": "Per-GT stock", "formula": "I*_g = draw_g x T_g / 2",
         "value": "12 h cover at T_g=24 h",
         "evidence": "measured median GT cover 12.6 h; p25 3.9 / p75 24.9"},
        {"item": "Zero-hold tier", "formula": "area_g <= one campaign",
         "value": "19% PCR / 10% TBR of GT-months",
         "evidence": "those GTs cure ~480 (PCR) a month = ~3 press-days"},
        {"item": "GT scrap", "formula": "build/cure - 1",
         "value": "PCR 0.479% / TBR 1.99%",
         "evidence": "built and NEVER cured. TBR TRENDS 1.09->2.87%: derive it"},
        {"item": "Campaigns", "formula": "campaigns = sum n_g",
         "value": "changeovers = sum n_g - |P|",
         "evidence": "closed form, no search. campaign == window"},
        {"item": "Rectangle", "formula": "n_g x D_g = area_g = N_g / rate",
         "value": "area fixed, shape free",
         "evidence": "strip packing; flatten = n_g - 1, never +1"},
        {"item": "Steady state", "formula": "E[dI] ~ 0",
         "value": "plant +38/day, sd 530",
         "evidence": "TREND test, not band: plant swings 3,400-6,200 daily"},
        {"item": "Build-cure lag", "formula": "same shift",
         "value": "p50 4.3 h; 69% within one shift",
         "evidence": "93% within a day. Build lead = ONE SHIFT, not one day"},
        {"item": "Size lock", "formula": "rim(g) == rim_lock(m)",
         "value": "99.89% PCR / 99.75% TBR",
         "evidence": "hard prefilter on machines, not a score term"},
    ])
    _sheet(wb, "plant_model", model, "The measured operating model of the plant")
    _index(wb, [("plant_model", "Formulas + measured values + evidence",
                 "Everything derived from 8 months of MES; nothing assumed")])
    f = out_dir / "05_plant_model.xlsx"; wb.save(f); written.append(f)

    for w in written:
        print(f"WROTE {w}  ({w.stat().st_size/1024:.0f} KB)")
    log.info("xlsx.done", files=len(written))
    return 0


if __name__ == "__main__":
    d = Path(sys.argv[1]) if len(sys.argv) > 1 else CONFIG.paths.root / "output" / "masters_xlsx"
    sys.exit(main(d))
