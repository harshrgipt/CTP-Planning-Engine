"""Export a planned month as shop-floor shift schedules + supporting sheets.

    python scripts/export_shift_schedule.py <run_id> <YYYY-MM> [out_dir]

Writes ONE Excel workbook with every sheet, plus the same sheets as individual
CSVs so the pack is usable without Excel.

REFUSES to export a stale arm. `arm_is_stale()` proves the scorecard in the run
directory describes the plan sitting beside it -- 15 directories once carried
another arm's result, so this check is not optional.

WHAT IS EXPORTED IS WHAT THE PLAN SAYS. Quantities are never smoothed, rounded
or tidied: if a run is 11 tyres it is exported as 11 tyres.

ONE QUANTITY, ONE NUMBER. Every fulfilment figure in this pack and in the
BTP-format workbooks comes from `plan_quantities()` below, and every
press-shift quantity from `press_shift_cure()`. `scripts/export_btp_format.py`
imports both. Where two sheets legitimately count different things -- press
capacity seated, tyres fed, tyres delivered, tyres cured in-month -- both are
kept, each under a name that says what it counts, and `9c_quantity_bridge` is
the ladder from any one to any other with a running total that adds up. See
the block above `month_bounds` for the measurement that forced this.

Sheets added 2026-08-21, all reporting-only:
    1c_build_after_month_end  the build slices the report window clips out
    9c_quantity_bridge        the ladder: nameplate -> ... -> fulfilment %
    10_press_crosswalk        MES wc_id <-> the plant's own press_no
"""
from __future__ import annotations

import json
import os
import shutil
import sys
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import polars as pl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from planner.cmbc.l11_validate_plan import arm_is_stale  # noqa: E402
from planner.config import PRESS_ROSTER  # noqa: E402
from planner.config import CONFIG  # noqa: E402
from planner import paths

# B12 lot floor. PARTITION §8: every enforced cap lives in `config.py` and is
# read from there -- "add a cap in config.py or nowhere". Do NOT restate 150/70.
CONFIG_FLOOR = dict(CONFIG.thresholds.min_lot_units)
_STRICT_NOTE = (
    "sub-floor run PLACED. Under PLANNER_STRICT_LOT_FLOOR=1 this is unreachable "
    "(`_place` refuses gq < floor on every machine, HARD_FLOOR is forced on and "
    "ATOMIC_SPLIT is force-disabled), so a row here means the run was let "
    "through by the plant-calibrated sub-floor BUDGET "
    "(PLANNER_SUBFLOOR_PCR/_TBR) with STRICT off, or by the atomic-split "
    "rescue. Check `0_settings` for which was active.")

SHIFT_START_H = 7          # plant day runs 07:00 -> 07:00
SHIFTS = ["A", "B", "C"]   # A 07-15, B 15-23, C 23-07
FMT = "%Y-%m-%d %H:%M"
# Gates PROVEN stricter than the plant's own executed behaviour (measured on
# runs/plant_2026-07 via scripts/plant_as_plan.py, identical denominators).
MISMINED = {
    "TBR build runs below min_lot (70)": "plant itself runs 40.3% -- gate too strict",
    "TBR build changeovers / machine-day": "plant itself runs 3.70 -- gate too strict",
    "TBR WEIGHTED build changeover min/machine-day": "plant itself 37.3 -- too strict",
    "TBR realised n_g (concurrent presses/GT)": "plant runs 1.69; this floor demands MORE dispersion than the plant uses",
    "PCR realised n_g (concurrent presses/GT)": "same mis-mined n_g floor",
}


def shift_cols(df: pl.DataFrame, ts_col: str, t0: datetime) -> pl.DataFrame:
    """Plant day + shift letter for a timestamp. Day boundary is 07:00.

    `date` IS THE PLANT-DAY DATE, not the wall-clock date of the timestamp.
    This used to be `ts_col.dt.date()`, which silently split every C shift in
    two: plant day 1 shift C runs 01 Jul 23:00 -> 02 Jul 07:00, so 7 of its 8
    hours carried the label `2026-07-02`. Filtering "01 Jul, shift C" then
    returned 352 of 4,241 PCR tyres and read as "nothing is being built on the
    night shift". 1,614 of 5,631 build rows (28.7 %, 135,379 tyres) were
    mislabelled that way, and it is also why the cure sheet appeared to stop a
    day before the build sheet. `date` + `shift` is now a true partition of the
    plant day; `cal_date` keeps the wall-clock date for anyone who needs it.
    """
    off = ((pl.col(ts_col) - pl.lit(t0)).dt.total_seconds() / 3600.0)
    return df.with_columns([
        (off // 24 + 1).cast(pl.Int64).alias("plant_day"),
        pl.col(ts_col).dt.strftime("%Y-%m-%d").alias("cal_date"),
        ((off % 24) // 8).cast(pl.Int64).alias("_s"),
    ]).with_columns([
        pl.when(pl.col("_s") == 0).then(pl.lit("A"))
        .when(pl.col("_s") == 1).then(pl.lit("B"))
        .otherwise(pl.lit("C")).alias("shift"),
        (pl.lit(t0.date()) + pl.duration(days=pl.col("plant_day") - 1))
        .dt.strftime("%Y-%m-%d").alias("date"),
    ]).drop("_s")


def fmt_ts(df: pl.DataFrame, cols) -> pl.DataFrame:
    for c in cols:
        if c in df.columns:
            df = df.with_columns(pl.col(c).dt.strftime(FMT).alias(c))
    return df


# =====================================================================
# ONE QUANTITY, ONE NUMBER -- the canonical basis block
# =====================================================================
# WHAT THIS DOES / WHY IT EXISTS -- a measured defect, found 2026-08-21.
#   This pack shipped FIVE different numbers all called "cured", and the one
#   the gate actually grades appeared in none of them. August PCR, run
#   SHIP2_aug, every figure re-derived from `runs/SHIP2_aug/`:
#
#     quantity                                    value  where it used to ship
#     campaign NAMEPLATE seated in-month        408,929  BTP curing workbook
#                                                        `Daily Cured tyres`
#                                                        TOTAL and
#                                                        `Planned_Units`
#     tyres DELIVERED, cure completes in-month  401,551  csv sheet 2 `qty`
#                                                        = sheet 6 = sheet 7
#     tyres FED IN-MONTH (campaign pro-rata)    397,326  nowhere
#     CURED IN-MONTH per L11 (fed in-month,
#       capped at each GT's own requirement)    396,636  nowhere -- and this
#                                                        IS the fulfilment
#                                                        numerator
#     the requirement (the denominator)         427,949  nowhere
#
#   So the BTP-format workbook published PCR Jul 98.37 / TBR Jul 99.30 /
#   PCR Aug 95.29 / TBR Aug 99.47 % while `9b_l11_invariants` in the SAME
#   PACK published 96.1 / 96.1 / 92.7 / 96.1 % FAIL. TBR shipped as FAIL in
#   one file and PASS in the other, from one run, because the workbook
#   divided press capacity seated by raw order-book demand.
#
#   FIX: every fulfilment figure in both exporters now comes from
#   `plan_quantities()` and every press-shift quantity from
#   `press_shift_cure()`. `scripts/export_btp_format.py` imports both, so
#   there is ONE implementation that can be wrong instead of two that can
#   disagree. `plan_quantities` re-derives L11's own arithmetic and then
#   ASSERTS its own percentage against the string L11 wrote into
#   `l11_invariants.parquet`; a mismatch prints `!! BASIS MISMATCH` and is
#   not silent.
#
#   THE TRAP IN READING THIS. These are not four errors and one truth. Four
#   of the five are correct answers to different questions and the plant
#   needs at least three of them -- nameplate minus fed IS the press
#   starvation signal, and deleting it to make totals agree would destroy
#   it. They are ALL kept, each under a name that says what it counts, and
#   `9c_quantity_bridge` prints the arithmetic from any one to any other,
#   per plant, as a ladder that adds up.
#
#   Ledger: PARTITION_AND_CHANGEOVER.md 1 (measurement errors) and 4af (the
#   denominator-basis class -- this is the seventh instance). L11's own
#   definition is at planner/cmbc/l11_validate_plan.py, section "--- demand
#   ---"; it is the definition the gate grades and therefore the one used.
# =====================================================================
PLANTS = ("PCR", "TBR")


def month_bounds(month: str):
    """(t0, n_plant_days, month_end). The plant day runs 07:00 -> 07:00."""
    y, m = int(month[:4]), int(month[5:7])
    t0 = datetime(y, m, 1, SHIFT_START_H)
    nd = (datetime(y + (m == 12), (m % 12) + 1, 1) - datetime(y, m, 1)).days
    return t0, nd, t0 + timedelta(days=nd)


def plant_closures(run: Path, month: str) -> dict:
    """Closed plant-days per plant, DERIVED FROM THE PLAN ITSELF.

    A plant-day on which not one of the plant's ~11 building machines is
    scheduled for a single minute is a plant closure (rule G3), not an idle
    day. Derived here rather than read from the calendar because the calendar
    file is mutable after the fact: `holiday.csv` at the wrapper root today
    declares a 2026-07-15 PCR closure created 2026-08-20, which the July plan
    in `runs/SHIP2_jul` does not contain -- its own log says
    "PLANT HOLIDAYS: none". Deriving from the plan cannot drift from the plan.

    The run's own log line is read as a CROSS-CHECK and returned under `_log`;
    if the two disagree the caller prints it. Only FULL-day closures are
    detectable this way -- a partial (shift-level) closure leaves build hours
    on the day, so `_log` stays the authority for those.

    Why it matters for the report and not only for the plan: a closed day is
    not available capacity. Dividing August machine hours by 744 h counts the
    24 h the plant was shut, which understated occupancy by 2.7 pt, and the
    old sheet 2 booked 3,576 press-hours and 16,059 pro-rata tyres onto
    2026-08-15 at zero real output.
    """
    t0, nd, _ = month_bounds(month)
    bs = pl.read_parquet(run / "build_schedule.parquet").filter(
        pl.col("machine") != "OPENING_STOCK")
    out: dict = {}
    for p in PLANTS:
        B = bs.filter(pl.col("plant") == p)
        days = []
        for d in range(nd):
            lo = t0 + timedelta(days=d)
            hi = lo + timedelta(days=1)
            h = float(B.select(
                ((pl.min_horizontal(pl.col("end_ts"), pl.lit(hi))
                  - pl.max_horizontal(pl.col("start_ts"), pl.lit(lo)))
                 .dt.total_seconds() / 3600.0)
                .clip(lower_bound=0).sum()).item()) if B.height else 0.0
            if h <= 1e-6:
                days.append(d + 1)
        out[p] = days
    logline = ""
    for f in ("log_l7_pull_release.txt", "log_l5_cure_master.txt"):
        fp = run / f
        if fp.exists():
            for ln in fp.read_text(errors="replace").splitlines():
                if "PLANT HOLIDAYS" in ln:
                    logline = ln.strip()
                    break
        if logline:
            break
    out["_log"] = logline
    out["_windows"] = {p: [(t0 + timedelta(days=d - 1), t0 + timedelta(days=d))
                           for d in out[p]] for p in PLANTS}
    return out


def _lr_round(vals: list, total: float) -> list:
    """Largest-remainder rounding: whole tyres that sum to `total` EXACTLY.

    A press-shift cannot make 0.3 of a tyre, and rounding 6,964 float rows
    independently drifts: `qty_run` rounded to one decimal per row summed to
    384,484.5 against the 384,406 it is supposed to reproduce -- 78 tyres
    manufactured by the CSV writer. Both exporters call this, so the csv pack
    and the BTP workbooks integerise the same rows the same way and their daily
    curves are identical rather than 15 tyres apart on the worst day.
    """
    tot = int(round(total))
    base = [int(v) for v in vals]
    rem = tot - sum(base)
    order = sorted(range(len(vals)), key=lambda i: (-(vals[i] - base[i]), i))
    k = 0
    while rem > 0 and order:
        base[order[k % len(order)]] += 1
        rem -= 1
        k += 1
    while rem < 0:
        i = max(range(len(base)), key=lambda j: base[j])
        if base[i] <= 0:
            break
        base[i] -= 1
        rem += 1
    return base


def _open_parts(lo: datetime, hi: datetime, wins) -> list:
    """[lo, hi) minus every closure window -- the press hours that exist."""
    parts = [(lo, hi)]
    for ws, we in wins:
        nxt = []
        for a, b in parts:
            if we <= a or ws >= b:
                nxt.append((a, b))
                continue
            if a < ws:
                nxt.append((a, min(b, ws)))
            if b > we:
                nxt.append((max(a, we), b))
        parts = [(a, b) for a, b in nxt if (b - a).total_seconds() > 1e-6]
    return parts


def press_shift_cure(run: Path, month: str, closed: dict | None = None):
    """One row per (plant, plant_day, shift, press, GT) -- THREE named bases.

    Returns (frame, residues). Columns:
      qty           tyres DELIVERED to this press in this shift and cured
                    in-month. Bucketed on the build slice's `cure_ts`, so it
                    is spiky by construction (a press works for two shifts on
                    a delivery it took in the first). Sums per plant to
                    `delivered_in_month`. THIS is the column sheets 6 and 7
                    and the verifier reconcile against.
      qty_run       the campaign's IN-MONTH FED quantity spread pro-rata over
                    the press hours it actually holds in this shift. Smooth,
                    reads like press output, and sums per plant EXACTLY to
                    `fed_in_month`, one documented step from the fulfilment
                    numerator.
      qty_planned   the campaign NAMEPLATE's share of this shift -- PRESS
                    CAPACITY SEATED, not tyres. Sums to `nameplate_in_month`.
                    Never quote it as output.

    TWO DEFECTS THIS REPLACES, both measured 2026-08-21 on the shipped pack.
      1. `qty_run` prorated over the campaign's CLIPPED span, so it summed to
         the WHOLE-PLAN fed quantity rather than the in-month one: August PCR
         409,110 against sheet 2's 401,551, 7,559 tyres of next month printed
         on this month's shift rows. The header comment claimed it "sums to
         `qty` per plant". It never did.
      2. Neither `hours` nor the pro-rata quantities knew about the plant
         closure. A campaign PAUSES over a shutdown, so its wall-clock span is
         24 h longer than the press hours it draws (planner/cmbc/holiday.py
         records the same defect arriving through `cure_ts`). August day 15
         carried 1,944 PCR + 1,632 TBR press-hours and 12,898 + 3,161 pro-rata
         tyres on a day the plant was shut and built nothing. Segments are now
         cut against the closure windows, so a closed press-shift produces no
         row at all -- which is also what `cure_by_shift.parquet` says, and
         that agreement is asserted below rather than assumed.
    """
    t0, nd, mend = month_bounds(month)
    if closed is None:
        closed = plant_closures(run, month)
    wins = closed.get("_windows", {p: [] for p in PLANTS})
    rec = pl.read_parquet(run / "cure_campaigns_reconciled.parquet")
    bs = pl.read_parquet(run / "build_schedule.parquet")

    seg = []
    for r in rec.iter_rows(named=True):
        s, e = r["start_ts"], r["end_ts"]
        if (e - s).total_seconds() <= 0:
            continue
        w = wins.get(r["plant"], [])
        # OPEN seconds of the WHOLE campaign -- the nameplate pro-rata
        # denominator. A paused campaign's wall-clock span is longer than the
        # press hours it draws, so this is not `end - start`.
        tot_open = sum((b - a).total_seconds() for a, b in _open_parts(s, e, w))
        if tot_open <= 0:
            continue
        cid = f"{r['plant']}-{r['press']}-{r['start_ts']:%m%d%H%M}"
        k0 = int((s - t0).total_seconds() // (8 * 3600))
        k1 = int(np.ceil((e - t0).total_seconds() / (8 * 3600)))
        mine = []
        for k in range(max(k0, 0), min(k1, 3 * nd)):
            lo = max(s, t0 + timedelta(hours=8 * k))
            hi = min(e, t0 + timedelta(hours=8 * (k + 1)))
            if (hi - lo).total_seconds() <= 1e-6:
                continue
            for a, b in _open_parts(lo, hi, w):
                mine.append((k, a, b, (b - a).total_seconds()))
        # `qty_run` is allocated over the IN-MONTH open segments only, with
        # L7's own `qty_fed_in_month` as the total, so the column sums to that
        # number EXACTLY per plant rather than approximately. `qty_planned`
        # keeps the whole-campaign denominator because the nameplate outside
        # the month is real press capacity that this month simply does not
        # report.
        in_open = sum(x[3] for x in mine) or 1.0
        for k, a, b, sec in mine:
            seg.append({
                "plant": r["plant"], "press": str(r["press"]),
                "gt_code": r["gt_code"], "mould_set": r.get("mould_set"),
                "plant_day": k // 3 + 1, "shift": "ABC"[k % 3],
                "start_ts": a, "end_ts": b, "hours": sec / 3600.0,
                "qty_planned": float(r["qty"]) * sec / tot_open,
                "qty_run": float(r["qty_fed_in_month"]) * sec / in_open,
                "campaign_id": cid})
    c = pl.DataFrame(seg)
    # The grain MUST be exactly (plant, plant_day, shift, press, GT) so the
    # join to the delivered tyres is 1:1. One campaign can re-enter a shift as
    # two segments (a closure now splits one too) and a press can switch
    # between two campaigns of the SAME GT inside one shift, so the quantities
    # are resolved PER SEGMENT, where the campaign is still unambiguous, and
    # only then summed. Keeping campaign_id in the group key survives the
    # first and not the second -- it once left 89 duplicate rows and
    # double-counted 13,095 fed tyres.
    c = c.group_by(["plant", "press", "gt_code", "plant_day", "shift"]).agg([
        pl.col("mould_set").first(),
        pl.col("start_ts").min(), pl.col("end_ts").max(),
        pl.col("hours").sum(), pl.col("qty_planned").sum(),
        pl.col("qty_run").sum(),
        pl.col("campaign_id").first(),
        pl.col("campaign_id").n_unique().alias("campaigns"),
        pl.len().alias("segments")])

    # tyres actually DELIVERED to this press in this shift, INCLUDING the
    # OPENING_STOCK pseudo-machine (a real feed to a press), and cured
    # in-month only -- a tyre whose cure completes after the boundary is next
    # month's output, not a day-1..N cure row.
    _off = (pl.col("cure_ts") - pl.lit(t0)).dt.total_seconds() / 3600.0
    _in = bs.filter(pl.col("cure_ts") <= pl.lit(mend))
    _orphan = _in.filter(pl.col("press").is_null())
    if _orphan.height:
        print(f"  !! {_orphan.height} build slices cure in-month with NO press "
              f"assigned ({float(_orphan['qty'].sum()):,.0f} tyres) -- bucketed "
              f"as press '(unassigned)'")
    fed_sh = (_in.with_columns([
        (_off // 24 + 1).cast(pl.Int64).alias("plant_day"),
        ((_off % 24) // 8).cast(pl.Int64).alias("_s"),
        pl.col("press").cast(pl.Utf8).fill_null("(unassigned)").alias("press")])
        .with_columns(pl.when(pl.col("_s") == 0).then(pl.lit("A"))
                      .when(pl.col("_s") == 1).then(pl.lit("B"))
                      .otherwise(pl.lit("C")).alias("shift"))
        .group_by(["plant", "press", "gt_code", "plant_day", "shift"])
        .agg(pl.col("qty").sum().alias("qty")))
    c = (c.join(fed_sh, on=["plant", "press", "gt_code", "plant_day", "shift"],
                how="full", coalesce=True)
         .with_columns([pl.col(x).fill_null(0.0) for x in
                        ("qty", "qty_planned", "hours", "qty_run")])
         .with_columns([pl.col("segments").fill_null(0),
                        pl.col("campaigns").fill_null(0)]))
    # A row with no press time AND no tyres is not a schedule row. Under the
    # closure this is exactly the 447 August day-15 rows.
    c = c.filter((pl.col("hours") > 1e-9) | (pl.col("qty") > 0))
    # SORT BEFORE ROUNDING, EXPLICITLY. `group_by` and a full join both leave
    # polars free to return rows in any order, and `_lr_round` hands its
    # remainder tyres to the rows with the largest fractions and breaks ties on
    # ROW INDEX. Two calls on the same run therefore produced two different
    # (still exactly-summing) allocations, and the csv pack and the BTP
    # workbook disagreed by up to 6 tyres on a plant-day for no reason but row
    # order. Measured 2026-08-21. A deterministic sort is the whole fix.
    c = c.sort(["plant", "plant_day", "shift", "press", "gt_code"])
    # WHOLE TYRES, SUMMING EXACTLY. Done here, once, so every consumer of this
    # frame integerises identically -- see `_lr_round`.
    parts = []
    for p in PLANTS:
        z = c.filter(pl.col("plant") == p)
        if not z.height:
            continue
        rec_p = rec.filter(pl.col("plant") == p)
        parts.append(z.with_columns([
            pl.Series("qty_run", _lr_round(
                z["qty_run"].to_list(), float(rec_p["qty_fed_in_month"].sum())),
                dtype=pl.Float64),
            pl.Series("qty_planned", _lr_round(
                z["qty_planned"].to_list(), float(z["qty_planned"].sum())),
                dtype=pl.Float64)]))
    c = pl.concat(parts) if parts else c

    # ---- ASSERT AGAINST L10's OWN DISCRETISATION --------------------------
    # `cure_by_shift.parquet` is L10's independent closure-aware view of the
    # same campaigns. If the segmentation here disagrees with it, one of the
    # two is wrong and the reader must be told which -- two artefacts that
    # disagree is the only reason the last closure defect was ever found.
    res: dict = {}
    cbs = run / "cure_by_shift.parquet"
    if cbs.exists():
        L10 = pl.read_parquet(cbs)
        for p in PLANTS:
            a = float(c.filter(pl.col("plant") == p)["qty_planned"].sum())
            b = float(L10.filter(pl.col("plant") == p)["qty"].sum())
            ha = float(c.filter(pl.col("plant") == p)["hours"].sum())
            hb = float(L10.filter(pl.col("plant") == p)["hours"].sum())
            res[p] = {"nameplate_here": a, "nameplate_L10": b,
                      "hours_here": ha, "hours_L10": hb}
    return c, res


def plan_quantities(run: Path, month: str) -> dict:
    """EVERY quantity this pack prints, derived once, per plant.

    The fulfilment numerator and denominator are L11's own definition,
    re-derived here line for line from `planner/cmbc/l11_validate_plan.py`
    and then CHECKED against the percentage string L11 itself wrote into
    `l11_invariants.parquet`. That check is the whole point: this function is
    a second route to a number the gate already computes, and measurement
    rule 2 in this project is that two routes to one quantity must be tied
    together or the downstream one is the one nobody checks.

      numerator    per GT: min(sum qty_fed_in_month, gross_build -
                   gross_build_la), lookahead-only GTs dropped entirely
      denominator  per GT: cure_requirement - cure_requirement_la over
                   ~residual, ~lookahead

    NOTE THE ASYMMETRY, IT IS L11's AND IT IS DELIBERATE: the numerator is
    capped at `gross_build` (what building must make) while the denominator
    is `cure_requirement` (what the presses must consume, opening stock
    included). The cap only ever removes cures BEYOND this month's own
    requirement for that GT -- real output that belongs in BUILT, never in
    this ratio. Do not "harmonise" the two without re-measuring L11.
    """
    t0, nd, mend = month_bounds(month)
    bs = pl.read_parquet(run / "build_schedule.parquet")
    rec = pl.read_parquet(run / "cure_campaigns_reconciled.parquet")
    req = pl.read_parquet(ROOT / "warehouse" / "derived" /
                          f"net_requirement_{month}.parquet")
    inv = pl.read_parquet(run / "l11_invariants.parquet")
    l11_pct = {r["invariant"].split()[0]: r["actual"] for r in
               inv.iter_rows(named=True)
               if r["invariant"].endswith("demand fulfilment")
               and " " in r["invariant"]}

    _fedcol = ("qty_fed_in_month" if "qty_fed_in_month" in rec.columns
               else "qty_fed")
    la_gts = set()
    if "lookahead" in req.columns:
        la_gts = {(r["plant"], r["gt_code"])
                  for r in req.filter(pl.col("lookahead")).iter_rows(named=True)}
    cap = {}
    for r in req.iter_rows(named=True):
        cap[(r["plant"], r["gt_code"])] = max(
            float(r["gross_build"] or 0.0) - float(r.get("gross_build_la") or 0.0),
            0.0)
    dcol = ("cure_requirement" if "cure_requirement" in req.columns
            else "gross_build")
    sc = req.filter(~pl.col("residual"))
    if "lookahead" in sc.columns:
        sc = sc.filter(~pl.col("lookahead"))

    out: dict = {"_month": month, "_run": run.name, "_l11": l11_pct,
                 "_ndays": nd, "_t0": t0, "_mend": mend}
    for p in PLANTS:
        B = bs.filter(pl.col("plant") == p)
        Bp = B.filter(pl.col("machine") != "OPENING_STOCK")
        R = rec.filter(pl.col("plant") == p)
        num_by_gt, fed_by_gt = {}, {}
        for r in (R.group_by("gt_code").agg(
                pl.col(_fedcol).sum().alias("t"),
                pl.col("qty_fed").sum().alias("f")).iter_rows(named=True)):
            g = r["gt_code"]
            fed_by_gt[g] = float(r["f"])
            if (p, g) in la_gts:
                continue
            num_by_gt[g] = min(float(r["t"]), cap.get((p, g), 1e18))
        S = sc.filter(pl.col("plant") == p)
        den_by_gt = {}
        for r in S.iter_rows(named=True):
            den_by_gt[r["gt_code"]] = (float(r[dcol] or 0.0)
                                       - float(r.get(dcol + "_la") or 0.0))
        numer = sum(num_by_gt.values())
        denom = sum(den_by_gt.values())
        nameplate = float(R["qty"].sum())
        unfed = float(R["qty_unfed"].sum())
        fed = float(R["qty_fed"].sum())
        fed_in = float(R[_fedcol].sum())
        built = float(Bp.filter(pl.col("end_ts") <= pl.lit(mend))["qty"].sum())
        built_late = float(Bp.filter(pl.col("end_ts") > pl.lit(mend))["qty"].sum())
        opening = float(B.filter(pl.col("machine") == "OPENING_STOCK")["qty"].sum())
        delivered = float(B.filter(pl.col("cure_ts") <= pl.lit(mend))["qty"].sum())
        closing = float(B.filter((pl.col("machine") != "OPENING_STOCK")
                                 & (pl.col("end_ts") <= pl.lit(mend))
                                 & (pl.col("cure_ts") > pl.lit(mend)))["qty"].sum())
        unassigned = float(B.filter(pl.col("press").is_null())["qty"].sum())
        out[p] = {
            "nameplate": nameplate, "unfed": unfed, "fed": fed,
            "tail_campaign": fed - fed_in, "fed_in_month": fed_in,
            "over_req_cap": fed_in - numer, "numerator": numer,
            "denominator": denom,
            "pct": (100.0 * numer / denom) if denom > 0 else 0.0,
            "built": built, "built_after_end": built_late, "opening": opening,
            "delivered_in_month": delivered, "closing_gt": closing,
            "closing_gt_unassigned": unassigned,
            "bs_total": float(B["qty"].sum()),
            "num_by_gt": num_by_gt, "den_by_gt": den_by_gt,
            "fed_by_gt": fed_by_gt,
        }
    # ---- the identities, ASSERTED, not hoped for -------------------------
    print("  BASIS CHECK  (fulfilment re-derived here vs the string L11 wrote)")
    for p in PLANTS:
        q = out[p]
        mine = f"{q['pct']:.1f}%"
        theirs = l11_pct.get(p, "(absent)")
        ok = (mine == theirs)
        print(f"     {p}  {q['numerator']:>9,.0f} / {q['denominator']:>9,.0f} "
              f"= {mine:>6}   L11 says {theirs:>6}   "
              f"{'OK' if ok else '!! BASIS MISMATCH'}")
        for name, lhs, rhs in (
            ("built + built_after_end + opening == build_schedule",
             q["built"] + q["built_after_end"] + q["opening"], q["bs_total"]),
            ("built + opening - closing_gt == delivered_in_month",
             q["built"] + q["opening"] - q["closing_gt"], q["delivered_in_month"]),
            ("nameplate - unfed == fed", q["nameplate"] - q["unfed"], q["fed"]),
            ("fed + press-unassigned closing == build_schedule",
             q["fed"] + q["closing_gt_unassigned"], q["bs_total"]),
        ):
            if abs(lhs - rhs) > 0.5:
                print(f"     !! {p} IDENTITY BROKEN  {name}: "
                      f"{lhs:,.1f} vs {rhs:,.1f}")
    return out


def quantity_bridge(q: dict) -> pl.DataFrame:
    """The ladder from press capacity to the fulfilment %, per plant.

    Every line is `running += value`, so a reader can add the column up and
    land on the number the gate grades. Nine of the ten quantities this pack
    prints appear here exactly once, under the name that says what they count.
    """
    rows = []
    for p in PLANTS:
        d = q[p]
        lad = [
            ("CURE / PRESS LADDER", None, "", ""),
            ("campaign NAMEPLATE (whole plan)", d["nameplate"], "=",
             "press capacity seated by L5. NOT tyres. 2b_cure_campaigns "
             "`qty_planned`"),
            ("- press slots never fed (starved presses)", -d["unfed"], "-",
             "2b_cure_campaigns `qty_unfed`, 6_press_summary `tyres_unfed`"),
            ("= tyres FED to presses (whole plan)", d["fed"], "=",
             "L7 `qty_fed`. Includes tyres whose cure completes next month"),
            ("- carry-out tail (cure completes NEXT month)",
             -d["tail_campaign"], "-",
             "campaign pro-rata, L7 `frac_in_month`. NOT the same number as "
             "9b's 'carry-out tail', which is this PLUS the cap below"),
            ("= tyres FED IN-MONTH", d["fed_in_month"], "=",
             "2_cure_schedule_shift `qty_run` sums to exactly this"),
            ("- cures beyond this GT's own requirement (capped)",
             -d["over_req_cap"], "-",
             "real output, counted in BUILT, excluded from this ratio only"),
            ("= CURED IN-MONTH  <- FULFILMENT NUMERATOR", d["numerator"], "=",
             "L11 `<plant> demand fulfilment` numerator. THE output number"),
            ("/ cure requirement  <- FULFILMENT DENOMINATOR", d["denominator"],
             "/", "net_requirement `cure_requirement` - `_la`, ~residual, "
             "~lookahead. 8_demand_vs_plan `requirement_L11` sums to this"),
            ("= demand fulfilment %", round(d["pct"], 2), "=",
             "identical in 9a, 9b, both BTP workbooks"),
            ("BUILD / GREEN-TYRE LADDER", None, "", ""),
            ("tyres BUILT this month", d["built"], "=",
             "1_build_schedule_shift, 5_machine_summary, 7 `built`"),
            ("+ tyres BUILT after month end (EXCLUDED from this report)",
             d["built_after_end"], "+",
             "1c_build_after_month_end. Next month's green tyres; every "
             "`carry_out` column in this pack is False BECAUSE of this clip"),
            ("+ OPENING GT stock consumed", d["opening"], "+",
             "9a 'of which OPENING STOCK carried in'"),
            ("= green tyres in the plan", d["bs_total"], "=",
             "build_schedule.parquet total"),
            ("- built after month end (excluded above)",
             -d["built_after_end"], "-", ""),
            ("- CLOSING GT stock (built now, cures next month)",
             -d["closing_gt"], "-",
             "7_daily_summary last `gt_inventory_close`; BTP curing "
             "`GT Gap Diagnostic` `Closing_Balance`"),
            ("= tyres DELIVERED, cure completes in-month",
             d["delivered_in_month"], "=",
             "2_cure_schedule_shift `qty`, 6_press_summary `tyres`, "
             "7_daily_summary `cured`"),
            ("CROSSWALK BETWEEN THE TWO LADDERS", None, "", ""),
            ("of the closing GT stock, never assigned to a press",
             d["closing_gt_unassigned"], "",
             "build_schedule rows with press = null. This is the whole gap "
             "between `fed` and the build_schedule total"),
            ("delivered in-month MINUS fed in-month",
             d["delivered_in_month"] - d["fed_in_month"], "",
             "slice `cure_ts` cut vs campaign time pro-rata -- two legitimate "
             "cuts of the same month boundary, not a leak"),
        ]
        run_tot = None
        for name, val, op, src in lad:
            if val is None:
                rows.append({"plant": p, "line": name, "value": None,
                             "op": "", "running": None, "where_it_appears": ""})
                run_tot = None
                continue
            if op == "=":
                run_tot = val
            elif op in ("+", "-") and run_tot is not None:
                run_tot = run_tot + val
            rows.append({"plant": p, "line": name, "value": round(float(val), 2),
                         "op": op,
                         "running": None if run_tot is None else round(float(run_tot), 2),
                         "where_it_appears": src})
    return pl.DataFrame(rows)


def main() -> int:
    run_id, month = sys.argv[1], sys.argv[2]
    out = Path(sys.argv[3]) if len(sys.argv) > 3 else \
        ROOT / "output" / f"{month.replace('-', '_')}_schedule"
    run = ROOT / "runs" / run_id
    why = arm_is_stale(run)
    if why and not why.startswith("no l11_provenance"):
        raise SystemExit(f"!! REFUSING to export a stale arm: {why}")
    y, m = int(month[:4]), int(month[5:7])
    t0 = datetime(y, m, 1, SHIFT_START_H)
    # NEVER rmtree the output folder. A workbook open in Excel holds a lock, and
    # rmtree deletes csv/ BEFORE it reaches the locked .xlsx -- which destroys
    # half the pack and then aborts. Clear only the files we are about to
    # rewrite, each one tolerantly.
    (out / "csv").mkdir(parents=True, exist_ok=True)
    for _f in (out / "csv").glob("*.csv"):
        try:
            _f.unlink()
        except OSError as e:                            # noqa: PERF203
            print(f"  !! could not remove {_f.name}: {e}")

    R = lambda n: pl.read_parquet(run / f"{n}.parquet")   # noqa: E731
    bs, cc = R("build_schedule"), R("cure_campaigns")

    # ---- REPORT WINDOW: DAYS 1..ndays ONLY --------------------------------
    # PLANT RULING 2026-08-10: plan on month + tail, REPORT only the month.
    # Applied HERE, at load, and nowhere else -- every sheet in this pack is
    # derived from `bs`/`cc`, and sheet 1 / 5 / 7 must reconcile to the tyre. A
    # per-sheet filter is how two sheets end up disagreeing.
    #   BUILD  cut WHOLE-SLICE on end_ts. That is the GT ledger's own convention
    #          (`gt_events` credits a slice whole at end_ts), so a slice
    #          finishing next month is next month's green tyre -- it is neither
    #          exported nor carried forward, and the two numbers agree.
    #   CURE   cut PRO-RATA. A campaign that crosses the boundary has a real
    #          in-month portion and its press really is running on day 31;
    #          dropping it whole would taper the presses at exactly the hour the
    #          ruling says not to. Same time-fraction L7 uses for
    #          `qty_fed_in_month`, so sheet 2 and the fulfilment KPI cannot
    #          disagree about the same campaign.
    ndays = (datetime(y + (m == 12), (m % 12) + 1, 1)
             - datetime(y, m, 1)).days      # plant month length
    _mend = t0 + timedelta(days=ndays)
    _b0, _c0 = float(bs["qty"].sum()), float(cc["qty"].sum())
    bs_full = bs          # UNCLIPPED -- sheet 1c exports what the clip removes
    bs = bs.filter(pl.col("end_ts") <= pl.lit(_mend))
    _frac = ((pl.lit(_mend) - pl.col("start_ts")).dt.total_seconds()
             / (pl.col("end_ts") - pl.col("start_ts")).dt.total_seconds()
             .clip(lower_bound=1.0)).clip(0.0, 1.0)
    cc = (cc.filter(pl.col("start_ts") < pl.lit(_mend))
          .with_columns([
              (pl.col("qty") * _frac).round(0).alias("qty"),
              (pl.col("hours") * _frac).alias("hours"),
              pl.min_horizontal(pl.col("end_ts"), pl.lit(_mend)).alias("end_ts")])
          .filter(pl.col("qty") > 0))
    # Closing GT stock = built in-month, cured after it. Same definition L7 uses
    # for `carry_forward_gt.parquet`; read here so the pack and the hand-off file
    # cannot drift apart.
    _cf_qty = {r["plant"]: float(r["q"]) for r in
               (bs.filter((pl.col("machine") != "OPENING_STOCK")
                          & (pl.col("cure_ts") > pl.lit(_mend)))
                .group_by("plant").agg(pl.col("qty").sum().alias("q"))
                ).iter_rows(named=True)}
    if abs(_b0 - float(bs["qty"].sum())) > 0.5 or abs(_c0 - float(cc["qty"].sum())) > 0.5:
        print(f"  REPORT WINDOW  days 1-{ndays} (to {_mend:%Y-%m-%d %H:%M}): "
              f"build {_b0:,.0f} -> {float(bs['qty'].sum()):,.0f}  "
              f"cure {_c0:,.0f} -> {float(cc['qty'].sum()):,.0f}  "
              f"-- the difference is next month's, not lost")
    st, up = R("build_starved"), R("cure_unplaced")
    inv = R("l11_invariants")
    # ---- THE CANONICAL BASIS. Everything numeric below reads from here. ----
    # See the "ONE QUANTITY, ONE NUMBER" block at the top of this file. Both
    # exporters call these three functions; nothing recomputes a fulfilment
    # figure locally any more.
    CLOSE = plant_closures(run, month)
    Q = plan_quantities(run, month)
    psc, _l10res = press_shift_cure(run, month, CLOSE)
    _closed_h = {p: 24.0 * len(CLOSE[p]) for p in PLANTS}
    H = ndays * 24                      # calendar hours in the plant month
    AVAIL = {p: float(H - _closed_h[p]) for p in PLANTS}   # hours that exist
    print("  PLANT CLOSURE  (derived from the plan; the run's own log beside it)")
    for p in PLANTS:
        print(f"     {p}  closed plant-days {CLOSE[p] or '(none)'}  "
              f"-> available {ndays * 24 - _closed_h[p]:,.0f} h of {ndays * 24:,} h")
    print(f"     run log: {CLOSE['_log'] or '(no PLANT HOLIDAYS line in the log)'}")
    _logdays = sorted({int(tok.split('-')[2][:2]) for tok in
                       CLOSE['_log'].replace('|', ' ').split()
                       if tok.count('-') == 2 and tok[:4].isdigit()})
    if _logdays and sorted(set(CLOSE["PCR"]) | set(CLOSE["TBR"])) != _logdays:
        print(f"     !! the plan and the log DISAGREE about the closure "
              f"(plan {sorted(set(CLOSE['PCR']) | set(CLOSE['TBR']))} vs log "
              f"{_logdays}) -- the plan is what ships and is what is used here")
    for p in PLANTS:
        _r = _l10res.get(p, {})
        if _r and (abs(_r["nameplate_here"] - _r["nameplate_L10"]) > 0.5
                   or abs(_r["hours_here"] - _r["hours_L10"]) > 0.5):
            print(f"     !! {p} press-shift segmentation disagrees with L10's "
                  f"cure_by_shift: nameplate {_r['nameplate_here']:,.0f} vs "
                  f"{_r['nameplate_L10']:,.0f}, hours {_r['hours_here']:,.0f} "
                  f"vs {_r['hours_L10']:,.0f}")
    # `press` in every cure artefact is the MES `wcID`. The plant's own press
    # NUMBER lives in wcmaster (`warehouse/derived/wc_master.parquet`, keyed by
    # `wc_id` = wcID, measured 175/175 in scripts/ingest_wcmaster.py). The two
    # namespaces have ZERO overlap -- 0 of 165 -- so a supervisor holding the
    # BTP workbook (press 4806) could not match a single row of this pack
    # (wc_id 120). Both identifiers now ship on every cure sheet and the
    # crosswalk itself ships as sheet 10.
    press_no: dict[str, str] = {}
    _wcp = ROOT / "warehouse" / "derived" / "wc_master.parquet"
    if _wcp.exists():
        for _r in pl.read_parquet(_wcp).iter_rows(named=True):
            if _r.get("press_no"):
                press_no[str(_r["wc_id"])] = str(_r["press_no"])
    req = pl.read_parquet(ROOT / "warehouse" / "derived" /
                          f"net_requirement_{month}.parquet")
    dem = pl.read_parquet(ROOT / "masters" / "demand" / f"demand_{month}.parquet")
    sz = pl.read_parquet(paths.INPUT_DERIVED / "gt_size.parquet")
    rim = {r["gt_code"]: str(r["rim"]) for r in sz.iter_rows(named=True)
           if r.get("gt_code") and r.get("rim")}

    # ---- HUMAN-READABLE DESCRIPTION per GT --------------------------------
    # A GT is a green-tyre spec and maps to MANY finished SKUs, so there is no
    # single "the" description. We show the MODAL SKU's description -- the one
    # the largest share of that GT's volume was actually sold as -- and say how
    # many other SKUs share the GT, e.g. "145 R12 ULTIMA XPC TL TATA (+14 SKUs)".
    # Source: warehouse/derived/gt_sku_from_recipe.parquet, built from the CURING
    # RECIPE (every cured tyre carries recipeID and that recipe's SAPMaterialCode
    # IS the finished SKU). 100 % of planned GTs resolve on both July and August,
    # so nothing is fabricated and nothing is blank.
    desc, modal_sku, n_sku, tsize = {}, {}, {}, {}
    _gsf = ROOT / "warehouse" / "derived" / "gt_sku_from_recipe.parquet"
    if _gsf.exists():
        for r in pl.read_parquet(_gsf).iter_rows(named=True):
            k = (r["plant"], r["gt_code"])
            n = int(r.get("n_skus") or 1)
            d = (r.get("sku_desc") or "").strip()
            desc[k] = f"{d} (+{n - 1} SKUs)" if d and n > 1 else d
            modal_sku[k] = r.get("sku_code") or ""
            n_sku[k] = n
            tsize[k] = r.get("tyre_size") or ""

    def add_desc(df: pl.DataFrame) -> pl.DataFrame:
        """gt_description / modal_sku / n_skus, keyed on (plant, gt_code)."""
        key = pl.concat_str([pl.col("plant"), pl.col("gt_code")], separator="|")
        mk = lambda d: {f"{a}|{b}": v for (a, b), v in d.items()}  # noqa: E731
        return df.with_columns([
            key.replace_strict(mk(desc), default="").alias("gt_description"),
            key.replace_strict(mk(modal_sku), default="").alias("modal_sku"),
            key.replace_strict(mk(n_sku), default=None,
                               return_dtype=pl.Int64).alias("n_skus"),
            key.replace_strict(mk(tsize), default="").alias("tyre_size"),
        ])

    # ---- CARRY-OUT: which rows are NOT this month's work -------------------
    # PLANNER_CARRY_OUT lets an L5 campaign START inside the horizon and FINISH
    # outside it, so a July plan legitimately contains rows dated into August.
    # That is correct planning and misleading presentation: a supervisor reading
    # "12-08" on a July sheet has no way to know it is deliberate. Every affected
    # row now carries `carry_out = true` and `plant_day > month length`.
    n_month_days = (datetime(y + (m == 12), (m % 12) + 1, 1) - datetime(y, m, 1)).days
    month_end = t0 + timedelta(days=n_month_days)
    try:
        _CC = pl.read_parquet(run / "cure_campaigns.parquet")
    except Exception:                                          # noqa: BLE001
        _CC = None
    sheets: dict[str, pl.DataFrame] = {}

    # ---- 0. SETTINGS -- the caps actually in force for THIS run -----------
    # Requested by the plant after 300/150 vs 150/70 vs the slice size were
    # confused for one another. They are THREE DIFFERENT OBJECTS at three
    # different levels, so each row says plainly what its number acts on.
    from planner.config import CONFIG, GT_SHELF_LIFE_H  # noqa: E402
    import planner.cmbc.l7_pull_release as L7           # noqa: E402
    th = CONFIG.thresholds
    try:
        _pr = json.loads((ROOT / "masters" /
                          f"press_list_{month}.json").read_text())
        roster = {k: len(v) for k, v in _pr.items()}
    except Exception:                                   # noqa: BLE001
        roster = {}
    try:
        _co = (pl.read_parquet(ROOT / "warehouse" / "derived" / "cap_changeover.parquet")
               .group_by(["plant", "machine_type", "same_min", "diff_min"])
               .agg(pl.len().alias("n")).sort(["plant", "machine_type"]))
        co_txt = " · ".join(
            f"{r['plant']} {r['machine_type']} x{r['n']}: same {r['same_min']} / "
            f"diff {r['diff_min']} min" for r in _co.iter_rows(named=True))
    except Exception:                                   # noqa: BLE001
        co_txt = "PCR BJ 28/60 · PCR CONTI 22/42 · TBR 10/24 (master fallback)"
    try:
        import planner.cmbc.l5_cure_master as L5        # noqa: E402
        early_stock, floor_basis = L5.EARLY_STOCK, L5.FLOOR_BASIS
    except Exception:                                   # noqa: BLE001
        early_stock, floor_basis = "(unavailable)", "(unavailable)"

    # The opening-GT file the PLAN used, resolved exactly as L4/L5/L7 resolve it.
    _ogov = os.environ.get("PLANNER_OPENING_GT", "").strip()
    _OGF = (Path(_ogov) if (_ogov and Path(_ogov).is_absolute())
            else (ROOT / "masters" / "opening_gt" / _ogov) if _ogov
            else ROOT / "masters" / "opening_gt" / f"opening_gt_{month}.parquet")
    _OG = (pl.read_parquet(_OGF) if _OGF.exists()
           else pl.DataFrame(schema={"plant": pl.Utf8, "age_h": pl.Float64}))

    S = lambda k, v, acts, src, note: {                  # noqa: E731
        "setting": k, "value": v, "acts_on": acts, "lives_in": src, "what_it_does": note}
    cfg = ROOT / "planner" / "config.py"
    l7f = ROOT / "planner" / "cmbc" / "l7_pull_release.py"
    settings = [
        # --- THE THREE NUMBERS THAT GET CONFUSED, stated plainly -------------
        S("min_demand_units", f"PCR {th.min_demand_units['PCR']} / TBR {th.min_demand_units['TBR']}",
          "DEMAND (per GT, per month)", "planner/config.py:265",
          "A GT whose WHOLE-MONTH demand is below this is not planned at all. "
          "It is routed to the residual policy, not built. ON for this run."),
        S("min_lot_units (B12 lot floor)", f"PCR {th.min_lot_units['PCR']} / TBR {th.min_lot_units['TBR']}",
          "RUN (one continuous same-GT block on one machine)", "planner/config.py:248",
          "Smallest economic build RUN. Judge it on sheet 1b_build_runs, NEVER on "
          "sheet 1 -- a sheet-1 row is a slice, not a run."),
        S("build slice size", "emergent, not a setting", "SLICE (one delivery to a press)",
          "planner/cmbc/l7_pull_release.py:866",
          "A slice is a JIT DELIVERY of green tyres to a press. It has NO minimum "
          "by design. n comes from R5 and B12: n >= (H+Q*cad)/71.7 and n <= Q/min_lot."),
        # --- run shaping ------------------------------------------------------
        S("SUBFLOOR_BUDGET", f"PCR {L7.SUBFLOOR_BUDGET['PCR']} / TBR {L7.SUBFLOOR_BUDGET['TBR']}",
          "RUN", "planner/cmbc/l7_pull_release.py:178",
          "How many below-floor setups are allowed, matched to the plant's own "
          "revealed rate. The floor is a BUDGET, not a gate -- the plant has no hard floor."),
        S("HARD_FLOOR mode", L7._HF, "RUN", "planner/cmbc/l7_pull_release.py:183",
          "'budget' = plant-calibrated allowance · '1' = absolute gate · 'off' = no floor."),
        S("RUN_MULT", L7.RUN_MULT, "RUN", "planner/cmbc/l7_pull_release.py:144",
          "Run-size target as a multiple of the lot floor."),
        S("SLICE_MULT", f"PCR {L7.SLICE_MULT['PCR']} / TBR {L7.SLICE_MULT['TBR']}",
          "SLICE", "planner/cmbc/l7_pull_release.py:98",
          "0 = the DERIVED R5/B12 slice rule. TBR 3.0 = legacy arm, kept because "
          "the derived rule costs TBR 8.67 points (EXPERT_AUDIT)."),
        S("SLICE_AGGR", L7.SLICE_AGGR, "SLICE", "planner/cmbc/l7_pull_release.py:112",
          "Where to sit in the legal window [n_R5, n_B12]. 1.0 = smallest legal slice."),
        S("LOT_INTERVAL_H (T)", L7.LOT_INTERVAL_H, "RUN", "planner/cmbc/l7_pull_release.py:282",
          "Replenishment interval. Q_g = r_g x T. THIS IS THE GT-INVENTORY DIAL: "
          "I = sum_g r_g (T/2 + tau*)."),
        S("SPAN_MULT", L7.SPAN_MULT, "RUN", "planner/cmbc/l7_pull_release.py:152",
          "How many intervals of cure demand one run may absorb. 99 = effectively off, "
          "so R5 is the outer bound."),
        # --- hard limits ------------------------------------------------------
        S("R5 GT shelf life", f"{GT_SHELF_LIFE_H:.0f} h", "TYRE (hard)",
          "planner/config.py:325 (GT_SHELF_LIFE_H)",
          "A green tyre not cured within this is scrap. Checked per slice in L7 and per tyre in L11."),
        S("R17 tau release floor", os.environ.get("PLANNER_TAU_RELEASE", "min"), "SLICE",
          "planner/cmbc/l7_pull_release.py:318",
          "'min' = physical floor tau_min. 'star' restores the old tau*-as-a-wall (cost 7.2 pt)."),
        S("gt_wip_rail", f"PCR {th.gt_wip_rail['PCR']:,} / TBR {th.gt_wip_rail['TBR']:,}",
          "PLANT GT STOCK (hard placement refusal)", "planner/config.py:192",
          "Runaway rail on the daily-mean GT stock. A placement breaching it is refused."),
        S("gt_wip_rail_margin", th.gt_wip_rail_margin, "PLANT GT STOCK",
          "planner/config.py:206",
          f"Rail is checked at {th.gt_wip_rail_margin:.0%} of the stated cap so the STATED "
          "cap survives post-plan reconciliation."),
        S("G8 GT inventory band", f"PCR {th.gt_wip_min['PCR']:,}-{th.gt_wip_max['PCR']:,} / "
          f"TBR {th.gt_wip_min['TBR']:,}-{th.gt_wip_max['TBR']:,}",
          "PLANT GT STOCK (reported, not enforced)", "planner/config.py:168",
          "Plant's stated steady-state band. A DETECTOR in L11, never a controller."),
        # --- assignment -------------------------------------------------------
        S("HARD_LOCK (rim lock)", L7.HARD_LOCK, "MACHINE", "planner/cmbc/l7_pull_release.py:187",
          "A GT may only build on its rim's locked machines. Measured, priced, kept."),
        S("HARD_PIN", L7.HARD_PIN, "MACHINE", "planner/cmbc/l7_pull_release.py:210",
          "Try the GT's own partitioned machine first."),
        S("PARTITION_PLANTS", ",".join(sorted(L7.PARTITION_PLANTS)), "MACHINE",
          "planner/cmbc/l7_pull_release.py PARTITION_PLANTS",
          "Static GT->machine partition applied to these plants. PCR only: TBR fails the two-month gate on BUILT (-168 Jul / +174 Aug)."),
        S("MACH_UTIL_CAP", L7.MACH_UTIL_CAP, "MACHINE", "planner/cmbc/l7_pull_release.py:117",
          "A GT stays on one machine until that machine reaches this occupancy."),
        S("CAD_BASIS", L7.CAD_BASIS, "MACHINE", "planner/cmbc/l7_pull_release.py:129",
          "'machine' = each machine's own cadence. 'plant' is the flat-cadence bug, A/B only."),
        S("B16 criterion", os.environ.get("PLANNER_B16_CRITERION", "coverage"),
          "MACHINE (TBR TT/TL split)", "planner/cmbc/l2_capability.py:277",
          "How TBR machines are split into tube-type / tubeless groups. No spill across groups."),
        # --- horizon / stock --------------------------------------------------
        S("EARLY_STOCK", early_stock, "CURE START",
          "planner/cmbc/l5_cure_master.py:76",
          "Lets a GT holding enough opening stock start curing at t0 instead of idling "
          "~11 h. Worth +2.3 pt."),
        S("L5 floor basis", floor_basis, "CURE START",
          "planner/cmbc/l5_cure_master.py:83",
          "'star' = tau* + build_band day-1 cure floor. Applies only where a GT has no "
          "opening stock to bridge the gap."),
        S("CARRY_OUT", os.environ.get("PLANNER_CARRY_OUT", "1") != "0", "HORIZON",
          "planner/cmbc/l5_cure_master.py:51",
          "A campaign starting inside the month may finish outside it; the tail is next "
          "month's opening state. This is why sheet 7 has carry_out days."),
        S("lookahead_days", os.environ.get("PLANNER_LOOKAHEAD_DAYS", "0"), "DEMAND HORIZON",
          "planner/cmbc/l4_net_requirement.py:103",
          "Next-month demand pulled into this month's build. 0 = off, so nothing pulls "
          "build on the last days -> the month-end WIP collapse."),
        # --- masters ----------------------------------------------------------
        S("changeover minutes", co_txt, "MACHINE (setup cost)",
          "masters/Master_Building_ChangeoverTime_*.csv",
          "Setup time IS changeover time and it is SIZE-DEPENDENT. Never hardcode these."),
        S("press roster", "  ".join(f"{k} {v}" for k, v in sorted(roster.items())) or "(none)",
          "PRESS", f"masters/press_list_{month}.json",
          "Presses available this month. July's roster is reused for August unless the "
          "plant sends a new one."),
        S("opening GT stock", f"PCR {int(_OG.filter(pl.col('plant')=='PCR').height):,} / "
          f"TBR {int(_OG.filter(pl.col('plant')=='TBR').height):,}"
          + (f"  (assumed age {float(_OG['age_h'].median()):.0f} h)" if _OG.height else ""),
          "OPENING STATE", f"masters/opening_gt/{_OGF.name}",
          "Green tyres on the floor at 07:00 on day 1, with their ages. Netted off BUILD, never off cure."),
        S("seed", CONFIG.seed, "DETERMINISM", "planner/config.py",
          "Same inputs give a byte-identical plan."),
        # --- what the numbers in this pack mean -------------------------------
        S("PLANT CLOSURE (rule G3)",
          " · ".join(f"{p}: days {CLOSE[p] or 'none'}" for p in PLANTS)
          + f"   [run log: {CLOSE['_log'] or 'no line'}]",
          "CALENDAR", "derived from the plan; cross-checked against the run log",
          "A closed plant-day is NOT available capacity. Machine and press "
          "occupancy on sheets 5/6 divide by "
          + " / ".join(f"{p} {AVAIL[p]:,.0f} h" for p in PLANTS)
          + f", not {H:,} h. Dividing August by 744 h understated occupancy "
          "2.7 pt and once booked 3,576 press-hours onto 2026-08-15 at zero "
          "output."),
        S("FULFILMENT DEFINITION", " · ".join(
            f"{p} {Q[p]['numerator']:,.0f} / {Q[p]['denominator']:,.0f} "
            f"= {Q[p]['pct']:.1f}%" for p in PLANTS),
          "THE GRADED NUMBER", "planner/cmbc/l11_validate_plan.py '--- demand ---'",
          "numerator = tyres fed in-month, per GT capped at that GT's own "
          "gross_build; denominator = cure_requirement over non-residual, "
          "non-lookahead GTs. This EXACT ratio is what 9b grades and is now "
          "printed identically in 9a, 9c and both BTP workbooks. Per-GT parts "
          "are columns on 8_demand_vs_plan."),
        S("THE SIX CURE QUANTITIES", " · ".join(
            f"{p}: nameplate {Q[p]['nameplate']:,.0f} / fed {Q[p]['fed']:,.0f} "
            f"/ fed in-month {Q[p]['fed_in_month']:,.0f} / delivered in-month "
            f"{Q[p]['delivered_in_month']:,.0f} / cured in-month "
            f"{Q[p]['numerator']:,.0f}" for p in PLANTS),
          "READING THE PACK", "9c_quantity_bridge",
          "Five of these shipped in this pack under names that did not "
          "distinguish them, and the BTP workbook published the NAMEPLATE as "
          "cured output. They are all correct answers to different questions; "
          "9c_quantity_bridge is the ladder from any one to any other."),
        S("REPORT WINDOW / carry_out columns",
          f"days 1-{ndays}, cut at {_mend:%Y-%m-%d %H:%M}",
          "EVERY SHEET", "this exporter, at load",
          "Build is cut WHOLE-SLICE on end_ts, so every `carry_out` column in "
          "this pack is structurally False. The clipped slices are NOT lost -- "
          "they are exported in 1c_build_after_month_end ("
          + " · ".join(f"{p} {Q[p]['built_after_end']:,.0f}" for p in PLANTS)
          + " tyres)."),
    ]
    sheets["0_settings"] = pl.DataFrame(
        settings, schema={"setting": pl.Utf8, "value": pl.Utf8, "acts_on": pl.Utf8,
                          "lives_in": pl.Utf8, "what_it_does": pl.Utf8},
        strict=False)

    # ---- 1. BUILD SCHEDULE, shift-wise -----------------------------------
    b = bs.filter(pl.col("machine") != "OPENING_STOCK")
    b = shift_cols(b, "start_ts", t0).with_columns([
        pl.col("gt_code").replace_strict(rim, default="(no rim)").alias("rim"),
        ((pl.col("end_ts") - pl.col("start_ts")).dt.total_seconds() / 3600.0)
        .round(2).alias("hours"),
    ])
    # Changeover flag: the machine's previous run was a different GT / rim.
    # SORT CHRONOLOGICALLY, EXPLICITLY. `shift(1).over(machine)` reads whatever
    # row order the frame happens to be in, and this used to run on a frame
    # sorted by (plant, date, shift, ...) where `date` was the WALL-CLOCK date.
    # Shift C spans midnight, so its post-midnight rows sorted into the NEXT
    # calendar date ahead of that date's A and B shifts: 297 PCR and 256 TBR
    # slices sat before a slice that started earlier. Each such inversion
    # invented a GT transition. Reported build changeovers were PCR 967 (true
    # 799, +21 %) and TBR 1,108 (true 800, +38 %). The true count is provable
    # independently: runs - machines = 810 - 11 = 799 and 809 - 9 = 800.
    b = b.sort(["plant", "machine", "start_ts"]).with_columns([
        pl.col("gt_code").shift(1).over(["plant", "machine"]).alias("_pg"),
        pl.col("rim").shift(1).over(["plant", "machine"]).alias("_pr"),
    ]).with_columns([
        (pl.col("gt_code") != pl.col("_pg")).fill_null(False).alias("changeover"),
        ((pl.col("gt_code") != pl.col("_pg")) & (pl.col("rim") != pl.col("_pr")))
        .fill_null(False).alias("size_change"),
        pl.col("_pg").alias("prev_gt"),
    ]).drop(["_pg", "_pr"])
    # A run_id can contain a >1 h gap (140 of 5,681 slices on July), so a naive
    # grouping would show a "run" that is not physically continuous. Split there.
    b = b.sort(["plant", "machine", "start_ts"]).with_columns(
        ((pl.col("start_ts") - pl.col("end_ts").shift(1).over("run_id"))
         .dt.total_seconds() / 3600.0).alias("_gap"))
    b = b.with_columns(
        ((pl.col("_gap") > 1.0) | pl.col("_gap").is_null()).cum_sum()
        .alias("_blk")).with_columns(
        (pl.col("run_id") + "#" + pl.col("_blk").cast(pl.Utf8)).alias("block_id"))
    b = b.with_columns([
        pl.col("qty").sum().over("block_id").alias("run_qty"),
        pl.col("qty").len().over("block_id").alias("_n"),
        (pl.col("start_ts").rank("ordinal").over("block_id")).alias("_i"),
    ]).with_columns(
        (pl.col("_i").cast(pl.Utf8) + " of " + pl.col("_n").cast(pl.Utf8))
        .alias("slice_i_of_n"))
    b = add_desc(b).with_columns(
        (pl.col("plant_day") > n_month_days).alias("carry_out"))
    sheets["1_build_schedule_shift"] = fmt_ts(b.select(
        ["plant", "date", "shift", "plant_day", "cal_date", "carry_out", "machine",
         "gt_code", "gt_description", "tyre_size", "rim", "modal_sku", "n_skus",
         "qty", "run_qty", "slice_i_of_n", "start_ts", "end_ts", "hours",
         "run_id", "block_id", "prev_gt", "changeover", "size_change",
         "press", "cure_ts", "wait_h"]),
        ["start_ts", "end_ts", "cure_ts"])

    # ---- 1b. RUN-LEVEL sheet -- the view a supervisor actually needs ------
    # One row per CONTINUOUS same-GT block on a machine. This is the object the
    # B12 lot floor acts on; the slice sheet above is the JIT feed to the press
    # and is NOT the right granularity for judging lot size.
    runs = (b.group_by(["plant", "machine", "block_id"]).agg([
        pl.col("gt_code").first(), pl.col("rim").first(),
        pl.col("qty").sum().alias("run_qty"),
        pl.len().alias("slices"),
        pl.col("start_ts").min().alias("start_ts"),
        pl.col("end_ts").max().alias("end_ts"),
        pl.col("date").first(), pl.col("shift").first(),
        pl.col("plant_day").first(),
        pl.col("changeover").any().alias("changeover"),
        pl.col("size_change").any().alias("size_change"),
        pl.col("prev_gt").first(),
    ]).with_columns(
        ((pl.col("end_ts") - pl.col("start_ts")).dt.total_seconds() / 3600.0)
        .round(2).alias("duration_h"))
        .sort(["plant", "start_ts", "machine"]))
    # ---- SETUP RESERVATION AUDIT, per run ---------------------------------
    # L7's `_place` only avoids INTERVAL OVERLAP on a machine; it never reserves
    # changeover time. Two different GTs can therefore sit exactly back-to-back
    # with a zero gap. These columns make that visible per run: what the plant
    # master says the changeover costs, what gap the plan actually left, and the
    # shortfall. See README section "known defect: changeover time is not
    # reserved". Reporting only -- nothing here changes the plan.
    _same = {r["machine"]: float(r["same_min"]) for r in
             pl.read_parquet(ROOT / "warehouse" / "derived" /
                             "cap_changeover.parquet").iter_rows(named=True)}
    _diff = {r["machine"]: float(r["diff_min"]) for r in
             pl.read_parquet(ROOT / "warehouse" / "derived" /
                             "cap_changeover.parquet").iter_rows(named=True)}
    runs = runs.with_columns([
        pl.col("end_ts").shift(1).over(["plant", "machine"]).alias("_pe"),
        pl.col("gt_code").shift(1).over(["plant", "machine"]).alias("_pg2"),
        pl.col("rim").shift(1).over(["plant", "machine"]).alias("_pr2"),
    ]).with_columns(
        ((pl.col("start_ts") - pl.col("_pe")).dt.total_seconds() / 60.0)
        .alias("gap_before_min"))
    runs = runs.with_columns(
        pl.when(pl.col("_pg2").is_null() | (pl.col("_pg2") == pl.col("gt_code")))
        .then(pl.lit(0.0))
        .when(pl.col("_pr2") == pl.col("rim"))
        .then(pl.col("machine").replace_strict(_same, default=22.0))
        .otherwise(pl.col("machine").replace_strict(_diff, default=42.0))
        .alias("setup_required_min"))
    runs = runs.with_columns(
        (pl.col("setup_required_min") - pl.col("gap_before_min").fill_null(1e9))
        .clip(lower_bound=0).round(1).alias("setup_shortfall_min"))
    runs = add_desc(runs).with_columns(
        (pl.col("plant_day") > n_month_days).alias("carry_out"))
    sheets["1b_build_runs"] = fmt_ts(runs.select(
        ["plant", "date", "shift", "plant_day", "carry_out", "machine",
         "gt_code", "gt_description", "tyre_size", "rim", "modal_sku", "n_skus",
         "run_qty", "slices", "start_ts", "end_ts", "duration_h",
         "changeover", "size_change", "prev_gt",
         "gap_before_min", "setup_required_min", "setup_shortfall_min",
         "block_id"]),
        ["start_ts", "end_ts"])

    # ---- 1c. BUILD CLIPPED OUT OF THIS REPORT -----------------------------
    # WHAT THIS DOES / WHY IT EXISTS -- a measured defect, found 2026-08-21.
    #   `bs` is cut at `end_ts <= month_end` when it is loaded, which makes the
    #   `carry_out` column on sheets 1, 1b, 2, 2b and 7 STRUCTURALLY False --
    #   six boolean columns that can never be true, and a KPI line that printed
    #   "BUILT after month end (carry-out) = 0" as a tautology rather than a
    #   measurement. The tyres are real: PCR 3,948 / TBR 546 on July and
    #   PCR 6,899 / TBR 940 on August. They were announced to stdout and then
    #   appeared in no exported file, so the pack could not be reconciled to
    #   `build_schedule.parquet` without re-reading the parquet.
    #   The clip itself is correct (PLANT RULING 2026-08-10: plan on month plus
    #   tail, REPORT only the month) and is NOT changed here. The rows are
    #   simply exported, in their own sheet, so the clip is visible.
    _late = bs_full.filter((pl.col("machine") != "OPENING_STOCK")
                           & (pl.col("end_ts") > pl.lit(_mend)))
    _late = shift_cols(_late, "start_ts", t0).with_columns([
        pl.col("gt_code").replace_strict(rim, default="(no rim)").alias("rim"),
        ((pl.col("end_ts") - pl.col("start_ts")).dt.total_seconds() / 3600.0)
        .round(2).alias("hours"),
        pl.lit("built after month end -- next month's green tyre. EXCLUDED "
               "from sheets 1/1b/2/5/6/7 and from every KPI, which is why "
               "their `carry_out` columns are all False.").alias("why_excluded"),
    ])
    sheets["1c_build_after_month_end"] = fmt_ts(
        add_desc(_late).select(
            ["plant", "date", "shift", "plant_day", "cal_date", "machine",
             "gt_code", "gt_description", "tyre_size", "rim", "qty",
             "start_ts", "end_ts", "hours", "press", "cure_ts", "wait_h",
             "run_id", "why_excluded"])
        if _late.height else
        pl.DataFrame([{"plant": "(none)", "date": "", "shift": "",
                       "plant_day": None, "cal_date": "", "machine": "",
                       "gt_code": "", "gt_description": "", "tyre_size": "",
                       "rim": "", "qty": None, "start_ts": None, "end_ts": None,
                       "hours": None, "press": "", "cure_ts": None,
                       "wait_h": None, "run_id": "",
                       "why_excluded": "no build slice finishes after month "
                                       "end in this plan"}]),
        ["start_ts", "end_ts", "cure_ts"])

    # ---- 2b. CURE CAMPAIGNS -- one row per campaign (what sheet 2 used to be)
    # A campaign is a MULTI-WEEK object: PCR p50 192.6 h, TBR p50 255.9 h. Dating
    # it by its start and calling the sheet "shift-wise" put up to 638 h of press
    # work on a single row labelled "01 Jul, shift A". Two consequences, both
    # reported from the floor: (a) the sheet appeared to stop on 30 Jul because
    # no campaign STARTS on the 31st, while the build sheet ran to the 31st;
    # (b) its `qty` sums to 485,647 -- the campaign NAMEPLATE -- against 475,271
    # tyres actually fed, so the pack read as if it cured 10,376 tyres it never
    # made. The campaign view is genuinely useful, so it is kept here, correctly
    # named and dated, and now carries the fed/unfed split. Sheet 2 below is the
    # real shift-wise cure schedule.
    camp = shift_cols(cc, "start_ts", t0).with_columns([
        pl.col("gt_code").replace_strict(rim, default="(no rim)").alias("rim"),
        (pl.col("plant") + "-" + pl.col("press") + "-" +
         pl.col("start_ts").dt.strftime("%m%d%H%M")).alias("campaign_id"),
    ]).sort(["plant", "start_ts", "press"])
    _rec = run / "cure_campaigns_reconciled.parquet"
    if _rec.exists():
        camp = camp.join(pl.read_parquet(_rec).select(
            ["plant", "gt_code", "press", "start_ts", "qty_fed", "qty_unfed"]),
            on=["plant", "gt_code", "press", "start_ts"], how="left")
    for _c in ("qty_fed", "qty_unfed"):
        if _c not in camp.columns:
            camp = camp.with_columns(pl.lit(None, pl.Float64).alias(_c))
    camp = add_desc(camp).with_columns([
        (pl.col("plant_day") > n_month_days).alias("carry_out"),
        (pl.col("end_ts") > pl.lit(month_end)).alias("finishes_next_month"),
    ]).rename({"qty": "qty_planned", "date": "start_date", "shift": "start_shift"})
    _sheet2b = fmt_ts(camp.select(
        ["plant", "start_date", "start_shift", "plant_day", "carry_out",
         "finishes_next_month", "press", "gt_code", "gt_description",
         "tyre_size", "rim", "modal_sku", "n_skus", "mould_set",
         "qty_planned", "qty_fed", "qty_unfed",
         "start_ts", "end_ts", "hours", "campaign_id"]),
        ["start_ts", "end_ts"])

    # ---- 2. CURE SCHEDULE, genuinely shift-wise --------------------------
    # ONE ROW PER (plant, plant_day, shift, press, GT) -- the same shape as the
    # build sheet, which is what the shop floor reads. Built by
    # `press_shift_cure()` (top of this file), which both exporters share.
    #
    # THREE QUANTITIES, THREE QUESTIONS. Read the one that answers yours; the
    # ladder between them is sheet `9c_quantity_bridge`.
    #   qty          tyres DELIVERED to this press in this shift whose cure
    #                completes in-month. Bucketed on the build slice's
    #                `cure_ts`, so it is SPIKY BY CONSTRUCTION -- p50 0, and
    #                65 % of rows read 0 while the press runs, because the
    #                press is still working through a delivery it took two
    #                shifts ago. It is the RECONCILING column: sheet 2 ->
    #                sheet 6 -> sheet 7 -> `delivered_in_month` at diff 0.
    #                Do arithmetic on this one.
    #   qty_run      the campaign's IN-MONTH FED quantity spread pro-rata over
    #                the press hours it actually holds in this shift. p50 ~55
    #                on PCR, which is what a press-shift physically makes.
    #                SUMS PER PLANT EXACTLY TO `fed_in_month` -- one documented
    #                step (the per-GT over-requirement cap) from the fulfilment
    #                numerator. The old column claimed in its header to sum to
    #                `qty` and actually summed to the WHOLE-PLAN fed quantity,
    #                7,559 tyres of next month printed on August PCR rows.
    #   qty_planned  the campaign NAMEPLATE's share of this shift -- PRESS
    #                CAPACITY SEATED, not tyres. Sums to `nameplate_in_month`
    #                and matches L10's own `cure_by_shift.parquet` to the tyre.
    #                NEVER quote it as output; that is defect 1.
    c = psc.with_columns([
        (pl.lit(t0.date()) + pl.duration(days=pl.col("plant_day") - 1))
        .dt.strftime("%Y-%m-%d").alias("date"),
        pl.col("gt_code").replace_strict(rim, default="(no rim)").alias("rim"),
        (pl.col("plant_day") > n_month_days).alias("carry_out"),
        pl.col("press").replace_strict(press_no, default="(unmapped)")
        .alias("press_no"),
    ])
    c = add_desc(c).sort(["plant", "plant_day", "shift", "press", "start_ts"])
    sheets["2_cure_schedule_shift"] = fmt_ts(c.select(
        ["plant", "date", "shift", "plant_day", "carry_out", "press",
         "press_no", "gt_code", "gt_description", "tyre_size", "rim",
         "modal_sku", "n_skus", "mould_set", "qty", "qty_run", "qty_planned",
         "start_ts", "end_ts", "hours", "segments", "campaigns",
         "campaign_id"]).with_columns([
             pl.col("qty_run").round(1), pl.col("qty_planned").round(1),
             pl.col("hours").round(3)]),
        ["start_ts", "end_ts"])
    sheets["2b_cure_campaigns"] = _sheet2b
    # PROVE all three columns tie to the canonical basis, per plant. The old
    # sheet 2 was only ever cross-checked against sheet 6, which was built from
    # the SAME campaign frame -- so the pair agreed with each other and
    # disagreed with the plan, and nothing caught it. Each column is now tied to
    # a DIFFERENT number that came from somewhere else.
    print("  RECONCILIATION  sheet 2, each column against its own basis")
    for _p in PLANTS:
        _z = c.filter(pl.col("plant") == _p)
        for _col, _want, _lbl in (
                ("qty", Q[_p]["delivered_in_month"], "delivered in-month"),
                ("qty_run", Q[_p]["fed_in_month"], "fed in-month (L7)"),
                ("qty_planned", Q[_p]["nameplate"] * 0 +
                 float(_l10res.get(_p, {}).get("nameplate_L10", 0.0)),
                 "nameplate in-month (L10 cure_by_shift)")):
            _got = float(_z[_col].sum())
            print(f"     {_p}  {_col:<12} {_got:>10,.0f}  vs {_lbl:<38}"
                  f"{_want:>10,.0f}   diff {_got - _want:>6,.0f}   "
                  f"{'OK' if abs(_got - _want) < 0.5 else '!! MISMATCH'}")


    # ---- 3. mould changes -------------------------------------------------
    _mc = pl.DataFrame()
    _f = run / "mould_changes.parquet"
    if _f.exists():
        _mc = pl.read_parquet(_f)
        d = shift_cols(_mc, "start_ts", t0).with_columns(
            pl.col("press").cast(pl.Utf8)
            .replace_strict(press_no, default="(unmapped)").alias("press_no"))
        sheets["3_mould_changes"] = fmt_ts(d, ["start_ts", "end_ts"])

    # ---- 4. CREW LOAD -- on TRUE CONCURRENCY, not on the starting shift ----
    # WHAT THIS DOES / WHY IT EXISTS -- a measured defect, found 2026-08-21.
    #   L10's `crew_load.parquet` charges every mould change wholly to the shift
    #   it STARTS in. A mould change runs 210-430 min against a 480 min shift
    #   and 141 of 184 July changes (153 of 197 August) cross a boundary, so
    #   the sheet understated the shifts the work runs INTO and printed no row
    #   at all for a shift whose only work was in progress:
    #
    #     plant-month   shifts touched by a change   shifts with a row (L10)
    #     Jul PCR                              52                        44
    #     Jul TBR                              55                        42
    #     Aug PCR                              49                        42
    #     Aug TBR                              45                        31
    #
    #   Worked example, PCR 16 July shift B: L10 says 2 changes / 6 fitters.
    #   11 changes overlap that window and 10 are concurrent at 15:00, which
    #   needs 30 fitters. A roster built on the L10 column is short by 24.
    #
    #   FIX, and it is a REPORTING fix -- the plan is untouched. Every column
    #   is derived from `mould_changes.parquet`'s own start/end/crew:
    #     changes_active   changes overlapping this shift at all
    #     changes_started  the L10 basis, KEPT so the two are comparable
    #     fitters_peak     max concurrent crew at any instant in the shift --
    #                      THE ROSTERING NUMBER
    #     fitters_L10      L10's `fitters` column, KEPT and labelled
    #     fitter_hours     crew x overlap hours, the workload not the peak
    #   A shift with work in progress and no change starting now gets a row.
    if _mc.height:
        _cl10 = {}
        _f2 = run / "crew_load.parquet"
        if _f2.exists():
            for r in pl.read_parquet(_f2).iter_rows(named=True):
                _cl10[(r["plant"], int(r["day"]), r["shift"])] = (
                    int(r["changes"]), int(r["fitters"]))
        _rows = []
        for p in PLANTS:
            M = [r for r in _mc.filter(pl.col("plant") == p).iter_rows(named=True)]
            # A MOULD CHANGE PAUSES OVER A PLANT CLOSURE, exactly as a cure
            # campaign does, so its wall-clock span is 24 h longer than the
            # fitter time it consumes. Charging the span booked 216 fitter-hours
            # on August TBR that nobody works, and put crew on a shut day. The
            # change is cut against the closure windows first and everything
            # below reads the OPEN intervals only.
            wins_p = CLOSE.get("_windows", {}).get(p, [])
            for r in M:
                r["_open"] = _open_parts(r["start_ts"], r["end_ts"], wins_p)
            for k in range(3 * n_month_days):
                w0 = t0 + timedelta(hours=8 * k)
                w1 = w0 + timedelta(hours=8)
                act = [r for r in M
                       if any(a < w1 and bnd > w0 for a, bnd in r["_open"])]
                started = [r for r in M if w0 <= r["start_ts"] < w1]
                if not act and not started:
                    continue
                # peak concurrency: crew is a step function, so it can only
                # change where an open interval begins -- evaluate there and
                # nowhere else. Sampling the shift boundary alone is what
                # produced the start-shift bug in the first place.
                pts = sorted({w0} | {a for r in act for a, _bnd in r["_open"]
                                     if w0 < a < w1})
                peak = 0
                for tpt in pts:
                    peak = max(peak, sum(
                        int(r["crew"]) for r in act
                        if any(a <= tpt < bnd for a, bnd in r["_open"])))
                fh = sum(int(r["crew"]) * max(
                    (min(bnd, w1) - max(a, w0)).total_seconds(), 0.0) / 3600.0
                    for r in act for a, bnd in r["_open"])
                l10 = _cl10.get((p, k // 3 + 1, "ABC"[k % 3]), (0, 0))
                _rows.append({
                    "plant": p, "day": k // 3 + 1,
                    "date": (t0 + timedelta(days=k // 3)).strftime("%Y-%m-%d"),
                    "shift": "ABC"[k % 3],
                    "changes_active": len(act), "changes_started": len(started),
                    "changes_L10": l10[0],
                    "fitters_peak": peak, "fitters_L10": l10[1],
                    "fitter_hours": round(fh, 2),
                    "understated_by": peak - l10[1],
                    "basis": ("in-progress only -- L10 emitted no row for this "
                              "shift" if not started else ""),
                })
        crew = pl.DataFrame(_rows).sort(["plant", "day", "shift"])
        sheets["4_crew_load"] = crew
        print("  CREW LOAD  true concurrency vs L10's start-shift basis")
        for p in PLANTS:
            z = crew.filter(pl.col("plant") == p)
            print(f"     {p}  shifts with a change ACTIVE {z.height:>3} · L10 "
                  f"rows {int((z['changes_L10'] > 0).sum()):>3} · shifts L10 "
                  f"missed entirely {int((z['changes_L10'] == 0).sum()):>3} · "
                  f"peak fitters {int(z['fitters_peak'].max())} (L10 max "
                  f"{int(z['fitters_L10'].max())}) · shifts understated "
                  f"{int((z['understated_by'] > 0).sum())}")

    # ---- 5/6. machine and press summaries --------------------------------
    # AVAILABLE HOURS ARE NOT CALENDAR HOURS. A closed plant-day (rule G3) is
    # not idle capacity, and dividing August by 744 h counted the 24 h the
    # plant was shut: PCR occupancy read 78.9 % against a true 81.6 %, 2.7 pt
    # understated, on a sheet whose whole job is to say how loaded the plant
    # is. `avail_h` is per plant because a closure can be per plant.
    _av = pl.col("plant").replace_strict(AVAIL, default=float(H))
    # CHANGEOVERS ARE COUNTED ON RUNS, NOT ON SLICES -- and both are shown.
    # This sheet used to count slice rows whose GT differs from the previous
    # slice (668 on July PCR) while `11_changeover_by_machine`, whose header
    # claims the two "RECONCILE BY CONSTRUCTION", counted setup BLOCKS (660).
    # They are different objects at different grains and neither was labelled.
    # A changeover is a SETUP: one per continuous same-GT block that follows a
    # different GT. That is the run-level count, so that is `changeovers`, and
    # the slice-level count is kept beside it under a name that says what it is.
    _co_run = (runs.group_by(["plant", "machine"]).agg(
        pl.col("changeover").sum().cast(pl.Int64).alias("changeovers"),
        pl.col("size_change").sum().cast(pl.Int64).alias("size_changes"),
        pl.len().cast(pl.Int64).alias("setup_blocks"))
        # THE EXACT IDENTITY, PER MACHINE:
        #   setup_blocks - 1 = changeovers + same_gt_restarts
        # The sheet-1 docstring states the independent check as
        # "runs - machines", which is only right when every block boundary is a
        # GT change. It is not: a block is split at a >1 h gap, so a machine can
        # restart the SAME GT after an idle hour and that is a resumption, not a
        # setup. July PCR: 735 block boundaries, 660 changeovers, 75 same-GT
        # restarts. Naming the third term makes the check exact instead of
        # approximately right.
        .with_columns((pl.col("setup_blocks") - 1 - pl.col("changeovers"))
                      .alias("same_gt_restarts")))
    mach = (b.group_by(["plant", "machine"]).agg(
        pl.col("qty").sum().alias("tyres"),
        pl.col("run_id").n_unique().alias("runs"),
        pl.col("gt_code").n_unique().alias("distinct_gts"),
        pl.col("changeover").sum().cast(pl.Int64).alias("slice_gt_transitions"),
        pl.col("size_change").sum().cast(pl.Int64).alias("slice_size_changes"),
        pl.col("hours").sum().round(1).alias("busy_h"))
        .join(_co_run, on=["plant", "machine"], how="left")
        .with_columns([
            _av.alias("avail_h"),
            (100 * pl.col("busy_h") / _av).round(1).alias("occupancy_pct"),
            (_av - pl.col("busy_h")).round(1).alias("idle_h")])
        .sort(["plant", "machine"]))
    # Setup the plan owes but never reserved on this machine's timeline.
    _sr = (runs.group_by(["plant", "machine"]).agg(
        (pl.col("setup_required_min").sum() / 60.0).round(1).alias("setup_required_h"),
        (pl.col("setup_shortfall_min").sum() / 60.0).round(1).alias("setup_unreserved_h"),
        (pl.col("setup_shortfall_min") > 0.001).sum().alias("runs_short_of_setup")))
    mach = (mach.join(_sr, on=["plant", "machine"], how="left")
            .with_columns([pl.col("setup_required_h").fill_null(0.0),
                           pl.col("setup_unreserved_h").fill_null(0.0),
                           pl.col("runs_short_of_setup").fill_null(0)])
            .with_columns((100 * (pl.col("busy_h") + pl.col("setup_unreserved_h")) / _av)
                          .round(1).alias("occupancy_pct_with_setup"))
            .select(["plant", "machine", "tyres", "runs", "setup_blocks",
                     "distinct_gts", "changeovers", "size_changes",
                     "same_gt_restarts", "slice_gt_transitions",
                     "slice_size_changes", "busy_h",
                     "avail_h", "occupancy_pct", "idle_h", "setup_required_h",
                     "setup_unreserved_h", "runs_short_of_setup",
                     "occupancy_pct_with_setup"]))
    sheets["5_machine_summary"] = mach
    # `tyres` is the FED figure so it reconciles with sheet 2 and the headline;
    # `tyres_planned` is the nameplate the presses were seated for. When it was
    # only the nameplate, this sheet agreed with the old sheet 2 and both were
    # 10,376 tyres above what the plan actually makes -- two sheets agreeing on
    # the same wrong number is exactly how that survived.
    press = (c.group_by(["plant", "press"]).agg(
        pl.col("qty").sum().alias("tyres"),
        pl.col("qty_run").sum().round(0).alias("tyres_fed_in_month"),
        pl.col("qty_planned").sum().round(0).alias("tyres_planned"),
        pl.col("gt_code").n_unique().alias("distinct_gts"),
        pl.col("hours").sum().round(1).alias("busy_h"))
        .join(camp.group_by(["plant", "press"]).agg(
            pl.len().alias("campaigns")), on=["plant", "press"], how="left")
        .with_columns([
            pl.col("press").replace_strict(press_no, default="(unmapped)")
            .alias("press_no"),
            (pl.col("tyres_planned") - pl.col("tyres")).alias("tyres_unfed"),
            _av.alias("avail_h"),
            (100 * pl.col("busy_h") / _av).round(1).alias("utilisation_pct"),
            (_av - pl.col("busy_h")).round(1).alias("idle_h"),
            pl.col("campaigns").fill_null(0)])
        .select(["plant", "press", "press_no", "tyres", "tyres_fed_in_month",
                 "tyres_planned", "tyres_unfed", "campaigns", "distinct_gts",
                 "busy_h", "avail_h", "utilisation_pct", "idle_h"])
        .sort(["plant", "press"]))
    sheets["6_press_summary"] = press

    # ---- 10. PRESS CROSSWALK -- the two namespaces, side by side ----------
    # WHAT THIS DOES / WHY IT EXISTS -- a measured defect, found 2026-08-21.
    #   The csv sheets carry the MES `wc_id` (e.g. `120`); the BTP-format
    #   workbooks carry the plant's own `press_no` from wcmaster (e.g. `4806`).
    #   They are the SAME PRESS -- verified by trace -- and 0 of 165 press
    #   identifiers overlap between the two namespaces. The crosswalk shipped
    #   in neither file, so a supervisor holding both could not match one row.
    #   It now ships as its own sheet AND as a column on sheets 2, 3 and 6, and
    #   the BTP curing workbook carries the wc_id beside its Machine column.
    _pk = sorted(set(c["press"].to_list()))
    sheets["10_press_crosswalk"] = pl.DataFrame([{
        "wc_id": k,
        "press_no": press_no.get(k, ""),
        "resolved": bool(press_no.get(k)),
        "source": ("warehouse/derived/wc_master.parquet `name` -> press_no, "
                   "keyed by iD = MES wcID (175/175, scripts/ingest_wcmaster.py)"
                   if press_no.get(k) else
                   "UNRESOLVED -- no wcmaster row for this wcID"),
        "appears_as_wc_id_in": "csv 2_cure_schedule_shift, 2b, 3, 6; "
                               "runs/*/cure_campaigns.parquet",
        "appears_as_press_no_in": "btp/optimizer_curing_schedule_full_*.xlsx "
                                  "`Machine`",
    } for k in _pk])
    print(f"  PRESS CROSSWALK  {sum(1 for k in _pk if press_no.get(k))}/"
          f"{len(_pk)} presses resolved wc_id -> plant press_no "
          f"(sheet 10_press_crosswalk)")

    # ---- 7. daily summary -------------------------------------------------
    # MUST RECONCILE WITH SHEET 1. Two bucketing defects have now been found here
    # and both silently DROPPED tyres, which is the worst failure mode for a
    # summary sheet -- it looks complete.
    #   1. bucketing by CALENDAR DATE lost day 31's C shift (968 + 765 tyres).
    #      Fixed by bucketing on `plant_day` (07:00 -> 07:00).
    #   2. looping `range(H // 24)` lost every row with plant_day > month length.
    #      Those are CARRY-OUT: L5 places campaigns that START inside the horizon
    #      and FINISH outside it (PLANNER_CARRY_OUT), so their feeding build
    #      slices legitimately land on plant_day 32..43. Measured on runs/f_solo:
    #      27 slices, 799 PCR + 734 TBR = 1,533 tyres, silently absent.
    # The loop now runs to the LAST plant_day present in the plan and flags
    # anything past the month end as carry-out, so sheet 7 sums to sheet 1 exactly.
    # THE REPORT IS THE MONTH. This loop used to run to the last plant_day in
    # the plan so carry-out rows (plant_day 32..43) still summed into sheet 7.
    # Under the extend ruling `bs`/`cc` are already cut at the boundary, so the
    # month IS the whole report and extending the loop only manufactures empty
    # day-32 rows. The GT balance below still walks the UNCLIPPED cure times --
    # a tyre held past the boundary must still drain, or day 31's closing stock
    # reads high.
    n_month_days = H // 24
    max_day = n_month_days
    H_ext = H
    rows = []
    for p in ("PCR", "TBR"):
        bp = bs.filter(pl.col("plant") == p)
        ivt = pl.concat([
            bp.select([pl.col("end_ts").alias("ts"), pl.col("qty").alias("d")]),
            bp.select([pl.col("cure_ts").alias("ts"), (-pl.col("qty")).alias("d")]),
        ]).sort("ts").with_columns(pl.col("d").cum_sum().alias("bal"))
        ts = np.array([(x - t0).total_seconds() / 3600.0 for x in ivt["ts"]])
        bal = np.array(ivt["bal"], float)
        idx = np.searchsorted(ts, np.arange(H_ext) + 0.5, side="right") - 1
        g = np.where(idx >= 0, bal[np.clip(idx, 0, len(bal) - 1)], 0.0)
        # THE STOCK AT 07:00 THE NEXT MORNING, sampled AT the boundary rather
        # than half an hour before it. `gt_inventory_close` samples hour k+0.5
        # and so misses everything credited in the last half hour of the plant
        # day: on July PCR it reads 2,200 on day 31 while the stock actually
        # handed to August is 4,357. Both are kept -- the sampled series is
        # what the G8 detector uses -- but the hand-off number is the exact one
        # and it is now a column instead of a KPI line nobody could tie to the
        # daily sheet.
        idxe = np.searchsorted(ts, np.arange(1, H_ext + 1) * 1.0,
                               side="right") - 1
        ge = np.where(idxe >= 0, bal[np.clip(idxe, 0, len(bal) - 1)], 0.0)
        bd = b.filter(pl.col("plant") == p)
        cd = shift_cols(bp.filter(pl.col("cure_ts") <= pl.lit(_mend)),
                        "cure_ts", t0)      # same basis as sheet 2
        # ---- CAPACITY REALITY CHECK ON THE DAILY CURE CURVE -----------------
        # `cured` buckets tyres on `cure_ts` with NO press-capacity constraint,
        # so it can publish a day the plant cannot physically run. Measured
        # 2026-08-21: August PCR day 3 reads 16,696, needing 2,389 press-hours
        # against a fleet of 86 x 24 = 2,064 -- 115.7 % of every press in the
        # plant. FOUR days exceed 100 % (16,696 / 15,482 / 15,001 / 14,749).
        # `cured_press_run` is the physical curve, max 14,286 = 99.1 %, and
        # never breaches. The PLAN is feasible; the COLUMN was not, and nothing
        # in the pack or in L11 graded it.
        # Denominator derived from THIS run's campaigns (tyres/press-hour x
        # roster x 24 h), never a constant -- a cap lives in config.py or
        # nowhere; this is a reporting yardstick, not an enforced limit.
        _caph = 0.0
        if _CC is not None:
            _ccp = _CC.filter(pl.col("plant") == p)
            if _ccp.height:
                _ph = float((_ccp["end_ts"] - _ccp["start_ts"])
                            .dt.total_seconds().sum()) / 3600.0
                if _ph > 0:
                    _caph = (float(_ccp["qty"].sum()) / _ph
                             * PRESS_ROSTER.get(p, 0) * 24.0)
        # THE DAILY CURVE, ON BOTH BASES AND LABELLED.
        # The BTP-format workbook drew a completely different day-1 (PCR Aug
        # 7,943 against this sheet's 11,974) because it prorated the campaign
        # NAMEPLATE over press hours while this sheet buckets DELIVERED tyres
        # on `cure_ts`. Two curves for the same plant-day, in one pack, neither
        # saying which it was. Both are now here, under their own names, and
        # both workbooks draw from these same three columns.
        cpd = (c.filter(pl.col("plant") == p)
               .group_by("plant_day").agg(pl.col("qty_run").sum().alias("r"),
                                          pl.col("qty_planned").sum().alias("n"),
                                          pl.col("hours").sum().alias("h")))
        cmap = {int(r["plant_day"]): r for r in cpd.iter_rows(named=True)}
        # green tyres CREDITED to the ledger that day = built + the opening
        # stock, which is credited whole at 07:00 on day 1. This is the
        # `gt_events` positive-delta basis the BTP building workbook draws as
        # `GT_Produced`; it is NOT sheet 1's `built` and the two differed by
        # 2,766 in the same workbook with no column saying so.
        opd = shift_cols(bs.filter((pl.col("plant") == p)
                                   & (pl.col("machine") == "OPENING_STOCK")),
                         "end_ts", t0) if bs.filter(
            (pl.col("plant") == p)
            & (pl.col("machine") == "OPENING_STOCK")).height else None
        for dnum in range(max_day):
            day = (t0 + timedelta(days=dnum)).strftime("%Y-%m-%d")
            cr = cmap.get(dnum + 1, {})
            _op = 0 if opd is None else int(
                opd.filter(pl.col("plant_day") == dnum + 1)["qty"].sum())
            _bt = int(bd.filter(pl.col("plant_day") == dnum + 1)["qty"].sum())
            rows.append({
                "plant": p, "plant_day": dnum + 1, "date": day,
                "carry_out": dnum + 1 > n_month_days,
                "closed": (dnum + 1) in CLOSE[p],
                "built": _bt,
                # ---- `cured` IS THE PHYSICAL CURVE. Swapped 2026-08-21. -----
                # It used to be the `cure_ts` bucketing, and that is an
                # INSTANT-STAMPING ARTEFACT, not a schedule. `cure_ts` stamps a
                # whole build slice at ONE moment, but a 154-tyre slice takes
                # ~22 press-hours to cure. Every tyre lands on the day the
                # stamp falls, and the 22 hours that spill into the next day
                # are counted today anyway.
                #
                # Measured, August PCR plant-day 3: 109 slices / 16,696 tyres
                # on only 73 distinct `cure_ts` instants across 81 presses.
                # The presses actually run 1,944 of 2,064 press-hours that day
                # = 13,589 tyres. 16,696 needs 115.7 % of every press in the
                # plant, and four other days were over 100 %.
                #
                # The press-run basis spreads each campaign over the hours its
                # press is actually occupied. It NEVER breaches the fleet on any
                # plant-month (max 99.0 %), and it sums to `fed_in_month` -- the
                # same basis the fulfilment numerator uses. So it is both
                # physical and consistent with the headline.
                #
                # The stamped series is KEPT, under a name that says what it is,
                # because sheets 2/6 and `verify_export.py` reconcile on it and
                # deleting information to make totals agree is how the three
                # "fulfilment numerators" happened.
                "cured": round(float(cr.get("r", 0.0)), 1),
                # Same series under its old name so every downstream reference
                # keeps working; `cured` is now the one a reader should use.
                "cured_press_run": round(float(cr.get("r", 0.0)), 1),
                "cured_gt_stamped_at_cure_ts": int(
                    cd.filter(pl.col("plant_day") == dnum + 1)["qty"].sum()),
                "cure_nameplate_seated": round(float(cr.get("n", 0.0)), 1),
                "press_hours": round(float(cr.get("h", 0.0)), 1),
                "press_capacity_tyres": round(_caph),
                "cured_pct_of_press_capacity": (
                    round(100.0 * float(cr.get("r", 0.0)) / _caph, 1)
                    if _caph else None),
                "stamped_pct_of_press_capacity": (
                    round(100.0 * int(cd.filter(pl.col("plant_day") == dnum + 1)["qty"].sum())
                          / _caph, 1) if _caph else None),
                "opening_gt_credited": _op,
                "gt_credited_incl_opening": _bt + _op,
                "gt_inventory_day_mean": round(float(g[dnum * 24:(dnum + 1) * 24].mean())),
                "gt_inventory_close": round(float(g[min((dnum + 1) * 24 - 1, H_ext - 1)])),
                "gt_stock_at_0700_next_day": round(float(ge[min((dnum + 1) * 24 - 1, H_ext - 1)])),
                "changeovers": int(bd.filter(pl.col("plant_day") == dnum + 1)["changeover"].sum()),
            })
    daily = pl.DataFrame(rows).sort(["plant", "plant_day"])
    sheets["7_daily_summary"] = daily
    # PROVE it ties. A summary that silently disagrees with its own detail sheet
    # is worse than no summary at all -- this refuses to be wrong quietly.
    # THE CURE SIDE OF THIS CHECK USED TO COMPARE THE WRONG PAIR. It set
    # `c1 = bs.qty.sum()` -- tyres fed inside the report window, 390,215 on
    # July PCR -- against sheet 7's `cured`, which is DELIVERED-and-cured
    # in-month, 385,858. The two differ by exactly the closing GT stock, so the
    # line printed `!! MISMATCH 4,357` on a pack that was internally
    # consistent, every month, for the whole project. A check that always
    # fails is a check nobody reads (EXPERT_AUDIT: "always-failing guards").
    # Each column is now compared against the canonical quantity it is built
    # from, and the closing stock is checked as its own identity.
    print("  RECONCILIATION  sheet 7 vs sheet 1 and the canonical basis")
    for p in PLANTS:
        s1 = int(b.filter(pl.col("plant") == p)["qty"].sum())
        s7 = int(daily.filter(pl.col("plant") == p)["built"].sum())
        c7 = int(daily.filter(pl.col("plant") == p)["cured"].sum())
        r7 = float(daily.filter(pl.col("plant") == p)["cured_press_run"].sum())
        gcl = int(daily.filter(pl.col("plant") == p)["gt_stock_at_0700_next_day"][-1])
        ok = (s1 == s7 and abs(c7 - Q[p]["delivered_in_month"]) < 0.5
              and abs(r7 - Q[p]["fed_in_month"]) < 1.0
              and abs(gcl - Q[p]["closing_gt"]) < 1.5)
        print(f"     {p}  built {s1:>9,} vs sheet7 {s7:>9,}  ·  cured "
              f"{c7:>9,} vs delivered-in-month "
              f"{Q[p]['delivered_in_month']:>9,.0f}  ·  cured_press_run "
              f"{r7:>9,.0f} vs fed-in-month {Q[p]['fed_in_month']:>9,.0f}  ·  "
              f"closing GT {gcl:>7,} vs {Q[p]['closing_gt']:>7,.0f}   "
              f"{'OK' if ok else '!! MISMATCH'}")

    # ---- 8. demand vs plan -------------------------------------------------
    # THE DENOMINATOR IS NOW A COLUMN. `9b` graded fulfilment against
    # `cure_requirement` and that number appeared in no sheet of this pack, so
    # nobody could check the headline against anything. `requirement_L11` sums
    # per plant to exactly the denominator in `9c_quantity_bridge`, and
    # `cured_in_month_L11` sums to exactly the numerator.
    #
    # `shortfall` IS SIGNED AND NETS OUT. A GT fed above its demand carries a
    # negative shortfall which cancelled a genuinely starved GT elsewhere: on
    # August TBR the column summed to 1,162 while 2,481 tyres of demand are
    # actually unmet. The signed column is kept -- it is the fed-vs-demanded
    # balance and some readers want it -- and `unmet_vs_demand` /
    # `surplus_vs_demand` are the one-sided halves that cannot cancel each
    # other out. `unmet_vs_requirement` / `surplus_vs_requirement` are the same
    # split against the GRADED denominator, which is the pair that ties to 9c.
    fed = bs.group_by(["plant", "gt_code"]).agg(pl.col("qty").sum().alias("fed"))
    stg = st.group_by(["plant", "gt_code"]).agg(
        pl.col("qty").sum().alias("starved"),
        pl.col("reason").first().alias("shortfall_reason"))
    upg = (up.group_by(["plant", "gt_code"]).agg(
        pl.col("qty").sum().alias("unplaced_L5")) if up.height else
        pl.DataFrame(schema={"plant": pl.Utf8, "gt_code": pl.Utf8,
                             "unplaced_L5": pl.Float64}))
    _numg = {f"{p}|{g}": v for p in PLANTS for g, v in Q[p]["num_by_gt"].items()}
    _deng = {f"{p}|{g}": v for p in PLANTS for g, v in Q[p]["den_by_gt"].items()}
    _key = pl.concat_str([pl.col("plant"), pl.col("gt_code")], separator="|")
    dv = (dem.group_by(["plant", "gt_code"]).agg(
        pl.col("qty").sum().alias("demanded"),
        pl.col("sku").n_unique().alias("skus"))
        .join(req.select(["plant", "gt_code", "gross_build", "residual"]),
              on=["plant", "gt_code"], how="left")
        .join(fed, on=["plant", "gt_code"], how="left")
        .join(stg, on=["plant", "gt_code"], how="left")
        .join(upg, on=["plant", "gt_code"], how="left")
        .with_columns([pl.col("fed").fill_null(0), pl.col("starved").fill_null(0),
                       pl.col("unplaced_L5").fill_null(0),
                       pl.col("gt_code").replace_strict(rim, default="(no rim)").alias("rim")])
        .with_columns([
            (pl.col("demanded") - pl.col("fed")).alias("shortfall"),
            _key.replace_strict(_numg, default=0.0, return_dtype=pl.Float64)
            .round(1).alias("cured_in_month_L11"),
            _key.replace_strict(_deng, default=0.0, return_dtype=pl.Float64)
            .round(1).alias("requirement_L11")])
        .with_columns([
            (pl.col("demanded") - pl.col("fed")).clip(lower_bound=0)
            .alias("unmet_vs_demand"),
            (pl.col("fed") - pl.col("demanded")).clip(lower_bound=0)
            .alias("surplus_vs_demand"),
            (pl.col("requirement_L11") - pl.col("cured_in_month_L11"))
            .clip(lower_bound=0).round(1).alias("unmet_vs_requirement"),
            (pl.col("cured_in_month_L11") - pl.col("requirement_L11"))
            .clip(lower_bound=0).round(1).alias("surplus_vs_requirement")])
        .sort(["plant", "unmet_vs_requirement"], descending=[False, True]))
    sheets["8_demand_vs_plan"] = add_desc(dv).select(
        ["plant", "gt_code", "gt_description", "tyre_size", "rim", "modal_sku",
         "n_skus", "skus", "demanded", "gross_build", "requirement_L11", "fed",
         "cured_in_month_L11", "shortfall", "unmet_vs_demand",
         "surplus_vs_demand", "unmet_vs_requirement", "surplus_vs_requirement",
         "starved", "unplaced_L5", "shortfall_reason", "residual"])
    print("  RECONCILIATION  sheet 8 vs the canonical basis")
    for _p in PLANTS:
        _z = dv.filter(pl.col("plant") == _p)
        for _col, _want, _lbl in (
                ("cured_in_month_L11", Q[_p]["numerator"], "9c numerator"),
                ("requirement_L11", Q[_p]["denominator"], "9c denominator"),
                ("fed", Q[_p]["built"] + Q[_p]["opening"],
                 "9a 'tyres fed' = built + opening")):
            _got = float(_z[_col].sum())
            print(f"     {_p}  {_col:<20}{_got:>10,.0f} vs {_lbl:<34}"
                  f"{_want:>10,.0f}  "
                  f"{'OK' if abs(_got - _want) < 1.5 else '!! MISMATCH'}")
        print(f"     {_p}  `shortfall` (signed, nets out) "
              f"{float(_z['shortfall'].sum()):>9,.0f}  vs one-sided "
              f"unmet_vs_demand {float(_z['unmet_vs_demand'].sum()):>9,.0f}  "
              f"(surplus_vs_demand {float(_z['surplus_vs_demand'].sum()):,.0f} "
              f"was cancelling it)")

    # ---- 9a. KPI SUMMARY ---------------------------------------------------
    # ONE FULFILMENT NUMBER, WITH ITS NUMERATOR AND ITS DENOMINATOR.
    #
    # WHAT THIS DOES / WHY IT EXISTS -- a measured defect, found 2026-08-21.
    #   This sheet printed THREE different "fulfilment numerators" and the real
    #   one was none of them. On July PCR it printed 384,406 labelled "tyres
    #   CURED in-month (fulfilment numerator)" and 385,858 labelled "D = ... =
    #   fulfilment numerator", while the 96.1 % on the line above it was
    #   computed from 383,099. August PCR spread 4,915 tyres across the three.
    #   The denominator (398,459 / 99,242 / 427,949 / 100,567) appeared nowhere
    #   in the pack, so the headline could not be checked against anything.
    #
    #   Every figure below now comes from `plan_quantities()`, which re-derives
    #   L11's arithmetic and asserts its own percentage against the string L11
    #   itself wrote. The A..F ladder is kept because the floor asked for it,
    #   but every line names the basis it is on, and `9c_quantity_bridge`
    #   carries the whole ladder with a running total that adds up.
    #
    #   TWO LABELS THIS SHEET IS CONTRACTUALLY BOUND TO -- do not rename them.
    #   `scripts/verify_export.py` finds the fed figure with
    #   `metric.starts_with("tyres fed")` and the opening stock with
    #   `metric.contains("OPENING STOCK")`, then re-derives fed == built +
    #   opening from the exported CSVs. Exactly ONE row per plant may match
    #   each pattern or the verifier silently reads whichever comes last.
    #
    # Sub-floor runs are counted HERE, on the setup-block basis sheet 12 uses,
    # so 9a can print it beside L11's own count. The pack shipped
    # `12_lot_size_violations` listing one TBR run below the floor while 9a
    # printed "runs below min lot 0.0 %" -- two answers to one question, and
    # neither said which object it was counting.
    _sub = runs.with_columns(
        pl.col("plant").replace_strict(CONFIG_FLOOR, default=0).alias("floor")
    ).filter(pl.col("run_qty") < pl.col("floor"))
    _sub_n = {p: int(_sub.filter(pl.col("plant") == p).height) for p in PLANTS}
    kpi = []
    for p in PLANTS:
        d = Q[p]
        bp = b.filter(pl.col("plant") == p)
        d11 = {r["invariant"]: r["actual"] for r in inv.iter_rows(named=True)}
        qq = bp.group_by("run_id").agg(pl.col("qty").sum())["qty"].to_numpy()
        mp = mach.filter(pl.col("plant") == p)
        rp = runs.filter(pl.col("plant") == p)
        dsum = daily.filter(pl.col("plant") == p)
        kpi += [
            {"plant": p, "metric": "DEMAND FULFILMENT (the graded number)",
             "value": f"{d['pct']:.1f}%"},
            {"plant": p, "metric": "  numerator: CURED IN-MONTH -- fed in-month, capped at each GT's own requirement",
             "value": f"{d['numerator']:,.0f}"},
            {"plant": p, "metric": "  denominator: CURE REQUIREMENT -- net_requirement, ~residual, ~lookahead",
             "value": f"{d['denominator']:,.0f}"},
            {"plant": p, "metric": "  the same ratio appears in 9b_l11_invariants, 9c_quantity_bridge and BOTH BTP workbooks",
             "value": f"{d['numerator']:,.0f} / {d['denominator']:,.0f}"},
            {"plant": p, "metric": "  per GT: 8_demand_vs_plan `cured_in_month_L11` / `requirement_L11`",
             "value": ""},
            {"plant": p, "metric": "unmet vs REQUIREMENT, one-sided (sheet 8 `unmet_vs_requirement`; does not cancel)",
             "value": f"{float(dv.filter(pl.col('plant') == p)['unmet_vs_requirement'].sum()):,.0f}"},
            {"plant": p, "metric": "unmet vs ORDER-BOOK DEMAND, one-sided (sheet 8 `unmet_vs_demand`)",
             "value": f"{float(dv.filter(pl.col('plant') == p)['unmet_vs_demand'].sum()):,.0f}"},
            {"plant": p, "metric": "  sheet 8 `shortfall` is SIGNED and nets these two out -- do not sum that column",
             "value": f"{float(dv.filter(pl.col('plant') == p)['shortfall'].sum()):,.0f}"},
            {"plant": p, "metric": "RECONCILIATION  build -> cure -> press capacity  (full ladder: 9c_quantity_bridge)", "value": ""},
            {"plant": p, "metric": "  A  BUILT this month = sheet 1 = sheet 5 = sheet 7 'built'", "value": f"{d['built']:,.0f}"},
            {"plant": p, "metric": "  A2 BUILT after month end -- EXCLUDED from every other sheet; the rows are in 1c", "value": f"{d['built_after_end']:,.0f}"},
            {"plant": p, "metric": "  B  + OPENING GT STOCK consumed (on the floor at 07:00 day 1)", "value": f"{d['opening']:,.0f}"},
            {"plant": p, "metric": "  C  - CLOSING GT stock at month end (built now, cures next month)", "value": f"{d['closing_gt']:,.0f}"},
            {"plant": p, "metric": "  D  = tyres DELIVERED, cure in-month (A+B-C) = sheet 2 'qty' = sheet 6 = sheet 7 'cured'", "value": f"{d['delivered_in_month']:,.0f}"},
            {"plant": p, "metric": "  D2 = tyres FED IN-MONTH, campaign pro-rata = sheet 2 'qty_run' = sheet 7 'cured_press_run'", "value": f"{d['fed_in_month']:,.0f}"},
            {"plant": p, "metric": "  D3 = CURED IN-MONTH = D2 minus cures beyond a GT's own requirement = THE NUMERATOR", "value": f"{d['numerator']:,.0f}"},
            {"plant": p, "metric": "  E  + press slots planned but NEVER FED (starved presses)", "value": f"{d['unfed']:,.0f}"},
            {"plant": p, "metric": "  F  = CAMPAIGN NAMEPLATE = sheet 2b 'qty_planned' = sheet 6 'tyres_planned'", "value": f"{d['nameplate']:,.0f}"},
            {"plant": p, "metric": "     F is PRESS CAPACITY SEATED, not tyres. Never quote it as output.", "value": ""},
            {"plant": p, "metric": "tyres fed to presses inside the report window (= A + B; the verifier re-derives this)",
             "value": f"{d['built'] + d['opening']:,.0f}"},
            {"plant": p, "metric": "  of which OPENING STOCK carried in", "value": f"{d['opening']:,.0f}"},
            {"plant": p, "metric": "THE FIVE NUMBERS THIS PACK USED TO CALL 'THE TAIL', each named once", "value": ""},
            {"plant": p, "metric": "  t1 carry-out tail: fed, cure completes NEXT month (campaign pro-rata)", "value": f"{d['tail_campaign']:,.0f}"},
            {"plant": p, "metric": "  t2 over-requirement cap: cured beyond that GT's own requirement -- real output, sits in BUILT", "value": f"{d['over_req_cap']:,.0f}"},
            {"plant": p, "metric": "  t3 = t1 + t2 = 9b's 'carry-out tail (excluded from fulfilment)'", "value": f"{d['tail_campaign'] + d['over_req_cap']:,.0f}"},
            {"plant": p, "metric": "  t4 CLOSING GT stock = sheet 7 last 'gt_stock_at_0700_next_day' = BTP 'Closing_Balance'", "value": f"{d['closing_gt']:,.0f}"},
            {"plant": p, "metric": "  t5 BUILT after month end, clipped out of the report entirely (= A2, sheet 1c)", "value": f"{d['built_after_end']:,.0f}"},
            {"plant": p, "metric": "  of t4, green tyres never assigned to a press at all", "value": f"{d['closing_gt_unassigned']:,.0f}"},
            {"plant": p, "metric": "same-size share", "value": d11.get(f"{p} same-size share of build changeovers")},
            {"plant": p, "metric": "build changeovers (SETUPS, run-level = sheet 5 = sheet 11)", "value": f"{int(rp['changeover'].sum()):,}"},
            {"plant": p, "metric": "  slice rows whose GT differs from the previous slice -- NOT setups", "value": f"{int(bp['changeover'].sum()):,}"},
            {"plant": p, "metric": "  independent check: setup blocks - machines - same-GT restarts", "value": f"{rp.height - mp.height - int(mp['same_gt_restarts'].sum()):,}"},
            {"plant": p, "metric": "  same-GT restarts after a >1 h gap (block boundaries that are NOT setups)", "value": f"{int(mp['same_gt_restarts'].sum()):,}"},
            {"plant": p, "metric": "changeovers / machine-day", "value": d11.get(f"{p} build changeovers / machine-day")},
            {"plant": p, "metric": "lot p50 (tyres per run)", "value": f"{np.median(qq):.0f}"},
            {"plant": p, "metric": "runs below min lot (L11 basis)", "value": d11.get(f"{p} build runs below min_lot ({'150' if p=='PCR' else '70'})")},
            {"plant": p, "metric": "  runs below min lot, SETUP-BLOCK basis = rows in 12_lot_size_violations", "value": f"{int(_sub_n.get(p, 0)):,}"},
            {"plant": p, "metric": "R5 GT wait max", "value": d11.get(f"{p} GT wait max (R5)")},
            {"plant": p, "metric": "GT wait p95", "value": d11.get(f"{p} GT wait p95")},
            {"plant": p, "metric": "GT inventory mean (time-wt)", "value": d11.get(f"{p} mean GT inventory (G8)")},
            {"plant": p, "metric": "GT inventory daily max", "value": f"{dsum['gt_inventory_day_mean'].max():,}"},
            {"plant": p, "metric": "GT inventory last day (sampled at hour 743.5, the G8 detector's basis)", "value": d11.get(f"{p} last-day GT inventory (G8)")},
            {"plant": p, "metric": "GT stock handed to next month (sampled AT 07:00, day 1 next month)", "value": f"{d['closing_gt']:,.0f}"},
            {"plant": p, "metric": "plant closed days, rule G3 (derived from the plan; see 0_settings)", "value": (", ".join(str(x) for x in CLOSE[p]) or "(none)")},
            {"plant": p, "metric": "machine hours AVAILABLE (month hours minus the closure)", "value": f"{AVAIL[p]:,.0f} h of {H:,} h"},
            {"plant": p, "metric": "machine occupancy %", "value": f"{100*mp['busy_h'].sum()/(mp.height*AVAIL[p]):.1f}%"},
            {"plant": p, "metric": "presses used", "value": press.filter(pl.col('plant') == p).height},
            {"plant": p, "metric": "realised n_g", "value": d11.get(f"{p} realised n_g (concurrent presses/GT)")},
        ]
    sheets["9a_kpi_summary"] = pl.DataFrame(kpi)
    # 9b: L11's own scorecard, untouched, plus the caveat column. The carry-out
    # tail line is caveated because it is t3 above -- two different quantities
    # summed under one name, which is why it reads larger than anything a
    # supervisor would call a tail.
    _cav = dict(MISMINED)
    for p in PLANTS:
        _cav[f"{p} carry-out tail (excluded from fulfilment)"] = (
            f"TWO NUMBERS SUMMED: {Q[p]['tail_campaign']:,.0f} cured next month "
            f"+ {Q[p]['over_req_cap']:,.0f} cured beyond that GT's own "
            f"requirement (real output, counted in BUILT). See 9a t1/t2/t3 and "
            f"9c_quantity_bridge.")
    sheets["9b_l11_invariants"] = inv.with_columns(
        pl.col("invariant").replace_strict(_cav, default="").alias("caveat"))
    # 9c: the ladder tying every quantity in this pack to every other one.
    sheets["9c_quantity_bridge"] = quantity_bridge(Q)


    # ---- 11. CHANGEOVER BY MACHINE ----------------------------------------
    # Per-machine changeover ledger, with the WHY of the single worst one and
    # the plant's own rate for the SAME machine beside ours.
    #
    # Every minute here is charged at THAT MACHINE's own rate from the plant
    # master (PCR 1-5 28/60, PCR 6-11 22/42, TBR 10/24). PARTITION §1c: a flat
    # 11.3/42.4 was once charged to both plants, which understated PCR same-size
    # ~2x and overstated TBR different-size 77 %, so the plants were not even
    # comparable. Never hardcode a changeover minute -- `_same`/`_diff` above are
    # read from `cap_changeover.parquet`.
    #
    # RECONCILES BY CONSTRUCTION: built from `runs`, the same frame behind sheet
    # 1b and the `mach` aggregate in sheet 5, so `changeovers` here sums to sheet
    # 5's column and `setup_required_min` to its `setup_required_h`.
    co = runs.filter(pl.col("changeover")).with_columns([
        pl.when(pl.col("size_change")).then(pl.lit("size-change"))
        .otherwise(pl.lit("same-size")).alias("co_type"),
        pl.col("_pr2").alias("from_rim"),
        pl.col("_pg2").alias("from_gt"),
    ])
    if co.height:
        # the single worst changeover on each machine, and why it cost that much
        _worst = (co.sort(["plant", "machine", "setup_required_min", "start_ts"],
                          descending=[False, False, True, False])
                  .group_by(["plant", "machine"]).first()
                  .select([
                      "plant", "machine",
                      pl.col("setup_required_min").round(1).alias("max_single_co_min"),
                      pl.col("from_gt").alias("max_co_from_gt"),
                      pl.col("gt_code").alias("max_co_to_gt"),
                      pl.col("from_rim").alias("max_co_from_rim"),
                      pl.col("rim").alias("max_co_to_rim"),
                      pl.col("co_type").alias("max_co_type"),
                      pl.col("start_ts").alias("max_co_at")]))
        cm = (co.group_by(["plant", "machine"]).agg(
            pl.len().alias("changeovers"),
            (pl.col("co_type") == "same-size").sum().alias("same_size"),
            (pl.col("co_type") == "size-change").sum().alias("size_changes"),
            pl.col("setup_required_min").sum().round(1).alias("total_co_min"),
            pl.col("setup_required_min").mean().round(2).alias("mean_co_min"),
            pl.col("setup_shortfall_min").sum().round(1).alias("unreserved_co_min"),
        ).join(_worst, on=["plant", "machine"], how="left"))
    else:
        cm = pl.DataFrame(schema={"plant": pl.Utf8, "machine": pl.Utf8,
                                  "changeovers": pl.UInt32})
    # machine-days on which this machine actually ran, so the rate has the same
    # denominator as the plant's (PARTITION §4d: a metric that divides plant
    # EVENT rows by our CAMPAIGN rows is a 10x error that still "passes").
    _md = (b.group_by(["plant", "machine"]).agg(
        pl.col("plant_day").n_unique().alias("machine_days_run")))
    cm = cm.join(_md, on=["plant", "machine"], how="left")
    # ---- the plant's OWN rate for the same machine, same month ----
    plant_co: dict = {}
    try:
        from planner.data.warehouse import duck            # noqa: E402
        _lo = f"{y}-{m:02d}-01"
        _hi = f"{y + (m == 12)}-{(m % 12) + 1:02d}-01"
        _pb = pl.DataFrame(duck().execute(
            "SELECT plant, machineCode m, itemCode g, event_ts ts FROM v_build "
            "WHERE stage = 2 AND itemCode IS NOT NULL AND machineCode IS NOT NULL "
            "AND date >= ?::DATE AND date < ?::DATE", [_lo, _hi]).fetchall(),
            schema=["plant", "m", "g", "ts"], orient="row")
        if _pb.height:
            _d = (_pb.sort(["m", "ts"])
                  .with_columns(pl.col("g").shift(1).over("m").alias("prev")))
            _runs = (_d.with_columns((pl.col("g") != pl.col("prev")).fill_null(True)
                                     .cum_sum().over("m").alias("rid"))
                     .group_by(["plant", "m", "rid"])
                     .agg(pl.col("ts").min().alias("t")))
            _pd = (_pb.with_columns(pl.col("ts").dt.date().alias("d"))
                   .select(["m", "d"]).unique()
                   .group_by("m").agg(pl.len().alias("pdays")))
            _pr = (_runs.group_by(["plant", "m"]).agg(pl.len().alias("pruns"))
                   .join(_pd, on="m", how="left"))
            for r in _pr.iter_rows(named=True):
                if r["pdays"]:
                    plant_co[r["m"]] = round((r["pruns"] - 1) / r["pdays"], 2)
    except Exception as _e:                                    # noqa: BLE001
        # A SWALLOWED EXCEPTION THAT LEAVES AN EMPTY COLUMN IS A LIE BY
        # OMISSION. `co_per_machine_day_PLANT` and `_DELTA` shipped blank on
        # all 20 machines of both months with nothing in the file saying why --
        # a reader cannot tell "the plant matches us" from "we could not look".
        # The raw MES drop is gitignored, so `v_build` is not registered in a
        # clone and this query cannot run. The reason now ships as a column.
        _plant_rate_status = f"UNAVAILABLE -- {type(_e).__name__}: {str(_e).splitlines()[0][:160]}"
        print(f"  !! plant per-machine changeover rate unavailable ({_e}) -- "
              f"column left null WITH A STATED REASON in "
              f"`plant_rate_status`; OURS is still exact")
    else:
        _plant_rate_status = (f"from v_build, {len(plant_co)} machines matched"
                              if plant_co else
                              "v_build query returned no rows for this month")
    cm = cm.with_columns([
        (pl.col("changeovers") / pl.col("machine_days_run"))
        .round(2).alias("co_per_machine_day_OURS"),
        pl.col("machine").replace_strict(plant_co, default=None)
        .alias("co_per_machine_day_PLANT"),
        pl.lit(_plant_rate_status).alias("plant_rate_status"),
    ])
    # The slice-level count beside the setup count, on BOTH sheets, so the
    # "RECONCILES BY CONSTRUCTION" claim above is actually checkable: sheet 5's
    # `changeovers` == this `changeovers`, and sheet 5's
    # `slice_gt_transitions` == this `slice_gt_transitions`.
    cm = cm.join(mach.select(["plant", "machine", "slice_gt_transitions"]),
                 on=["plant", "machine"], how="left")
    cm = cm.with_columns(
        (pl.col("co_per_machine_day_OURS") - pl.col("co_per_machine_day_PLANT"))
        .round(2).alias("co_per_machine_day_DELTA")).sort(["plant", "machine"])
    sheets["11_changeover_by_machine"] = fmt_ts(cm, ["max_co_at"])
    for _p in PLANTS:
        _a = int(cm.filter(pl.col("plant") == _p)["changeovers"].sum())
        _bb = int(mach.filter(pl.col("plant") == _p)["changeovers"].sum())
        print(f"  RECONCILIATION  sheet 11 vs sheet 5 changeovers  {_p}: "
              f"{_a:,} vs {_bb:,}  {'OK' if _a == _bb else '!! MISMATCH'}")

    # ---- 12. LOT SIZE VIOLATIONS ------------------------------------------
    # Every build run below its plant's B12 floor, with the reason it was let
    # through. PARTITION §4f: a ROW is a SLICE, not a setup -- 100 % of slice
    # rows sit below the floor and that is meaningless. This sheet judges
    # `block_id`, i.e. a run split at >1 h gaps, which is the physical setup.
    #
    # Under PLANNER_STRICT_LOT_FLOOR=1 (default, plant instruction §4m) the
    # correct content is ZERO ROWS. The sheet is still emitted, with a stated
    # zero line, because an absent sheet cannot be told apart from an
    # unmeasured one -- an empty sheet is evidence, not an omission.
    _floor_of = CONFIG_FLOOR
    viol = (runs.with_columns(
        pl.col("plant").replace_strict(_floor_of, default=0).alias("floor"))
        .filter(pl.col("run_qty") < pl.col("floor"))
        .with_columns([
            (pl.col("floor") - pl.col("run_qty")).alias("shortfall"),
            pl.lit(_STRICT_NOTE).alias("why_allowed"),
        ])
        .select(["plant", "machine", "gt_code", "gt_description", "rim",
                 "run_qty", "floor", "shortfall", "slices", "start_ts",
                 "end_ts", "block_id", "why_allowed"])
        .sort(["plant", "shortfall"], descending=[False, True]))
    if viol.height == 0:
        viol = pl.DataFrame([{
            "plant": "(none)", "machine": "", "gt_code": "", "gt_description": "",
            "rim": "", "run_qty": None, "floor": None, "shortfall": None,
            "slices": None, "start_ts": "", "end_ts": "", "block_id": "",
            "why_allowed":
            f"ZERO VIOLATIONS. Every build run on both plants is at or above "
            f"its B12 floor (PCR {_floor_of.get('PCR')} / TBR "
            f"{_floor_of.get('TBR')}), judged on setup blocks split at >1 h "
            f"gaps. PLANNER_STRICT_LOT_FLOOR=1."}],
            schema_overrides={"run_qty": pl.Float64, "floor": pl.Int64,
                              "shortfall": pl.Float64, "slices": pl.UInt32})
        print(f"  LOT SIZE VIOLATIONS: none -- sheet 12 emitted with a stated "
              f"zero line (floors PCR {_floor_of.get('PCR')} / "
              f"TBR {_floor_of.get('TBR')})")
    else:
        viol = fmt_ts(viol, ["start_ts", "end_ts"])
        print(f"  LOT SIZE VIOLATIONS: {viol.height} run(s) below floor -- see "
              f"sheet 12")
    sheets["12_lot_size_violations"] = viol

    # ---- write -------------------------------------------------------------
    # Explicit order so the workbook tabs and the csv listing read in the same
    # sequence a person works through the pack, whatever order the code built
    # them in.
    _ORDER = ["0_settings", "1_build_schedule_shift", "1b_build_runs",
              "1c_build_after_month_end", "2_cure_schedule_shift",
              "2b_cure_campaigns", "3_mould_changes", "4_crew_load",
              "5_machine_summary", "6_press_summary", "7_daily_summary",
              "8_demand_vs_plan", "9a_kpi_summary", "9b_l11_invariants",
              "9c_quantity_bridge", "10_press_crosswalk",
              "11_changeover_by_machine", "12_lot_size_violations"]
    sheets = {k: sheets[k] for k in _ORDER if k in sheets} | {
        k: v for k, v in sheets.items() if k not in _ORDER}
    counts = {}
    for name, df in sheets.items():
        df.write_csv(out / "csv" / f"{name}.csv")
        counts[name] = df.height
    # ---- CARRY-OUT + SETUP report, printed so it cannot be missed ----------
    # CARRY-OUT, ON THE BASIS THAT CAN ACTUALLY BE NON-ZERO. `b.carry_out` is
    # structurally False (bs is clipped at load), so this used to print
    # "build rows 0 (0 tyres)" as a tautology while 3,948 PCR tyres really were
    # clipped. The clipped rows are what is reported now, and they ship in
    # sheet 1c.
    print("  CARRY-OUT (what the report window excludes -- exported in 1c)")
    for p in PLANTS:
        _lb = _late.filter(pl.col("plant") == p)
        _cf = camp.filter((pl.col("plant") == p) & pl.col("finishes_next_month"))
        print(f"     {p}  build slices finishing after month end {_lb.height:>4} "
              f"({int(_lb['qty'].sum()):>6,} tyres) · closing GT stock "
              f"{Q[p]['closing_gt']:>6,.0f} · cure campaigns finishing next "
              f"month {_cf.height:>3} · `carry_out` columns in this pack are "
              f"structurally False, by the report-window ruling")
    print("  SETUP RESERVATION (L7 does not reserve changeover time -- see README)")
    for p in ("PCR", "TBR"):
        _r = runs.filter(pl.col("plant") == p)
        _need = float(_r["setup_required_min"].sum()) / 60.0
        _short = float(_r["setup_shortfall_min"].sum()) / 60.0
        _n = int((_r["setup_shortfall_min"] > 0.001).sum())
        _tr = int((_r["setup_required_min"] > 0).sum())
        print(f"     {p}  owed {_need:7.1f} h · NOT reserved {_short:7.1f} h "
              f"({100*_short/max(_need,1e-9):4.1f}%) · runs short {_n:,} of {_tr:,}")

    xl = out / f"schedule_{month}.xlsx"
    with pl.Config():
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        hdr = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", start_color="1F3864")
        for name, df in sheets.items():
            ws = wb.create_sheet(name[:31])
            ws.append(df.columns)
            for cell in ws[1]:
                cell.font, cell.fill = hdr, fill
            for row in df.iter_rows():
                ws.append([None if isinstance(v, float) and np.isnan(v) else v
                           for v in row])
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for i, c in enumerate(df.columns, 1):
                w = max(len(str(c)) + 2, 12)
                ws.column_dimensions[get_column_letter(i)].width = min(w, 34)
        # A workbook open in Excel cannot be overwritten. Never lose the export
        # over it: save beside the locked file and say so loudly.
        try:
            wb.save(xl)
        except PermissionError:
            alt = out / f"schedule_{month}__NEW.xlsx"
            wb.save(alt)
            print(f"  !! {xl.name} is LOCKED (open in Excel). Wrote {alt.name} "
                  f"instead -- close Excel, delete the old file and rename this one.")
            xl = alt
    print(f"  -> {xl}")
    for k, v in counts.items():
        print(f"     {k:<28}{v:>8,} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
