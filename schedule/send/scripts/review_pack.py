"""REVIEW PACK -- six sheets per month, feasibility before performance.

    python scripts/review_pack.py runs/frozen/2026-07

KPIs computed on an infeasible plan are decoration, so Sheet 1 is exceptions and
the rule is: if it has rows, stop -- nothing below it is meaningful.

  1 Exceptions        unplaced / shelf-life rows / R_g<1 / past-horizon / overlap
  2 Daily balance     demand, build, cure, closing stock, cumulative gap
  3 Press-day grid    presses x days, GT per cell -- thrash and orphan mounts
  4 Machine-day grid  machines x days, GT per cell, runs/machine-day
  5 Aging histogram   with the 72h line and the tail listed individually
  6 Traceability      20 sampled lots: which machine, which press, why

Sheets 3 and 4 are the ones a planner reads in ten seconds. Everything else is
arithmetic they can check afterwards.
"""
from __future__ import annotations

import sys
from datetime import timedelta
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SHELF_H = 72.0
PALETTE = ["FFD9E1F2", "FFE2EFDA", "FFFCE4D6", "FFFFF2CC", "FFEDEDED",
           "FFDDEBF7", "FFF8CBAD", "FFD9D9D9", "FFC6E0B4", "FFFFE699"]


def _sheet(wb, title):
    ws = wb.create_sheet(title[:31])
    return ws


def _put(ws, df: pl.DataFrame, r0=1, title=None):
    r = r0
    if title:
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=12)
        r += 2
    if df is None or df.height == 0:
        ws.cell(row=r, column=1, value="(none)").font = Font(italic=True)
        return r + 2
    for j, c in enumerate(df.columns, 1):
        cell = ws.cell(row=r, column=j, value=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFDDDDDD")
    for i, row in enumerate(df.iter_rows(), 1):
        for j, v in enumerate(row, 1):
            ws.cell(row=r + i, column=j,
                    value=v if v is None or isinstance(v, (int, float, str))
                    else str(v))
    for j, c in enumerate(df.columns, 1):
        w = max(len(str(c)) + 2,
                min(28, int(df[c].cast(pl.Utf8).str.len_chars().max() or 8) + 2))
        ws.column_dimensions[get_column_letter(j)].width = w
    return r + df.height + 3


def build(run: Path) -> Path:
    bs = pl.read_parquet(run / "build_schedule.parquet")
    cs = pl.read_parquet(run / "cure_schedule.parquet")
    ev = pl.read_parquet(run / "gt_events.parquet")
    camp = (pl.read_parquet(run / "press_campaigns.parquet")
            if (run / "press_campaigns.parquet").exists() else pl.DataFrame())
    origin = bs["start_ts"].min().replace(hour=0, minute=0, second=0, microsecond=0)

    wb = Workbook()
    wb.remove(wb.active)

    # ---- 1 EXCEPTIONS ---------------------------------------------------
    ws = _sheet(wb, "1 Exceptions")
    ws.cell(row=1, column=1,
            value="IF THIS SHEET HAS ROWS, STOP. Nothing below is meaningful."
            ).font = Font(bold=True, size=13, color="FF990000")
    r = 3
    for name, label in [
            ("unplaced", "UNPLACED LOTS (no slot inside the horizon)"),
            ("past_horizon", "LOTS ENDING PAST HORIZON"),
            ("machine_overlap", "MACHINE DOUBLE-BOOKING"),
            ("supply_ratio", "SUPPLY RATIO R_g < 1  (presses out-throughput building "
                             "-> starve by construction)")]:
        f = run / f"exc_{name}.parquet"
        d = pl.read_parquet(f) if f.exists() else pl.DataFrame()
        if name == "supply_ratio" and d.height:
            d = d.filter(pl.col("R") < 1.0).head(40)
        r = _put(ws, d, r, label)
    f = run / "exc_shelf_life.parquet"
    sl = pl.read_parquet(f) if f.exists() else pl.DataFrame()
    _put(ws, sl.head(60), r, f"SHELF-LIFE BREACHES >{SHELF_H:.0f}h "
                             f"({sl.height} rows, worst first)")

    # ---- 2 DAILY BALANCE ------------------------------------------------
    ev2 = ev.with_columns(pl.col("ts").dt.date().alias("day"))
    b = (ev2.filter(pl.col("source") == "build").group_by(["plant", "day"])
         .agg(pl.col("qty_delta").sum().alias("build")))
    c = (ev2.filter(pl.col("source") == "cure").group_by(["plant", "day"])
         .agg((-pl.col("qty_delta")).sum().alias("cure")))
    iv = (ev2.with_columns(pl.when(pl.col("source") == "cure")
                           .then(-pl.col("qty_delta").abs())
                           .otherwise(pl.col("qty_delta").abs()).alias("d"))
          .sort(["plant", "ts"])
          .with_columns(pl.col("d").cum_sum().over("plant").alias("I"))
          .group_by(["plant", "day"]).agg(pl.col("I").last().alias("close_gt")))
    bal = (b.join(c, on=["plant", "day"], how="full", coalesce=True)
           .join(iv, on=["plant", "day"], how="full", coalesce=True)
           .fill_null(0).sort(["plant", "day"]))
    bal = bal.with_columns(
        (pl.col("build") - pl.col("cure")).alias("build_minus_cure"),
        (pl.col("build") - pl.col("cure")).cum_sum().over("plant").alias("cum_gap"))
    ws = _sheet(wb, "2 Daily balance")
    r = 1
    for p in sorted(bal["plant"].unique().to_list()):
        r = _put(ws, bal.filter(pl.col("plant") == p), r, f"{p} -- daily balance")

    # ---- 3 PRESS-DAY GRID ----------------------------------------------
    ws = _sheet(wb, "3 Press-day grid")
    r = 1
    if camp.height:
        rows = []
        for x in camp.iter_rows(named=True):
            for d in range(int(x["start_day"]), int(x["end_day"])):
                rows.append({"plant": x["plant"], "press": x["press"],
                             "day": d, "gt": x["gt_code"]})
        M = pl.DataFrame(rows)
        cured = (cs.with_columns(
            ((pl.col("start_ts") - pl.lit(origin)).dt.total_seconds() // 86400)
            .cast(pl.Int64).alias("day"))
            .group_by(["plant", "press", "day"]).agg(pl.len().alias("n")))
        M = M.join(cured, on=["plant", "press", "day"], how="left").fill_null(0)
        for p in sorted(M["plant"].unique().to_list()):
            s = M.filter(pl.col("plant") == p)
            days = sorted(s["day"].unique().to_list())
            presses = sorted(s["press"].unique().to_list())
            ws.cell(row=r, column=1,
                    value=f"{p} -- press x day, GT mounted (blank = unmounted, "
                          f"'!' = mounted but cured 0)").font = Font(bold=True, size=12)
            r += 1
            ws.cell(row=r, column=1, value="press").font = Font(bold=True)
            for j, d in enumerate(days, 2):
                ws.cell(row=r, column=j, value=d + 1).font = Font(bold=True)
            look = {(x["press"], x["day"]): (x["gt"], x["n"])
                    for x in s.iter_rows(named=True)}
            gts = sorted({x["gt"] for x in s.iter_rows(named=True)})
            colour = {g: PALETTE[i % len(PALETTE)] for i, g in enumerate(gts)}
            for i, pr in enumerate(presses, 1):
                ws.cell(row=r + i, column=1, value=pr)
                for j, d in enumerate(days, 2):
                    hit = look.get((pr, d))
                    if not hit:
                        continue
                    g, n = hit
                    lbl = "".join(ch for ch in g if ord(ch) < 128)[:10]
                    cell = ws.cell(row=r + i, column=j,
                                   value=lbl + ("" if n else " !"))
                    cell.fill = PatternFill("solid", fgColor=colour[g])
                    cell.alignment = Alignment(horizontal="center")
                    if not n:
                        cell.font = Font(bold=True, color="FF990000")
            r += len(presses) + 3
        ws.freeze_panes = "B2"

    # ---- 4 MACHINE-DAY GRID --------------------------------------------
    ws = _sheet(wb, "4 Machine-day grid")
    bd = bs.with_columns(
        ((pl.col("start_ts") - pl.lit(origin)).dt.total_seconds() // 86400)
        .cast(pl.Int64).alias("day"))
    md = (bd.group_by(["plant", "machine", "day"])
          .agg(pl.col("gt_code").n_unique().alias("gts"), pl.len().alias("runs"),
               pl.col("qty").sum().alias("qty"),
               pl.col("gt_code").sort().first().alias("gt1")))
    r = 1
    for p in sorted(md["plant"].unique().to_list()):
        s = md.filter(pl.col("plant") == p)
        days = sorted(s["day"].unique().to_list())
        machines = sorted(s["machine"].unique().to_list())
        ws.cell(row=r, column=1,
                value=f"{p} -- machine x day, cell = runs (GTs). plant p50 is "
                      f"3.13 runs/machine-day").font = Font(bold=True, size=12)
        r += 1
        ws.cell(row=r, column=1, value="machine").font = Font(bold=True)
        for j, d in enumerate(days, 2):
            ws.cell(row=r, column=j, value=d + 1).font = Font(bold=True)
        look = {(x["machine"], x["day"]): x for x in s.iter_rows(named=True)}
        for i, m in enumerate(machines, 1):
            ws.cell(row=r + i, column=1, value=m)
            for j, d in enumerate(days, 2):
                x = look.get((m, d))
                if not x:
                    continue
                cell = ws.cell(row=r + i, column=j,
                               value=f"{x['runs']} ({x['gts']})")
                cell.alignment = Alignment(horizontal="center")
                if x["runs"] >= 6:
                    cell.fill = PatternFill("solid", fgColor="FFF8CBAD")
        r += len(machines) + 2
        rpm = float(s["runs"].mean())
        ws.cell(row=r, column=1,
                value=f"{p} mean runs/machine-day = {rpm:.2f}  "
                      f"(plant 3.13 PCR / 4.15 TBR)").font = Font(bold=True)
        r += 3
    ws.freeze_panes = "B2"

    # ---- 5 AGING --------------------------------------------------------
    ws = _sheet(wb, "5 Aging")
    bt = (ev.filter(pl.col("source").is_in(["build", "opening"])
                    & (pl.col("qty_delta") > 0))
          .with_columns(pl.col("qty_delta").cast(pl.Int64))
          .with_columns(pl.int_ranges(pl.col("qty_delta")).alias("_i")).explode("_i")
          .select(["plant", "gt_code", "ts", "lot_id"])
          .sort(["plant", "gt_code", "ts", "lot_id"])
          .with_columns(pl.int_range(pl.len()).over(["plant", "gt_code"]).alias("k")))
    ct = (cs.select(["plant", "gt_code", "start_ts"])
          .sort(["plant", "gt_code", "start_ts"])
          .with_columns(pl.int_range(pl.len()).over(["plant", "gt_code"]).alias("k")))
    lag = (ct.rename({"start_ts": "c_ts"})
           .join(bt, on=["plant", "gt_code", "k"], how="inner")
           .with_columns(((pl.col("c_ts") - pl.col("ts")).dt.total_seconds() / 3600)
                         .alias("age_h")))
    edges = [0, 2, 4, 8, 12, 16, 24, 36, 48, 72, 1e9]
    rows = []
    for p in sorted(lag["plant"].unique().to_list()):
        s = lag.filter(pl.col("plant") == p)
        n = s.height
        for lo, hi in zip(edges[:-1], edges[1:]):
            cnt = s.filter((pl.col("age_h") >= lo) & (pl.col("age_h") < hi)).height
            rows.append({"plant": p,
                         "bucket": f"{lo:g}-{'inf' if hi > 1e8 else f'{hi:g}'}h",
                         "tyres": cnt, "pct": round(100 * cnt / max(n, 1), 2),
                         "over_72h_line": "<<< SHELF LIFE" if lo >= 72 else ""})
    r = _put(ws, pl.DataFrame(rows), 1, "Aging histogram (72h = shelf life)")
    _put(ws, lag.filter(pl.col("age_h") > SHELF_H)
         .sort("age_h", descending=True)
         .select(["plant", "gt_code", "lot_id", "ts", "c_ts", "age_h"]).head(80),
         r, "The tail, listed individually")

    # ---- 6 TRACEABILITY -------------------------------------------------
    ws = _sheet(wb, "6 Traceability")
    step = max(bs.height // 20, 1)
    samp = bs.sort(["plant", "start_ts", "lot_id"]).gather_every(step).head(20)
    keep = [c for c in ["plant", "gt_code", "lot_id", "machine", "qty",
                        "start_ts", "end_ts", "setup_s", "decision_trace"]
            if c in samp.columns]
    _put(ws, samp.select(keep), 1,
         "20 sampled lots -- machine choice and its reason (decision_trace)")

    out = run / f"review_{run.name}.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    for a in sys.argv[1:]:
        p = Path(a)
        try:
            print(f"  {p.name}: {build(p)}")
        except Exception as e:                                  # noqa: BLE001
            print(f"  {p.name}: FAILED {type(e).__name__}: {e}")
