"""RUNNING MOULDS -- which mould set is actually seated on which press.

    python -m scripts.ingest_running_moulds --month 2026-08 \
        --pcr "../../running molds/august.xlsx" \
        --tbr "../../running molds/august (1).xlsx"

WHAT THESE FILES ARE
  Raw MES curing events, one row per cured tyre:

      iD, wcID, gtbarCode, pressbarCode, serialNo, recipeID, mouldNo,
      manningID, dtandTime

  `wcID` is the press. `mouldNo` is the seated mould PAIR ("HM07#HM24",
  "GC03#HM16") -- a press holds two mould halves and the plant names both.
  `recipeID` joins to Recipemaster.iD, which is the SKU -> GT bridge.

  PLANT IS BY FILE, NOT BY COLUMN. There is no plant column. The PCR file
  carries 86 distinct presses and the TBR file 72-76, matching the two rosters
  (PCR 86 presses, TBR 79), and the PCR `pressbarCode` values read 4510LHS /
  4609LHS -- the 45xx/46xx press numbers. So the caller names which is which.

WHY IT MATTERS FOR PLANNING
  A mould change is 6.0 h of press time. A campaign seated on a press that
  ALREADY holds the mould it needs starts immediately; one seated elsewhere pays
  the change first. L5 currently picks the eligible press that frees earliest
  and is blind to what is mounted, so it buys mould changes it did not need.
  This file is what lets it prefer a press already running the mould.

  `is_current` marks, per press, the LAST mould set seen -- that is what is
  physically in the press at the end of the window, i.e. what the next month
  starts with.

DATE CAVEAT -- READ BEFORE TRUSTING THE TIMESTAMPS
  `dtandTime` in the shipped files reads **2024-06 / 2024-07**, while the
  planning months are 2026-07 / 2026-08 and the MES history elsewhere in this
  project runs 2025-12..2026-07. The month is taken from the FILENAME the caller
  passes, never from the timestamps, and the observed range is printed so the
  discrepancy stays visible. The (press -> mould) pairing is what we consume and
  it does not depend on the year; the ordering within a file does.
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import polars as pl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from planner import paths                                          # noqa: E402


def _scan(f: Path, plant: str) -> list[dict]:
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h) if h is not None else "" for h in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    need = ("wcID", "mouldNo", "dtandTime")
    if not all(k in ix for k in need):
        raise SystemExit(f"{f.name}: expected columns {need}, got {hdr}")
    agg: dict[tuple, dict] = defaultdict(
        lambda: {"tyres": 0, "first": None, "last": None, "recipes": set(),
                 "rc_count": {}})
    for r in it:
        w, mo, ts = r[ix["wcID"]], r[ix["mouldNo"]], r[ix["dtandTime"]]
        if w is None or not mo:
            continue
        k = (plant, str(int(w)), str(mo).strip())
        a = agg[k]
        a["tyres"] += 1
        if isinstance(ts, datetime):
            a["first"] = ts if a["first"] is None or ts < a["first"] else a["first"]
            a["last"] = ts if a["last"] is None or ts > a["last"] else a["last"]
        rc = r[ix["recipeID"]] if "recipeID" in ix else None
        if rc is not None:
            a["recipes"].add(str(rc))
            a["rc_count"][str(rc).strip()] = a["rc_count"].get(str(rc).strip(), 0) + 1
    wb.close()
    return [{"plant": p, "press": pr, "mould_set": ms, "tyres": v["tyres"],
             "first_ts": v["first"], "last_ts": v["last"],
             "n_recipes": len(v["recipes"]),
             "modal_recipe": (max(v["rc_count"].items(), key=lambda x: x[1])[0]
                              if v["rc_count"] else None)}
            for (p, pr, ms), v in agg.items()]


def _target_day(y: int, m: int) -> int:
    """ONE DAY ONLY, three days before month end (plant instruction 2026-08-11).

        31-day month -> day 28        30-day month -> day 27

    Why a single day and not the window: a press holds ONE mould set at a time,
    so the planning input is a STATE, not a history. Aggregating three days makes
    a press that changed mould mid-window look like it holds two sets at once --
    7 PCR presses and 1 TBR press did exactly that in the shipped files, and
    `is_current` had to break the tie. Reading one day removes the ambiguity
    instead of resolving it after the fact.

    Three days back, rather than the last day, because the final days of a month
    carry end-of-month changeovers that are not representative of the steady
    state the next month starts from.
    """
    import calendar
    return calendar.monthrange(y, m)[1] - 3


def run(month: str, pcr: Path | None, tbr: Path | None,
        *, write: bool = True, day: int | None = None) -> pl.DataFrame:
    rows: list[dict] = []
    for f, plant in ((pcr, "PCR"), (tbr, "TBR")):
        if f is None:
            continue
        if not f.exists():
            raise SystemExit(f"missing {f}")
        got = _scan(f, plant)
        rows += got
        print(f"  {plant}  {f.name}: {len(got)} (press, mould_set) pairs")
    if not rows:
        raise SystemExit("no input files given")

    df = pl.DataFrame(rows)

    # ---- SINGLE-DAY SNAPSHOT ------------------------------------------------
    ts = df.filter(pl.col("last_ts").is_not_null())
    if ts.height:
        d0 = ts["first_ts"].min()
        want = day if day is not None else _target_day(d0.year, d0.month)
        have = sorted({v.day for v in
                       df.select(pl.col("last_ts").dt.date())["last_ts"]
                       if v is not None})
        use = want if want in have else min(have, key=lambda x: abs(x - want))
        if use != want:
            print(f"  !! requested day {want} is NOT in the file "
                  f"(present: {have}) -- using day {use}, the nearest available")
        before = df.height
        df = df.filter((pl.col("first_ts").dt.day() <= use)
                       & (pl.col("last_ts").dt.day() >= use))
        print(f"  single-day snapshot: day {use} -> {df.height} of {before} "
              f"(press, mould_set) pairs seated that day")

    # is_current: per press, the mould set seen LAST. Ties -> most tyres.
    df = (df.sort(["plant", "press", "last_ts", "tyres"], descending=[False, False, True, True])
            .with_columns((pl.int_range(pl.len()).over(["plant", "press"]) == 0)
                          .alias("is_current"))
            .with_columns(pl.lit(month).alias("month"))
            .sort(["plant", "press", "tyres"], descending=[False, False, True]))

    print(f"\n  RUNNING MOULDS  {month}")
    print(f"  {'-' * 68}")
    for p, g in df.group_by("plant"):
        cur = g.filter(pl.col("is_current"))
        print(f"    {p[0]}: {g['press'].n_unique():>3} presses · "
              f"{g['mould_set'].n_unique():>3} distinct mould sets · "
              f"{g.height:>4} pairings · {g['tyres'].sum():>8,} tyres observed")
        multi = (g.group_by("press").agg(pl.len().alias("n"))
                  .filter(pl.col("n") > 1))
        print(f"         {cur.height} presses have a CURRENT mould; "
              f"{multi.height} changed mould at least once in the window")
        ts = g.filter(pl.col("last_ts").is_not_null())
        if ts.height:
            print(f"         observed timestamps {ts['first_ts'].min()} .. "
                  f"{ts['last_ts'].max()}   <- NOT {month}; month is from the filename")
    # ---- RESOLVE THE PRESS'S CURRENT GT --------------------------------
    # `mouldNo` ("HM07#HM24") has ZERO string overlap with the engine's
    # mould_set ("G::PCR::GT 1503 NEO MSIL") -- the namespace trap again. The
    # bridge is recipeID -> gt_sku_master.recipe_id -> gt_code, which IS the
    # engine's planning key. This is what lets L5 know which GT each press was
    # already running at month start, so it can CONTINUE it instead of paying a
    # mould change and waiting for fresh supply.
    # THREE-TIER, AND THE OBVIOUS TABLE IS THE ONE THAT DOES NOT WORK.
    # gt_sku_master.recipe_id matched 0 of 30 TBR press recipes -- every TBR
    # press came out unresolved on both months, which is why TBR never got the
    # warm-press correction PCR did. The recipe IDs on the presses are CURING
    # recipe ids ("787", "1246"); gt_sku_master keys something else.
    #
    #   1. recipe_gt_sku.curing_recipe_id -> gt_name   ALREADY the MES namespace
    #   2. recipe_bridge.curingRecipeID  -> gt_code    BOM short code ("GT 5002"),
    #      expanded through gt_namespace.resolve_gt_label, which returns None on
    #      ambiguity so a head is never guessed onto the wrong GT
    #   3. gt_sku_master                               last, and for PCR only in
    #      practice -- it is the BOM namespace for TBR (the documented trap)
    #
    # Measured 2026-08: 13 of 30 recipes resolve (11 direct + 2 numeric-head),
    # covering 43 of 75 TBR presses, against 0 before. The other 17 recipes are
    # absent from every bridge -- those GTs have never been cured in the mined
    # window, so no evidence exists to map them and nothing is invented.
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from gt_namespace import resolve_gt_label                   # noqa: E402

    ns: dict[str, list[str]] = {}
    for _f in ("cap_machine", "cap_press"):
        for _m in sorted(paths.WH_DERIVED.glob(f"{_f}_*.parquet")):
            for _p, _g in pl.read_parquet(_m).select("plant", "gt_code").unique().rows():
                ns.setdefault(_p, []).append(_g)
    ns = {k: sorted(set(v)) for k, v in ns.items()}

    rgs = pl.read_parquet(paths.wh_derived("recipe_gt_sku.parquet"))
    direct = {str(r["curing_recipe_id"]).strip(): r["gt_name"]
              for r in rgs.iter_rows(named=True)
              if r.get("curing_recipe_id") is not None and r.get("gt_name")}
    rb = pl.read_parquet(paths.wh_derived("recipe_bridge.parquet"))
    bom = {str(r["curingRecipeID"]).strip(): r["gt_code"]
           for r in rb.iter_rows(named=True) if r.get("curingRecipeID") is not None}
    gsm = pl.read_parquet(paths.wh_derived("gt_sku_master.parquet"))
    r2g = {str(r["recipe_id"]).strip(): r["gt_code"]
           for r in gsm.iter_rows(named=True) if r.get("recipe_id") is not None}

    _tier: dict = {}

    def _resolve(rec: str | None, plant: str) -> str | None:
        if not rec:
            return None
        known = ns.get(plant, [])
        g = direct.get(rec)
        if g and g in known:
            _tier[rec] = "direct"
            return g
        b = bom.get(rec)
        if b:
            c, t = resolve_gt_label(b, plant, ns)
            if c:
                _tier[rec] = f"expand:{t}"
                return c
        g = r2g.get(rec)
        if g and g in known:
            _tier[rec] = "master"
            return g
        _tier[rec] = "unresolved"
        return None

    df = df.with_columns(
        pl.struct(["modal_recipe", "plant"]).map_elements(
            lambda r: _resolve(r["modal_recipe"], r["plant"]),
            return_dtype=pl.Utf8).alias("current_gt"))
    from collections import Counter as _C
    print(f"  recipe->GT tiers: {dict(_C(_tier.values()))}")
    cur = df.filter(pl.col("is_current"))
    res = cur.filter(pl.col("current_gt").is_not_null())
    print(f"  press -> current GT resolved: {res.height} of {cur.height} presses "
          f"({100*res.height/max(cur.height,1):.0f} %) via recipeID")
    for p, g in res.group_by("plant"):
        print(f"     {p[0]}: {g['press'].n_unique()} presses · "
              f"{g['current_gt'].n_unique()} distinct GTs")

    if write:
        f = paths.INPUT_DERIVED / f"running_moulds_{month}.parquet"
        df.write_parquet(f)
        print(f"  -> {f}")
    return df


def main() -> None:
    ap = argparse.ArgumentParser(description="MES curing events -> seated moulds")
    ap.add_argument("--month", required=True)
    ap.add_argument("--pcr", default=None)
    ap.add_argument("--tbr", default=None)
    ap.add_argument("--day", type=int, default=None,
                    help="day-of-month to snapshot; default = month length - 3 "
                         "(31d -> 28, 30d -> 27)")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()
    run(a.month, Path(a.pcr) if a.pcr else None,
        Path(a.tbr) if a.tbr else None, write=not a.dry_run, day=a.day)


if __name__ == "__main__":
    main()
