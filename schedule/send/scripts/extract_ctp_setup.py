"""Mine `CTP Set up building ,curing and inspection.xlsx` -- closes GAP-4 and GAP-9.

    python scripts/extract_ctp_setup.py

This workbook sat in the inputs unmined.  It contains four things the engine was
otherwise blocked on:

  Sheet1 / Sheet2      per-PRESS mould-change minutes + platen diameter  -> GAP-4, R14
  PCR Curing           press class, platen dia, warm-up, PCI rim change  -> GAP-4, R14
  TBR Curing           mould change 133.7 + dummy 47 + warm-up 180       -> GAP-4
  PCR BUILDING         within-inch / inch-to-inch setup catalogue        -> R7 (PCR)
  TBR Building         setup-change catalogue by job type               -> R7 (TBR)
  Change_set up        mixing / calendar / cutter / extruder changeovers -> GAP-9, L8

Formula cells are read with data_only=True; where Excel stored no cached result
the cell is skipped rather than guessed.
"""
from __future__ import annotations

import warnings
from pathlib import Path

import polars as pl

warnings.filterwarnings("ignore")
from openpyxl import load_workbook                            # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT.parent.parent
BOOK = SRC / "INPUT" / "raw" / "CTP Set up building ,curing and inspection (1) 2.xlsx"
OUT = ROOT / "warehouse" / "derived"
INP = SRC / "INPUT" / "derived"


def num(x):
    try:
        v = float(str(x).strip())
        return v if v == v else None
    except (TypeError, ValueError):
        return None


def main() -> None:
    wb = load_workbook(BOOK, read_only=True, data_only=True)
    made: list[tuple[str, int, str]] = []

    # ---- 1. per-press mould change + platen (GAP-4, R14) ----------------
    press = []
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    for r in rows[1:]:
        if r[0] is None:
            continue
        press.append({"plant": "PCR", "press": str(r[0]).strip(),
                      "platen_in": num(r[1]),
                      "descr": str(r[2] or "").strip(),
                      "mould_change_min": num(r[3])})
    rows = list(wb["Sheet2"].iter_rows(values_only=True))
    for r in rows[1:]:
        if r[0] is None:
            continue
        press.append({"plant": "TBR", "press": str(r[0]).strip(),
                      "platen_in": None, "descr": "",
                      "mould_change_min": num(r[1])})
    pf = pl.DataFrame(press).filter(pl.col("mould_change_min").is_not_null())
    # The CTP sheet keys presses by SAP number (3615, 4501); the schedule keys
    # them by wcID (119, 123). Those are different namespaces, so every
    # per-press lookup resolved 0 of 165 and each layer silently fell back to a
    # plant-wide median -- discarding a 2x spread (PCR 210-430 min). Carry the
    # bridge in the file itself so no consumer has to know about it.
    _wc = OUT / "wc_master.parquet"
    if _wc.exists():
        pf = pf.join(pl.read_parquet(_wc).select(["wc_id", "press_no"]),
                     left_on="press", right_on="press_no", how="left")
    pf.write_parquet(OUT / "press_mould_change.parquet")
    made.append(("press_mould_change", pf.height, "GAP-4 / R14"))

    # ---- 2. PCR curing press classes (GAP-4, R14) -----------------------
    cls = []
    rows = list(wb["PCR Curing"].iter_rows(values_only=True))
    hdr = [str(x or "").strip() for x in rows[1]]
    def gi(name):
        for i, h in enumerate(hdr):
            if name.lower() in h.lower():
                return i
        return None
    ix = {k: gi(k) for k in ["Changeover Type", "Platen dia", "Press No",
                             "No. of Press", "Tyre Size", "Warm up",
                             "Mould change", "PCI Rim"]}
    for r in rows[2:]:
        if not any(r):
            continue
        def g(k):
            i = ix[k]
            return r[i] if i is not None and i < len(r) else None
        cls.append({"changeover_type": str(g("Changeover Type") or "").strip(),
                    "platen_in": str(g("Platen dia") or "").strip(),
                    "press_range": str(g("Press No") or "").strip(),
                    "n_press": num(g("No. of Press")),
                    "tyre_size": str(g("Tyre Size") or "").strip(),
                    "warm_up_min": num(g("Warm up")),
                    "mould_change_min": num(g("Mould change")),
                    "pci_rim_min": num(g("PCI Rim"))})
    cf = pl.DataFrame(cls).filter(pl.col("changeover_type") != "")
    cf.write_parquet(OUT / "press_class_pcr.parquet")
    made.append(("press_class_pcr", cf.height, "GAP-4 / R14"))

    # ---- 3. building setup catalogue (R7) -------------------------------
    cat = []
    for sh, plant in [("PCR BUILDING", "PCR"), ("TBR Building", "TBR")]:
        rows = list(wb[sh].iter_rows(values_only=True))
        for r in rows:
            if len(r) < 4 or r[1] is None:
                continue
            d = str(r[1]).strip()
            t = num(r[3])
            if t is None or not d or d.lower().startswith(("s.no", "description")):
                continue
            kind = ("within_inch" if "within inch" in d.lower() else
                    "inch_to_inch" if "inch to inch" in d.lower() else "other")
            cat.append({"plant": plant, "descr": d, "job": str(r[2] or "").strip(),
                        "minutes": t, "kind": kind})
    bf = pl.DataFrame(cat).unique()
    bf.write_parquet(OUT / "build_setup_catalogue.parquet")
    made.append(("build_setup_catalogue", bf.height, "R7"))

    # ---- 4. prep-shop changeovers (GAP-9, L8) ---------------------------
    prep, section = [], ""
    rows = list(wb["Change_set up"].iter_rows(values_only=True))
    for r in rows:
        cells = [str(x).strip() if x is not None else "" for x in r[:6]]
        nonempty = [c for c in cells if c]
        if len(nonempty) == 1 and not num(nonempty[0]):
            section = nonempty[0]
            continue
        if len(cells) < 5:
            continue
        machine, descr, uom, t = cells[1], cells[2], cells[3], num(r[4])
        if t is None or not descr:
            continue
        mins = t / 60.0 if uom.lower().startswith("sec") else t
        prep.append({"section": section, "machine": machine, "operation": descr,
                     "uom": uom, "value": t, "minutes": round(mins, 3)})
    pr = pl.DataFrame(prep)
    pr = pr.filter(~pl.col("operation").str.to_lowercase().str.starts_with("description"))
    pr.write_parquet(OUT / "prep_changeover.parquet")
    made.append(("prep_changeover", pr.height, "GAP-9 / L8"))
    wb.close()

    INP.mkdir(parents=True, exist_ok=True)
    for n, _c, _k in made:
        import shutil
        shutil.copy2(OUT / f"{n}.parquet", INP / f"{n}.parquet")

    print(f"{'asset':<26}{'rows':>7}   closes")
    for n, c, k in made:
        print(f"  {n:<24}{c:>7}   {k}")
    print()
    print("press mould-change minutes:")
    for p in ["PCR", "TBR"]:
        s = pf.filter(pl.col("plant") == p)
        if s.height:
            print(f"  {p}: {s.height} presses, "
                  f"min {s['mould_change_min'].min():.0f} "
                  f"p50 {s['mould_change_min'].median():.0f} "
                  f"max {s['mould_change_min'].max():.0f} min")
    print("\nbuild setup catalogue:")
    for r in bf.group_by(["plant", "kind"]).agg(
            pl.col("minutes").min().alias("lo"),
            pl.col("minutes").max().alias("hi"),
            pl.len().alias("n")).sort(["plant", "kind"]).iter_rows(named=True):
        print(f"  {r['plant']} {r['kind']:<14} {r['n']:>3} entries  "
              f"{r['lo']:.0f}-{r['hi']:.0f} min")
    print(f"\nprep shop: {pr.height} operations across "
          f"{pr['section'].n_unique()} sections, {pr['machine'].n_unique()} machines")


if __name__ == "__main__":
    main()
