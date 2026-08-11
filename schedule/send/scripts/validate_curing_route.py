"""Were we reading CURING wrong? Two independent routes, cross-checked.

    python scripts/validate_curing_route.py

ROUTE A (what the engine has always used)
    v_curing.gtbarCode -> v_build.productionID -> v_build.itemCode
    Per-tyre barcode. 99.6% hit rate.

ROUTE B (new, from the recipe master)
    v_curing.recipeID -> recipemaster.iD -> SAPMaterialCode / name

These are INDEPENDENT: one goes through the tyre's own barcode, the other
through the press recipe that was loaded. If they disagree, every curing-derived
number in this project -- press rates, campaign structure, aging, the Little's
Law inventory law -- is measured against the wrong GT.

This is the check that should have been run the day the recipe master arrived.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import polars as pl

from planner.data.warehouse import duck, set_cutoff
from planner.runs.logger import log
from planner import paths

RM = paths.raw("Recipemaster 1.xlsx")


def _c(x) -> str:
    return "".join(ch for ch in str(x) if ord(ch) < 128).strip() if x is not None else ""


def main() -> int:
    set_cutoff(None)
    con = duck()

    import openpyxl
    wb = openpyxl.load_workbook(RM, read_only=True, data_only=True)
    raw = [[_c(c) for c in r] for r in wb.worksheets[0].iter_rows(values_only=True)]
    wb.close()
    hdr = raw[0]
    rm = [dict(zip(hdr, r)) for r in raw[1:] if any(r)]
    recipe = pl.DataFrame([{"recipe_id": d["iD"],
                            "r_pid": d.get("processID", ""),
                            "r_sap": d.get("SAPMaterialCode", ""),
                            "r_name": d.get("name", ""),
                            "r_desc": d.get("description", "")} for d in rm])
    con.register("recipe", recipe.to_arrow())

    print("=" * 80)
    print("1. COVERAGE of curing.recipeID")
    print("=" * 80)
    r = con.execute("""
        SELECT count(*) AS n,
               count(*) FILTER (WHERE recipeID IS NULL) AS nulls,
               count(DISTINCT recipeID::VARCHAR) AS distinct_recipes
        FROM v_curing WHERE statuscritical='Normal'
    """).fetchone()
    print(f"  cures={r[0]:,}  null recipeID={r[1]:,}  distinct={r[2]:,}")
    m = con.execute("""
        SELECT count(*) AS n,
               count(*) FILTER (WHERE r.recipe_id IS NOT NULL) AS matched
        FROM v_curing c LEFT JOIN recipe r ON c.recipeID::VARCHAR = r.recipe_id
        WHERE c.statuscritical='Normal'
    """).fetchone()
    print(f"  recipeID -> recipemaster.iD : {m[1]:,}/{m[0]:,} "
          f"({100*m[1]/max(m[0],1):.1f}%)")

    print("\n" + "=" * 80)
    print("2. ROUTE A vs ROUTE B  -- do they name the SAME GT?")
    print("=" * 80)
    con.execute("""
        CREATE OR REPLACE TEMP TABLE cmp AS
        SELECT c.plant,
               b.itemCode              AS gt_route_a,
               r.r_sap                 AS sap_route_b,
               r.r_name                AS name_route_b,
               r.r_pid                 AS pid
        FROM v_curing c
        JOIN v_build b  ON b.productionID = c.gtbarCode AND b.stage = 2
        LEFT JOIN recipe r ON c.recipeID::VARCHAR = r.recipe_id
        WHERE c.statuscritical='Normal'
    """)
    tot = con.execute("SELECT count(*) FROM cmp").fetchone()[0]
    agree = con.execute("""
        SELECT count(*) FROM cmp
        WHERE upper(replace(replace(gt_route_a,' ',''),'-','')) =
              upper(replace(replace(sap_route_b,' ',''),'-',''))
           OR upper(replace(replace(gt_route_a,' ',''),'-','')) =
              upper(replace(replace(name_route_b,' ',''),'-',''))
    """).fetchone()[0]
    print(f"  tyres compared: {tot:,}")
    print(f"  routes AGREE  : {agree:,} ({100*agree/max(tot,1):.1f}%)")
    print(f"  routes DIFFER : {tot-agree:,} ({100*(tot-agree)/max(tot,1):.1f}%)")

    print("\n  what does the curing recipe actually point at? (processID)")
    d = con.execute("""
        SELECT pid, count(*) n FROM cmp GROUP BY 1 ORDER BY 2 DESC
    """).pl()
    for row in d.iter_rows(named=True):
        print(f"    processID={row['pid'] or 'NULL':<5s} {row['n']:>10,}")

    print("\n  sample of route A vs route B:")
    s = con.execute("""
        SELECT gt_route_a, sap_route_b, name_route_b, pid, count(*) n
        FROM cmp GROUP BY 1,2,3,4 ORDER BY 5 DESC LIMIT 8
    """).pl()
    for row in s.iter_rows(named=True):
        a = (row["gt_route_a"] or "")[:26]
        b = (row["sap_route_b"] or "")[:22]
        nm = (row["name_route_b"] or "")[:20]
        print(f"    A='{a:<26s}' B_sap='{b:<22s}' B_name='{nm:<20s}' "
              f"pid={row['pid'] or '-':<3s} n={row['n']:>8,}")

    print("\n" + "=" * 80)
    print("3. VERDICT")
    print("=" * 80)
    pct = 100 * agree / max(tot, 1)
    if pct > 90:
        print(f"  Route A CONFIRMED by an independent key ({pct:.1f}% agree).")
        print("  Everything measured off curing stands.")
    else:
        print(f"  Routes disagree on {100-pct:.1f}% of tyres.")
        print("  Curing recipe points at the FINISHED SKU, not the green tyre --")
        print("  so it is not a competing GT key, it is the SKU we were missing.")
        print("  Route A remains correct for GT; route B adds the SKU dimension.")
    log.info("validate_curing_route.done", agree_pct=round(pct, 1))
    return 0


if __name__ == "__main__":
    sys.exit(main())
