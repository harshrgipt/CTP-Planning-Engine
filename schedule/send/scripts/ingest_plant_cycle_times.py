"""Ingest the plant's AUTHORITATIVE cycle-time workbooks (cycletime/).

    python -m scripts.ingest_plant_cycle_times --month 2026-07

These files REPLACE the mined cadence. They arrive at a finer grain than the
tables they replace, and that granularity change is the whole point:

    build   was  plant x machine            (build_cadence_*.parquet, 20 rows)
            now  plant x GT x MACHINE-MAKE  (CONTI vs BJ)
    cure    was  plant x press (a re-mined median dwell)
            now  plant x SKU                (minutes)

WHAT THE FILES ACTUALLY CONTAIN -- measured, not assumed
--------------------------------------------------------
PCR BUILDING CYCLE TIME.xlsx / Sheet1, 267 GT rows:
  `Actual Cycle Time CONTI` and `Actual Cycle Time BJ` are **100 % EMPTY**
  (0 of 267 populated). The instruction "prefer Actual, fall back to Norms"
  therefore resolves to Norms for every single GT. Said plainly rather than
  silently.
  `Norms CT CONTI` / `Norms CT BJ` are populated on 221 of 267 (46 are #N/A).
  `CONTI` / `BJ` are the plant's own *operative* columns -- `Cont Min` and
  `BJ Min` are exactly CONTI/60 and BJ/60, so CONTI/BJ (not the raw norm) are
  what the plant itself plans with. We take CONTI/BJ and carry `Norms CT` as
  provenance.

TBR BUILDING CYCLE TIME.xlsx: two sheets, `SKU Cycle Time` (SKU->GT->CT) and
  `GT Cycle Master` (GT->CT). One CT per GT, no machine-make split -- correct,
  because TBR's nine machines are SAV-1..3 + MESNAC-1..6 and the changeover
  master gives all nine identical 10/24 min, i.e. the plant treats them as one
  make for planning.

PCR Curing cycle time 2.xlsx / `PCR CYCLE TIME`: SKU -> cure minutes.
TBR CURING CYCLE TIME.xlsx / `Bladder list`: SKU -> TOTAL CURE TIME (Min),
  plus a `Missing curing cycle time` sheet naming 50 SKUs with no time at all.

THE CONT/BJ MACHINE SPLIT -- derived, with its evidence
------------------------------------------------------
    BJ    = TBMPCR1..5   (machine codes 3401-3405)
    CONTI = TBMPCR6..11  (machine codes 3406-3411)

Three independent sources agree:
  1. `masters/Master_Building_ChangeoverTime_pcr.csv` splits the eleven PCR
     machines into exactly two tiers at the same boundary -- 3401-3405 at
     28/60 min, 3406-3411 at 22/42 min. Two changeover regimes = two machine
     makes.
  2. `PROJECT_STATE.md` line 176 and `CMBC_BUILD_LOG.md` lines 145-146 name
     them: "PCR **BJ** (3401-3405) | 28 min | 60 min" and
     "PCR **CONTI** (3406-11) | 22 | 42".
  3. The machine-code identity 34NN -> TBMPCR<NN>Stage2 is itself measured at
     98.0-100 % purity (PARTITION_AND_CHANGEOVER.md, derive_machine_map).
TBR has no CONT/BJ split and its file supplies none.

CAVITIES: THE PRESS RUNS TWO TYRES PER CYCLE, BOTH PLANTS
---------------------------------------------------------
Measured on July 2026 `v_curing`, clustering tyre events on a press into loads
at a 120 s gap:

    tyres per press load   PCR              TBR
    1                       6.2 %            7.7 %
    2                      92.5 %           92.3 %
    >=3                     1.3 %            0.0 %

and the inter-load gap -- the real press cycle -- is p50 **17.1 min on PCR** and
**59.6 min on TBR**, which is the plant's stated cure time (PCR 13-16, TBR
45-57) plus load/unload. `MouldCountLH`/`MouldCountRH` and the `HM01#HM02`
mould-pair labelling say the same thing.

So a cure time in these files is a PER-CYCLE time and a cycle yields 2 tyres.
Anything that converts cure minutes to a press rate must divide by 2, or it
will halve press capacity.
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl
import polars as pl

ROOT = Path(__file__).resolve().parent.parent
CT = ROOT.parent.parent / "cycletime"
D = ROOT / "warehouse" / "derived"

# ONE definition of the machine-make map, the cavity count and the load/unload
# adder -- `planner/cmbc/plant_ct.py`. Duplicating a constant across two files is
# the defect that produced PARTITION §1g twice; do not re-declare them here.
from planner.cmbc.plant_ct import CAVITIES, LOAD_UNLOAD_MIN, MAKE_OF  # noqa: E402


def _rows(path: Path, sheet: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    out = [r for r in ws.iter_rows(values_only=True)]
    wb.close()
    return out


def _num(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if not v or v.startswith("#"):
            return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f if f > 0 else None


def _clean_gt(s) -> str | None:
    """GT labels in these workbooks carry stray whitespace ('GT 2868 RAN XAT ')
    and the engine's own gt_code does not. Normalise on read, never on write."""
    if s is None:
        return None
    s = re.sub(r"\s+", " ", str(s)).strip()
    return s or None


def _clean_sku(s) -> str | None:
    if s is None:
        return None
    s = str(s).strip()
    return s or None


# --------------------------------------------------------------------------
def pcr_build() -> pl.DataFrame:
    rs = _rows(CT / "PCR BUILDING CYCLE TIME.xlsx", "Sheet1")
    hdr = [str(h).strip() if h else "" for h in rs[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rs[1:]:
        gt = _clean_gt(r[ix["GT code"]])
        if not gt:
            continue
        a_c, a_b = _num(r[ix["Actual Cycle Time CONTI"]]), _num(r[ix["Actual Cycle Time BJ"]])
        n_c, n_b = _num(r[ix["Norms CT CONTI"]]), _num(r[ix["Norms CT BJ"]])
        o_c, o_b = _num(r[ix["CONTI"]]), _num(r[ix["BJ"]])
        # priority: Actual > operative CONTI/BJ > raw norm
        c = a_c or o_c or n_c
        b = a_b or o_b or n_b
        basis = ("actual" if (a_c or a_b) else
                 "operative" if (o_c or o_b) else
                 "norms" if (n_c or n_b) else "none")
        if c is None and b is None:
            basis = "none"
        out.append({"plant": "PCR", "gt_code": gt,
                    "sku": _clean_sku(r[ix["Product Code"]]),
                    "ct_s_CONTI": c, "ct_s_BJ": b,
                    "norm_s_CONTI": n_c, "norm_s_BJ": n_b,
                    "basis": basis})
    df = pl.DataFrame(out)
    # A GT can repeat (variant rows); keep the first row that carries a time.
    return (df.sort("basis").unique(subset=["plant", "gt_code"], keep="first")
              .sort("gt_code"))


def tbr_build() -> tuple[pl.DataFrame, pl.DataFrame]:
    """Returns (by GT-label, by SKU). TBR's GT labels are 'GT 5001'-style and do
    NOT match the engine's size-led TBR gt_code, so the SKU route is the one
    that joins -- the GT route is kept only for the coverage report."""
    rs = _rows(CT / "TBR BUILDING CYCLE TIME.xlsx", "SKU Cycle Time")
    hdr = [str(h).strip() if h else "" for h in rs[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    sku_rows, gt_rows = [], []
    for r in rs[1:]:
        sku = _clean_sku(r[ix["SKU Code"]])
        s = _num(r[ix["CT in Sec"]])
        if not sku or s is None:
            continue
        sku_rows.append({"plant": "TBR", "sku": sku,
                         "plant_gt_label": _clean_gt(r[ix["GT Code"]]),
                         "ct_s": s, "basis": "actual"})
    rs2 = _rows(CT / "TBR BUILDING CYCLE TIME.xlsx", "GT Cycle Master")
    h2 = [str(h).strip() if h else "" for h in rs2[0]]
    j = {h: i for i, h in enumerate(h2)}
    for r in rs2[1:]:
        g = _clean_gt(r[j["GT Code"]])
        s = _num(r[j["CT in Sec"]])
        if g and s is not None:
            gt_rows.append({"plant": "TBR", "plant_gt_label": g, "ct_s": s})
    return (pl.DataFrame(gt_rows).unique(subset=["plant_gt_label"]),
            pl.DataFrame(sku_rows).unique(subset=["plant", "sku"]))


def pcr_cure() -> pl.DataFrame:
    rs = _rows(CT / "PCR Curing cycle time 2.xlsx", "PCR CYCLE TIME")
    hdr = [str(h).strip() if h else "" for h in rs[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rs[1:]:
        sku = _clean_sku(r[ix["SKU Code"]])
        mn = _num(r[ix["Cure time"]])
        if not sku or mn is None:
            continue
        out.append({"plant": "PCR", "sku": sku,
                    "recipe_name": _clean_sku(r[ix["Curing code"]]),
                    "plant_gt_label": _clean_gt(r[ix["GT Code"]]),
                    "cure_min": mn, "basis": "plant"})
    return pl.DataFrame(out).unique(subset=["plant", "sku"])


def tbr_cure() -> tuple[pl.DataFrame, pl.DataFrame]:
    rs = _rows(CT / "TBR CURING CYCLE TIME.xlsx", "Bladder list")
    hdr = None
    out = []
    for r in rs:
        if hdr is None:
            if r and any(str(x).strip().upper() == "SAP CODE" for x in r if x):
                hdr = [str(h).strip().upper() if h else "" for h in r]
            continue
        ix = {h: i for i, h in enumerate(hdr)}
        sku = _clean_sku(r[ix["SAP CODE"]])
        mn = _num(r[ix["TOTAL CURE TIME (MIN)"]])
        if not sku or mn is None:
            continue
        out.append({"plant": "TBR", "sku": sku, "plant_gt_label": None,
                    "recipe_name": _clean_sku(r[ix["RECIPE CODE"]]),
                    "cure_min": mn, "basis": "plant"})
    miss = []
    rs2 = _rows(CT / "TBR CURING CYCLE TIME.xlsx", "Missing curing cycle time")
    for r in rs2[1:]:
        sku = _clean_sku(r[0])
        if sku:
            miss.append({"plant": "TBR", "sku": sku,
                         "description": str(r[1] or ""),
                         "plant_gt_label": _clean_gt(r[2])})
    return (pl.DataFrame(out).unique(subset=["plant", "sku"]),
            pl.DataFrame(miss))


# --------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", default="2026-07")
    a = ap.parse_args()

    print("=" * 92)
    print(f"PLANT CYCLE-TIME INGEST  --  {a.month}")
    print("=" * 92)

    pb = pcr_build()
    tb_gt, tb_sku = tbr_build()
    pc = pcr_cure()
    tc, tc_miss = tbr_cure()

    print(f"\n  read  PCR build {pb.height:>4} GT rows   "
          f"TBR build {tb_sku.height:>4} SKU / {tb_gt.height} GT rows")
    print(f"        PCR cure  {pc.height:>4} SKU rows   "
          f"TBR cure  {tc.height:>4} SKU rows  (+{tc_miss.height} named missing)")
    print("\n  PCR build basis (Actual is 100 % EMPTY in the source file):")
    for r in pb.group_by("basis").agg(pl.len()).sort("basis").iter_rows():
        print(f"      {r[0]:<12} {r[1]:>4} GTs")

    # ---- SKU -> engine gt_code bridge -----------------------------------
    gsm = pl.read_parquet(D / "gt_sku_master.parquet").select(
        ["plant", "sku_code", "gt_code"]).rename({"sku_code": "sku"}).unique()
    dem = pl.read_parquet(ROOT / "masters" / "demand" / f"demand_{a.month}.parquet")
    dem_gt = dem.group_by(["plant", "gt_code"]).agg(pl.col("qty").sum().alias("qty"))
    # ORDER MATTERS AND COST TBR 100 % OF ITS COVERAGE ONCE.
    # `gt_sku_master` maps a TBR SKU to a "GT 5001"-style label; the ENGINE's TBR
    # gt_code is size-led ("10.00 R 20 JDE"). Both are legitimate names for the
    # same tyre and they share not one character. `unique(keep="first")` keeps
    # whichever frame is concatenated first, so putting the master first silently
    # bridged every TBR SKU to a gt_code the engine has never heard of -- TBR
    # coverage read 0/56 GTs, 0.0 % volume, and looked like missing plant data.
    # The demand file's own (sku -> gt_code) is authoritative because it IS the
    # key the planner runs on; the master is the fallback for SKUs not in demand.
    bridge = pl.concat([
        dem.select(["plant", "sku", "gt_code"]).unique(),
        gsm,
    ]).unique(subset=["plant", "sku"], keep="first")

    # PCR GT labels are matched on a WHITESPACE-NORMALISED key and then written
    # back as the ENGINE's own gt_code. The engine carries 'GT  T1457 STAR' with
    # a double space; the workbook carries one. A literal join drops it.
    eng_gt = pl.concat([
        dem.select(["plant", "gt_code"]).unique(),
        pl.read_parquet(D / f"cap_machine_{a.month}.parquet")
          .select(["plant", "gt_code"]).unique()]).unique()
    key = {(r["plant"], re.sub(r"\s+", " ", r["gt_code"]).strip()): r["gt_code"]
           for r in eng_gt.iter_rows(named=True)}

    # ---- BUILD: plant x gt_code x make ----------------------------------
    rows = []
    for r in pb.iter_rows(named=True):
        gt = key.get(("PCR", r["gt_code"]), r["gt_code"])
        for make, col in (("CONTI", "ct_s_CONTI"), ("BJ", "ct_s_BJ")):
            if r[col]:
                rows.append({"plant": "PCR", "gt_code": gt,
                             "make": make, "ct_s": float(r[col]),
                             "basis": r["basis"]})
    tbj = tb_sku.join(bridge, on=["plant", "sku"], how="left")
    for r in tbj.filter(pl.col("gt_code").is_not_null()).iter_rows(named=True):
        rows.append({"plant": "TBR", "gt_code": r["gt_code"], "make": "TBR",
                     "ct_s": float(r["ct_s"]), "basis": "actual"})
    build = (pl.DataFrame(rows)
             .group_by(["plant", "gt_code", "make"])
             .agg(pl.col("ct_s").median(), pl.col("basis").first()))

    # ---- CURE: plant x sku, and a volume-weighted plant x gt_code -------
    cure_sku = pl.concat([
        pc.select(["plant", "sku", "recipe_name", "cure_min", "basis"]),
        tc.select(["plant", "sku", "recipe_name", "cure_min", "basis"])])

    # ---- TWO ROUTES TO gt_code, MEASURED AGAINST EACH OTHER --------------
    # (a) SKU STRING  -- the demand file's own (sku -> gt_code), master fallback
    # (b) RECIPE CODE -- the plant's `Curing code` / `RECIPE CODE` joined to
    #     `recipe_name` in gt_sku_share_<month>, which comes from
    #     Recipemaster 1.xlsx via v_curing.recipeID. That chain is documented at
    #     100 % of cured volume on both plants (scripts/build_gt_sku.py).
    # Route (b) is preferred where it fires: the SKU namespace carries plant-
    # specific variants and has produced two silent join failures on this
    # project, while the recipe code is one flat namespace.
    shr_f = D / f"gt_sku_share_{a.month}.parquet"
    shr = pl.read_parquet(shr_f) if shr_f.exists() else pl.DataFrame(
        schema={"plant": pl.Utf8, "gt_code": pl.Utf8, "sku_code": pl.Utf8,
                "recipe_name": pl.Utf8, "tyres": pl.Int64, "share": pl.Float64})
    rec_bridge = (shr.select(["plant", "recipe_name", "gt_code", "tyres"])
                  .filter(pl.col("recipe_name").is_not_null())
                  .sort("tyres", descending=True)
                  .unique(subset=["plant", "recipe_name"], keep="first")
                  .select(["plant", "recipe_name", "gt_code"]))
    by_sku = (cure_sku.join(bridge, on=["plant", "sku"], how="left")
              .rename({"gt_code": "gt_sku"}))
    both = (by_sku.join(rec_bridge, on=["plant", "recipe_name"], how="left")
            .rename({"gt_code": "gt_rec"}))
    print("\n  CURE gt_code BRIDGE -- recipe-code route vs SKU-string route")
    print(f"    {'plant':<6}{'rows':>6}{'by SKU':>9}{'by RECIPE':>11}"
          f"{'either':>8}{'both':>7}{'DISAGREE':>10}")
    for p in ("PCR", "TBR"):
        z = both.filter(pl.col("plant") == p)
        ns = int(z["gt_sku"].is_not_null().sum())
        nr = int(z["gt_rec"].is_not_null().sum())
        ne = int((z["gt_sku"].is_not_null() | z["gt_rec"].is_not_null()).sum())
        nb = int((z["gt_sku"].is_not_null() & z["gt_rec"].is_not_null()).sum())
        nd = int(z.filter(pl.col("gt_sku").is_not_null()
                          & pl.col("gt_rec").is_not_null()
                          & (pl.col("gt_sku") != pl.col("gt_rec"))).height)
        print(f"    {p:<6}{z.height:>6}{ns:>9}{nr:>11}{ne:>8}{nb:>7}{nd:>10}")
    # RECIPE FIRST, SKU as the fallback -- and record which fired.
    both = both.with_columns([
        pl.coalesce(["gt_rec", "gt_sku"]).alias("gt_code"),
        pl.when(pl.col("gt_rec").is_not_null()).then(pl.lit("recipe"))
          .when(pl.col("gt_sku").is_not_null()).then(pl.lit("sku"))
          .otherwise(pl.lit("none")).alias("bridge_route")])
    cure_sku = both.select(["plant", "sku", "recipe_name", "gt_code",
                            "bridge_route", "cure_min", "basis"])
    cure_gt = (cure_sku.filter(pl.col("gt_code").is_not_null())
               .group_by(["plant", "gt_code"])
               .agg(pl.col("cure_min").median(),
                    pl.col("sku").n_unique().alias("n_sku"),
                    pl.col("bridge_route").first()))

    # ---- COVERAGE against the month's actual demand ---------------------
    print(f"\n  COVERAGE vs {a.month} demand")
    print(f"    {'plant':<6}{'GTs':>6}{'tyres':>10}   build CT      cure CT")
    cov = {}
    for p in ("PCR", "TBR"):
        d = dem_gt.filter(pl.col("plant") == p)
        b = set(build.filter(pl.col("plant") == p)["gt_code"].to_list())
        c = set(cure_gt.filter(pl.col("plant") == p)["gt_code"].to_list())
        tot = float(d["qty"].sum())
        bq = float(d.filter(pl.col("gt_code").is_in(list(b)))["qty"].sum())
        cq = float(d.filter(pl.col("gt_code").is_in(list(c)))["qty"].sum())
        print(f"    {p:<6}{d.height:>6}{tot:>10,.0f}   "
              f"{len(b & set(d['gt_code'].to_list())):>3}/{d.height} GT "
              f"{100 * bq / tot:>5.1f}% vol   "
              f"{len(c & set(d['gt_code'].to_list())):>3}/{d.height} GT "
              f"{100 * cq / tot:>5.1f}% vol")
        cov[p] = {
            "build_missing": sorted(set(d["gt_code"].to_list()) - b),
            "cure_missing": sorted(set(d["gt_code"].to_list()) - c)}
    for p in ("PCR", "TBR"):
        for k in ("build_missing", "cure_missing"):
            if cov[p][k]:
                print(f"    {p} {k}: " + ", ".join(cov[p][k][:12])
                      + (" ..." if len(cov[p][k]) > 12 else ""))

    # ---- machine-level roll-up, demand weighted -------------------------
    cm = pl.read_parquet(D / f"cap_machine_{a.month}.parquet").select(
        ["plant", "gt_code", "machine"]).unique()
    mk = pl.DataFrame([{"machine": m, "make": k} for m, k in MAKE_OF.items()])
    mach = (cm.join(mk, on="machine", how="inner")
            .join(build, on=["plant", "gt_code", "make"], how="inner")
            .join(dem_gt, on=["plant", "gt_code"], how="left")
            .with_columns(pl.col("qty").fill_null(1.0)))
    mach_cad = (mach.group_by(["plant", "machine"])
                .agg(((pl.col("ct_s") * pl.col("qty")).sum()
                      / pl.col("qty").sum()).alias("cadence_s_plant"),
                     pl.len().alias("n_gt")))
    mined = pl.read_parquet(ROOT / "warehouse" / "params"
                            / "build_cadence_2026-08-01.parquet")
    cmp_ = (mach_cad.join(mined.select(["machine", "cadence_s_p50"]),
                          on="machine", how="left")
            .with_columns((100.0 * (pl.col("cadence_s_plant")
                                    / pl.col("cadence_s_p50") - 1)).alias("delta_pct"))
            .sort(["plant", "machine"]))
    print("\n  BUILD CADENCE  plant file (demand-weighted per machine) vs mined")
    print(f"    {'machine':<18}{'make':<7}{'GTs':>5}{'plant s':>10}"
          f"{'mined s':>10}{'delta':>9}")
    for r in cmp_.iter_rows(named=True):
        print(f"    {r['machine']:<18}{MAKE_OF.get(r['machine'], '?'):<7}"
              f"{r['n_gt']:>5}{r['cadence_s_plant']:>10.1f}"
              f"{r['cadence_s_p50']:>10.1f}{r['delta_pct']:>8.1f}%")

    # ---- CURE: validate the plant file against the MES, per GT ----------
    # A plant master is a claim until it is checked against the event stream.
    # Compare the plant's stated cure time to the OBSERVED press cycle (the
    # inter-load gap, loads clustered at a 120 s gap). If the file is right the
    # residual is a near-constant load/unload offset across the whole range.
    from planner.data.warehouse import duck, set_cutoff
    set_cutoff(None)
    obs = duck().execute(f"""
      WITH e AS (SELECT b.plant, b.itemCode gt_code, c.wcID, c.event_ts, c.gtbarCode
          FROM v_build b JOIN v_curing c ON b.productionID=c.gtbarCode
          WHERE b.stage=2 AND c.statuscritical='Normal' AND b.itemCode IS NOT NULL
            AND c.date >= DATE '{a.month}-01'
            AND c.date <  DATE '{a.month}-01' + INTERVAL 1 MONTH),
       s AS (SELECT *, coalesce(date_diff('second', lag(event_ts) OVER
              (PARTITION BY plant,wcID ORDER BY event_ts,gtbarCode), event_ts),
              99999) g FROM e),
       f AS (SELECT *, sum(CASE WHEN g>120 THEN 1 ELSE 0 END) OVER
              (PARTITION BY plant,wcID ORDER BY event_ts,gtbarCode
               ROWS UNBOUNDED PRECEDING) lid FROM s),
       L AS (SELECT plant, gt_code, wcID, lid, min(event_ts) t0, count(*) sz
              FROM f GROUP BY 1,2,3,4),
       G AS (SELECT plant, gt_code, sz, date_diff('second',
              lag(t0) OVER (PARTITION BY plant,wcID ORDER BY t0), t0)/60.0 gmin
              FROM L)
      SELECT plant, gt_code, count(*) n, median(gmin) obs_cycle_min,
             avg(sz) load_sz
      FROM G WHERE gmin BETWEEN 2 AND 300 GROUP BY 1,2 HAVING count(*)>=30""").pl()
    val = (obs.join(cure_gt, on=["plant", "gt_code"], how="inner")
           .with_columns((pl.col("obs_cycle_min") - pl.col("cure_min"))
                         .alias("overhead_min")))
    print("\n  VALIDATION  plant cure time vs OBSERVED press cycle, per GT")
    print(f"    {'plant':<6}{'GTs':>5}{'load sz':>9}{'cure p50':>10}"
          f"{'obs p50':>9}   overhead p25/p50/p75")
    for p in ("PCR", "TBR"):
        s = val.filter(pl.col("plant") == p)
        if not s.height:
            continue
        print(f"    {p:<6}{s.height:>5}{s['load_sz'].mean():>9.2f}"
              f"{s['cure_min'].median():>10.1f}{s['obs_cycle_min'].median():>9.1f}"
              f"   {s['overhead_min'].quantile(0.25):.2f} / "
              f"{s['overhead_min'].median():.2f} / "
              f"{s['overhead_min'].quantile(0.75):.2f}"
              f"   (module uses {LOAD_UNLOAD_MIN[p]})")

    # ---- press rate: new vs the model it replaces -----------------------
    import json as _json
    P0 = _json.loads(sorted((ROOT / "warehouse" / "params")
                            .glob("params_*.json"))[-1].read_text())
    av = {p: float(P0["press_availability"][p]["availability"])
          for p in ("PCR", "TBR")}
    cav = pl.read_parquet(D / "l3_cavities.parquet")
    print("\n  PRESS RATE  new (plant CT x 2 cavities x availability) vs old "
          "(cavities x 3600/cycle_s)")
    print(f"    {'plant':<6}{'presses':>8}{'old t/p-h':>11}{'new t/p-h':>11}"
          f"{'old /p-day':>12}{'new /p-day':>12}{'delta':>8}   observed /p-day")
    for p in ("PCR", "TBR"):
        c = cav.filter(pl.col("plant") == p)
        old = float(c["cavities"].median()) * 3600.0 / float(c["cycle_s"].median())
        g = cure_gt.filter(pl.col("plant") == p).join(
            dem_gt.filter(pl.col("plant") == p), on=["plant", "gt_code"])
        if not g.height:
            continue
        w = float((g["cure_min"] * g["qty"]).sum() / g["qty"].sum())
        new = CAVITIES * 60.0 / (w + LOAD_UNLOAD_MIN[p]) * av[p]
        print(f"    {p:<6}{c.height:>8}{old:>11.3f}{new:>11.3f}"
              f"{old * 24:>12.1f}{new * 24:>12.1f}{100 * (new / old - 1):>7.1f}%"
              f"   mean {float(c['tyres_per_day'].mean()):.1f} "
              f"p50 {float(c['tyres_per_day'].median()):.0f}")

    # ---- write ----------------------------------------------------------
    build.write_parquet(D / "plant_ct_build.parquet")
    cure_sku.write_parquet(D / "plant_ct_cure_sku.parquet")
    cure_gt.write_parquet(D / "plant_ct_cure_gt.parquet")
    tc_miss.write_csv(D / "plant_ct_cure_missing_TBR.csv")
    mk.write_parquet(D / "plant_machine_make.parquet")
    print(f"\n  -> {D/'plant_ct_build.parquet'}      {build.height} rows")
    print(f"  -> {D/'plant_ct_cure_sku.parquet'}   {cure_sku.height} rows")
    print(f"  -> {D/'plant_ct_cure_gt.parquet'}    {cure_gt.height} rows")
    print(f"  -> {D/'plant_ct_cure_missing_TBR.csv'} {tc_miss.height} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
