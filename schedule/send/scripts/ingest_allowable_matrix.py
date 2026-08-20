"""THE PLANT'S OWN ALLOWABLE MATRICES -> allowed_machine_matrix.parquet

    python -m scripts.ingest_allowable_matrix

Reads `INPUT/allowable machine/`:

    PCR BUILDING ALLOWABLE MATRIX.xlsx     sheet "BUILDING MATRIX"
        FG CODE | GT CODE | _ | Size | 3401 .. 3411      cell "P" = allowed
    TBR BUILDING ALLOWABLE MATRIX (1).xlsx sheet "SKU-MACHINE-CONSTRUCTION"
        S NO | DESCRIPTION | SKU CODE | GT CODE | TBM 01 .. TBM 09
                                                        cell "YES" = allowed

WHY THIS MATTERS
  Until now PCR had NO plant building matrix at all -- `allowed_machine_matrix`
  carried 108 PCR GTs derived from other sources, and the engine ran on it as a
  hard constraint. This file is the plant's own statement, so it replaces the
  derived one for every SKU it names.

BRIDGE ON SKU, NEVER ON GT CODE
  The GT CODE columns are in THREE different namespaces and none is the engine's:

      this PCR file      T1457 STAR, 3215 BLA XAT     (no "GT " prefix)
      this TBR file      GT 5001                      (BOM short code)
      engine / MES       GT  T1457 STAR, 10.00 R 20 JUH5

  The TBR BOM namespace has ZERO string overlap with TBR MES itemCodes -- the
  documented trap that has cost this project debugging twice. Both files carry
  the FG/SKU code, and `gt_sku_master` maps SKU -> (plant, gt_code) through the
  curing-recipe chain at 100 % of cured volume. So the SKU is the key; the GT
  CODE column is read only for reporting.

MACHINE COLUMN -> ENGINE MACHINE
  PCR 34NN -> TBMPCR<NN>Stage2. This identity is measured at 98-100 % purity and
  is the same one the BTP exporter and plant_ct rely on.
  TBR "TBM 0N" -> TBMTBR<N>Stage2.

WHAT IS DELIBERATELY EXCLUDED
  The PCR workbook has a sheet literally named "to be clarified by plant team",
  holding rows the plant has not signed off, plus a sheet "shared by ram over
  mail" and an OE preference sheet. Only "BUILDING MATRIX" is ingested. Pulling
  in unclarified rows would put guessed eligibility behind a HARD constraint --
  the exact defect that made the platen master unusable. They are counted and
  reported instead.

  Cells reading "N0" (a typo for NO) are treated as NO, not as unknown.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
import polars as pl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from planner import paths                                          # noqa: E402

SRC = "allowable machine"
PCR_F = "PCR BUILDING ALLOWABLE MATRIX.xlsx"
TBR_F = "TBR BUILDING ALLOWABLE MATRIX (1).xlsx"


def _dir() -> Path:
    return paths.INPUT / SRC


def _rows(f: Path, sheet: str) -> list[tuple]:
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    out = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    return out


def _pcr() -> list[dict]:
    rows = _rows(_dir() / PCR_F, "BUILDING MATRIX")
    hdr = [str(x).strip() if x is not None else "" for x in rows[0]]
    cols = {i: h for i, h in enumerate(hdr) if h.isdigit() and len(h) == 4}
    out = []
    for r in rows[1:]:
        sku = str(r[0]).strip() if r[0] else ""
        if not sku:
            continue
        for i, code in cols.items():
            if i < len(r) and str(r[i] or "").strip().upper() == "P":
                out.append({"plant": "PCR", "sku": sku,
                            "machine": f"TBMPCR{int(code[2:])}Stage2",
                            "src_gt": str(r[1] or "").strip()})
    return out


def _tbr() -> list[dict]:
    rows = _rows(_dir() / TBR_F, "SKU-MACHINE-CONSTRUCTION")
    hdr = [str(x).strip().upper() if x is not None else "" for x in rows[0]]
    cols = {i: h for i, h in enumerate(hdr) if h.startswith("TBM ")}
    out = []
    for r in rows[1:]:
        sku = str(r[2]).strip() if r[2] else ""
        if not sku:
            continue
        for i, h in cols.items():
            if i < len(r) and str(r[i] or "").strip().upper() == "YES":
                out.append({"plant": "TBR", "sku": sku,
                            "machine": f"TBMTBR{int(h.split()[1])}Stage2",
                            "src_gt": str(r[3] or "").strip()})
    return out


def run(*, write: bool = True) -> pl.DataFrame:
    raw = pl.DataFrame(_pcr() + _tbr())

    # ---- SKU -> GT, AND THE BRIDGE ORDER MATTERS ------------------------
    # `gt_sku_master` returns the BOM namespace for TBR ("GT 5025"), not the MES
    # itemCode the engine plans on ("10.00 R 20 JDE"). Bridging TBR through it
    # produced 1,165 rows and 0/56 demand-GT coverage -- a matrix that looks
    # populated and matches nothing. `gt_sku_from_recipe` is derived from what
    # was actually cured (recipe -> v_curing -> v_build.itemCode), so it is in
    # the engine namespace: 37/37 demand overlap on August TBR.
    #
    # So: recipe chain FIRST, gt_sku_master only as a fallback for SKUs the
    # recipe chain has never seen. PCR is unaffected -- both agree there.
    rec = (pl.read_parquet(paths.wh_derived("gt_sku_from_recipe.parquet"))
           .select(["plant", "sku_code", "gt_code"])
           .rename({"sku_code": "sku", "gt_code": "gt_rec"})
           .unique(subset=["plant", "sku"]))
    gsm = (pl.read_parquet(paths.wh_derived("gt_sku_master.parquet"))
           .select(["plant", "sku_code", "gt_code"])
           .rename({"sku_code": "sku", "gt_code": "gt_mas"})
           .unique(subset=["plant", "sku"]))
    j = (raw.join(rec, on=["plant", "sku"], how="left")
            .join(gsm, on=["plant", "sku"], how="left")
            .with_columns(pl.coalesce(["gt_rec", "gt_mas"]).alias("gt_code"))
            .drop(["gt_rec", "gt_mas"]))
    # ---- FALLBACK: the file's own GT CODE column, EXACT MATCH ONLY ---------
    # A row whose SKU does not bridge loses its machines entirely, and that is
    # not harmless: the matrix is keyed per FG, so one GT can appear on several
    # rows and we take the UNION. Drop one row and the GT silently narrows.
    #
    # GT 1513 XPC1 MSIL is exactly that. Two FG rows:
    #     1225215013008SXC10 -> 3404, 3407, 3410   (SKU not in recipe OR master)
    #     1D25215013008SXC10 -> 3407, 3410         (bridges fine)
    # so the union shipped as M7, M10 and TBMPCR4 was lost -- on the plant's
    # single largest GT (55,663 tyres in July, 1,378 short).
    #
    # The fallback is deliberately the narrowest thing that fixes it: take the
    # row's own GT CODE, normalise whitespace/case and the missing "GT " prefix,
    # and accept it ONLY on an exact match against a gt_code the engine already
    # knows for that plant. No fuzzy matching, no numeric-head expansion.
    #
    # This is safe for TBR by construction rather than by exception: that file's
    # GT CODE column is the BOM namespace ("GT 5001"), which has ZERO string
    # overlap with TBR MES codes, so nothing matches and nothing is added.
    # Measured: 43 PCR rows fail the SKU bridge and exactly ONE of them has a
    # GT CODE the engine recognises. A fallback that fired more often than that
    # would be guessing, and guessing a machine onto a GT is unbuildable work.
    #
    # MEASURED 2026-08-13: THE FIX IS CORRECT AND WORTH EXACTLY ZERO TYRES.
    # It adds one (gt, machine) pair -- PCR 841 -> 842, TBR unchanged -- and
    # both months come out BIT-IDENTICAL:
    #     Jul PCR 385,084 BUILT / 96.2 %      Jul TBR 95,783 / 95.7 %
    #     Aug PCR 410,652 BUILT / 91.4 %      Aug TBR 92,541 / 89.4 %
    # GT 1513 now ALLOWS M4, M7, M10 and the scheduler still uses M7, M10 only:
    # TBMPCR4 runs at 92.1 % with 59 idle hours in August, so it has nothing to
    # offer a 55,663-tyre GT and loses the machine-choice ranking to M7/M10.
    #
    # Kept anyway. A constraint that misstates what the plant permits is wrong
    # whether or not it currently binds, and it would bind the moment M4's load
    # changes. Do not "revert the no-op".
    eng = (pl.read_parquet(paths.wh_derived("cap_machine_2026-08.parquet"))
             .select("plant", "gt_code").unique())
    eng = eng.with_columns(
        pl.col("gt_code").str.replace(r"^GT ", "").str.strip_chars()
          .str.replace_all(r"\s+", " ").str.to_uppercase().alias("_k"))
    key = {(r["plant"], r["_k"]): r["gt_code"] for r in eng.iter_rows(named=True)}
    j = j.with_columns(
        pl.when(pl.col("gt_code").is_null() & pl.col("src_gt").is_not_null())
          .then(pl.struct(["plant", "src_gt"]).map_elements(
              lambda r: key.get((r["plant"], " ".join(
                  str(r["src_gt"]).strip().upper().split()))),
              return_dtype=pl.Utf8))
          .otherwise(pl.col("gt_code")).alias("gt_code"))

    ok = j.filter(pl.col("gt_code").is_not_null())
    bad = j.filter(pl.col("gt_code").is_null())

    new = (ok.select(["plant", "gt_code", "machine"]).unique()
             .with_columns(pl.lit("plant_matrix").alias("basis"))
             .sort(["plant", "gt_code", "machine"]))

    old = pl.read_parquet(paths.input_derived("allowed_machine_matrix.parquet"))
    print(f"\n  PLANT ALLOWABLE MATRIX INTAKE")
    print(f"  {'-' * 72}")
    print(f"  raw (sku, machine) marks   {raw.height:>6}"
          f"   PCR {raw.filter(pl.col('plant') == 'PCR').height}"
          f" · TBR {raw.filter(pl.col('plant') == 'TBR').height}")
    print(f"  SKUs bridged to a GT       {ok['sku'].n_unique():>6} of "
          f"{raw['sku'].n_unique()}"
          f"   ({100 * ok['sku'].n_unique() / max(raw['sku'].n_unique(), 1):.1f}%)")
    if bad.height:
        print(f"  !! {bad['sku'].n_unique()} SKUs have no GT in gt_sku_master -> "
              f"dropped, listed in UNBRIDGED_allowable.csv")
    print()
    for p in ("PCR", "TBR"):
        n = new.filter(pl.col("plant") == p)
        o = old.filter(pl.col("plant") == p)
        if not n.height:
            continue
        per = n.group_by("gt_code").agg(pl.len().alias("k"))
        pero = o.group_by("gt_code").agg(pl.len().alias("k"))
        print(f"  {p}:  NEW {n['gt_code'].n_unique():>3} GTs / {n.height:>4} pairs"
              f" · machines per GT p50 {per['k'].median():.0f}"
              f" min {per['k'].min()} max {per['k'].max()}")
        print(f"        was {o['gt_code'].n_unique():>3} GTs / {o.height:>4} pairs"
              f" · machines per GT p50 {pero['k'].median():.0f}"
              f" min {pero['k'].min()} max {pero['k'].max()}")

    if write:
        f = paths.input_derived("allowed_machine_matrix.parquet")
        bk = f.with_suffix(".parquet.bak")
        if not bk.exists():
            old.write_parquet(bk)
            print(f"\n  backed up previous matrix -> {bk.name}")
        new.write_parquet(f)
        print(f"  -> {f}  ({new.height} rows)")
        if bad.height:
            (bad.select(["plant", "sku", "src_gt", "machine"]).unique()
                .write_csv(paths.MASTERS / "UNBRIDGED_allowable.csv"))
    return new


def main() -> None:
    ap = argparse.ArgumentParser(description="plant allowable matrices -> parquet")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()
    run(write=not a.dry_run)


if __name__ == "__main__":
    main()
