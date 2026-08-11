"""Assemble INPUT/ -- every file the engine needs, in one place, with a manifest.

    python scripts/build_input.py

Copies the raw masters and DERIVES the four that were missing or wrong:
  * tt_tl.parquet          R8 -- was declared unimplementable; it was a regex bug
  * gt_size.parquet        R6/R7 -- raises size coverage from 71/64%
  * aging_limits.parquet   R5/R17 -- per-SKU MaxAging/Minaging, never used before
  * mould_cavity.parquet   R3/R14 -- curingCapacity per SKU

TT/TL PARSE NOTE
  `\\bTT\\b` does NOT match "10.00R20_JDC3_16PR_K_TT" because underscore is a word
  character.  That single bug made R8 look impossible.  Use an explicit delimiter
  class instead.
"""
from __future__ import annotations

import re
import shutil
import warnings
from pathlib import Path

import polars as pl

warnings.filterwarnings("ignore")
from openpyxl import load_workbook                            # noqa: E402
from planner import paths

ROOT = Path(__file__).resolve().parent.parent
SRC = paths.RAW
OUT = paths.INPUT

# underscore-, slash-, space- and dash-delimited.  NOT \b -- see module docstring.
TTTL = re.compile(r"(?:^|[\s_/\-])(TT|TL)(?:$|[\s_/\-])", re.I)
SIZE = re.compile(r"(\d{1,3}(?:\.\d+)?)\s*(?:/\s*(\d{2,3}))?\s*[R/]\s*(\d{2}(?:\.\d)?)", re.I)


def tag(text: str | None) -> str | None:
    if not text:
        return None
    m = TTTL.search(str(text))
    return m.group(1).upper() if m else None


def size_of(text: str | None) -> str | None:
    if not text:
        return None
    m = SIZE.search(str(text).replace(" ", ""))
    return f"R{m.group(3)}" if m else None


def sheet(path: Path, name: str | None = None) -> list[tuple]:
    wb = load_workbook(path, read_only=True)
    ws = wb[name] if name else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def main() -> None:
    OUT.mkdir(exist_ok=True)
    (OUT / "raw").mkdir(exist_ok=True)
    (OUT / "derived").mkdir(exist_ok=True)
    man: list[dict] = []

    # ---- 1. raw masters, copied verbatim --------------------------------
    raws = ["Master_Building_ChangeoverTime_pcr.csv",
            "Master_Building_ChangeoverTime_tbr.csv",
            "Master_Mapping_Mould_SKU.csv",
            "TBR BUILDING ALLOWABLE MATRIX.xlsx",
            "ALL PCR CTP SKUS.xlsx",
            "Recipemaster 1.xlsx", "recipelookup 1.xlsx", "wcmaster 1.xlsx",
            "CTP Set up building ,curing and inspection (1) 2.xlsx",
            "Building Business Rules.docx"]
    for r in raws:
        p = SRC / r
        if p.exists():
            shutil.copy2(p, OUT / "raw" / r)
            man.append({"file": f"raw/{r}", "kind": "raw", "status": "copied"})
        else:
            man.append({"file": f"raw/{r}", "kind": "raw", "status": "MISSING"})

    # ---- 2. TT / TL  (R8) -----------------------------------------------
    rec = []
    rows = sheet(paths.raw("TBR BUILDING ALLOWABLE MATRIX.xlsx"),
                 "SKU-MACHINE-CONSTRUCTION")
    h = [str(x) if x else "" for x in rows[0]]
    ig, isk, idc = h.index("GT CODE"), h.index("SKU CODE"), h.index("DESCRIPTION")
    for r in rows[1:]:
        if r[ig] is None:
            continue
        rec.append({"plant": "TBR", "gt_code": str(r[ig]).strip(),
                    "sku": str(r[isk] or "").strip(),
                    "descr": str(r[idc] or ""), "src": "TBR allowable matrix"})
    mo = pl.read_csv(paths.raw("Master_Mapping_Mould_SKU.csv"), infer_schema_length=0)
    for x in mo.iter_rows(named=True):
        rec.append({"plant": "", "gt_code": "", "sku": str(x["Matl.Code"] or ""),
                    "descr": str(x["Matl.Description"] or ""),
                    "src": "Master_Mapping_Mould_SKU"})
    tt = pl.DataFrame(rec)
    tt = tt.with_columns(
        pl.col("descr").map_elements(tag, return_dtype=pl.Utf8).alias("tt_tl"),
        pl.col("descr").map_elements(size_of, return_dtype=pl.Utf8).alias("rim"))
    tt = tt.filter(pl.col("tt_tl").is_not_null())
    tt.write_parquet(OUT / "derived" / "tt_tl.parquet")
    man.append({"file": "derived/tt_tl.parquet", "kind": "R8",
                "status": f"{tt.height} rows, "
                          f"{tt.filter(pl.col('tt_tl')=='TT').height} TT / "
                          f"{tt.filter(pl.col('tt_tl')=='TL').height} TL"})

    # ---- 3. size / rim inch  (R6, R7) -----------------------------------
    sz = tt.filter(pl.col("rim").is_not_null()).select(
        ["plant", "gt_code", "sku", "rim"]).unique()
    sz.write_parquet(OUT / "derived" / "gt_size.parquet")
    man.append({"file": "derived/gt_size.parquet", "kind": "R6/R7",
                "status": f"{sz.height} rows, {sz['rim'].n_unique()} rims"})

    # ---- 4. aging limits per SKU  (R5, R17) -----------------------------
    rows = sheet(paths.raw("Recipemaster 1.xlsx"))
    h = [str(x) if x else "" for x in rows[0]]
    def col(n):
        return h.index(n) if n in h else None
    ci = {n: col(n) for n in ["SAPMaterialCode", "description", "tyreSize",
                              "MaxAging", "Minaging", "AgingInterlock",
                              "curingCapacity", "SpecWeight", "recipeNo"]}
    ag = []
    for r in rows[1:]:
        def g(n):
            i = ci[n]
            return r[i] if i is not None and i < len(r) else None
        ag.append({"sku": str(g("SAPMaterialCode") or "").strip(),
                   "descr": str(g("description") or ""),
                   "tyre_size": str(g("tyreSize") or ""),
                   "max_aging": g("MaxAging"), "min_aging": g("Minaging"),
                   "aging_interlock": g("AgingInterlock"),
                   "curing_capacity": g("curingCapacity"),
                   "spec_weight": g("SpecWeight")})
    agd = pl.DataFrame(ag).with_columns(
        pl.col("max_aging").cast(pl.Float64, strict=False),
        pl.col("min_aging").cast(pl.Float64, strict=False),
        pl.col("aging_interlock").cast(pl.Int64, strict=False),
        pl.col("curing_capacity").cast(pl.Float64, strict=False))
    agd = agd.with_columns(
        pl.col("descr").map_elements(tag, return_dtype=pl.Utf8).alias("tt_tl"),
        pl.col("descr").map_elements(size_of, return_dtype=pl.Utf8).alias("rim"))
    agd.write_parquet(OUT / "derived" / "aging_limits.parquet")
    n48 = agd.filter(pl.col("max_aging") == 48).height
    n72 = agd.filter(pl.col("max_aging") == 72).height
    man.append({"file": "derived/aging_limits.parquet", "kind": "R5/R17",
                "status": f"{agd.height} SKUs; MaxAging 48h={n48} 72h={n72}; "
                          f"interlock on={agd.filter(pl.col('aging_interlock')==1).height}"})

    # ---- 5. mould cavity  (R3, R14) -------------------------------------
    cav = agd.filter(pl.col("curing_capacity") > 0).select(
        ["sku", "curing_capacity"]).unique()
    cav.write_parquet(OUT / "derived" / "mould_cavity.parquet")
    man.append({"file": "derived/mould_cavity.parquet", "kind": "R3/R14",
                "status": f"{cav.height} SKUs with a cavity/capacity value"})

    # ---- 6. existing derived warehouse assets ---------------------------
    for f in ["allowed_press_matrix", "allowed_machine_matrix", "capacity_press_day",
              "capacity_machine_day", "calendar_shifts", "cycle_time_curing",
              "cycle_time_building", "opening_gt_inventory", "sku_construction",
              "tbr_machine_certified", "gt_sku_master", "lot_size",
              "press_platen_master"]:
        p = ROOT / "warehouse" / "derived" / f"{f}.parquet"
        if p.exists():
            shutil.copy2(p, OUT / "derived" / f"{f}.parquet")
            man.append({"file": f"derived/{f}.parquet", "kind": "warehouse",
                        "status": f"{pl.read_parquet(p).height} rows"})
    dm = ROOT / "masters" / "demand"
    if dm.exists():
        (OUT / "demand").mkdir(exist_ok=True)
        n = 0
        for p in dm.glob("*.parquet"):
            shutil.copy2(p, OUT / "demand" / p.name); n += 1
        man.append({"file": "demand/", "kind": "R1", "status": f"{n} months"})

    pl.DataFrame(man).write_csv(OUT / "MANIFEST.csv")
    print(f"INPUT assembled at {OUT}")
    for m in man:
        print(f"  {m['file']:<44}{m['kind']:<12}{m['status']}")


if __name__ == "__main__":
    main()
