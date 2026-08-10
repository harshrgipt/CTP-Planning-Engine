"""Ingest the WORK CENTRE MASTER -- the platen master we have been missing.

    python scripts/ingest_wcmaster.py [path/to/wcmaster.xlsx]

This closes the biggest input gap in the engine. Until now press eligibility was
mined from history, which can only say what HAS run, never what MAY run -- and
36-45% of press-GT pairs are new every month, so history is structurally
incomplete as a feasibility set.

The work centre master carries PLATEN SIZE in the description:

    processID  5   84 rows   65.5" GRM TYRE CURING PRESS      -> TBR
    processID  8   95 rows   48" / 45" / 36" LTM CURING PRESS -> PCR
    processID 27    5 rows   36" LTM CURING PRESS             -> PCR
    processID  4    9 rows   UNISTAGE MACHINE / TBM           -> building
    processID  7   12 rows   2nd STAGE TBM MODULE             -> building

Platen size is the physical gate (v2 Retraction 2): a 36" platen takes rim
12-14, a 48" platen rim 14-20. Presses of the same platen are interchangeable
for a GT of the right rim, whether or not that pair appears in history.

Many rows have a BLANK description (56 of 84 on processID 5, 12 of 95 on 8), so
platen is filled from the block the press number falls in -- presses are
numbered in contiguous platen blocks, which the populated rows establish.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import polars as pl

from planner.config import CONFIG
from planner.data.warehouse import duck, set_cutoff
from planner.runs.logger import log

SRC = Path("C:/Users/91810/Downloads/send/wcmaster 1.xlsx")
PRESS_PIDS = {"5", "8", "27"}
BUILD_PIDS = {"4", "7"}


def _clean(x) -> str:
    return "".join(ch for ch in str(x) if ord(ch) < 128).strip() if x is not None else ""


def load(src: Path) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [_clean(x) for x in rows[0]]
    return [dict(zip(hdr, [_clean(x) for x in r])) for r in rows[1:] if any(r)]


def main(src: Path) -> int:
    set_cutoff(None)
    data = load(src)
    log.info("wcmaster.loaded", rows=len(data), src=str(src))

    # ---- presses ---------------------------------------------------------
    # THE CROSSWALK: MES `wcID` is the work-centre master's `iD`, NOT its
    # `name`. `name` is the plant-floor press number (4504, 5401); `wcID` is
    # the row id (1-294). Joining on name matched 0 of 175; joining on iD
    # matches 175 of 175. Getting this backwards is why the platen master
    # looked unusable.
    rows = []
    for d in data:
        if d.get("processID") not in PRESS_PIDS:
            continue
        m = re.match(r'^\s*([\d.]+)"', d.get("description", ""))
        rows.append({"press": d["iD"], "press_no": d["name"],
                     "processID": d["processID"],
                     "platen_in": float(m.group(1)) if m else None,
                     "description": d.get("description", "")})
    p = pl.DataFrame(rows)
    # Fill blanks from the surrounding numeric block: presses are numbered in
    # contiguous platen runs, which the populated rows establish.
    p = p.with_columns(pl.col("press_no").cast(pl.Int64, strict=False).alias("_n")).sort(["processID", "_n"])
    p = p.with_columns(
        pl.col("platen_in").forward_fill().backward_fill().over("processID"))
    # plant follows the platen: 65.5" is TBR geometry, 36-48" is PCR
    p = p.with_columns(
        pl.when(pl.col("platen_in") >= 60).then(pl.lit("TBR"))
          .otherwise(pl.lit("PCR")).alias("plant"))
    # rim range from platen (v2 Retraction 2)
    p = p.with_columns(
        pl.when(pl.col("platen_in") <= 36).then(pl.lit(12))
          .when(pl.col("platen_in") <= 48).then(pl.lit(14))
          .otherwise(pl.lit(16)).alias("rim_lo"),
        pl.when(pl.col("platen_in") <= 36).then(pl.lit(14))
          .when(pl.col("platen_in") <= 48).then(pl.lit(20))
          .otherwise(pl.lit(24)).alias("rim_hi"))
    p = p.drop("_n").sort(["plant", "platen_in", "press_no"])

    print("PRESS MASTER from work centre file:")
    for r in (p.group_by(["plant", "platen_in", "rim_lo", "rim_hi"]).len()
              .sort(["plant", "platen_in"]).iter_rows(named=True)):
        print(f"  {r['plant']}  platen {r['platen_in']:5.1f}\"  rim "
              f"{r['rim_lo']}-{r['rim_hi']}  -> {r['len']:3d} presses")

    # ---- reconcile against MES -------------------------------------------
    mes = duck().execute(
        "SELECT DISTINCT plant, wcID::VARCHAR AS press FROM v_curing").pl()
    j = mes.join(p.select(["press", "press_no", "plant", "platen_in"]).rename(
        {"plant": "wc_plant"}), on="press", how="left")
    matched = j.filter(pl.col("platen_in").is_not_null())
    print(f"\nRECONCILIATION vs MES: {matched.height}/{mes.height} presses matched "
          f"({100*matched.height/mes.height:.1f}%)")
    agree = matched.filter(pl.col("plant") == pl.col("wc_plant")).height
    print(f"  plant agrees on {agree}/{matched.height} "
          f"({100*agree/max(matched.height,1):.1f}%)")
    miss = j.filter(pl.col("platen_in").is_null())
    if miss.height:
        print(f"  {miss.height} MES presses NOT in the work centre master: "
              f"{sorted(miss['press'].to_list())[:12]}")
    extra = p.join(mes.select("press"), on="press", how="anti")
    if extra.height:
        print(f"  {extra.height} master presses never seen in MES "
              f"(spare/idle capacity)")

    # ---- machines --------------------------------------------------------
    mrows = [{"wc_id": d["iD"], "machine": d["name"], "processID": d["processID"],
              "description": d.get("description", "")}
             for d in data if d.get("processID") in BUILD_PIDS]
    m = pl.DataFrame(mrows).sort("machine")
    print(f"\nBUILDING MACHINE MASTER: {m.height} work centres "
          f"(processID 4 = UNISTAGE, 7 = 2nd stage TBM)")

    out = CONFIG.paths.warehouse / "derived"
    out.mkdir(parents=True, exist_ok=True)
    p.write_parquet(out / "press_platen_master.parquet", compression="zstd")
    m.write_parquet(out / "machine_wc_master.parquet", compression="zstd")
    print(f"\nWROTE {out/'press_platen_master.parquet'}  ({p.height} presses)")
    print(f"WROTE {out/'machine_wc_master.parquet'}  ({m.height} machines)")
    log.info("wcmaster.done", presses=p.height, machines=m.height,
             matched=matched.height)
    return 0


if __name__ == "__main__":
    sys.exit(main(Path(sys.argv[1]) if len(sys.argv) > 1 else SRC))
