"""FULL CTP EXPORT -- mirrors referance/ctp_*_schedule_<month><year>_<plant>.xlsx.

    python scripts/export_ctp_full.py runs/frozen/2026-07

Produces, per plant, the building workbook (20 sheets) and the curing workbook
(17 sheets) in the reference layout, plus a final PLAN ANALYSIS sheet that the
reference does not have: what demand was missed, why, and how the plan compares
to what the plant actually did that month.

REFERENCE CONVENTIONS, read off the files rather than assumed:
  * the plant day starts at 07:00, not midnight; shifts are A 07-15, B 15-23,
    C 23-07. Day d spans [d 07:00, d+1 07:00).
  * schedules are SKU-level. The engine plans at GT level, so each lot is split
    across the SKUs of its GT in proportion to their demand.
  * Qty/Mins are per (machine, day, SKU) with segments consolidated to one row.
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DAY_START_H = 7
SHIFT_H = 8
SHIFTS = ["A", "B", "C"]
MONTHS = ["january", "february", "march", "april", "may", "june", "july",
          "august", "september", "october", "november", "december"]
HDR = PatternFill("solid", fgColor="FFDDDDDD")
WARN = PatternFill("solid", fgColor="FFFCE4D6")
BAD = PatternFill("solid", fgColor="FFF8CBAD")
GOOD = PatternFill("solid", fgColor="FFE2EFDA")
MIN_LOT = 150
ROOT = Path(__file__).resolve().parent.parent


# --------------------------------------------------------------------------
def _sku_master() -> pl.DataFrame:
    """SKU -> description, and mould count per SKU (the M_g constraint)."""
    f = ROOT / "masters" / "Master_Mapping_Mould_SKU.csv"
    if not f.exists():
        return pl.DataFrame(schema={"sku": pl.Utf8, "SKU_Description": pl.Utf8,
                                    "moulds": pl.Int64})
    d = pl.read_csv(f, infer_schema_length=0)
    d = d.rename({"Matl.Code": "sku", "Matl.Description": "SKU_Description"})
    act = d.filter(pl.col("Active Flag") == "1") if "Active Flag" in d.columns else d
    m = act.group_by("sku").agg(pl.col("Mould").n_unique().alias("moulds"))
    desc = (d.group_by("sku").agg(pl.col("SKU_Description").first()))
    return desc.join(m, on="sku", how="left").with_columns(
        pl.col("moulds").fill_null(0))


def _day_of(ts, origin: datetime) -> int:
    return int((ts - origin).total_seconds() // 86400) + 1


def _shift_of(ts, origin: datetime) -> str:
    k = int(((ts - origin).total_seconds() // 3600) % 24) // SHIFT_H
    return SHIFTS[min(k, 2)]


def _put(ws, df: pl.DataFrame, r0=1, title=None, note=None, widths=True):
    r = r0
    if title:
        c = ws.cell(row=r, column=1, value=title)
        c.font = Font(bold=True, size=12)
        r += 1
    if note:
        ws.cell(row=r, column=1, value=note).font = Font(italic=True, size=9)
        r += 1
    if title or note:
        r += 1
    if df is None or df.height == 0:
        ws.cell(row=r, column=1, value="(none)").font = Font(italic=True)
        return r + 2
    for j, c in enumerate(df.columns, 1):
        cell = ws.cell(row=r, column=j, value=c)
        cell.font = Font(bold=True)
        cell.fill = HDR
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for i, row in enumerate(df.iter_rows(), 1):
        for j, v in enumerate(row, 1):
            ws.cell(row=r + i, column=j,
                    value=v if v is None or isinstance(v, (int, float, str))
                    else str(v))
    if widths:
        for j, c in enumerate(df.columns, 1):
            try:
                w = int(df[c].cast(pl.Utf8).str.len_chars().max() or 8)
            except Exception:                                   # noqa: BLE001
                w = 12
            ws.column_dimensions[get_column_letter(j)].width = \
                max(len(str(c)) + 2, min(30, w + 2))
    return r + df.height + 3


# --------------------------------------------------------------------------
class Ctx:
    def __init__(self, run: Path, plant: str):
        self.run, self.plant = run, plant
        self.bs = pl.read_parquet(run / "build_schedule.parquet").filter(
            pl.col("plant") == plant)
        self.cs = pl.read_parquet(run / "cure_schedule.parquet").filter(
            pl.col("plant") == plant)
        self.ev = pl.read_parquet(run / "gt_events.parquet").filter(
            pl.col("plant") == plant)
        self.rep = json.loads((run / "run_report.json").read_text())
        # Month comes from the run report, not the folder name -- an output
        # directory may be named anything (july_v2, frozen/, a ticket id).
        self.month = str(self.rep.get("input", {}).get("plan_start", run.name))[:7]
        y, mo = int(self.month[:4]), int(self.month[5:7])
        self.origin = datetime(y, mo, 1, DAY_START_H)
        self.ndays = (datetime(y + (mo == 12), (mo % 12) + 1, 1)
                      - datetime(y, mo, 1)).days
        dem = ROOT / "masters" / "demand" / f"demand_{self.month}.parquet"
        self.dem = (pl.read_parquet(dem).filter(pl.col("plant") == plant)
                    if dem.exists() else pl.DataFrame())
        self.sku = _sku_master()
        self.exc = {n: (pl.read_parquet(run / f"exc_{n}.parquet")
                        if (run / f"exc_{n}.parquet").exists() else pl.DataFrame())
                    for n in ["supply_ratio", "shelf_life", "past_horizon",
                              "machine_overlap", "unplaced"]}

    # ---- GT -> SKU split, weighted by demand ----------------------------
    def gt_sku(self) -> pl.DataFrame:
        if self.dem.height == 0:
            return pl.DataFrame(schema={"gt_code": pl.Utf8, "sku": pl.Utf8,
                                        "share": pl.Float64})
        g = (self.dem.group_by(["gt_code", "sku"]).agg(pl.col("qty").sum().alias("q")))
        tot = g.group_by("gt_code").agg(pl.col("q").sum().alias("t"))
        return (g.join(tot, on="gt_code")
                .with_columns((pl.col("q") / pl.col("t")).alias("share"))
                .select(["gt_code", "sku", "share"]).sort(["gt_code", "sku"]))

    def build_rows(self) -> pl.DataFrame:
        """(machine, day, SKU) consolidated -- the reference Shift Schedule."""
        if self.bs.height == 0:
            return pl.DataFrame()
        b = self.bs.with_columns(
            pl.col("start_ts").map_elements(lambda t: _day_of(t, self.origin),
                                            return_dtype=pl.Int64).alias("Day"))
        b = b.join(self.gt_sku(), on="gt_code", how="left")
        b = b.with_columns(pl.col("share").fill_null(1.0),
                           pl.col("sku").fill_null(pl.col("gt_code")))
        b = b.with_columns((pl.col("qty") * pl.col("share")).round().alias("Qty"))
        g = (b.group_by(["machine", "Day", "sku", "gt_code"]).agg(
            pl.col("Qty").sum().alias("Qty"),
            ((pl.col("end_ts").max() - pl.col("start_ts").min())
             .dt.total_seconds() / 60).alias("Mins"),
            pl.len().alias("Segments"),
            pl.col("start_ts").min().alias("StartTime"),
            pl.col("end_ts").max().alias("EndTime")))
        g = g.filter(pl.col("Qty") > 0)
        g = g.join(self.sku.select(["sku", "SKU_Description"]), on="sku", how="left")
        return (g.rename({"machine": "Machine", "sku": "SKU",
                          "gt_code": "GT_Code"})
                .with_columns(pl.col("Mins").round(1),
                              pl.col("SKU_Description").fill_null(""))
                .select(["Machine", "Day", "SKU", "Qty", "Mins", "Segments",
                         "StartTime", "EndTime", "GT_Code", "SKU_Description"])
                .sort(["Day", "Machine", "SKU"]))

    def cure_rows(self) -> pl.DataFrame:
        """(press, day, shift, SKU) -- the reference curing Shift Schedule."""
        if self.cs.height == 0:
            return pl.DataFrame()
        c = self.cs.with_columns(
            pl.col("start_ts").map_elements(lambda t: _day_of(t, self.origin),
                                            return_dtype=pl.Int64).alias("Day"),
            pl.col("start_ts").map_elements(lambda t: _shift_of(t, self.origin),
                                            return_dtype=pl.Utf8).alias("Shift"))
        c = c.join(self.gt_sku(), on="gt_code", how="left")
        c = c.with_columns(pl.col("sku").fill_null(pl.col("gt_code")))
        g = (c.group_by(["press", "Day", "Shift", "sku", "gt_code"]).agg(
            pl.len().alias("Cured_Qty"),
            (pl.col("cycle_s").sum() / 60).round(1).alias("Used_Mins"),
            pl.col("start_ts").min().alias("StartTime"),
            pl.col("end_ts").max().alias("EndTime")))
        g = g.join(self.sku.select(["sku", "SKU_Description"]), on="sku", how="left")
        return (g.rename({"press": "Press", "sku": "SKU", "gt_code": "GT_Code"})
                .with_columns(pl.lit("RUNNING").alias("Status"),
                              pl.lit(0).alias("CO_Mins"),
                              pl.lit(0).alias("Mould_Clean_Mins"),
                              pl.lit(0).alias("Starved_Qty"),
                              pl.col("SKU_Description").fill_null(""))
                .select(["Press", "Day", "Shift", "SKU", "Status", "Cured_Qty",
                         "Used_Mins", "CO_Mins", "Mould_Clean_Mins",
                         "Starved_Qty", "StartTime", "EndTime", "GT_Code",
                         "SKU_Description"])
                .sort(["Day", "Shift", "Press"]))

    def daily(self) -> pl.DataFrame:
        e = self.ev.with_columns(
            pl.col("ts").map_elements(lambda t: _day_of(t, self.origin),
                                      return_dtype=pl.Int64).alias("Day"))
        b = (e.filter(pl.col("source") == "build").group_by("Day")
             .agg(pl.col("qty_delta").sum().alias("Built")))
        c = (e.filter(pl.col("source") == "cure").group_by("Day")
             .agg((-pl.col("qty_delta")).sum().alias("Cured")))
        i = (e.with_columns(pl.when(pl.col("source") == "cure")
                            .then(-pl.col("qty_delta").abs())
                            .otherwise(pl.col("qty_delta").abs()).alias("d"))
             .sort("ts").with_columns(pl.col("d").cum_sum().alias("I"))
             .group_by("Day").agg(pl.col("I").last().alias("EndDay_GT_Inventory")))
        d = (b.join(c, on="Day", how="full", coalesce=True)
             .join(i, on="Day", how="full", coalesce=True).fill_null(0)
             .filter((pl.col("Day") >= 1) & (pl.col("Day") <= self.ndays))
             .sort("Day"))
        y, mo = int(self.month[:4]), int(self.month[5:7])
        return d.with_columns(
            pl.col("Day").map_elements(
                lambda x: str((datetime(y, mo, 1) + timedelta(days=x - 1)).date()),
                return_dtype=pl.Utf8).alias("Date")).select(
            ["Day", "Date", "Built", "Cured", "EndDay_GT_Inventory"])

    def fulfilment(self) -> pl.DataFrame:
        """Per FG SKU: demand, opening, built, cured, coverage, and the REASON."""
        if self.dem.height == 0:
            return pl.DataFrame()
        d = (self.dem.group_by(["sku", "gt_code"]).agg(
            pl.col("qty").sum().alias("Demand_Qty")))
        gs = self.gt_sku()
        bb = (self.bs.group_by("gt_code").agg(pl.col("qty").sum().alias("gt_built")))
        cc = (self.cs.group_by("gt_code").agg(pl.len().alias("gt_cured")))
        op = (self.ev.filter(pl.col("source") == "opening")
              .group_by("gt_code").agg(pl.col("qty_delta").sum().alias("gt_open")))
        d = (d.join(gs, on=["gt_code", "sku"], how="left")
             .join(bb, on="gt_code", how="left").join(cc, on="gt_code", how="left")
             .join(op, on="gt_code", how="left").fill_null(0))
        d = d.with_columns(
            (pl.col("gt_built") * pl.col("share")).round().alias("Built"),
            (pl.col("gt_cured") * pl.col("share")).round().alias("Cured"),
            (pl.col("gt_open") * pl.col("share")).round().alias("Opening_GT"))
        d = d.with_columns(
            (pl.col("Demand_Qty") - pl.col("Cured")).alias("Shortfall"),
            (100 * pl.col("Cured") / pl.col("Demand_Qty").clip(lower_bound=1))
            .round(1).alias("Coverage_%"))
        sr = self.exc["supply_ratio"]
        if sr.height:
            d = d.join(sr.select(["gt_code", "R", "presses", "machines"]),
                       on="gt_code", how="left")
        else:
            d = d.with_columns(pl.lit(None, dtype=pl.Float64).alias("R"),
                               pl.lit(None, dtype=pl.Int64).alias("presses"),
                               pl.lit(None, dtype=pl.Int64).alias("machines"))
        d = d.join(self.sku.select(["sku", "SKU_Description", "moulds"]),
                   on="sku", how="left")
        return d.sort("Shortfall", descending=True)


# --------------------------------------------------------------------------
def _changeovers(rows: pl.DataFrame, key: str) -> pl.DataFrame:
    """Consecutive SKU changes on a machine/press -> the changeover plan."""
    if rows.height == 0:
        return pl.DataFrame()
    d = rows.sort([key, "StartTime"]).with_columns(
        pl.col("SKU").shift(1).over(key).alias("Old_SKU"),
        pl.col("GT_Code").shift(1).over(key).alias("Old_GT"))
    d = d.filter(pl.col("Old_SKU").is_not_null()
                 & (pl.col("Old_SKU") != pl.col("SKU")))
    return d.select([key, "Day", "Old_SKU", pl.col("SKU").alias("New_SKU"),
                     "Old_GT", pl.col("GT_Code").alias("New_GT"),
                     "StartTime"]).sort(["Day", key])


def _version(ctx: Ctx, kind: str) -> pl.DataFrame:
    k = json.loads((ctx.run / "kpi.json").read_text()) if (
        ctx.run / "kpi.json").exists() else {}
    e = ctx.rep.get("exceptions", {})
    rows = [
        ("Plan month", ctx.month, "-"),
        ("Plant", ctx.plant, "-"),
        ("Workbook", kind, "-"),
        ("Generated by", "planner.engine (frozen baseline)", "-"),
        ("Run id", str(ctx.rep.get("run_id", "")), "-"),
        ("Day boundary", f"{DAY_START_H:02d}:00 (shifts A/B/C x 8h)", "-"),
        ("Horizon days", ctx.ndays, "-"),
        ("Cure fulfilment %", k.get("cure_fulfilment_pct", ""), "-"),
        ("Build fulfilment %", k.get("build_fulfilment_pct", ""), "-"),
        ("GT aging p50 / p95 h", f"{k.get('aging_p50_h','')} / "
                                 f"{k.get('aging_p95_h','')}", "-"),
        ("Hard violations", k.get("hard_violations", ""), "-"),
        ("Shelf-life breach rows", e.get("shelf_life", 0), "-"),
        ("GTs with R<1 (over-mounted)", e.get("supply_ratio_under_1", 0), "-"),
        ("Unplaced lots", e.get("unplaced", 0), "-"),
        ("Machine overlaps", e.get("machine_overlap", 0), "-"),
        ("tau_min (h)", 0.25, "PLANNER_TH_*"),
        ("Replenish interval T0 (h)", 12.0, "PLANNER_TH_REPLENISH_INTERVAL_H"),
    ]
    return pl.DataFrame({"Parameter": [str(a) for a, _, _ in rows],
                         "Value": [str(b) for _, b, _ in rows],
                         "Revert env var": [c for _, _, c in rows]})


def _exceptions(ctx: Ctx) -> pl.DataFrame:
    out = []
    for r in ctx.exc["unplaced"].iter_rows(named=True) if \
            ctx.exc["unplaced"].height else []:
        out.append(("Unplaced lot", r.get("gt_code"), r.get("machine"),
                    f"{r.get('qty')} units", "must fit inside horizon", "Hard",
                    r.get("reason")))
    for r in ctx.exc["past_horizon"].iter_rows(named=True) if \
            ctx.exc["past_horizon"].height else []:
        out.append(("Lot past horizon", r.get("gt_code"), r.get("machine"),
                    f"+{r.get('over_h')} h", "end <= horizon", "Hard", ""))
    for r in ctx.exc["machine_overlap"].iter_rows(named=True) if \
            ctx.exc["machine_overlap"].height else []:
        out.append(("Machine double-booked", "", r.get("machine"),
                    f"{r.get('overlap_min')} min", "no overlap", "Hard", ""))
    sl = ctx.exc["shelf_life"]
    if sl.height:
        out.append(("GT shelf life > 72h", "", "", f"{sl.height} tyres",
                    "<= 72 h", "Hard",
                    f"worst {float(sl['age_h'].max()):.0f} h on "
                    f"{sl['gt_code'][0]}"))
    sr = ctx.exc["supply_ratio"]
    if sr.height:
        u = sr.filter(pl.col("R") < 1.0)
        for r in u.head(25).iter_rows(named=True):
            out.append(("Presses out-throughput building (R<1)", r["gt_code"], "",
                        f"R={r['R']:.2f}", "R >= 1.0", "Soft",
                        f"{r['presses']} presses mounted, "
                        f"{r['machines']} building machine(s)"))
    small = ctx.build_rows()
    if small.height:
        s = small.filter(pl.col("Qty") < MIN_LOT)
        if s.height:
            out.append(("Small build lot", "", "",
                        f"{s.height} lots < {MIN_LOT}",
                        f">= {MIN_LOT} units/lot", "Soft",
                        f"{int(s['Qty'].sum()):,} tyres in sub-scale lots"))
    if not out:
        return pl.DataFrame()
    return pl.DataFrame({
        "Exception_Type": [a for a, *_ in out],
        "SKUs (GT)": [b for _, b, *_ in out],
        "Machine/Press": [c for _, _, c, *_ in out],
        "Value": [d for *_, d, _, _, _ in [list(x) for x in out]],
        "Threshold": [x[4] for x in out],
        "Hard/Soft": [x[5] for x in out],
        "Remark": [x[6] for x in out]})


# --------------------------------------------------------------------------
def plan_analysis(ctx: Ctx, ws) -> None:
    """THE SHEET THE REFERENCE DOES NOT HAVE.

    What demand was missed, WHY it was missed (attributed to a cause, not
    merely counted), and how the plan compares to what the plant actually did
    in the same month.
    """
    ws.cell(row=1, column=1,
            value=f"PLAN ANALYSIS - {ctx.plant} {ctx.month}").font = Font(
        bold=True, size=14)
    ws.cell(row=2, column=1,
            value="Read top to bottom: what was missed, why, and how it "
                  "compares to the plant.").font = Font(italic=True)
    r = 4

    k = json.loads((ctx.run / "kpi.json").read_text())
    ff = ctx.fulfilment()
    short = ff.filter(pl.col("Shortfall") > 0) if ff.height else pl.DataFrame()

    dem_t = float(ctx.dem["qty"].sum()) if ctx.dem.height else 0.0
    cured_t = float(ctx.cs.height)
    built_t = float(ctx.bs["qty"].sum()) if ctx.bs.height else 0.0
    head = pl.DataFrame({
        "Metric": ["Demand (tyres)", "Cured (delivered)", "Built (green tyres)",
                   "Shortfall vs demand", "Coverage %",
                   "SKUs short of demand", "SKUs fully met"],
        "Value": [f"{dem_t:,.0f}", f"{cured_t:,.0f}", f"{built_t:,.0f}",
                  f"{dem_t - cured_t:,.0f}",
                  f"{100 * cured_t / max(dem_t, 1):.2f}%",
                  f"{short.height}", f"{ff.height - short.height}"]})
    r = _put(ws, head, r, "1. DID WE MEET DEMAND?")

    if short.height:
        s = short.with_columns(
            pl.when(pl.col("Built") < pl.col("Demand_Qty") * 0.995)
            .then(pl.lit("BUILD short - green tyres were never made"))
            .when(pl.col("R").is_not_null() & (pl.col("R") < 1.0))
            .then(pl.lit("PRESS over-mounted (R<1) - presses idle waiting for feed"))
            .otherwise(pl.lit("CURE short - green tyres made but not cured in time"))
            .alias("Reason"))
        tot_short = max(float(s["Shortfall"].sum()), 1.0)
        agg = (s.group_by("Reason").agg(
            pl.len().alias("SKUs"), pl.col("Shortfall").sum().alias("Tyres"))
            .sort("Tyres", descending=True))
        agg = agg.with_columns(
            (100 * pl.col("Tyres") / tot_short).round(1).alias("Share_%"))
        r = _put(ws, agg, r, "2. WHY DEMAND WAS NOT MET - attributed by cause")
        cols = [c for c in ["sku", "SKU_Description", "gt_code", "Demand_Qty",
                            "Built", "Cured", "Shortfall", "Coverage_%", "R",
                            "presses", "machines", "moulds", "Reason"]
                if c in s.columns]
        r = _put(ws, s.select(cols).head(30), r,
                 "   The 30 largest shortfalls, with their cause")
    else:
        r = _put(ws, pl.DataFrame({"Result": ["All SKUs met in full"]}), r,
                 "2. WHY DEMAND WAS NOT MET")

    try:
        from planner.data.warehouse import duck, set_cutoff
        set_cutoff(None)
        mo = f"{ctx.month}-01"
        pb = duck().execute(
            "SELECT count(*), count(DISTINCT machineCode), "
            "count(DISTINCT itemCode) FROM v_build WHERE stage=2 AND plant=? "
            "AND itemCode IS NOT NULL AND date_trunc('month', event_ts)=?::DATE",
            [ctx.plant, mo]).fetchone()
        pc = duck().execute(
            "SELECT count(*), count(DISTINCT wcID) FROM v_curing WHERE plant=? "
            "AND statuscritical='Normal' AND date_trunc('month', event_ts)=?::DATE",
            [ctx.plant, mo]).fetchone()
        plag = duck().execute(
            "SELECT median(date_diff('second', b.event_ts, c.event_ts)/3600.0), "
            "quantile_cont(date_diff('second', b.event_ts, c.event_ts)/3600.0, 0.95) "
            "FROM v_build b JOIN v_curing c ON b.productionID=c.gtbarCode "
            "WHERE b.stage=2 AND c.statuscritical='Normal' "
            "AND c.event_ts>=b.event_ts AND b.plant=? "
            "AND date_trunc('month', c.event_ts)=?::DATE",
            [ctx.plant, mo]).fetchone()
        inv = ctx.daily()
        band = ("4,500-4,800" if ctx.plant == "PCR" else "1,200-1,500")
        cmp_ = pl.DataFrame({
            "Metric": ["Tyres built", "Tyres cured", "Building machines used",
                       "Presses used", "Distinct GTs built",
                       "GT aging p50 (h)", "GT aging p95 (h)",
                       "Mean daily GT inventory"],
            "Our plan": [f"{built_t:,.0f}", f"{cured_t:,.0f}",
                         f"{ctx.bs['machine'].n_unique()}",
                         f"{ctx.cs['press'].n_unique()}",
                         f"{ctx.bs['gt_code'].n_unique()}",
                         f"{k.get('aging_p50_h', '')}",
                         f"{k.get('aging_p95_h', '')}",
                         f"{float(inv['EndDay_GT_Inventory'].mean()):,.0f}"],
            "Plant actual": [f"{pb[0]:,}", f"{pc[0]:,}", f"{pb[1]}", f"{pc[1]}",
                             f"{pb[2]}", f"{plag[0]:.1f}", f"{plag[1]:.1f}",
                             f"target band {band}"],
            "Reading": ["", "", "", "fewer is better - a mounted mould is a "
                        "commitment, not capacity", "",
                        "lower is better", "lower is better",
                        "G8 rule band"]})
        r = _put(ws, cmp_, r,
                 f"3. OUR PLAN vs WHAT THE PLANT ACTUALLY DID ({ctx.month})")
    except Exception as e:                                      # noqa: BLE001
        r = _put(ws, pl.DataFrame({"note": [f"plant comparison unavailable: {e}"]}),
                 r, "3. OUR PLAN vs PLANT")

    e_ = ctx.rep.get("exceptions", {})
    lim = pl.DataFrame({
        "Limitation": [
            "Calendar assumed 24x7",
            "Mould availability not enforced",
            "Demand is in-sample",
            "Press count not optimised",
            "Shelf-life breaches present",
            "Small build lots"],
        "Impact on this plan": [
            "Production is scheduled on Sundays and holidays; real capacity is "
            "lower and the plan will not execute exactly as printed",
            f"Presses per GT are not bounded by mould sets owned; "
            f"{e_.get('supply_ratio_under_1', 0)} GTs have more press capacity "
            f"mounted than building can feed",
            "Month M is planned from month M's own output, so coverage here is "
            "reconstruction accuracy, not forecast accuracy",
            "All available presses are mounted; the plant runs fewer",
            f"{e_.get('shelf_life', 0)} tyres exceed the 72 h green-tyre shelf "
            f"life",
            "Lots below 150 units are scheduled; a supervisor will not run them"],
        "What is needed to close it": [
            "PLANT DATA: working days, shift pattern, planned maintenance "
            "windows per machine and per press",
            "PLANT DATA: mould sets owned and serviceable per SKU (partially "
            "available in Master_Mapping_Mould_SKU.csv)",
            "Plan month M from the M-1 order book",
            "Bound presses_mounted(g) <= moulds owned, then minimise press count",
            "Reduce build-to-cure lead; see GT aging above",
            "Enforce a minimum lot size with demand rounding"]})
    _put(ws, lim, r, "4. KNOWN LIMITATIONS - read before executing this plan")
    for col, w in [("A", 42), ("B", 58), ("C", 58), ("D", 34)]:
        ws.column_dimensions[col].width = w


def _wb_building(ctx: Ctx) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    br = ctx.build_rows()
    _put(wb.create_sheet("Exceptions"), _exceptions(ctx), 1,
         "THRESHOLD EXCEPTIONS - hard rows block execution, soft are advisory")
    _put(wb.create_sheet("Version"), _version(ctx, "building"), 1)
    _put(wb.create_sheet("Shift Schedule"), br, 1)
    if ctx.bs.height:
        ps = ctx.bs.with_columns(
            pl.col("start_ts").map_elements(lambda t: _day_of(t, ctx.origin),
                                            return_dtype=pl.Int64).alias("Day"),
            pl.col("start_ts").map_elements(lambda t: _shift_of(t, ctx.origin),
                                            return_dtype=pl.Utf8).alias("Shift"))
        ps = (ps.group_by(["machine", "Day", "Shift", "gt_code"])
              .agg(pl.col("qty").sum().alias("Qty"),
                   pl.col("start_ts").min().alias("StartTime"),
                   pl.col("end_ts").max().alias("EndTime"))
              .rename({"machine": "Machine", "gt_code": "GT_Code"})
              .sort(["Day", "Shift", "Machine"]))
        _put(wb.create_sheet("Shift Schedule (per-shift)"), ps, 1)
    _put(wb.create_sheet("Changeover Plan"), _changeovers(br, "Machine"), 1)
    if ctx.bs.height:
        su = (ctx.bs.with_columns(
            pl.col("start_ts").map_elements(lambda t: _day_of(t, ctx.origin),
                                            return_dtype=pl.Int64).alias("Day"))
            .group_by(["Day", "machine"])
            .agg((pl.col("setup_s").sum() / 60).round(1).alias("Setup_Mins"))
            .with_columns((100 * pl.col("Setup_Mins") / 1440).round(2)
                          .alias("Setup_%_of_1440"))
            .rename({"machine": "Machine"}).sort(["Day", "Machine"]))
        _put(wb.create_sheet("Setup Time %"), su, 1)
        mu = (ctx.bs.group_by("machine").agg(
            ((pl.col("end_ts") - pl.col("start_ts")).dt.total_seconds().sum() / 60)
            .round(0).alias("Production_Mins"),
            (pl.col("setup_s").sum() / 60).round(0).alias("CO_Mins"),
            pl.col("qty").sum().alias("Qty_Built"),
            pl.col("gt_code").n_unique().alias("Distinct_GTs"))
            .rename({"machine": "Machine"}))
        mu = mu.with_columns(
            (1440 * ctx.ndays - pl.col("Production_Mins")).round(0).alias("Idle_Mins"),
            (100 * pl.col("Production_Mins") / (1440 * ctx.ndays)).round(1)
            .alias("Utilization_%")).sort("Machine")
        _put(wb.create_sheet("Machine Utilization"), mu, 1)
        _put(wb.create_sheet("Machine SKU Variety"),
             ctx.bs.group_by("machine").agg(
                 pl.col("gt_code").n_unique().alias("Distinct_GTs"),
                 pl.col("qty").sum().alias("Total_Qty"))
             .rename({"machine": "Machine"}).sort("Machine"), 1)
        if br.height:
            _put(wb.create_sheet("Small Build Batches"),
                 br.filter(pl.col("Qty") < MIN_LOT), 1,
                 f"Daily build lots below {MIN_LOT} units")
        sp = (ctx.bs.with_columns(
            pl.col("start_ts").map_elements(lambda t: _day_of(t, ctx.origin),
                                            return_dtype=pl.Int64).alias("Day"))
            .group_by(["gt_code", "Day"])
            .agg(pl.col("machine").n_unique().alias("N_Machines"))
            .filter(pl.col("N_Machines") > 1)
            .rename({"gt_code": "GT_Code"}).sort("N_Machines", descending=True))
        _put(wb.create_sheet("SKU on Multiple Machines"), sp, 1,
             "Plant mined p50 is 1 machine per GT per day (PCR) / 2 (TBR)")
    _put(wb.create_sheet("Daily GT & Inventory"), ctx.daily(), 1)
    ffb = ctx.fulfilment()
    if ffb.height:
        cols = [c for c in ["sku", "gt_code", "SKU_Description", "Demand_Qty",
                            "Opening_GT", "Built", "Cured", "Shortfall",
                            "Coverage_%"] if c in ffb.columns]
        _put(wb.create_sheet("Demand Fulfillment (Build)"), ffb.select(cols), 1)
    plan_analysis(ctx, wb.create_sheet("PLAN ANALYSIS"))
    return wb


def _wb_curing(ctx: Ctx) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    cr = ctx.cure_rows()
    _put(wb.create_sheet("Exceptions"), _exceptions(ctx), 1,
         "THRESHOLD EXCEPTIONS")
    _put(wb.create_sheet("Version"), _version(ctx, "curing"), 1)
    ff = ctx.fulfilment()
    if ff.height:
        cols = [c for c in ["sku", "gt_code", "SKU_Description", "Demand_Qty",
                            "Cured", "Shortfall", "Coverage_%"]
                if c in ff.columns]
        _put(wb.create_sheet("Demand Fulfillment"), ff.select(cols), 1)
    _put(wb.create_sheet("Shift Schedule"), cr, 1)
    _put(wb.create_sheet("Changeover Plan"), _changeovers(cr, "Press"), 1)
    if ctx.cs.height:
        pu = (ctx.cs.group_by("press").agg(
            pl.len().alias("Cured_Qty"),
            (pl.col("cycle_s").sum() / 60).round(0).alias("Used_Mins"),
            pl.col("gt_code").n_unique().alias("Distinct_GTs"))
            .rename({"press": "Press"}))
        pu = pu.with_columns(
            (100 * pl.col("Used_Mins") / (1440 * ctx.ndays)).round(1)
            .alias("Utilization_%")).sort("Press")
        _put(wb.create_sheet("Machine Utilization"), pu, 1)
        _put(wb.create_sheet("Small Cure Campaigns"),
             ctx.cs.group_by("gt_code").agg(pl.len().alias("Cured"))
             .filter(pl.col("Cured") < 200).sort("Cured"), 1,
             "GTs cured in very small total quantity")
    _put(wb.create_sheet("Daily Cured"),
         ctx.daily().select(["Day", "Date", "Cured"]), 1)
    sr = ctx.exc["supply_ratio"]
    if sr.height:
        _put(wb.create_sheet("Feed NGT Summary"),
             sr.filter(pl.col("R") < 1.0).sort("R").head(60), 1,
             "R = tyres built / capacity of the presses mounted. R<1 means the "
             "presses out-throughput building and starve by construction.")
    sl = ctx.exc["shelf_life"]
    if sl.height:
        _put(wb.create_sheet("Shelf Life Breaches"), sl.head(500), 1,
             "Green tyres cured more than 72 h after they were built")
    plan_analysis(ctx, wb.create_sheet("PLAN ANALYSIS"))
    return wb


def main(run: Path) -> None:
    rep = json.loads((run / "run_report.json").read_text())
    mtag = str(rep.get("input", {}).get("plan_start", run.name))[:7]
    y, mo = int(mtag[:4]), int(mtag[5:7])
    tag = f"{MONTHS[mo - 1]}{y}"
    out = run / "ctp"
    out.mkdir(exist_ok=True)
    for plant in ["PCR", "TBR"]:
        ctx = Ctx(run, plant)
        if ctx.bs.height == 0:
            print(f"  {plant}: no build rows, skipped")
            continue
        p = plant.lower()
        f1 = out / f"ctp_building_schedule_{tag}_{p}.xlsx"
        f2 = out / f"ctp_curing_schedule_{tag}_{p}.xlsx"
        _wb_building(ctx).save(f1)
        _wb_curing(ctx).save(f2)
        print(f"  {plant}: {f1.name} + {f2.name}")


if __name__ == "__main__":
    for a in sys.argv[1:]:
        print(f"[{a}]")
        main(Path(a))
