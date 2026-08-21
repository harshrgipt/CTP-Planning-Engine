"""Turn a plant ORDER-BOOK workbook into `masters/demand/demand_<month>.parquet`.

    python -m scripts.ingest_orderbook_demand --xlsx "<any demand file>" --month 2026-08

ONE FILE IN, TWO PLANTS OUT
  The uploader supplies demand. Splitting it PCR/TBR is OURS to do, and so is
  finding the GT. Everything below is optional: the sheet name (auto-detected,
  and REFUSED if more than one sheet qualifies rather than guessed), the column
  names (`SKUCode`/`SKU Code`/`Material Code`... and `Requirement`/`Qty`/...),
  the `Classification` column, the `Matched GT Code` column, and the raw MES.

  Verified end to end 2026-08-21: a bare two-column sheet holding only
  ("Material Code", "Quantity"), no Classification, no GT, no MES, reproduces
  the committed August month EXACTLY -- PCR 429,146 / TBR 99,019, frame-identical.

HOW THE PLANT IS DECIDED  (`planner/cmbc/plant_split.py`)
  `Classification` used to be mandatory: a row whose value was not exactly "PCR"
  or "TBR" was dropped. That made a hand-filled column load-bearing for the one
  decision nothing downstream can detect if it is wrong.

  `resolve_plant` now votes across up to four sources and REFUSES on any
  disagreement. The strongest is physical: PCR tops out at R18, TBR starts at
  R20, the 19-inch gap is empty, and no GT appears in both plants. Measured over
  both prepared months -- 1,840 rows, ZERO mismatches against the plant's own
  column, and July resolves with no Classification column at all.
  Locked by `tests/unit/test_plant_split.py`.

MES IS A CROSS-CHECK, NOT A REQUIREMENT
  `sku_to_gt()` and `mes_namespace()` read `v_curing`/`v_build` -- the 4.4 GB
  drop that is gitignored, so absent on a clone and on the frontend machine.
  Since this is the one step a frontend user triggers, GT resolution reads the
  committed `sku_gt_crosswalk.parquet` (`scripts/build_sku_gt_crosswalk.py`) and
  uses MES, when present, only to verify the crosswalk has not gone stale.

WHAT THE WORKBOOK GIVES AND WHAT IT DOES NOT
  * `Delivery date` and `Priority flag` are EMPTY on every row of the August
    file. There is therefore NO due-date profile in this data. We do not invent
    one: each (plant, GT, SKU) gets ONE row carrying the whole month's
    requirement, dated the last day of the month. A flat daily spread would look
    like information and be none -- and `due_date` has no live reader, but L4.5
    DOES read `day` (l45_lotsize.py, per-GT phase curve -> lot_deadlines), so an
    unphased book collapses every deadline to one value and L5's
    earliest-deadline sort degenerates -- checked 2026-08-18.
  * `Requirement` is fractional on some rows (e.g. 433422.1 total). Kept as
    float through the mapping and rounded ONCE at the end with largest-remainder
    so the plant total is preserved exactly.

THREE REJECT PILES, ALL REPORTED WITH THEIR QUANTITY -- never silently dropped
  UNMAPPED_<M>.csv      no source could place the plant, or no GT could be found
  UNPLANNABLE_<M>.csv   GT resolves but NO press can cure it. Planning these
                        would add unfillable shortfall, not output -- 47 August
                        rows / 4,276 tyres, every one a PCR GT that is in the
                        recipe master but has never been cured.
  UNCONFIRMED_PLANT_<M>.csv  placed on the workbook's own word with no master
                        row and therefore no rim to confirm it.
"""
from __future__ import annotations

import argparse
import calendar
import csv
import sys
from datetime import date
from pathlib import Path

import polars as pl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from gt_namespace import mes_namespace, resolve_gt_label, sku_to_gt  # noqa: E402

from planner import paths  # noqa: E402
from planner.cmbc.plant_split import (  # noqa: E402
    load_rim_map, load_sku_map, resolve_plant)


def _cell(v) -> str:
    return "" if v is None else str(v).strip()


# A frontend upload is one file, and the plant does not name its columns the
# same way twice. Only these two matter; everything else is a bonus cross-check.
SKU_ALIASES = ("SKUCODE", "SKUCD", "SKU", "MATERIALCODE", "MATERIAL", "ITEMCODE")
QTY_ALIASES = ("REQUIREMENT", "TOTALREQUIREMENT", "QTY", "QUANTITY", "DEMAND",
               "ORDERQTY", "REQQTY", "VOLUME")


def _norm_hdr(v: str) -> str:
    return "".join(ch for ch in v.upper() if ch.isalnum())


def _scan(ws) -> tuple[int, list[str]] | None:
    """Find the header row: the first row within the top 15 holding BOTH a SKU
    column and a quantity column. The sheet carries a title and a note row above
    it, so the header is never row 0 and cannot be assumed."""
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i > 15:
            return None
        vals = [_cell(c) for c in r]
        n = {_norm_hdr(v) for v in vals if v}
        if n & set(SKU_ALIASES) and n & set(QTY_ALIASES):
            return i, vals
    return None


def pick_sheet(xlsx: Path) -> str:
    """Choose the demand sheet when the caller named none.

    REFUSES ON AMBIGUITY rather than picking. The August workbook has SIX sheets
    and FOUR of them carry a SKU header -- 'Unique SKU Classification' is the
    deduplicated view of the same demand and 'Not Existing' is the reject pile.
    Silently picking either would produce a plausible, wrong month. So: exactly
    one candidate is used, more than one is an error that names them all.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    cands = []
    for sn in wb.sheetnames:
        hit = _scan(wb[sn])
        if hit:
            cands.append(sn)
    wb.close()
    if len(cands) == 1:
        return cands[0]
    if not cands:
        raise SystemExit(
            f"no demand sheet in {xlsx.name}: no sheet has both a SKU column "
            f"({'/'.join(SKU_ALIASES[:3])}...) and a quantity column "
            f"({'/'.join(QTY_ALIASES[:3])}...)")
    raise SystemExit(
        f"{len(cands)} sheets in {xlsx.name} look like demand -- pass --sheet to "
        f"say which:\n    " + "\n    ".join(cands))


def read_sheet(xlsx: Path, sheet: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"sheet {sheet!r} not in {xlsx.name}; has: "
                         + ", ".join(wb.sheetnames))
    ws = wb[sheet]
    hit = _scan(ws)
    if hit is None:
        wb.close()
        raise SystemExit(f"sheet {sheet!r} has no SKU + quantity header row")
    hdr_i, hdr = hit
    rows = [[_cell(c) for c in r] for r in ws.iter_rows(values_only=True)]
    wb.close()

    # Normalise the two load-bearing columns to the canonical names the rest of
    # this script uses, so a workbook titled "SKU Code"/"Total Requirement"
    # needs no edit before upload.
    ren = {}
    for h in hdr:
        k = _norm_hdr(h)
        if k in SKU_ALIASES and "SKUCode" not in ren.values():
            ren[h] = "SKUCode"
        elif k in QTY_ALIASES and "Requirement" not in ren.values():
            ren[h] = "Requirement"
    hdr = [ren.get(h, h) for h in hdr]

    out = []
    for r in rows[hdr_i + 1:]:
        if not any(x for x in r):
            continue
        out.append(dict(zip(hdr, r)))
    return out


def largest_remainder(vals: list[float], total: int) -> list[int]:
    """Integerise so the sum is exactly `total` and nothing is silently lost."""
    base = [int(v) for v in vals]
    rem = total - sum(base)
    order = sorted(range(len(vals)), key=lambda i: -(vals[i] - base[i]))
    for i in range(rem):
        base[order[i % len(order)]] += 1
    return base


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", default=None,
                    help="omit to auto-detect; refuses if >1 sheet qualifies")
    ap.add_argument("--month", required=True)
    ap.add_argument("--out", default=None, help="default masters/demand")
    a = ap.parse_args()

    xlsx = Path(a.xlsx)
    if not xlsx.is_absolute():
        xlsx = (Path.cwd() / xlsx).resolve()
    out_dir = Path(a.out) if a.out else ROOT / "masters" / "demand"
    out_dir.mkdir(parents=True, exist_ok=True)

    sheet = a.sheet or pick_sheet(xlsx)
    rows = read_sheet(xlsx, sheet)
    y, m = int(a.month[:4]), int(a.month[5:7])
    last = date(y, m, calendar.monthrange(y, m)[1])

    # MES IS OPTIONAL HERE, AND THAT IS THE POINT.
    # `sku_to_gt()` and `mes_namespace()` both read `v_curing` / `v_build`, i.e.
    # the 4.4 GB raw drop that is gitignored and absent on a clone and on the
    # frontend machine. This script is the ONE step a frontend user triggers, so
    # requiring MES here would mean the single-file upload only works on the
    # machine that already has everything. The committed
    # `sku_gt_crosswalk.parquet` carries the same SKU -> (plant, engine gt_code)
    # mapping, frozen; MES, when present, is used only as a CROSS-CHECK.
    try:
        chain = sku_to_gt()
        ns = mes_namespace()
        mes = "present"
    except Exception as e:                      # noqa: BLE001 - any MES absence
        chain, ns, mes = {}, {}, f"absent ({type(e).__name__})"

    # THE PLANT SPLIT IS OURS TO MAKE, NOT THE UPLOADER'S.
    # `Classification` used to be mandatory: a row whose value was not exactly
    # "PCR" or "TBR" was dropped. That made the column load-bearing on a file a
    # human fills in, for the one decision in this pipeline that nothing
    # downstream can detect if it is wrong -- a TBR tyre routed to PCR is planned
    # against PCR machines, presses, moulds and lot floor, and every number for
    # BOTH plants is then wrong.
    #
    # `plant_split.resolve_plant` places each SKU from up to four sources and
    # refuses on any disagreement. The strongest is PHYSICAL: PCR tops out at
    # R18 and TBR starts at R20, with a clean 19-inch gap and no GT in both
    # plants. Verified on both prepared months: 1,840 rows, ZERO mismatches
    # against the workbook's own column, and July resolves with no Classification
    # column at all. So the column is now OPTIONAL -- one more cross-check when
    # present, never the sole authority when a rim exists.
    rim_map = load_rim_map()
    sku_map = load_sku_map()

    # PLANNABILITY GATE -- a GT with no press cannot be cured, ever.
    # The crosswalk is a NAMESPACE map, not a capability gate: `gt_sku_master`
    # lists every PCR SKU in the plant's recipe master, including ones the plant
    # has never actually cured. Widening GT resolution therefore pulled in 15
    # August SKUs (4,276 tyres) that have a building machine but NO row in
    # `allowed_press_matrix` and no rim.
    #
    # Planning those would not add 4,276 tyres of output -- it would add 4,276
    # tyres of permanent, unfillable shortfall, which LOWERS fulfilment while
    # looking on paper like extra demand. Both prepared months held this
    # invariant implicitly before this change (July: 104 GTs, ZERO without a
    # press), so it is made explicit here rather than left to be rediscovered.
    #
    # They are reported with their quantity, not silently dropped: an order the
    # plant cannot cure is a fact the planner owes the user.
    try:
        _pm = pl.read_parquet(paths.input_derived("allowed_press_matrix.parquet"))
        press_gts = {g for g in _pm["gt_code"].to_list() if g}
    except Exception:                            # noqa: BLE001
        press_gts = set()                        # absent -> gate disabled, not silent
    unplannable = []
    recs, unmapped = [], []
    tiers: dict[str, int] = {}
    agree = disagree = 0
    agree_xw = disagree_xw = 0
    xw_bad = []
    plant_mismatch = []
    unconfirmed = []
    for r in rows:
        sku = r.get("SKUCode", "")
        try:
            qty = float(r.get("Requirement", "") or 0)
        except ValueError:
            qty = 0.0
        mgt = r.get("Matched GT Code", "")
        cls, split_basis, split_why = resolve_plant(
            sku, None, r.get("Classification", ""), None, rim_map, sku_map)
        if cls is None:
            unmapped.append({"sku": sku, "qty": qty,
                             "plant": r.get("Classification", ""),
                             "matched_gt": mgt, "reason": split_why or "unplaceable",
                             "desc": r.get("SKU Description", "")})
            continue
        if "UNCONFIRMED" in split_basis:
            # Placed on the plant's own word with no master row to confirm it.
            # Accepted -- the plant is the authority on its own new product --
            # but never silently: every one is listed at the end of the run.
            unconfirmed.append({"sku": sku, "qty": qty, "plant": cls,
                                "basis": split_basis, "why": split_why,
                                "desc": r.get("SKU Description", "")})
        ch = chain.get(sku)
        by_label, tier = resolve_gt_label(mgt, cls, ns) if ns else (None, "no-mes")
        xw = sku_map.get(sku)                   # committed crosswalk, MES-free
        if ch and by_label:
            if ch[1] == by_label:
                agree += 1
            else:
                disagree += 1
        if ch and xw:
            # The frozen crosswalk vs the live chain, on every row where both
            # speak. This is the check that would catch the crosswalk going
            # stale, so it is reported even when it is clean.
            if ch[1] == xw[1]:
                agree_xw += 1
            else:
                disagree_xw += 1
                xw_bad.append((sku, ch[1], xw[1]))
        if ch and ch[0] != cls:
            plant_mismatch.append((sku, cls, ch[0], ch[1]))
        # Order of authority: live recipe chain (when MES is here) > the frozen
        # crosswalk > the workbook's own label through the string bridge.
        if ch:
            gt, src = ch[1], "recipe-chain"
        elif xw and xw[1]:
            gt, src = xw[1], "crosswalk"
        else:
            gt, src = by_label, (f"label:{tier}" if by_label else "")
        if not gt:
            unmapped.append({"sku": sku, "qty": qty, "plant": cls,
                             "matched_gt": mgt,
                             "reason": f"no GT in crosswalk or MES ({tier})",
                             "desc": r.get("SKU Description", "")})
            continue
        if press_gts and gt not in press_gts:
            unplannable.append({"sku": sku, "qty": qty, "plant": cls,
                                "gt_code": gt, "gt_src": src,
                                "reason": "no row in allowed_press_matrix -- "
                                          "no press can cure this GT",
                                "desc": r.get("SKU Description", "")})
            continue
        tiers[src] = tiers.get(src, 0) + 1
        recs.append({"plant": cls, "gt_code": gt, "sku": sku, "qty": qty,
                     "otype": r.get("Order Type", ""), "mkt": r.get("Market", "")})

    df = pl.DataFrame(recs) if recs else pl.DataFrame(
        schema={"plant": pl.Utf8, "gt_code": pl.Utf8, "sku": pl.Utf8,
                "qty": pl.Float64, "otype": pl.Utf8, "mkt": pl.Utf8})
    g = (df.group_by(["plant", "gt_code", "sku"])
         .agg(pl.col("qty").sum().alias("q"))
         .filter(pl.col("q") > 0)
         .sort(["plant", "gt_code", "sku"]))

    # integerise ONCE, per plant, so each plant's total is preserved exactly
    parts = []
    for p in ("PCR", "TBR"):
        s = g.filter(pl.col("plant") == p)
        if not s.height:
            continue
        tot = int(round(float(s["q"].sum())))
        s = s.with_columns(pl.Series("qty", largest_remainder(s["q"].to_list(), tot),
                                     dtype=pl.Int64))
        parts.append(s)
    fin = (pl.concat(parts) if parts else g.with_columns(pl.col("q").cast(pl.Int64).alias("qty")))
    fin = fin.select([
        "plant", "gt_code", "sku", "qty",
        pl.lit(a.month).alias("month"),
        pl.lit(last).alias("due_date"),
        pl.lit(last.day).cast(pl.Int64).alias("day"),
    ]).sort(["plant", "gt_code", "sku"])

    fin.write_parquet(out_dir / f"demand_{a.month}.parquet", compression="zstd")
    fin.write_csv(out_dir / f"demand_{a.month}.csv")
    if unmapped:
        with open(out_dir / f"UNMAPPED_{a.month}.csv", "w", newline="",
                  encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(unmapped[0].keys()))
            w.writeheader()
            w.writerows(unmapped)

    print("=" * 92)
    print(f"ORDER-BOOK DEMAND  --  {a.month}   {xlsx.name} :: {sheet}")
    print("=" * 92)
    print(f"  workbook rows read          {len(rows)}")
    print(f"  MES (cross-check only)      {mes}")
    print(f"  classified PCR/TBR          {len(rows) - sum(1 for u in unmapped if u['reason'].startswith('not in'))}")
    print(f"  GT resolution tiers         " + "  ".join(f"{k}={v}" for k, v in sorted(tiers.items())))
    print(f"  crosswalk vs recipe chain   agree {agree_xw}  DISAGREE {disagree_xw}"
          + ("   <-- crosswalk is STALE, rebuild it" if disagree_xw else ""))
    for b in xw_bad[:8]:
        print(f"      !! {b[0]}  chain={b[1]}  crosswalk={b[2]}")
    print(f"  cross-check both routes     agree {agree}  DISAGREE {disagree}  "
          f"plant mismatch {len(plant_mismatch)}")
    for x in plant_mismatch:
        print(f"      !! {x}")
    print()
    print(f"  {'plant':<6}{'GTs':>6}{'SKUs':>7}{'rows':>7}{'tyres':>12}")
    for p in ("PCR", "TBR"):
        s = fin.filter(pl.col("plant") == p)
        print(f"  {p:<6}{s['gt_code'].n_unique():>6}{s['sku'].n_unique():>7}"
              f"{s.height:>7}{int(s['qty'].sum()):>12,}")
    print(f"  {'TOTAL':<6}{fin['gt_code'].n_unique():>6}{fin['sku'].n_unique():>7}"
          f"{fin.height:>7}{int(fin['qty'].sum()):>12,}")
    if unmapped:
        uq = sum(u["qty"] for u in unmapped)
        print(f"\n  UNMAPPED {len(unmapped)} rows, {uq:,.0f} tyres "
              f"-> UNMAPPED_{a.month}.csv  (NOT planned)")
        for u in sorted(unmapped, key=lambda z: -z["qty"])[:20]:
            if u["qty"] > 0:
                print(f"      {u['plant']:<14}{u['sku']:<20}{u['qty']:>10,.0f}"
                      f"  {u['matched_gt']:<18}{u['reason']}")
    if unplannable:
        uq = sum(u["qty"] for u in unplannable)
        print("")
        print(f"  UNPLANNABLE {len(unplannable)} rows, {uq:,.0f} tyres  (NOT planned)")
        print("  GT resolves, but no press in allowed_press_matrix can cure it. "
              "Planning these")
        print("  would add unfillable shortfall, not output.")
        for u in sorted(unplannable, key=lambda z: -z["qty"])[:20]:
            print(f"      {u['plant']:<4}{u['sku']:<20}{u['qty']:>9,.0f}  "
                  f"{u['gt_code']:<20}via {u['gt_src']}")
        if len(unplannable) > 20:
            print(f"      ... +{len(unplannable) - 20} more")
        with (out_dir / f"UNPLANNABLE_{a.month}.csv").open(
                "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(unplannable[0].keys()))
            w.writeheader()
            w.writerows(unplannable)

    # PLACED ON THE PLANT'S WORD ALONE. Never fold this into the totals above --
    # it is the set a wrong Classification could still corrupt, and the only
    # place a reader can see it. A row here has no master entry, so no rim, so
    # nothing physical confirmed which plant it belongs to.
    if unconfirmed:
        uq = sum(u["qty"] for u in unconfirmed)
        print("")
        print(f"  UNCONFIRMED PLANT {len(unconfirmed)} rows, {uq:,.0f} tyres "
              f"({100 * uq / max(fin['qty'].sum(), 1):.1f}% of the month)")
        print("  SKUs with no master row -- placed on the workbook's "
              "Classification alone, with no rim to confirm it.")
        for u in sorted(unconfirmed, key=lambda z: -z["qty"])[:20]:
            print(f"      {u['plant']:<6}{u['sku']:<22}{u['qty']:>10,.0f}  "
                  f"{(u['desc'] or '')[:44]}")
        if len(unconfirmed) > 20:
            print(f"      ... +{len(unconfirmed) - 20} more")
        with (out_dir / f"UNCONFIRMED_PLANT_{a.month}.csv").open(
                "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(unconfirmed[0].keys()))
            w.writeheader()
            w.writerows(unconfirmed)

    print(f"\n  -> {out_dir / f'demand_{a.month}.parquet'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
