"""SHEET 10 -- DAILY BUILD / CURE / GT, one row per plant-day.

    python scripts/export_daily_bcg.py <run_id> <YYYY-MM> <pack_dir>

Adds `10_daily_build_cure_gt` and `10b_daily_curve` to an ALREADY EXPORTED pack
(it appends to the workbook and writes the CSVs beside the others). It never
re-plans and never rewrites another sheet.

WHY IT IS A SEPARATE SCRIPT
  `export_shift_schedule.py` is the shared exporter used by every month and both
  packs. This sheet was asked for on one pack, so it is added beside that
  exporter rather than inside it -- nothing already shipped changes shape.

THE DAY BOUNDARY IS 07:00, AND HOURS ARE CLIPPED INTO THE DAY THEY ARE SPENT
  A build run or a cure campaign that crosses 07:00 belongs to BOTH days, in
  proportion. Bucketing a whole interval into the day it STARTED is the defect
  that has now bitten this pack three times (sheet 7 §1 and §2 in
  export_shift_schedule.py). Hours here are integrated, not attributed.

WHAT TIES TO WHAT -- stated, because the two streams are different objects
  built_tyres    build slices (machine != OPENING_STOCK) bucketed at start_ts.
                 Identical convention to sheets 1 and 7, so it ties EXACTLY.
  cured_tyres    tyres FED to a press, bucketed at cure_ts. This is delivered
                 quantity, not planned campaign quantity -- campaigns were
                 planned larger than building could feed, so the campaign frame
                 sums HIGHER and is not the fulfilment numerator.
  press_h        from the L5 campaign plan (sheet 2 / 6 basis) -- the press is
                 occupied for the campaign whether or not every tyre arrived.
  gt_in_built    build slices bucketed at END time. A tyre enters GT stock when
                 the run finishes, not when it starts, so the GT identity uses
                 end_ts while `built_tyres` uses start_ts. The two totals are
                 equal; only the daily split differs.
  gt_day_mean    TRUE time-weighted mean over the 24 h (DO-NOT #9: mean over
                 EVENTS is not mean over TIME). Sheet 7's column samples 24
                 hourly midpoints instead; both are printed and compared.
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import polars as pl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

SHIFT_START_H = 7


def clip_hours(s: np.ndarray, e: np.ndarray, w: np.ndarray, nday: int) -> np.ndarray:
    """Hours of each [s, e) interval that fall inside each plant-day.

    s, e are hours since t0. Returns an array of length `nday`. `w` weights each
    interval (1.0 for hours; qty for a prorated quantity)."""
    out = np.zeros(nday)
    if len(s) == 0:
        return out
    lo = np.arange(nday) * 24.0
    hi = lo + 24.0
    for d in range(nday):
        ov = np.minimum(e, hi[d]) - np.maximum(s, lo[d])
        np.clip(ov, 0.0, None, out=ov)
        out[d] = float((ov * w).sum())
    return out


def bucket(ts_h: np.ndarray, q: np.ndarray, nday: int) -> np.ndarray:
    """Sum q by the plant-day containing ts_h."""
    out = np.zeros(nday)
    if len(ts_h) == 0:
        return out
    d = np.floor(ts_h / 24.0).astype(int)
    ok = (d >= 0) & (d < nday)
    np.add.at(out, d[ok], q[ok])
    return out


def tw_mean_day(ts: np.ndarray, bal: np.ndarray, d0: float, d1: float) -> float:
    """Time-weighted mean of a step function on [d0, d1)."""
    # value at d0
    i = np.searchsorted(ts, d0, side="right") - 1
    cur = bal[i] if i >= 0 else 0.0
    acc, t = 0.0, d0
    j = np.searchsorted(ts, d0, side="right")
    while j < len(ts) and ts[j] < d1:
        acc += cur * (ts[j] - t)
        t = ts[j]
        cur = bal[j]
        j += 1
    acc += cur * (d1 - t)
    return acc / (d1 - d0)


def main() -> int:
    run_id, month, pack = sys.argv[1], sys.argv[2], Path(sys.argv[3])
    run = ROOT / "runs" / run_id
    y, m = int(month[:4]), int(month[5:7])
    t0 = datetime(y, m, 1, SHIFT_START_H)
    n_month_days = (datetime(y + (m == 12), (m % 12) + 1, 1)
                    - datetime(y, m, 1)).days

    from planner.config import CONFIG                       # noqa: E402
    th = CONFIG.thresholds
    BAND_LO, BAND_HI = th.gt_wip_min, th.gt_wip_max
    RAIL = th.gt_wip_rail

    bs = pl.read_parquet(run / "build_schedule.parquet")
    cc = pl.read_parquet(run / "cure_campaigns.parquet")
    try:
        roster = {k: len(v) for k, v in json.loads(
            (ROOT / "masters" / f"press_list_{month}.json").read_text()).items()}
    except Exception:                                       # noqa: BLE001
        roster = {}

    H = lambda ts: np.array(                                 # noqa: E731
        [(x - t0).total_seconds() / 3600.0 for x in ts], dtype=float)

    # How many days the plan actually touches (carry-out can exceed the month).
    last = 0.0
    for col, fr in (("end_ts", bs), ("cure_ts", bs), ("end_ts", cc)):
        if fr.height:
            last = max(last, float(H(fr[col])[-len(fr):].max()))
    nday = max(n_month_days, int(np.ceil(last / 24.0)))

    # ---- changeover flag, IDENTICAL derivation to sheet 1 -----------------
    b = bs.filter(pl.col("machine") != "OPENING_STOCK")
    off = ((pl.col("start_ts") - pl.lit(t0)).dt.total_seconds() / 3600.0)
    b = (b.with_columns([(off // 24 + 1).cast(pl.Int64).alias("plant_day"),
                         pl.col("start_ts").dt.strftime("%Y-%m-%d").alias("date"),
                         ((off % 24) // 8).cast(pl.Int64).alias("_s")])
         .with_columns(pl.when(pl.col("_s") == 0).then(pl.lit("A"))
                       .when(pl.col("_s") == 1).then(pl.lit("B"))
                       .otherwise(pl.lit("C")).alias("shift")).drop("_s"))
    # CHRONOLOGICAL sort, explicitly. This used to sort by (plant, date, shift,
    # machine, start_ts) on the WALL-CLOCK date, which puts shift C's
    # post-midnight rows ahead of the next date's A and B shifts and invents a GT
    # transition at every inversion. Same defect as sheet 1's; it inflated the
    # count to PCR 967 / TBR 1,108 against a true 799 / 800 (= runs - machines).
    b = b.sort(["plant", "machine", "start_ts"]).with_columns(
        pl.col("gt_code").shift(1).over(["plant", "machine"]).alias("_pg")
    ).with_columns(
        (pl.col("gt_code") != pl.col("_pg")).fill_null(False).alias("changeover"))
    # continuous same-GT block = a RUN (split at a >1 h gap, sheet 1b's rule)
    b = b.sort(["plant", "machine", "start_ts"]).with_columns(
        ((pl.col("start_ts") - pl.col("end_ts").shift(1).over("run_id"))
         .dt.total_seconds() / 3600.0).alias("_gap"))
    b = b.with_columns(((pl.col("_gap") > 1.0) | pl.col("_gap").is_null())
                       .cum_sum().alias("_blk")).with_columns(
        (pl.col("run_id") + "#" + pl.col("_blk").cast(pl.Utf8)).alias("block_id"))

    rows = []
    for p in ("PCR", "TBR"):
        bp = bs.filter(pl.col("plant") == p)                 # incl. OPENING_STOCK
        bb = b.filter(pl.col("plant") == p)                  # built this month
        cp = cc.filter(pl.col("plant") == p)
        if not bp.height:
            continue
        n_mach = int(bb["machine"].n_unique())
        n_press_plan = int(cp["press"].n_unique()) if cp.height else 0
        n_press_roster = int(roster.get(p, n_press_plan) or n_press_plan)

        bstart, bend = H(bb["start_ts"]), H(bb["end_ts"])
        bqty = np.array(bb["qty"], float)
        built = bucket(bstart, bqty, nday)                   # ties to sheets 1/7
        gt_in_built = bucket(bend, bqty, nday)               # GT credit at run end
        mach_h = clip_hours(bstart, bend, np.ones(len(bstart)), nday)
        co = bucket(bstart, np.array(bb["changeover"], float), nday)
        # first slice of each run -> a run started
        rs = (bb.group_by("block_id").agg(pl.col("start_ts").min()))
        runs_started = bucket(H(rs["start_ts"]), np.ones(rs.height), nday)

        cured = bucket(H(bp["cure_ts"]), np.array(bp["qty"], float), nday)
        op = bp.filter(pl.col("machine") == "OPENING_STOCK")
        gt_in_open = bucket(H(op["end_ts"]), np.array(op["qty"], float), nday) \
            if op.height else np.zeros(nday)

        if cp.height:
            cs, ce = H(cp["start_ts"]), H(cp["end_ts"])
            press_h = clip_hours(cs, ce, np.ones(len(cs)), nday)
            camp_started = bucket(cs, np.ones(len(cs)), nday)
            act = np.zeros(nday)
            for d in range(nday):
                act[d] = int(((ce > d * 24.0) & (cs < (d + 1) * 24.0)).sum())
        else:
            press_h = camp_started = act = np.zeros(nday)

        # ---- GT ledger: +qty at end_ts, -qty at cure_ts (L11's own stream) ---
        ivt = pl.concat([
            bp.select([pl.col("end_ts").alias("ts"), pl.col("qty").alias("d")]),
            bp.select([pl.col("cure_ts").alias("ts"), (-pl.col("qty")).alias("d")]),
        ]).sort("ts").with_columns(pl.col("d").cum_sum().alias("bal"))
        ts = H(ivt["ts"])
        bal = np.array(ivt["bal"], float)
        # collapse simultaneous events to the LAST balance at that instant
        keep = np.append(np.diff(ts) > 1e-9, True)
        tsk, balk = ts[keep], bal[keep]

        def at(x: float) -> float:
            i = np.searchsorted(tsk, x, side="right") - 1
            return float(balk[i]) if i >= 0 else 0.0

        # sheet 7's estimator, for comparison only
        idx = np.searchsorted(tsk, np.arange(nday * 24) + 0.5, side="right") - 1
        g_h = np.where(idx >= 0, balk[np.clip(idx, 0, len(balk) - 1)], 0.0)

        for d in range(nday):
            gopen = at(d * 24.0 - 1e-9) if d else 0.0
            gclose = at((d + 1) * 24.0 - 1e-9)
            lo, hi, rail = BAND_LO.get(p, 0), BAND_HI.get(p, 0), RAIL.get(p, 0)
            flag = ("ABOVE_RAIL" if gclose > rail else
                    "ABOVE_BAND" if gclose > hi else
                    "BELOW_BAND" if gclose < lo else "")
            dmean = tw_mean_day(tsk, balk, d * 24.0, (d + 1) * 24.0)
            rows.append({
                "plant": p, "plant_day": d + 1,
                "date": (t0 + timedelta(days=d)).strftime("%Y-%m-%d"),
                "carry_out": d + 1 > n_month_days,
                # ---- BUILD ----
                "built_tyres": int(round(built[d])),
                "build_machine_h": round(float(mach_h[d]), 2),
                "build_machines": n_mach,
                "build_occupancy_pct": round(100 * mach_h[d] / (n_mach * 24), 1),
                "build_changeovers": int(round(co[d])),
                "build_runs_started": int(round(runs_started[d])),
                # ---- CURE ----
                "cured_tyres": int(round(cured[d])),
                "press_h": round(float(press_h[d]), 2),
                "presses_in_plan": n_press_plan,
                "presses_roster": n_press_roster,
                "press_occupancy_pct": round(
                    100 * press_h[d] / (n_press_roster * 24), 1),
                "campaigns_active": int(act[d]),
                "campaigns_started": int(round(camp_started[d])),
                # ---- GT ----
                "gt_open": int(round(gopen)),
                "gt_in_opening_stock": int(round(gt_in_open[d])),
                "gt_in_built": int(round(gt_in_built[d])),
                "gt_out_cured": int(round(cured[d])),
                "gt_close": int(round(gclose)),
                "gt_day_mean": int(round(dmean)),
                "gt_day_mean_sheet7_basis": int(round(
                    float(g_h[d * 24:(d + 1) * 24].mean()))),
                "gt_band_lo": lo, "gt_band_hi": hi, "gt_rail": rail,
                "gt_in_band": lo <= gclose <= hi,
                "gt_flag": flag,
                # The rail is a DAILY-MEAN control (config.gt_wip_rail, checked
                # in L7 at 94 % of the stated cap). `gt_close` is an INSTANT at
                # 07:00 on a Q/2 sawtooth, so it crosses the band on days the
                # enforced metric does not. Both are shown; only this one is
                # what L7 refused placements against.
                "gt_day_mean_flag": ("ABOVE_RAIL" if dmean > rail else
                                     "ABOVE_BAND" if dmean > hi else
                                     "BELOW_BAND" if dmean < lo else ""),
                "gt_balance_check": int(round(
                    gclose - (gopen + gt_in_open[d] + gt_in_built[d] - cured[d]))),
            })

    daily = pl.DataFrame(rows).sort(["plant", "plant_day"])

    # ---- RECONCILE against the pack's own sheets --------------------------
    C = pack / "csv"
    s1 = pl.read_csv(C / "1_build_schedule_shift.csv")
    s7 = pl.read_csv(C / "7_daily_summary.csv")
    s5 = pl.read_csv(C / "5_machine_summary.csv")
    s6 = pl.read_csv(C / "6_press_summary.csv")
    print("  RECONCILIATION  sheet 10 vs sheets 1 / 5 / 6 / 7 and the run")
    ok = True
    for p in ("PCR", "TBR"):
        d10 = daily.filter(pl.col("plant") == p)
        b1 = int(s1.filter(pl.col("plant") == p)["qty"].sum())
        b10 = int(d10["built_tyres"].sum())
        c7 = int(s7.filter(pl.col("plant") == p)["cured"].sum())
        c10 = int(d10["cured_tyres"].sum())
        b7 = int(s7.filter(pl.col("plant") == p)["built"].sum())
        x7 = int(s7.filter(pl.col("plant") == p)["changeovers"].sum())
        x10 = int(d10["build_changeovers"].sum())
        h5 = float(s5.filter(pl.col("plant") == p)["busy_h"].sum())
        h10 = float(d10["build_machine_h"].sum())
        p6 = float(s6.filter(pl.col("plant") == p)["busy_h"].sum())
        p10 = float(d10["press_h"].sum())
        gin = int(d10["gt_in_built"].sum())
        bad = int((d10["gt_balance_check"] != 0).sum())
        ok &= (b1 == b10 == b7) and (c7 == c10) and (x7 == x10) and bad == 0
        print(f"     {p}  built  s1 {b1:>9,}  s7 {b7:>9,}  s10 {b10:>9,}  "
              f"diff {b10 - b1:>4,}")
        print(f"          cured  s7 {c7:>9,}  s10 {c10:>9,}  diff {c10 - c7:>4,}"
              f"   ·  gt_in_built {gin:>9,} (= built, split at run END)")
        print(f"          chgovr s7 {x7:>9,}  s10 {x10:>9,}  diff {x10 - x7:>4,}"
              f"   ·  GT balance identity fails on {bad} day(s)")
        print(f"          mach_h s5 {h5:>9,.1f}  s10 {h10:>9,.1f}  "
              f"diff {h10 - h5:>6.2f}   ·  press_h s6 {p6:>9,.1f}  "
              f"s10 {p10:>9,.1f}  diff {p10 - p6:>6.2f}")
    print(f"     -> {'ALL TIE (diff 0)' if ok else '!! MISMATCH'}")

    # ---- chart-ready long form -------------------------------------------
    curve = daily.select(
        ["plant", "plant_day", "date", "built_tyres", "cured_tyres",
         "gt_close", "gt_day_mean", "gt_band_lo", "gt_band_hi", "gt_rail"]
    ).unpivot(index=["plant", "plant_day", "date"],
              variable_name="series", value_name="value").sort(
        ["plant", "series", "plant_day"])

    out = {"10_daily_build_cure_gt": daily, "10b_daily_curve": curve}
    for name, df in out.items():
        df.write_csv(C / f"{name}.csv")
        print(f"     {name:<28}{df.height:>8,} rows")

    # ---- append to the workbook ------------------------------------------
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    xl = pack / f"schedule_{month}.xlsx"
    wb = openpyxl.load_workbook(xl)
    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", start_color="1F3864")
    for name, df in out.items():
        if name in wb.sheetnames:
            wb.remove(wb[name])
        ws = wb.create_sheet(name[:31])
        ws.append(df.columns)
        for cell in ws[1]:
            cell.font, cell.fill = hdr, fill
        for row in df.iter_rows():
            ws.append(list(row))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i, c in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(
                max(len(str(c)) + 2, 12), 34)
    try:
        wb.save(xl)
    except PermissionError:
        alt = pack / f"schedule_{month}__NEW.xlsx"
        wb.save(alt)
        print(f"  !! {xl.name} is LOCKED. Wrote {alt.name} instead.")
        xl = alt
    print(f"  -> {xl}")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
