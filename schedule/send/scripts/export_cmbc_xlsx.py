"""Export a CMBC run to per-plant XLSX workbooks.

    python scripts/export_cmbc_xlsx.py runs/july_cmbc_v1 --month 2026-07

One workbook per plant, plus a combined audit workbook. Sheets follow the
reference CTP convention where it applies: the plant day starts at 07:00 and
shifts are A 07-15, B 15-23, C 23-07.

Every sheet that carries a number carries its basis too. A schedule a planner
cannot audit is a schedule a planner will override, and each override corrupts
what L0 learns next month.
"""
from __future__ import annotations

import argparse
import json
from datetime import datetime, timedelta
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
D = ROOT / "warehouse" / "derived"
PARAMS = ROOT / "warehouse" / "params"
HDR = PatternFill("solid", fgColor="FFD9E1F2")
WARN = PatternFill("solid", fgColor="FFFCE4D6")
BAD = PatternFill("solid", fgColor="FFF8CBAD")
GOOD = PatternFill("solid", fgColor="FFE2EFDA")


def put(ws, df: pl.DataFrame, title: str = "", note: str = "", r0: int = 1) -> int:
    r = r0
    if title:
        c = ws.cell(row=r, column=1, value=title)
        c.font = Font(bold=True, size=13)
        r += 1
    if note:
        ws.cell(row=r, column=1, value=note).font = Font(italic=True, size=9)
        r += 1
    r += 1
    if df is None or df.height == 0:
        ws.cell(row=r, column=1, value="(none)").font = Font(italic=True)
        return r + 2
    for j, col in enumerate(df.columns, 1):
        c = ws.cell(row=r, column=j, value=col)
        c.font = Font(bold=True)
        c.fill = HDR
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, row in enumerate(df.iter_rows(), 1):
        for j, v in enumerate(row, 1):
            if isinstance(v, datetime):
                v = v.strftime("%Y-%m-%d %H:%M")
            elif v is not None and not isinstance(v, (int, float, str)):
                v = str(v)
            ws.cell(row=r + i, column=j, value=v)
    for j, col in enumerate(df.columns, 1):
        try:
            w = int(df[col].cast(pl.Utf8).str.len_chars().max() or 8)
        except Exception:                                       # noqa: BLE001
            w = 14
        ws.column_dimensions[get_column_letter(j)].width = max(
            len(str(col)) + 2, min(34, w + 2))
    return r + df.height + 3


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("run")
    ap.add_argument("--month", default="2026-07")
    a = ap.parse_args()
    run = Path(a.run) if Path(a.run).is_absolute() else ROOT / a.run
    y, m = int(a.month[:4]), int(a.month[5:7])
    t0 = datetime(y, m, 1, 7, 0)

    P = json.loads(sorted(PARAMS.glob("params_*.json"))[-1].read_text())
    camp = pl.read_parquet(run / "cure_campaigns.parquet")
    bs = pl.read_parquet(run / "build_schedule.parquet")
    rec = pl.read_parquet(run / "cure_campaigns_reconciled.parquet")
    req = pl.read_parquet(D / f"net_requirement_{a.month}.parquet")
    mc = pl.read_parquet(run / "mould_changes.parquet")
    inv = pl.read_parquet(run / "l11_invariants.parquet")
    gap = pl.read_parquet(run / "l12_gap.parquet")
    cbs = pl.read_parquet(run / "cure_by_shift.parquet")
    bbs = pl.read_parquet(run / "build_by_shift.parquet")
    unp = pl.read_parquet(run / "cure_unplaced.parquet")
    prep = pl.read_parquet(run / "prep_requirement.parquet")
    comp = (pl.read_parquet(run / "rule_compliance.parquet")
            if (run / "rule_compliance.parquet").exists() else None)
    dem = pl.read_parquet(ROOT / "masters" / "demand" / f"demand_{a.month}.parquet")

    # RESOLVE PRESS IDENTITY. `press` in the plan is v_curing.wcID -- an internal
    # key (133, 134), not a press number. wcmaster carries the real identity
    # (wcID 133 -> press 5435; 119 -> 4805, 48" LTM CURING PRESS NO 5). Shipping
    # the key to a planner asks them to translate it by hand every time.
    wcf = D / "wc_master.parquet"
    if wcf.exists():
        wc = pl.read_parquet(wcf)
        def with_press(df):
            if "press" not in df.columns:
                return df
            return (df.join(wc, left_on="press", right_on="wc_id", how="left")
                    .with_columns(pl.col("press_no").fill_null(pl.col("press"))
                                  .alias("press_no"),
                                  pl.col("description").fill_null("").alias("press_desc"))
                    .drop("press").rename({"press_no": "press"}))
        camp = with_press(camp)
        bs = with_press(bs)
        rec = with_press(rec)
        mc = with_press(mc)
        unp = with_press(unp) if unp.height else unp

    out_dir = run / "xlsx"
    out_dir.mkdir(exist_ok=True)
    made = []

    for plant in ["PCR", "TBR"]:
        wb = Workbook()
        wb.remove(wb.active)

        # ---- 1. summary -------------------------------------------------
        ws = wb.create_sheet("SUMMARY")
        rr = rec.filter(pl.col("plant") == plant)
        need = float(req.filter((pl.col("plant") == plant)
                                & ~pl.col("residual"))["gross_build"].sum())
        fed = float(rr["qty_fed"].sum())
        w = bs.filter(pl.col("plant") == plant)["wait_h"]
        g = gap.filter(pl.col("plant") == plant).row(0, named=True)
        kpi = pl.DataFrame({
            "metric": ["plannable requirement (tyres)", "delivered (tyres)",
                       "fulfilment", "cure campaigns", "build slices",
                       "presses used", "machines used",
                       "GT head p50 (h)", "GT head p95 (h)", "GT head max (h)",
                       "tau* target (h)", "mould changes", "press-h lost to changes",
                       "press utilisation", "capacity (press-h)"],
            "value": [f"{need:,.0f}", f"{fed:,.0f}", f"{100*fed/max(need,1):.1f}%",
                      f"{camp.filter(pl.col('plant')==plant).height:,}",
                      f"{bs.filter(pl.col('plant')==plant).height:,}",
                      f"{camp.filter(pl.col('plant')==plant)['press'].n_unique()}",
                      f"{bs.filter((pl.col('plant')==plant)&(pl.col('machine')!='OPENING_STOCK'))['machine'].n_unique()}",
                      f"{float(w.median()):.2f}", f"{float(w.quantile(0.95)):.2f}",
                      f"{float(w.max()):.2f}",
                      f"{float(P['tau'][plant]['tau_star_h']):.2f}",
                      f"{mc.filter(pl.col('plant')==plant).height:,}",
                      f"{float(mc.filter(pl.col('plant')==plant)['minutes'].sum())/60:,.0f}",
                      f"{100*g['productive_h']/g['capacity_h']:.1f}%",
                      f"{g['capacity_h']:,.0f}"]})
        r = put(ws, kpi, f"{plant} — JULY {y} PRODUCTION PLAN (CMBC v3)",
                "curing planned first; building released backwards at "
                "t_cure - tau* - build_duration")
        r = put(ws, pl.DataFrame({
            "gap cause": ["productive", "mould changes", "cold start",
                          "horizon tail", "unattributed"],
            "press-hours": [g["productive_h"], g["mould_change_h"],
                            g["cold_start_h"], g["horizon_tail_h"],
                            g["unattributed_h"]],
            "% of capacity": [f"{100*g[k]/g['capacity_h']:.1f}%" for k in
                              ["productive_h", "mould_change_h", "cold_start_h",
                               "horizon_tail_h", "unattributed_h"]]}),
                "WHERE THE CAPACITY WENT", "every press-hour attributed to one cause", r)

        # ---- 2. cure schedule ------------------------------------------
        ws = wb.create_sheet("CURE SCHEDULE")
        cs = (camp.filter(pl.col("plant") == plant)
              .select(["press", "press_desc", "gt_code", "start_ts", "end_ts",
                       "hours", "qty", "mould_set"])
              .sort(["press", "start_ts"]))
        put(ws, cs, f"{plant} — CURE CAMPAIGNS",
            "mould-set x press x [start, end]; the plan is born here (L5)")

        # ---- 3. build schedule -----------------------------------------
        ws = wb.create_sheet("BUILD SCHEDULE")
        bsl = (bs.filter(pl.col("plant") == plant)
               .select(["machine", "gt_code", "start_ts", "end_ts", "qty",
                        "press", "press_desc", "cure_ts", "wait_h"])
               .sort(["machine", "start_ts"]))
        put(ws, bsl, f"{plant} — BUILD SLICES (pull-released)",
            "wait_h is the GT's own head: cure_ts - build end. "
            "OPENING_STOCK rows are existing stock, not new builds")

        # ---- 4. by shift ------------------------------------------------
        ws = wb.create_sheet("BY SHIFT")
        cq = (cbs.filter(pl.col("plant") == plant)
              .group_by(["day", "shift"]).agg(pl.col("qty").sum().round(0).alias("cure_qty"))
              .sort(["day", "shift"]))
        bq = (bbs.filter(pl.col("plant") == plant)
              .group_by(["day", "shift"]).agg(pl.col("qty").sum().round(0).alias("build_qty"))
              .sort(["day", "shift"]))
        put(ws, cq.join(bq, on=["day", "shift"], how="outer_coalesce")
            .fill_null(0).sort(["day", "shift"]),
            f"{plant} — SHIFT PLAN", "day starts 07:00 · A 07-15 · B 15-23 · C 23-07")

        # ---- 5. mould changes -------------------------------------------
        ws = wb.create_sheet("MOULD CHANGES")
        put(ws, mc.filter(pl.col("plant") == plant)
            .select(["press", "press_desc", "from_gt", "to_gt", "start_ts",
                     "end_ts", "minutes", "crew", "day", "shift",
                     "spans_boundary"])
            .sort(["day", "shift", "press"]),
            f"{plant} — MOULD CHANGES",
            "the dominant loss: ~6 h each. spans_boundary=TRUE takes two crews (R13)")

        # ---- 6. demand vs plan ------------------------------------------
        ws = wb.create_sheet("DEMAND vs PLAN")
        dv = (req.filter(pl.col("plant") == plant)
              .select(["gt_code", "demand", "from_stock", "net_cure",
                       "gross_build", "residual"])
              .join(rec.filter(pl.col("plant") == plant)
                    .group_by("gt_code").agg(pl.col("qty_fed").sum().round(0)),
                    on="gt_code", how="left").fill_null(0)
              .sort("demand", descending=True))
        put(ws, dv, f"{plant} — DEMAND vs DELIVERED",
            "residual=TRUE is below min_demand (B12) and routed, not dropped")

        # ---- 7. exceptions ----------------------------------------------
        ws = wb.create_sheet("EXCEPTIONS")
        r = put(ws, unp.filter(pl.col("plant") == plant) if unp.height else None,
                f"{plant} — CURE THAT DID NOT FIT THE HORIZON", "")
        r = put(ws, rec.filter((pl.col("plant") == plant) & (pl.col("qty_unfed") > 0))
                .select(["gt_code", "press", "qty", "qty_fed", "qty_unfed"]),
                "CURE PLACED BUT NOT FED BY BUILDING",
                "these would show as press starvation on the floor", r)

        f = out_dir / f"ctp_{a.month}_{plant}.xlsx"
        wb.save(f)
        made.append(f)

    # ---- audit workbook -------------------------------------------------
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("RULE COMPLIANCE")
    if comp is not None:
        put(ws, comp, f"RULE COMPLIANCE — {a.month}",
            "FOLLOWED = enforced and holds · PARTIAL = biased but not constrained "
            "· SKIPPED = not implemented · BLOCKED = needs plant data")
    ws = wb.create_sheet("INVARIANTS")
    put(ws, inv, f"L11 PLAN INVARIANTS — {a.month}",
        "rewritten for a pull plant: 3 old invariants deleted, "
        "build-slice minimum deliberately NOT checked")
    ws = wb.create_sheet("GAP TO CEILING")
    put(ws, gap, "L12 CAPACITY ATTRIBUTION",
        "each press-hour assigned to exactly one named cause")
    ws = wb.create_sheet("PREP REQUIREMENT")
    put(ws, prep, "L8 SEMI-FINISHED REQUIREMENT BY SHIFT",
        "tread and bead apex bind at 24 h -- tighter than the GT's own 72 h")
    f = out_dir / f"ctp_{a.month}_AUDIT.xlsx"
    wb.save(f)
    made.append(f)

    print(f"written to {out_dir}")
    for f in made:
        print(f"  {f.name:<34}{f.stat().st_size/1024:>8,.0f} KB")


if __name__ == "__main__":
    main()
