"""HISTORICAL MACHINE SHARE -> machine_gt_share.parquet

    python -m scripts.ingest_machine_share

Reads the plant's 12-month running share per GT x building machine:

    INPUT/raw/percentage weightage on tbmmachine ... pcr skus_ (1).xlsx
    INPUT/raw/percentage weightage on tbmmachine ... tbr skus_.xlsx

    GT / Item Code | Item Name | SKU Code | Description | Tyre Size
        | TBMPCR1Stage2 | TBMPCR2Stage2 | ...        <- percent, rows sum ~100

NO NAMESPACE BRIDGE IS NEEDED, WHICH IS UNUSUAL HERE
  Column A is already the ENGINE key -- "GT 1402 XPC TATA", "10.00 R 20 JDC3" --
  i.e. `v_build.itemCode`, and the machine columns are already the engine's own
  machine names. Every other plant workbook in this project needed a SKU bridge
  because its GT column was in the BOM or workbook namespace. This one does not,
  so it is read directly. The SKU column is kept only for provenance.

  Several SKUs map to one GT and repeat the same row; the share is a property of
  the GT, so rows are deduplicated per (plant, gt_code, machine) on the max.

WHAT IT IS FOR -- PREFERENCE, NEVER PERMISSION
  A GT with 98 / 1 / 1 should run ~98 % on its dominant machine, but must SPILL
  the moment that machine has no legal pre-deadline window. Waiting for the
  preferred machine starves a press, and a press-hour is gone for good while an
  early tyre only ages (it has 72 h of shelf life to spend).

  So the share ORDERS the candidate list and nothing else:
    * it never adds a machine -- candidates are already filtered by the plant's
      allowable matrix, so a 0 % machine that is allowable stays available as
      spill, and a machine absent from the allowable matrix is never used
      whatever its historical share;
    * it never blocks -- L7 walks the ordered list and takes the first machine
      with a legal window, so fallthrough is immediate.

  This is deliberately weaker than `machine_rim_lock`, which was tried as a HARD
  constraint and cost 14 points of PCR fulfilment by stranding 8 GTs entirely.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
import polars as pl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from planner import paths                                          # noqa: E402

FILES = {
    "PCR": "percentage weightage on tbmmachine for last 1 year for pcr skus_ (1).xlsx",
    "TBR": "percentage weightage on tbmmachine for last 1 year for tbr skus_.xlsx",
}


def _read(plant: str, f: Path) -> list[dict]:
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(x).strip() if x is not None else "" for x in rows[0]]
    mcols = {i: h for i, h in enumerate(hdr) if h.upper().startswith("TBM")}
    out = []
    for r in rows[1:]:
        gt = str(r[0]).strip() if r[0] else ""
        if not gt:
            continue
        for i, mc in mcols.items():
            if i >= len(r):
                continue
            try:
                v = float(r[i] or 0)
            except (TypeError, ValueError):
                continue
            if v > 0:
                out.append({"plant": plant, "gt_code": gt, "machine": mc,
                            "share_pct": round(v, 4)})
    return out


def run(*, write: bool = True) -> pl.DataFrame:
    rows: list[dict] = []
    for plant, name in FILES.items():
        f = paths.raw(name)
        if not f.exists():
            print(f"  !! missing {f}")
            continue
        got = _read(plant, f)
        rows += got
        print(f"  {plant}  {f.name[:52]}: {len(got)} (GT, machine) shares")
    if not rows:
        raise SystemExit("no share rows read")

    df = (pl.DataFrame(rows)
          .group_by(["plant", "gt_code", "machine"])
          .agg(pl.col("share_pct").max())
          .sort(["plant", "gt_code", "share_pct"], descending=[False, False, True]))

    print(f"\n  MACHINE SHARE  (12-month running history)")
    print(f"  {'-' * 68}")
    am = pl.read_parquet(paths.input_derived("allowed_machine_matrix.parquet"))
    for p, g in df.group_by("plant"):
        per = g.group_by("gt_code").agg(pl.len().alias("n"),
                                        pl.col("share_pct").max().alias("top"))
        dom = per.filter(pl.col("top") >= 95)
        a = am.filter(pl.col("plant") == p[0])
        # a share row on a machine the plant does not allow is history we cannot use
        bad = g.join(a.with_columns(pl.lit(True).alias("_ok")),
                     on=["plant", "gt_code", "machine"], how="left")
        ruled = a.select(["plant", "gt_code"]).unique().with_columns(
            pl.lit(True).alias("_r"))
        bad = (bad.join(ruled, on=["plant", "gt_code"], how="left")
                  .filter(pl.col("_r").is_not_null() & pl.col("_ok").is_null()))
        print(f"    {p[0]}: {per.height} GTs · machines/GT p50 {per['n'].median():.0f} "
              f"max {per['n'].max()} · GTs with a >=95 % dominant machine "
              f"{dom.height} ({100 * dom.height / per.height:.0f} %)")
        print(f"         top-machine share: p50 {per['top'].median():.1f} % · "
              f"min {per['top'].min():.1f} %")
        if bad.height:
            print(f"         !! {bad.height} share rows name a machine the allowable "
                  f"matrix forbids -- IGNORED (allowable wins)")
    if write:
        f = paths.input_derived("machine_gt_share.parquet")
        df.write_parquet(f)
        print(f"  -> {f}  ({df.height} rows)")
    return df


def main() -> None:
    ap = argparse.ArgumentParser(description="12-month machine share -> parquet")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()
    run(write=not a.dry_run)


if __name__ == "__main__":
    main()
