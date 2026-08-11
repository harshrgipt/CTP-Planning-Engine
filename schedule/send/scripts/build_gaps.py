"""Close GAP-8 (semi-finished ageing), GAP-2 (PCR inch eligibility), GAP-12.

    python scripts/build_gaps.py

Outputs to warehouse/derived/ and INPUT/derived/:
  semi_finished_ageing.parquet   min/max age per component      GAP-8
  pcr_inch_eligibility.parquet   machine x rim-inch capability   GAP-2
  opening_gt_*.parquet           copied from masters/opening_gt  GAP-12

PCR INCH CAPABILITY, read off `CTP Set up ...xlsx` -> [PCR BUILDING]:
    machines 1,2            12 to 20 inch
    machines 3,4,5          12 to 16 inch
    machines 6..11          13 to ?      <- upper bound TRUNCATED in the source
The upper bound for 6-11 is inferred from what those machines actually ran and
flagged `inferred`; it must be confirmed by the plant, not assumed.

THE HABIT-vs-PHYSICS TEST
  PCR building HHI = 1.00: every GT ran on exactly one machine for 8 months.
  If a GT's rim fits N machines by capability but it only ever ran on 1, the
  dedication is HABIT, not physics -- and that is the largest throughput
  question on site.  This script measures it instead of assuming either way.
"""
from __future__ import annotations

import re
import shutil
import warnings
from pathlib import Path

import polars as pl

warnings.filterwarnings("ignore")
import pdfplumber                                             # noqa: E402
from planner import paths

ROOT = Path(__file__).resolve().parent.parent
SRC = paths.RAW
OUT = ROOT / "warehouse" / "derived"
INP = paths.INPUT_DERIVED
PDF = paths.raw("Ageing spec-20.01.2024 (2).pdf")

# machine number -> (min inch, max inch, source)
PCR_INCH = {1: (12, 20, "spec"), 2: (12, 20, "spec"),
            3: (12, 16, "spec"), 4: (12, 16, "spec"), 5: (12, 16, "spec"),
            6: (13, None, "spec-truncated"), 7: (13, None, "spec-truncated"),
            8: (13, None, "spec-truncated"), 9: (13, None, "spec-truncated"),
            10: (13, None, "spec-truncated"), 11: (13, None, "spec-truncated")}
PCR_TYPE = {n: ("BJ" if n <= 5 else "CONTI") for n in range(1, 12)}
PCR_CHG = {"BJ": (28, 60), "CONTI": (22, 42)}


def to_hours(s: str | None) -> float | None:
    if not s:
        return None
    t = str(s).strip().lower()
    if t in ("na", "n/a", "-", ""):
        return None
    m = re.search(r"([\d.]+)\s*(min|hr|hour|day|month|year)", t)
    if not m:
        return None
    v, u = float(m.group(1)), m.group(2)
    return v * {"min": 1 / 60, "hr": 1, "hour": 1,
                "day": 24, "month": 720, "year": 8760}[u]


def ageing() -> pl.DataFrame:
    rows = []
    with pdfplumber.open(PDF) as pdf:
        section = ""
        for page in pdf.pages:
            for tbl in page.extract_tables():
                for r in tbl:
                    c = [(" ".join(str(x).split()) if x else "") for x in r]
                    c += [""] * (6 - len(c))
                    sr, grp, _idx, comp, mn, mx = c[0], c[1], c[2], c[3], c[4], c[5]
                    if sr.strip().isdigit() and grp:
                        section = grp
                    if grp and not sr.strip().isdigit():
                        section = grp or section
                    if not comp or comp.upper().startswith("COMPONENT"):
                        continue
                    rows.append({"section": section or grp, "component": comp,
                                 "min_raw": mn, "max_raw": mx,
                                 "min_h": to_hours(mn), "max_h": to_hours(mx)})
    return pl.DataFrame(rows).filter(
        pl.col("min_h").is_not_null() | pl.col("max_h").is_not_null()).unique()


def pcr_rim_map() -> dict[str, int]:
    """PCR GT codes ("GT 1402 XPC TATA") carry no rim -- the rim lives in
    ALL PCR CTP SKUS.xlsx as a tyre size ("145 R 12").  Parsing the GT code
    itself yields nothing, which is why the first run returned an empty test."""
    from openpyxl import load_workbook
    wb = load_workbook(paths.raw("ALL PCR CTP SKUS.xlsx"), read_only=True)
    rows = list(wb["pcr skus running in ctp"].iter_rows(values_only=True))
    wb.close()
    out: dict[str, int] = {}
    for r in rows[1:]:
        if not r[0] or not r[1]:
            continue
        m = re.search(r"R\s*(\d{2})", str(r[1]).upper())
        if not m:
            continue
        code = str(r[0]).strip()
        rim = int(m.group(1))
        out[code] = rim
        out[f"GT {code}"] = rim          # MES prefixes "GT "
    return out


RIM = pcr_rim_map()


def rim_of(gt: str) -> int | None:
    g = str(gt).strip()
    return RIM.get(g) or RIM.get(g.replace("GT ", "", 1))


def main() -> None:
    INP.mkdir(parents=True, exist_ok=True)

    # ---- GAP-8 ----------------------------------------------------------
    ag = ageing()
    ag.write_parquet(OUT / "semi_finished_ageing.parquet")
    shutil.copy2(OUT / "semi_finished_ageing.parquet",
                 INP / "semi_finished_ageing.parquet")
    print(f"GAP-8  semi_finished_ageing : {ag.height} components")
    tight = ag.filter(pl.col("max_h").is_not_null()).sort("max_h").head(6)
    print("  tightest MAX shelf lives (these bind prep planning):")
    for r in tight.iter_rows(named=True):
        print(f"    {r['component'][:52]:<52} {r['max_h']:>6.0f} h "
              f"(min {r['min_h'] or 0:.0f} h)")

    # ---- GAP-2 ----------------------------------------------------------
    from planner.data.warehouse import duck, set_cutoff
    set_cutoff(None)
    obs = duck().execute("""
        SELECT machineCode m, itemCode gt, count(*) n FROM v_build
        WHERE stage=2 AND plant='PCR' AND itemCode IS NOT NULL
          AND machineCode IS NOT NULL GROUP BY 1,2""").pl()
    obs = obs.with_columns(
        pl.col("m").str.extract(r"PCR(\d+)").cast(pl.Int64).alias("mn"),
        pl.col("gt").map_elements(rim_of, return_dtype=pl.Int64).alias("rim"))
    # infer the truncated upper bound for machines 6-11 from what they ran
    hi6 = obs.filter((pl.col("mn") >= 6) & pl.col("rim").is_not_null())["rim"].max()
    print(f"\nGAP-2  machines 6-11 upper bound: spec truncated; "
          f"observed max rim = {hi6}\"  -> inferred, needs plant confirmation")

    elig = []
    for mn, (lo, hi, src) in PCR_INCH.items():
        hi_use = hi if hi is not None else int(hi6 or 20)
        for rim in range(lo, hi_use + 1):
            elig.append({"plant": "PCR", "machine": f"TBMPCR{mn}Stage2",
                         "machine_no": mn, "machine_type": PCR_TYPE[mn],
                         "rim": rim, "inch_lo": lo, "inch_hi": hi_use,
                         "source": src,
                         "chg_same_min": PCR_CHG[PCR_TYPE[mn]][0],
                         "chg_diff_min": PCR_CHG[PCR_TYPE[mn]][1]})
    ef = pl.DataFrame(elig)
    ef.write_parquet(OUT / "pcr_inch_eligibility.parquet")
    shutil.copy2(OUT / "pcr_inch_eligibility.parquet",
                 INP / "pcr_inch_eligibility.parquet")
    print(f"  pcr_inch_eligibility : {ef.height} (machine, rim) pairs, "
          f"{ef['rim'].n_unique()} rims, 11 machines")

    # ---- the habit-vs-physics test --------------------------------------
    per_gt = (obs.filter(pl.col("rim").is_not_null())
              .group_by(["gt", "rim"])
              .agg(pl.col("mn").n_unique().alias("machines_used"),
                   pl.col("n").sum().alias("tyres")))
    cap = ef.group_by("rim").agg(pl.col("machine_no").n_unique().alias("capable"))
    t = per_gt.join(cap, on="rim", how="left").with_columns(
        pl.col("capable").fill_null(0))
    print("\n  HABIT vs PHYSICS  (PCR, 8 months)")
    print(f"    {'rim':>5}{'GTs':>6}{'tyres':>12}{'machines capable':>18}"
          f"{'machines used p50':>19}")
    for r in t.group_by("rim").agg(
            pl.len().alias("gts"), pl.col("tyres").sum().alias("ty"),
            pl.col("capable").first().alias("cap"),
            pl.col("machines_used").median().alias("used")).sort("rim").iter_rows(named=True):
        print(f"    {r['rim']:>5}{r['gts']:>6}{r['ty']:>12,}{r['cap']:>18}"
              f"{r['used']:>19.0f}")
    locked = t.filter((pl.col("machines_used") == 1) & (pl.col("capable") > 1))
    print(f"\n    GTs locked to 1 machine but capable on >1 : "
          f"{locked.height} of {t.height}")
    print(f"    volume in those GTs                       : "
          f"{int(locked['tyres'].sum()):,} tyres "
          f"({100*locked['tyres'].sum()/max(t['tyres'].sum(),1):.1f}%)")
    print("    >>> " + ("HABIT -- dedication is not explained by inch capability"
                        if locked.height > t.height * 0.5 else
                        "PHYSICS -- capability explains most dedication"))

    # ---- GAP-12 ---------------------------------------------------------
    og = ROOT / "masters" / "opening_gt"
    d = paths.OPENING_GT
    d.mkdir(parents=True, exist_ok=True)
    n = 0
    for p in og.glob("*"):
        shutil.copy2(p, d / p.name); n += 1
    print(f"\nGAP-12 opening_gt : {n} files copied to INPUT/opening_gt/")


if __name__ == "__main__":
    main()
