"""Build ``sku_con_cluster.parquet`` — per-GT construction cluster and the
workbook's own machine assignment, from the pre-built SKU_Construction_Clusters
workbooks.

WHAT THESE WORKBOOKS ARE
------------------------
Two files, one per plant, each with five sheets:

    Methodology & Findings   the author's own summary (READ IT — it disagrees
                             with the "use clusters to group PCR" reading)
    Machine Groups           machine code -> SKUs / rims / cluster count
    Construction Clusters    ConCluster · N_SKUs · Rims · Assigned_Machine ·
                             Construction_Signature · SKUs
    SKU Assignment           one row per SKU: Assigned_Machine · Rim ·
                             ConCluster · GT · SKU · Size · Hist_Machines ·
                             Hist_Dominant · Feb..Jul · component slots
    Triangular Map           SKU x machine eligibility grid, Y = assigned,
                             y = other machines the GT historically used

They were produced by average-linkage clustering on Hamming distance over the
component codes, cut at 0.3.

THE FIVE CORRECTNESS RULES, EACH ONE A DEFECT THAT WAS FOUND
------------------------------------------------------------
1. **The unit of a cluster is a SKU; the unit of the planner is a GT.**
   PCR runs 1.23 SKUs per GT and TBR 1.95, so a "cluster of 2 SKUs" is very
   often ONE green tyre listed twice. Every size figure here is therefore
   reported in GT space, and a cluster is only "groupable" if it holds >1
   DISTINCT planner GT that is co-active in the month being planned.

2. **The TBR GT-code namespace trap.** The TBR workbook's ``GT`` column is
   ``GT 5003``-style and matches the planner's key ``0 %`` of the time — the
   planner plans on ``v_build`` stage-2 ``itemCode`` (``10.00 R 20 JUH5``).
   TBR must be bridged SKU -> gt_code through the warehouse bridge tables
   (99.1 % of volume). PCR's ``GT`` column IS planner-style and matches
   directly; both routes agree on PCR, so both are used and cross-checked.

3. **Never match a PCR GT on its 4-digit numeric core.** Those digits encode
   SIZE. ``GT 1513 XPC1 MSIL`` (the largest July GT) and the workbook's
   ``GT1513 NEO`` are different tread patterns. The core is a size, not a key.

4. **The machine codes 3401-3411 / 3801-3809 appear NOWHERE in the MES.**
   They are derived here, not assumed, by cross-tabulating the workbook's own
   ``Feb``..``Jul`` per-month machine column against each GT's MES dominant
   machine, and the derivation is ASSERTED at build time. If purity drops
   below ``--min-purity`` the build fails loudly rather than shipping a wrong
   mapping. (Measured: identity, 34NN -> TBMPCR<NN>Stage2 and
   38NN -> TBMTBR<NN>Stage2, at 98-100 % over 222 PCR and 415 TBR GT-months.)

5. **A cluster that spans more than one RIM is not a construction family.**
   PCR cluster 132 holds 15 GTs across all seven rims and seven machines — it
   is the dustbin average-linkage leaves at cut 0.3, and using it as an
   adjacency signal would actively encourage different-size changeovers.
   Such clusters are flagged ``multi_rim`` and, by default, their members are
   emitted with ``con_cluster = None`` so no consumer can group on them.
   ``--keep-multi-rim`` disables the filter.

Output ``INPUT/derived/sku_con_cluster.parquet``:

    plant             str   PCR | TBR
    gt_code           str   the PLANNER's key (v_build stage-2 itemCode)
    con_cluster       str   workbook ConCluster id, NULL if dropped by rule 5
    raw_cluster       str   the id before rule 5, always populated
    multi_rim         bool  the raw cluster spans >1 rim
    assigned_machine  str   workbook Assigned_Machine, MAPPED to the MES name
    assigned_code     str   the raw workbook code (3401 ...)
    hist_dominant     str   workbook Hist_Dominant, mapped
    hist_machines     str   workbook Hist_Machines, mapped, comma-joined
    eligible_machines str   Triangular Map Y+y, mapped, comma-joined
    rim               str   rim from gt_size.parquet (the planner's own rim)
    wb_rim            str   the workbook's Rim label (often blank)
    signature         str   Construction_Signature of the raw cluster
    n_sku             int   workbook SKUs that resolved to this GT
    bridge            str   how the GT was reached: direct | sku | both

Usage
-----
    .venv/Scripts/python.exe scripts/build_sku_con_cluster.py
    .venv/Scripts/python.exe scripts/build_sku_con_cluster.py --keep-multi-rim
    .venv/Scripts/python.exe scripts/build_sku_con_cluster.py --min-purity 0.9
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import duckdb
import openpyxl
import polars as pl

ROOT = Path(__file__).resolve().parent.parent          # schedule/send
WORKSPACE = ROOT.parent.parent                         # repo wrapper root
WAREHOUSE = ROOT / "warehouse"
IN_DERIVED = WORKSPACE / "INPUT" / "derived"
DEFAULT_OUT = IN_DERIVED / "sku_con_cluster.parquet"

WB = {
    "PCR": WORKSPACE / "SKU_Construction_Clusters_PCR (1).xlsx",
    "TBR": WORKSPACE / "SKU_Construction_Clusters_TBR (1).xlsx",
}
PREFIX = {"PCR": "TBMPCR", "TBR": "TBMTBR"}
# The workbook's per-month columns. Used ONLY to derive the machine mapping.
MONTH_COLS = {"Feb": "2026-02", "Mar": "2026-03", "Apr": "2026-04",
              "May": "2026-05", "Jun": "2026-06", "Jul": "2026-07"}
BLANKS = {"", "nan", "None", "-", "NaN", "#N/A", "NA", "N/A"}


# --------------------------------------------------------------------------- io
def _c(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s in BLANKS else s


def _grid(path: Path, sheet: str) -> list[list]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    g = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    while g and all(v is None for v in g[-1]):
        g.pop()
    return g


def _sheet(path: Path, sheet: str, key: str) -> tuple[list[dict], list[str]]:
    g = _grid(path, sheet)
    hdr = [_c(v) for v in g[0]]
    out = []
    for row in g[1:]:
        d = {hdr[i]: (_c(row[i]) if i < len(row) else None) for i in range(len(hdr))}
        if d.get(key):
            out.append(d)
    return out, hdr


# --------------------------------------------------------------------------- warehouse
def _bpath(*parts: str) -> str:
    return str(WAREHOUSE.joinpath(*parts)).replace("\\", "/")


def load_bridge(con: duckdb.DuckDBPyConnection) -> dict[str, set[str]]:
    """workbook SKU -> candidate planner gt_code. Exact keys only (rule 3)."""
    queries = [
        ("gt_sku_master", f"SELECT gt_code, sku_code FROM read_parquet('{_bpath('derived','gt_sku_master.parquet')}')"),
        ("gt_sku_from_recipe", f"SELECT gt_code, sku_code FROM read_parquet('{_bpath('derived','gt_sku_from_recipe.parquet')}')"),
        ("gt_size", f"SELECT gt_code, sku FROM read_parquet('{_bpath('derived','gt_size.parquet')}')"),
        ("bom_gt_map", f"SELECT gt_code, super_parent_sku FROM read_parquet('{_bpath('bom','bom_gt_map.parquet')}')"),
        ("recipe_bridge", f"SELECT gt_code, sku FROM read_parquet('{_bpath('derived','recipe_bridge.parquet')}')"),
        ("recipe_gt_sku.gt", f"SELECT gt_name, gt_code FROM read_parquet('{_bpath('derived','recipe_gt_sku.parquet')}')"),
        ("recipe_gt_sku.sku", f"SELECT gt_name, sku_code FROM read_parquet('{_bpath('derived','recipe_gt_sku.parquet')}')"),
    ]
    out: dict[str, set[str]] = defaultdict(set)
    for name, q in queries:
        try:
            rows = con.execute(q).fetchall()
        except Exception as exc:                                   # noqa: BLE001
            print(f"  !! bridge source {name} unavailable: {str(exc)[:80]}")
            continue
        n = 0
        for gt, sku in rows:
            if gt and sku:
                out[str(sku).strip()].add(str(gt).strip())
                n += 1
        print(f"  bridge {name:22s} {n:6d} pairs")
    return dict(out)


def load_mes(con: duckdb.DuckDBPyConnection):
    """(plant,gt,month,machine) -> tyres, over the workbook's own Feb-Jul window."""
    g = str(WAREHOUSE / "building_output" / "plant=*" / "stage=2" / "date=*"
            / "**" / "*.parquet").replace("\\", "/")
    rows = con.execute(f"""
        SELECT plant, itemCode gt, strftime(date,'%Y-%m') m, machineName mach, count(*) n
        FROM read_parquet('{g}', hive_partitioning=1)
        WHERE date >= DATE '2026-02-01' AND date < DATE '2026-08-01'
        GROUP BY 1,2,3,4
    """).fetchall()
    gm: dict = defaultdict(int)
    vol: dict = defaultdict(int)
    for plant, gt, m, mach, n in rows:
        gm[(plant, gt, m, mach)] += n
        vol[(plant, gt)] += n
    return gm, vol


# --------------------------------------------------------------------------- mapping
def derive_machine_map(plant: str, assign: list[dict], sku2gt: dict[str, str],
                       gm: dict, min_purity: float) -> dict[str, str]:
    """DERIVE workbook machine code -> MES machineName. Never assume identity.

    Evidence: for every (SKU, month) where the workbook names a machine, take
    that GT's MES dominant machine in that month and cross-tabulate. Assert the
    winner's purity so a silent mis-map is impossible.
    """
    cont: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for d in assign:
        gt = sku2gt.get(d["SKU"])
        if not gt:
            continue
        for col, mon in MONTH_COLS.items():
            code = d.get(col)
            if not code:
                continue
            cand = {mach: n for (p, g2, m2, mach), n in gm.items()
                    if p == plant and g2 == gt and m2 == mon}
            if cand:
                cont[code][max(cand, key=cand.get)] += 1
    mapping: dict[str, str] = {}
    bad: list[str] = []
    print(f"  {plant} machine-code derivation "
          f"({sum(sum(v.values()) for v in cont.values())} GT-months of evidence)")
    for code in sorted(cont):
        d1 = cont[code]
        best = max(d1, key=d1.get)
        pur = d1[best] / sum(d1.values())
        mapping[code] = best
        ident = f"{PREFIX[plant]}{int(code) % 100}Stage2"
        tag = "identity" if best == ident else f"NOT identity (ident={ident})"
        print(f"      {code} -> {best:16s} purity {pur:6.1%} on {sum(d1.values()):>4d}   {tag}")
        if pur < min_purity:
            bad.append(f"{code} purity {pur:.1%}")
    if bad:
        raise SystemExit(
            f"!! {plant}: machine-code mapping is not trustworthy: {'; '.join(bad)}.\n"
            f"   Refusing to emit. Re-derive before using Assigned_Machine.")
    # Codes with no evidence at all: fill by ELIMINATION only, and say so.
    all_codes = {c for d in assign for c in (
        [d.get("Assigned_Machine"), d.get("Hist_Dominant")]
        + list((d.get("Hist_Machines") or "").split(","))) if c and c.strip().isdigit()}
    for code in sorted(all_codes - set(mapping)):
        ident = f"{PREFIX[plant]}{int(code) % 100}Stage2"
        taken = set(mapping.values())
        mapping[code] = ident
        print(f"      {code} -> {ident:16s} NO EVIDENCE — by elimination"
              f"{' (COLLIDES with a derived code!)' if ident in taken else ''}")
    return mapping


def _map_list(s: str | None, mapping: dict[str, str]) -> str | None:
    if not s:
        return None
    out = [mapping.get(x.strip(), x.strip()) for x in s.split(",") if x.strip()]
    return ",".join(sorted(set(out))) or None


# --------------------------------------------------------------------------- main
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--min-purity", type=float, default=0.85,
                    help="fail if a machine code's dominant MES machine is less pure")
    ap.add_argument("--keep-multi-rim", action="store_true",
                    help="do NOT null out clusters that span >1 rim (rule 5)")
    args = ap.parse_args()
    out_path = Path(args.out)

    print("=" * 82)
    print("  build_sku_con_cluster")
    print("=" * 82)
    for p, f in WB.items():
        print(f"  {p} workbook: {f}   {'OK' if f.exists() else '!! MISSING'}")
        if not f.exists():
            raise SystemExit(f"!! missing workbook {f}")
    print(f"  out       : {out_path}")
    print()

    rim_of = {r["gt_code"]: str(r["rim"]) for r in
              pl.read_parquet(IN_DERIVED / "gt_size.parquet").iter_rows(named=True)
              if r.get("gt_code") and r.get("rim")}

    con = duckdb.connect()
    sku2gt_all = load_bridge(con)
    gm, vol = load_mes(con)
    con.close()
    print()

    records: list[dict] = []
    for plant in ("PCR", "TBR"):
        assign, _ = _sheet(WB[plant], "SKU Assignment", "SKU")
        clusters, _ = _sheet(WB[plant], "Construction Clusters", "ConCluster")
        tri, tri_hdr = _sheet(WB[plant], "Triangular Map", "SKU")
        mcols = [h for h in tri_hdr if h and h.isdigit()]
        active = {gt for (p, gt) in vol if p == plant}
        tot_vol = sum(v for (p, gt), v in vol.items() if p == plant)

        # ---- SKU -> planner gt_code. Two routes, cross-checked (rule 2).
        sku2gt: dict[str, str] = {}
        how: dict[str, str] = {}
        n_amb = 0
        for d in assign:
            sku = d["SKU"]
            via_sku = {x for x in sku2gt_all.get(sku, set()) if x in active}
            via_gt = {d["GT"]} if d.get("GT") in active else set()
            if len(via_sku) > 1:
                n_amb += 1
                via_sku = set()
            cand = via_sku | via_gt
            if len(cand) == 1:
                gt = next(iter(cand))
                sku2gt[sku] = gt
                how[gt] = ("both" if (via_sku and via_gt) else
                           ("sku" if via_sku else "direct"))
            elif len(cand) > 1:
                n_amb += 1
        disagree = sum(1 for d in assign
                       if d.get("GT") in active
                       and sku2gt.get(d["SKU"]) not in (None, d["GT"]))
        print(f"--- {plant}: {len(assign)} workbook SKU rows")
        print(f"      resolved to an active planner GT : {len(sku2gt)}  "
              f"({n_amb} ambiguous, {disagree} where the two routes disagree)")

        mapping = derive_machine_map(plant, assign, sku2gt, gm, args.min_purity)

        # ---- GT -> cluster (modal over its SKUs, then lexically smallest)
        cl_votes: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        mach_votes: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        nsku: dict[str, int] = defaultdict(int)
        wbrim: dict[str, str] = {}
        histdom: dict[str, str] = {}
        histm: dict[str, str] = {}
        for d in assign:
            gt = sku2gt.get(d["SKU"])
            if not gt:
                continue
            nsku[gt] += 1
            if d.get("ConCluster"):
                cl_votes[gt][d["ConCluster"]] += 1
            if d.get("Assigned_Machine"):
                mach_votes[gt][d["Assigned_Machine"]] += 1
            wbrim.setdefault(gt, d.get("Rim"))
            histdom.setdefault(gt, d.get("Hist_Dominant"))
            histm.setdefault(gt, d.get("Hist_Machines"))
        elig: dict[str, set[str]] = defaultdict(set)
        for d in tri:
            gt = sku2gt.get(d["SKU"])
            if gt:
                elig[gt] |= {m for m in mcols if d.get(m)}

        gt2cl = {gt: sorted(v.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]
                 for gt, v in cl_votes.items()}
        gt2mach = {gt: sorted(v.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]
                   for gt, v in mach_votes.items()}
        n_clconf = sum(1 for v in cl_votes.values() if len(v) > 1)
        n_mconf = sum(1 for v in mach_votes.values() if len(v) > 1)
        print(f"      GTs reached {len(gt2cl)}; SKUs of one GT disagree on cluster "
              f"{n_clconf}, on machine {n_mconf}")

        sig = {c["ConCluster"]: c.get("Construction_Signature") for c in clusters}

        # ---- rule 5: a cluster spanning >1 rim is not a family
        cl_rims: dict[str, set[str]] = defaultdict(set)
        for gt, c in gt2cl.items():
            r = rim_of.get(gt)
            if r:
                cl_rims[c].add(r)
        multi_rim = {c for c, rs in cl_rims.items() if len(rs) > 1}
        if multi_rim:
            print(f"      MULTI-RIM clusters (rule 5): {sorted(multi_rim)}")
            for c in sorted(multi_rim):
                mem = [gt for gt, cc in gt2cl.items() if cc == c]
                print(f"          cluster {c}: {len(mem)} GTs across rims "
                      f"{sorted(cl_rims[c])}  "
                      f"{sum(vol[(plant, gt)] for gt in mem):,} tyres"
                      f"{'  -> DROPPED' if not args.keep_multi_rim else '  -> KEPT'}")

        for gt, c in sorted(gt2cl.items()):
            keep = args.keep_multi_rim or c not in multi_rim
            records.append({
                "plant": plant,
                "gt_code": gt,
                "con_cluster": c if keep else None,
                "raw_cluster": c,
                "multi_rim": c in multi_rim,
                "assigned_code": gt2mach.get(gt),
                "assigned_machine": mapping.get(gt2mach.get(gt) or "", None),
                "hist_dominant": mapping.get(histdom.get(gt) or "", None),
                "hist_machines": _map_list(histm.get(gt), mapping),
                "eligible_machines": ",".join(sorted(mapping.get(m, m)
                                                     for m in elig.get(gt, set()))) or None,
                "rim": rim_of.get(gt),
                "wb_rim": wbrim.get(gt),
                "signature": sig.get(c),
                "n_sku": nsku.get(gt, 0),
                "bridge": how.get(gt),
            })

        # ---- coverage, printed EVERY run
        cov_v = sum(vol[(plant, gt)] for gt in gt2cl if (plant, gt) in vol)
        print(f"      COVERAGE Feb-Jul MES: {len(gt2cl)}/{len(active)} GTs, "
              f"{cov_v:,}/{tot_vol:,} tyres = {cov_v / tot_vol:.1%}")
        miss = sorted(active - set(gt2cl), key=lambda gt: -vol[(plant, gt)])[:8]
        if miss:
            print("      largest UNCOVERED GTs (no workbook row):")
            for gt in miss:
                print(f"          {gt!r:36s} {vol[(plant, gt)]:>9,}")
        # groupability, in GT space, among covered GTs
        cl2gt: dict[str, list[str]] = defaultdict(list)
        for gt, c in gt2cl.items():
            if args.keep_multi_rim or c not in multi_rim:
                cl2gt[c].append(gt)
        sizes = sorted((len(v) for v in cl2gt.values()), reverse=True)
        ngroup = sum(s for s in sizes if s > 1)
        print(f"      GT-space clusters {len(sizes)}, sizes {sizes[:14]}"
              f"{' ...' if len(sizes) > 14 else ''}")
        print(f"      GTs in a multi-GT cluster: {ngroup}/{len(gt2cl)} = "
              f"{ngroup / max(len(gt2cl), 1):.0%}")
        print()

    df = pl.DataFrame(records)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.write_parquet(out_path)
    print("=" * 82)
    print(f"  wrote {out_path}  ({df.height} rows)")
    print(df.group_by("plant").agg(
        pl.len().alias("gts"),
        pl.col("con_cluster").n_unique().alias("clusters"),
        pl.col("assigned_machine").n_unique().alias("machines")))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
