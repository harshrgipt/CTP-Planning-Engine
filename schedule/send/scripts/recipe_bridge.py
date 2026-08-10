"""RECIPE BRIDGE -- the authoritative GT <-> SKU resolution (plant-specified).

    python scripts/recipe_bridge.py

THE JOIN PATH, exactly as the plant describes it:

  1. curing.recipeID            ->  recipemaster.iD
     recipemaster.SAPMaterialCode is the SAP code for that recipe.

  2. If SAPMaterialCode resolves to a GT CODE (not a finished SKU), then that
     row's `iD` is a BUILDING recipe id.  Bridge it to curing:

         recipemaster.iD  ==  recipelookup.tbmRecipeID
                              recipelookup.curingRecipeID
                          ==  recipemaster.iD   ->  SAPMaterialCode = SKU CODE

  So the full chain is:

     GT code --(SAPMaterialCode)--> iD --(tbmRecipeID)--> curingRecipeID
             --(iD)--> SAPMaterialCode == SKU

This supersedes name-matching heuristics: it is the plant's own key path, so it
is authoritative wherever it resolves.
"""
from __future__ import annotations

import warnings
from pathlib import Path

import polars as pl

warnings.filterwarnings("ignore")
from openpyxl import load_workbook                            # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT.parent.parent
OUT = ROOT / "warehouse" / "derived"


def sheet_df(path: Path) -> pl.DataFrame:
    wb = load_workbook(path, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(x).strip() if x is not None else f"c{i}"
           for i, x in enumerate(rows[0])]
    data = {h: [] for h in hdr}
    for r in rows[1:]:
        for i, h in enumerate(hdr):
            data[h].append(str(r[i]).strip() if i < len(r) and r[i] is not None
                           else None)
    return pl.DataFrame(data)


def main() -> None:
    rm = sheet_df(SRC / "Recipemaster 1.xlsx")
    rl = sheet_df(SRC / "recipelookup 1.xlsx")
    print(f"recipemaster {rm.height} rows   recipelookup {rl.height} rows")

    rm = rm.select(["iD", "SAPMaterialCode", "description", "tyreSize",
                    "MaxAging", "Minaging", "curingCapacity"]).filter(
        pl.col("iD").is_not_null())
    rl = rl.select(["tbmRecipeID", "curingRecipeID"]).filter(
        pl.col("tbmRecipeID").is_not_null()
        & pl.col("curingRecipeID").is_not_null()).unique()
    print(f"  recipelookup usable pairs: {rl.height}")

    # ---- step 1: recipe id -> SAP code ---------------------------------
    code = rm.select(["iD", "SAPMaterialCode", "description"]).rename(
        {"SAPMaterialCode": "sap", "description": "descr"})

    # ---- step 2: build recipe -> cure recipe -> SKU ---------------------
    chain = (code.rename({"iD": "tbmRecipeID", "sap": "gt_sap",
                          "descr": "gt_descr"})
             .join(rl, on="tbmRecipeID", how="inner")
             .join(code.rename({"iD": "curingRecipeID", "sap": "sku",
                                "descr": "sku_descr"}),
                   on="curingRecipeID", how="inner"))
    print(f"  resolved GT->SKU chains: {chain.height}")

    # which side is the GT? compare against the MES GT universe
    from planner.data.warehouse import duck, set_cutoff
    set_cutoff(None)
    mes = duck().execute(
        "SELECT DISTINCT plant, itemCode AS gt FROM v_build "
        "WHERE stage=2 AND itemCode IS NOT NULL").pl()
    gts = set(mes["gt"].to_list())
    dem = pl.read_parquet(ROOT / "masters" / "demand" / "demand_2026-07.parquet")
    skus = set(dem["sku"].unique().to_list())
    dem_gts = {p: set(dem.filter(pl.col("plant") == p)["gt_code"].unique().to_list())
               for p in ["PCR", "TBR"]}

    hit_gt = chain.filter(pl.col("gt_sap").is_in(list(gts))).height
    hit_gt_d = chain.filter(pl.col("gt_descr").is_in(list(gts))).height
    hit_sku = chain.filter(pl.col("sku").is_in(list(skus))).height
    print(f"\n  chain.gt_sap  matches an MES GT code : {hit_gt}")
    print(f"  chain.gt_descr matches an MES GT code : {hit_gt_d}")
    print(f"  chain.sku     matches a July SKU      : {hit_sku}")

    # TBR MES codes carry a suffix ("GT 5117 - 10.00R20 CUC (BD)") and some
    # bridge codes carry a "GT" prefix on a bare size ("GT295/90R20JUHXF").
    # Normalise both sides to a join key before matching.
    import re as _re

    def norm(x: str | None) -> str | None:
        if not x:
            return None
        t = str(x).upper().strip()
        m = _re.match(r"^GT\s*(\d{3,5})", t)
        if m:
            return f"GT{m.group(1)}"
        t = _re.sub(r"^GT\s*", "", t)
        return _re.sub(r"[^A-Z0-9./]", "", t)

    gtcol = "gt_descr" if hit_gt_d >= hit_gt else "gt_sap"
    out = (chain.select([pl.col(gtcol).alias("gt_code"), "sku", "sku_descr",
                         "tbmRecipeID", "curingRecipeID"])
           .filter(pl.col("gt_code").is_not_null() & pl.col("sku").is_not_null())
           .unique())
    print(f"\n  bridge rows: {out.height}   using '{gtcol}' as the GT key")

    # add both raw and normalised keys; downstream joins on gt_key
    out = out.with_columns(
        pl.col("gt_code").map_elements(norm, return_dtype=pl.Utf8).alias("gt_key"))
    for p in ["PCR", "TBR"]:
        raw = set(out["gt_code"].to_list())
        nk = set(out["gt_key"].drop_nulls().to_list())
        dn = {g: norm(g) for g in dem_gts[p]}
        hit_raw = len(dem_gts[p] & raw)
        hit_norm = sum(1 for g, k in dn.items() if k in nk)
        print(f"    {p}: raw {100*hit_raw/max(len(dem_gts[p]),1):>3.0f}%  "
              f"normalised {100*hit_norm/max(len(dem_gts[p]),1):>3.0f}%  "
              f"({hit_norm}/{len(dem_gts[p])})")

    out.write_parquet(OUT / "recipe_bridge.parquet")
    print(f"\n  -> {OUT / 'recipe_bridge.parquet'}")

    # ---- curing recipeID coverage in MES --------------------------------
    cov = duck().execute("""
        SELECT plant, count(*) n,
               avg(CASE WHEN recipeID IS NULL THEN 1.0 ELSE 0 END) null_frac,
               count(DISTINCT recipeID) distinct_recipes
        FROM v_curing GROUP BY 1""").pl()
    print("\n  v_curing.recipeID availability:")
    for r in cov.iter_rows(named=True):
        print(f"    {r['plant']}: {r['n']:,} rows, "
              f"{r['distinct_recipes']} distinct recipeIDs, "
              f"{100*r['null_frac']:.1f}% null")


if __name__ == "__main__":
    main()
