"""GT -> SKU SPLIT SHARES and the RECIPE-CODE bridge, for one month.

    python -m scripts.build_gt_sku_share --month 2026-07

`gt_sku_from_recipe.parquet` keeps only the MODAL SKU per GT (`r = 1`), which is
what the demand file inherited -- one SKU per GT, 48 PCR and 56 GTs in July. That
is enough to LABEL a GT and not enough to SPLIT one, and the BTP output format is
SKU-level. This writes the same query without the `r = 1` filter, so the full
distribution survives.

    plant · gt_code · sku_code · recipe_name · tyres · share

`recipe_name` is the plant's own curing recipe code -- `CTCP1201` on PCR,
`CTCT7.502001` on TBR -- and is the join key the plant asked for. It comes from
`Recipemaster 1.xlsx` via `v_curing.recipeID`, the same chain
`scripts/build_gt_sku.py` documents at 100 % of cured volume on both plants.

WHY THE RECIPE CODE AND NOT THE SKU STRING
  The SKU namespace has misled this project twice (TBR's `gt_sku_master` maps a
  SKU to a "GT 5001" label while the engine's TBR gt_code is size-led, which cost
  TBR 100 % of its cycle-time coverage until the join order was fixed). The
  recipe code is a single flat namespace with no such variants. Coverage of both
  routes is measured and printed here; take the better one and say which.
"""
from __future__ import annotations

import argparse
from pathlib import Path

import polars as pl

from planner.config import CONFIG
from planner.data.warehouse import duck, set_cutoff
from planner import paths

ROOT = Path(__file__).resolve().parent.parent
RM = paths.raw("Recipemaster 1.xlsx")


def _c(x) -> str:
    return "".join(ch for ch in str(x) if ord(ch) < 128).strip() if x is not None else ""


def build(month: str) -> pl.DataFrame:
    set_cutoff(None)
    con = duck()
    import openpyxl
    wb = openpyxl.load_workbook(RM, read_only=True, data_only=True)
    raw = [[_c(c) for c in r] for r in wb.worksheets[0].iter_rows(values_only=True)]
    wb.close()
    hdr = raw[0]
    rm = [dict(zip(hdr, r)) for r in raw[1:] if any(r)]
    con.register("recipe", pl.DataFrame([{
        "recipe_id": d["iD"], "sku_code": d.get("SAPMaterialCode", ""),
        "recipe_name": d.get("name", ""), "sku_desc": d.get("description", ""),
    } for d in rm]).to_arrow())
    d = con.execute(f"""
        WITH x AS (
            SELECT b.plant, b.itemCode AS gt_code, r.sku_code, r.sku_desc,
                   r.recipe_name, count(*) AS n
            FROM v_curing c
            JOIN v_build b ON b.productionID = c.gtbarCode AND b.stage = 2
            JOIN recipe r  ON c.recipeID::VARCHAR = r.recipe_id
            WHERE c.statuscritical = 'Normal' AND b.itemCode IS NOT NULL
              AND c.date >= DATE '{month}-01'
              AND c.date <  DATE '{month}-01' + INTERVAL 1 MONTH
            GROUP BY 1,2,3,4,5)
        SELECT plant, gt_code, sku_code, sku_desc, recipe_name, n AS tyres,
               sum(n) OVER (PARTITION BY plant, gt_code) AS gt_tyres
        FROM x ORDER BY plant, gt_code, n DESC, sku_code""").pl()
    d = d.with_columns((pl.col("tyres") / pl.col("gt_tyres")).alias("share"))
    con.unregister("recipe")
    return d


def build_from_demand(month: str) -> pl.DataFrame:
    """Shares taken from the ORDER BOOK itself. The only correct source forward.

    The curing-recipe route above measures a month that has already happened. For
    a month being PLANNED there is no cured volume to measure, and borrowing the
    previous month's share would attribute a GT's output to SKUs nobody has
    ordered -- a fabricated split that reads as fact on the shop-floor sheet.

    An order book keyed on (plant, gt_code, sku) already IS the split: the plant
    asked for N of this SKU and M of that one off the same green tyre. Use it,
    and label the source so a reader can tell the two apart.
    """
    dem = pl.read_parquet(CONFIG.paths.masters / "demand" / f"demand_{month}.parquet")
    d = (dem.group_by(["plant", "gt_code", "sku"])
         .agg(pl.col("qty").sum().alias("tyres"))
         .rename({"sku": "sku_code"}))
    d = d.with_columns(
        pl.col("tyres").sum().over(["plant", "gt_code"]).alias("gt_tyres"))
    d = d.with_columns((pl.col("tyres") / pl.col("gt_tyres")).alias("share"))
    # descriptions + the plant's own curing recipe code, from the same masters
    # the recipe route uses, so both variants carry identical columns
    desc, rname = {}, {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(RM, read_only=True, data_only=True)
        raw = [[_c(c) for c in r] for r in wb.worksheets[0].iter_rows(values_only=True)]
        wb.close()
        hdr = raw[0]
        for row in raw[1:]:
            r = dict(zip(hdr, row))
            k = r.get("SAPMaterialCode", "")
            if k:
                desc.setdefault(k, r.get("description", ""))
                rname.setdefault(k, r.get("name", ""))
    except Exception as exc:                                    # noqa: BLE001
        print(f"  (recipe master unreadable, descriptions blank: {exc})")
    return (d.with_columns([
        pl.col("sku_code").replace_strict(desc, default="").alias("sku_desc"),
        pl.col("sku_code").replace_strict(rname, default="").alias("recipe_name"),
    ]).select(["plant", "gt_code", "sku_code", "sku_desc", "recipe_name",
               "tyres", "gt_tyres", "share"])
      .sort(["plant", "gt_code", "tyres"], descending=[False, False, True]))


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", default="2026-07")
    ap.add_argument("--from-demand", action="store_true",
                    help="derive the split from masters/demand/demand_<month> "
                         "instead of from cured MES volume. REQUIRED for a "
                         "forward month, which has no cured volume to measure.")
    a = ap.parse_args()
    d = build_from_demand(a.month) if a.from_demand else build(a.month)
    out = CONFIG.paths.warehouse / "derived" / f"gt_sku_share_{a.month}.parquet"
    d.write_parquet(out, compression="zstd")
    src = "ORDER BOOK (demand file)" if a.from_demand else "cured MES volume"
    print(f"GT->SKU shares, {a.month}, from {src}: {d.height} (plant, GT, SKU) "
          f"rows, {d['sku_code'].n_unique()} SKUs, {int(d['tyres'].sum()):,} tyres")
    for p in sorted(d["plant"].unique().to_list()):
        s = d.filter(pl.col("plant") == p)
        g = s.group_by("gt_code").agg(pl.len().alias("n"))
        print(f"  {p}: {g.height} GTs -> {s.height} GT-SKU pairs   "
              f"SKUs/GT p50 {g['n'].median():.0f} max {g['n'].max()}   "
              f"{int((g['n'] > 1).sum())} GTs split")
    print(f"\nWROTE {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
