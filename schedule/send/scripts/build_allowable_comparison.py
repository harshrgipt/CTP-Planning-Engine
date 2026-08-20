"""ALLOWABLE MACHINE COMPARISON SHEET -- ours vs the plant's, with a verdict.

    PYTHONPATH=. python scripts/build_allowable_comparison.py

Writes ALLOWABLE_COMPARISON.xlsx at the wrapper root: one sheet per plant, one
row per demanded GT, showing three sources side by side.

THE THREE SOURCES, AND WHY THEY DIFFER
  1. PLANT FILE (raw)   INPUT/allowable machine/*.xlsx, read cell by cell and
                        UNIONED across every FG row that names the same GT.
                        PCR only -- the TBR file keys GT in the BOM namespace
                        ("GT 5001"), which has zero string overlap with the MES
                        codes the engine plans in, so it is bridged on SKU
                        inside the ingest and cannot be re-joined here by name.
  2. OUR ALLOWABLE      allowed_machine_matrix.parquet -- what the engine
                        actually enforces through allowable.restrict().
  3. PLANT OBSERVED     cap_machine_<M> where basis is OBSERVED / BOTH /
                        CERTIFIED, plus history_tyres_8mo from
                        machine_gt_preference -- the machines the plant's own
                        MES shows ACTUALLY building that GT over 8 months.

     DO NOT FILTER ON n_used > 0. `n_used` counts rows in the capability table
     and 547 of 813 PCR rows carry basis INCH -- machines theoretically capable
     by RIM that the plant has never run that GT on. Filtering n_used > 0 made
     GT 2258 RAN HPE look like it ran on 8 machines when the 8-month history
     shows ONE (TBMPCR2, 37,650 tyres, 99.7 % share) -- exactly what our
     allowable already says. That error overstated the plant-vs-file conflict
     3x (26 GTs / 12,154 tyres claimed, 12 GTs / 3,848 actual).

  Two verdicts, and they answer different questions:
    INGEST OK?          does 2 match 1?  A cross is OUR bug.
    OBSERVED WITHIN?    is 3 a subset of 2?  A cross means the plant HAS RUN
                        that GT on a machine its own file forbids.

  Measured 2026-08-13 on July, observed basis, across 45 short GTs:
      INGEST mismatch          1 GT  · 1,378 tyres  (GT 1513, union bug)
      OBSERVED outside ours   12 GTs · 3,848 tyres
  The plant's file agrees with the plant's own 8-month behaviour on most GTs:
  GT 2258 RAN HPE, GT 1402 XPC TATA, GT 2568 HT2 and GT 2267 ROYL HYU all match
  exactly. Those shortfalls are NOT an allowable problem.
"""
from __future__ import annotations

import collections
import re
import sys
from pathlib import Path

import openpyxl
import polars as pl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from planner import paths                                          # noqa: E402

OUT = paths.ROOT.parent.parent / "ALLOWABLE_COMPARISON.xlsx"
# Excel holds an exclusive lock on an open workbook; fall back rather than die.
_ALT = paths.ROOT.parent.parent / "ALLOWABLE_COMPARISON_v2.xlsx"
MONTHS = ("2026-07", "2026-08")

HDR = PatternFill("solid", fgColor="1F3864")
OKF = PatternFill("solid", fgColor="C6EFCE")
BADF = PatternFill("solid", fgColor="FFC7CE")
WARN = PatternFill("solid", fgColor="FFEB9C")


def _short(m: str) -> str:
    return (m.replace("Stage2", "").replace("Stage", "")
             .replace("TBMPCR", "M").replace("TBMTBR", "T"))


def _norm(g: str) -> str:
    return re.sub(r"\s+", " ", str(g).replace("GT ", "", 1).strip().upper())


def pcr_source() -> dict[str, set[str]]:
    """PCR xlsx -> per-GT UNION of allowed machines.

    The union matters: the matrix is keyed per FG/SKU and a GT can appear on
    several rows with different machines. We plan at GT level, so the GT is
    allowed on a machine if ANY of its SKUs is. Taking one row instead of the
    union is exactly the bug this sheet exposes on GT 1513 XPC1 MSIL.
    """
    f = paths.INPUT / "allowable machine" / "PCR BUILDING ALLOWABLE MATRIX.xlsx"
    if not f.exists():
        return {}
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb["BUILDING MATRIX"]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    out: dict[str, set[str]] = collections.defaultdict(set)
    for r in it:
        if not r or not r[1]:
            continue
        k = _norm(r[1])
        for i in range(4, len(hdr)):
            if r[i] and str(r[i]).strip().upper() in ("P", "Y", "YES", "X", "1"):
                out[k].add(f"TBMPCR{int(hdr[i]) - 3400}Stage2")
    wb.close()
    return dict(out)


def main() -> None:
    src = pcr_source()
    am = pl.read_parquet(paths.input_derived("allowed_machine_matrix.parquet"))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    cols = ["GT code", "Jul demand", "Aug demand", "Jul short",
            "OUR ALLOWABLE (engine)", "n", "PLANT FILE (raw xlsx)", "n",
            "PLANT OBSERVED 8-mo MES", "n", "8-mo tyres by machine",
            "INCH-only (capable, NEVER used)",
            "INGEST OK?", "OBSERVED WITHIN OUR ALLOWABLE?",
            "machines the PLANT RAN that we FORBID"]

    for plant in ("PCR", "TBR"):
        ws = wb.create_sheet(plant)
        ours = {g: {r["machine"] for r in
                    am.filter((pl.col("plant") == plant) & (pl.col("gt_code") == g))
                      .iter_rows(named=True)}
                for g in am.filter(pl.col("plant") == plant)["gt_code"].unique()}

        dem: dict[str, dict] = {}
        for mth in MONTHS:
            nr = pl.read_parquet(paths.wh_derived(f"net_requirement_{mth}.parquet"))
            for r in nr.filter((pl.col("plant") == plant) & ~pl.col("residual")).iter_rows(named=True):
                dem.setdefault(r["gt_code"], {})[mth] = r["demand"]

        cm = pl.read_parquet(paths.wh_derived("cap_machine_2026-07.parquet"))
        # OBSERVED basis ONLY -- see the module docstring. INCH is theoretical.
        obs = cm.filter((pl.col("plant") == plant)
                        & pl.col("basis").is_in(["OBSERVED", "BOTH", "CERTIFIED"]))
        inch = cm.filter((pl.col("plant") == plant) & (pl.col("basis") == "INCH"))
        mes = {g: set(obs.filter(pl.col("gt_code") == g)["machine"].to_list())
               for g in obs["gt_code"].unique()}
        inchm = {g: set(inch.filter(pl.col("gt_code") == g)["machine"].to_list())
                 for g in inch["gt_code"].unique()}
        mpf = paths.wh_derived("machine_gt_preference_2026-07.parquet")
        hist: dict = {}
        if mpf.exists():
            mp = pl.read_parquet(mpf).filter(pl.col("plant") == plant)
            for r in mp.iter_rows(named=True):
                hist.setdefault(r["gt_code"], {})[r["machine"]] = r.get("history_tyres_8mo") or 0

        run = Path("runs") / "jul_v13" / "gt_events.parquet"
        shortf: dict[str, float] = {}
        if run.exists():
            ge = pl.read_parquet(run).filter(pl.col("plant") == plant)
            cured = (ge.filter(pl.col("d") < 0).group_by("gt_code")
                       .agg((-pl.col("d").sum()).alias("eng")))
            cd = {r["gt_code"]: r["eng"] for r in cured.iter_rows(named=True)}
            for g, d in dem.items():
                if "2026-07" in d:
                    shortf[g] = max(d["2026-07"] - cd.get(g, 0.0), 0.0)

        ws.append(cols)
        for i, c in enumerate(ws[1], 1):
            c.fill, c.font = HDR, Font(bold=True, color="FFFFFF", size=9)
            c.alignment = Alignment(wrap_text=True, vertical="center",
                                    horizontal="center")

        n_ing = n_mes = 0
        for g in sorted(dem, key=lambda x: -shortf.get(x, 0.0)):
            o = ours.get(g, set())
            s = src.get(_norm(g)) if plant == "PCR" else None
            m = mes.get(g, set())
            ing = "n/a" if s is None else ("YES" if s == o else "NO")
            wit = "n/a" if not m or not o else ("YES" if m <= o else "NO")
            extra = sorted(_short(x) for x in (m - o)) if (m and o) else []
            if ing == "NO":
                n_ing += 1
            if wit == "NO":
                n_mes += 1
            ws.append([
                g,
                dem[g].get("2026-07", 0) or None,
                dem[g].get("2026-08", 0) or None,
                round(shortf.get(g, 0.0)) or None,
                ", ".join(sorted(_short(x) for x in o)) or "(not in file)",
                len(o) or None,
                ", ".join(sorted(_short(x) for x in s)) if s else "",
                len(s) if s else None,
                ", ".join(sorted(_short(x) for x in m)),
                len(m) or None,
                ", ".join(f"{_short(k)}:{v:,}" for k, v in
                          sorted(hist.get(g, {}).items(), key=lambda x: -x[1]) if v),
                ", ".join(sorted(_short(x) for x in inchm.get(g, set()) - m)),
                ing, wit, ", ".join(extra)])

        for row in ws.iter_rows(min_row=2):
            for j in (12, 13):                         # the two verdict columns
                c = row[j]
                if c.value == "YES":
                    c.fill, c.value = OKF, "OK"
                elif c.value == "NO":
                    c.fill, c.value = BADF, "X"
                else:
                    c.fill = WARN
                c.alignment = Alignment(horizontal="center")
            for j in (1, 2, 3, 5, 7, 9):
                row[j].alignment = Alignment(horizontal="right")

        for i, w in enumerate([26, 11, 11, 10, 22, 5, 22, 5, 24, 5, 30, 26, 11, 15, 26], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

        ws.append([])
        ws.append([f"{plant}: {len(dem)} demanded GTs · INGEST mismatches {n_ing} "
                   f"· OBSERVED outside our allowable {n_mes}"])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        print(f"  {plant}: {len(dem)} GTs · ingest X {n_ing} · MES-outside X {n_mes}")

    try:
        wb.save(OUT)
        print(f"  -> {OUT}")
    except PermissionError:
        wb.save(_ALT)
        print(f"  !! {OUT.name} is open in Excel -> wrote {_ALT.name}")


if __name__ == "__main__":
    main()
