"""One XLSX per MASTER FILE -- plant-style, individually shareable.

    python scripts/export_master_files.py [out_dir]

Writes to `output/master_files/`. Each file is a single master with a HEADER
block (what it is, how it was derived, and any known limitation) followed by the
data, so a file can be sent to the plant on its own and still be interpretable.

    Master_Line_Speed_Building.xlsx     s/tyre and tyres/h per machine
    Master_Line_Speed_Curing.xlsx       s/tyre and tyres/h per press
    Master_Allowable_Machines.xlsx      GT -> building machines
    Master_Allowable_Presses.xlsx       GT -> curing presses
    Master_Capacity_Machine_Day.xlsx    tyres/machine-day
    Master_Capacity_Press_Day.xlsx      tyres/press-day
    Master_Changeover_Building.xlsx     machine changeover minutes
    Master_Changeover_Curing.xlsx       press mould-change minutes
    Master_Mould.xlsx                   GT x press -> mould, and mould counts
    Master_Lot_Size.xlsx                observed run lengths per GT
    Master_GT_Size.xlsx                 GT -> rim size
    Master_Scrap_Rate.xlsx              GT loss rate per plant per month
    Master_Calendar_Shifts.xlsx         shift grid
    Master_Opening_GT_Inventory.xlsx    WIP at each month open
"""
from __future__ import annotations

import sys
from pathlib import Path

import polars as pl

from planner.config import CONFIG
from planner.data.warehouse import duck, set_cutoff
from planner.runs.logger import log

DER = CONFIG.paths.warehouse / "derived"


def write_master(out: Path, name: str, df: pl.DataFrame, title: str,
                 derivation: str, caveat: str = "") -> Path | None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    if df is None or df.height == 0:
        log.warning("master.empty", name=name)
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:31]
    hdr = Font(bold=True)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Derivation: {derivation}"
    ws["A3"] = f"Rows: {df.height}    Source: JK Tyre MES, 8 months (Dec 2025 - Jul 2026)"
    r = 4
    if caveat:
        ws[f"A{r}"] = f"LIMITATION: {caveat}"
        ws[f"A{r}"].font = Font(bold=True, color="9C0006")
        r += 1
    for c in ("A1", "A2", "A3"):
        ws[c].alignment = Alignment(vertical="center")
    start = r + 1
    ws.append([]) if start > r else None
    ws.cell(row=start, column=1)
    for i, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start, column=i, value=col)
        cell.font = hdr
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
        ws.column_dimensions[cell.column_letter].width = max(
            12, min(34, len(str(col)) + 6))
    for row in df.iter_rows():
        ws.append([v if isinstance(v, (int, float, str, bool)) or v is None
                   else str(v) for v in row])
    ws.freeze_panes = f"A{start + 1}"
    p = out / f"{name}.xlsx"
    wb.save(p)
    return p


def main(out: Path) -> int:
    set_cutoff(None)
    out.mkdir(parents=True, exist_ok=True)
    con = duck()
    made: list[Path] = []

    def rd(n):
        p = DER / f"{n}.parquet"
        return pl.read_parquet(p) if p.exists() else pl.DataFrame()

    # ---- line speed -----------------------------------------------------
    b = rd("cycle_time_building")
    if b.height:
        b = b.with_columns(
            (3600.0 / pl.col("s_per_tyre")).round(2).alias("tyres_per_hour"),
            (86400.0 / pl.col("s_per_tyre")).round(0).alias("tyres_per_day"))
        made.append(write_master(
            out, "Master_Line_Speed_Building", b.sort(["plant", "machine"]),
            "BUILDING LINE SPEED (per machine)",
            "Median observed seconds per tyre from stage-2 build events.",
            "A machine's speed varies with the SKU on it; this is the machine median."))
    c = rd("cycle_time_curing")
    if c.height:
        c = c.with_columns(
            (3600.0 / pl.col("s_per_tyre")).round(2).alias("tyres_per_hour"),
            (28800.0 // pl.col("s_per_tyre")).alias("tyres_per_shift"),
            (3 * (28800.0 // pl.col("s_per_tyre"))).alias("tyres_per_day"))
        made.append(write_master(
            out, "Master_Line_Speed_Curing", c.sort(["plant", "press"]),
            "CURING LINE SPEED (per press)",
            "Throughput cadence, NOT in-press dwell. eff_CT = (raw_dwell + 2.3)/0.94; "
            "tyres/shift = floor(480/eff_CT) x slots. Reproduces the plant's actual "
            "156 (PCR) / 48 (TBR) tyres per press-day exactly.",
            "Dwell time understates capacity ~3x -- do not substitute it."))

    # ---- allowables -----------------------------------------------------
    am = rd("allowed_machine_matrix")
    if am.height:
        made.append(write_master(
            out, "Master_Allowable_Machines", am.sort(["plant", "gt_code", "machine"]),
            "ALLOWABLE BUILDING MACHINES (GT -> machine)",
            "basis=direct: the GT actually ran on that machine. "
            "basis=size: widened to machines that ran the same rim size. "
            "VALIDATED against all 8 months: covers 98.7-99.5% of the "
            "machine-GT pairs the plant actually used. The median of 2 machines "
            "per GT is the plant's real behaviour, not a gap in the data.",
            "New machine-GT pairs settle over time (41% in Jan -> 9% by Jul), so "
            "this is usable as a gate ONLY after ~4 months of history. Still not "
            "a certification list: it records what HAS run, not what MAY run."))
    ap = rd("allowed_press_matrix")
    if ap.height:
        made.append(write_master(
            out, "Master_Allowable_Presses", ap.sort(["plant", "gt_code", "press"]),
            "ALLOWABLE CURING PRESSES (GT -> press)",
            "basis=direct: the GT actually cured on that press. "
            "basis=size: widened to presses that cured the same rim size. "
            "Cross-plant press IDs have been REMOVED (56 rows). Validated "
            "against all 8 months: covers 96.5-99.1% of the press-GT pairs "
            "actually used.",
            "NEVER USE AS A HARD GATE. 36-45% of press-GT pairs are NEW every "
            "single month, with no sign of settling (Jan 45% -> Jul 36%). "
            "History RANKS presses; the platen rim range GATES them, and we do "
            "not have that master."))

    # ---- capacity -------------------------------------------------------
    for n, lbl, unit in [("capacity_machine_day", "BUILDING MACHINE", "machine"),
                         ("capacity_press_day", "CURING PRESS", "press")]:
        d = rd(n)
        if d.height:
            made.append(write_master(
                out, f"Master_Capacity_{'Machine' if unit=='machine' else 'Press'}_Day",
                d.sort(["plant", unit]),
                f"{lbl} DAILY CAPACITY (tyres/day)",
                "p50 / p95 / max of observed daily output per resource.",
                "p95 is the planning figure; max is a single best day and is not "
                "sustainable."))

    # ---- changeovers ----------------------------------------------------
    try:
        cob = con.execute("""
            WITH s AS (
                SELECT plant, machineCode m, event_ts, itemCode gt,
                       lag(itemCode) OVER (PARTITION BY plant, machineCode
                                           ORDER BY event_ts) prev,
                       lag(event_ts) OVER (PARTITION BY plant, machineCode
                                           ORDER BY event_ts) prev_ts
                FROM v_build WHERE stage=2 AND itemCode IS NOT NULL)
            SELECT plant, m AS machine, count(*) changeovers,
                   round(quantile_cont(date_diff('second', prev_ts, event_ts)/60.0, 0.5), 1) gap_p50_min
            FROM s WHERE prev IS NOT NULL AND prev <> gt GROUP BY 1,2 ORDER BY 1,2
        """).pl()
        made.append(write_master(
            out, "Master_Changeover_Building", cob,
            "BUILDING CHANGEOVERS (per machine, 8 months)",
            "Count of SKU transitions and the median gap across the transition.",
            "The gap includes queueing, so it is an UPPER bound on pure setup time. "
            "Use the CTP setup master for the true changeover minutes."))
    except Exception as e:  # noqa: BLE001
        log.warning("master.co_build_failed", err=str(e))

    try:
        coc = con.execute("""
            WITH s AS (
                SELECT c.plant, c.wcID::VARCHAR press, c.event_ts, b.itemCode gt,
                       lag(b.itemCode) OVER (PARTITION BY c.plant, c.wcID
                                             ORDER BY c.event_ts) prev,
                       lag(c.event_ts) OVER (PARTITION BY c.plant, c.wcID
                                             ORDER BY c.event_ts) prev_ts
                FROM v_curing c JOIN v_build b ON b.productionID = c.gtbarCode
                WHERE b.stage=2 AND c.statuscritical='Normal')
            SELECT plant, press, count(*) changeovers,
                   round(quantile_cont(date_diff('second', prev_ts, event_ts)/60.0, 0.5), 1) gap_p50_min
            FROM s WHERE prev IS NOT NULL AND prev <> gt GROUP BY 1,2 ORDER BY 1,2
        """).pl()
        made.append(write_master(
            out, "Master_Changeover_Curing", coc,
            "CURING MOULD CHANGES (per press, 8 months)",
            "Count of GT transitions per press and the median gap across them.",
            "Planning charges one full 480-min shift per mould change."))
    except Exception as e:  # noqa: BLE001
        log.warning("master.co_cure_failed", err=str(e))

    # ---- mould ----------------------------------------------------------
    try:
        mo = con.execute("""
            WITH m AS (
                SELECT b.plant, b.itemCode gt, c.wcID::VARCHAR press,
                       c.MouldCodeLH mould FROM v_build b
                JOIN v_curing c ON b.productionID=c.gtbarCode
                WHERE b.stage=2 AND c.statuscritical='Normal' AND c.MouldCodeLH IS NOT NULL
                UNION ALL
                SELECT b.plant, b.itemCode, c.wcID::VARCHAR, c.MouldCodeRH FROM v_build b
                JOIN v_curing c ON b.productionID=c.gtbarCode
                WHERE b.stage=2 AND c.statuscritical='Normal' AND c.MouldCodeRH IS NOT NULL)
            SELECT plant, gt AS gt_code, press, mould, count(*) tyres
            FROM m GROUP BY 1,2,3,4 ORDER BY 1,2,3,5 DESC
        """).pl()
        made.append(write_master(
            out, "Master_Mould", mo,
            "MOULD MASTER (GT x press -> mould)",
            "Observed mould codes (LH and RH) per GT per press.",
            "Moulds are PER (plant, GT, press) -- each press holds its own physical "
            "copy. Forcing one primary mould per GT produced 416k phantom "
            "double-booking violations."))
        cnt = (mo.group_by(["plant", "gt_code"])
               .agg(pl.col("mould").n_unique().alias("moulds_observed"),
                    pl.col("press").n_unique().alias("presses_used"))
               .sort(["plant", "gt_code"]))
        made.append(write_master(
            out, "Master_Mould_Count", cnt,
            "MOULD COUNT PER GT (M_g)",
            "Distinct mould codes seen for each GT across 8 months.",
            "This is a LOWER BOUND on moulds owned -- a mould never mounted in the "
            "window is invisible. M_g caps how many presses can run a GT at once."))
    except Exception as e:  # noqa: BLE001
        log.warning("master.mould_failed", err=str(e))

    # ---- scrap ----------------------------------------------------------
    try:
        sc = con.execute("""
            WITH b AS (SELECT plant, date_trunc('month', event_ts)::DATE mo,
                              productionID pid FROM v_build
                       WHERE stage=2 AND QualityStatus='1' AND productionID IS NOT NULL),
                 c AS (SELECT DISTINCT gtbarCode pid FROM v_curing
                       WHERE statuscritical='Normal')
            SELECT b.plant, b.mo AS month, count(*) built,
                   count(*) FILTER (WHERE c.pid IS NULL) never_cured,
                   round(100.0*count(*) FILTER (WHERE c.pid IS NULL)/count(*), 3) loss_pct
            FROM b LEFT JOIN c ON b.pid=c.pid GROUP BY 1,2 ORDER BY 1,2
        """).pl()
        made.append(write_master(
            out, "Master_Scrap_Rate", sc,
            "GREEN TYRE LOSS / SCRAP RATE",
            "Tyres built and NEVER cured anywhere in the history. This is what "
            "build/cure ratio - 1 actually is: the plant runs 1.0032 (PCR) while "
            "its inventory stays trend-flat, so the excess must leave the system.",
            "The LAST month is right-censored (its tyres may cure next month) -- "
            "ignore it. TBR is NOT stationary: 1.09% (Jan) -> 2.87% (May); derive "
            "it from the trailing window, never fix it as a constant."))
    except Exception as e:  # noqa: BLE001
        log.warning("master.scrap_failed", err=str(e))

    # ---- lot size / gt size / calendar / opening ------------------------
    ls = rd("lot_size")
    if ls.height:
        made.append(write_master(
            out, "Master_Lot_Size", ls.sort(["plant", "gt_code"]),
            "OBSERVED RUN LENGTH PER GT",
            "Gaps-and-islands over consecutive same-GT production on a machine.",
            "Descriptive, not a target. The engine sizes lots from the replenishment "
            "interval (Q_g = draw_g x T_0), because holding QUANTITY constant makes "
            "the replenishment GAP swing ~11x across GTs and starves slow movers."))
    gs = rd("gt_size")
    if gs.height:
        made.append(write_master(
            out, "Master_GT_Size", gs.sort("gt_code"),
            "GT -> RIM SIZE",
            "Parsed from the GT code and the SKU construction mapping.",
            "Drives the building rim-size lock (measured 99.89% PCR / 99.75% TBR)."))
    cal = rd("calendar_shifts")
    if cal.height:
        made.append(write_master(
            out, "Master_Calendar_Shifts", cal,
            "SHIFT GRID",
            "3 shifts x 480 min, plant day boundary 07:00.",
            "NO PLANT CALENDAR MASTER EXISTS -- 24x7 is assumed. January shows a "
            "near-shutdown day (3,068 vs 12,666 typical) that is NOT modelled. "
            "This is the single biggest input gap."))
    od = CONFIG.paths.masters / "opening_gt"
    if od.exists():
        frames = [pl.read_csv(p) for p in sorted(od.glob("opening_gt_2*.csv"))]
        if frames:
            made.append(write_master(
                out, "Master_Opening_GT_Inventory", pl.concat(frames),
                "OPENING GREEN TYRE INVENTORY (per month)",
                "Built before the month opens and not yet cured, as of the 1st at "
                "07:00. Age-bounded at the measured build->cure p99 lag.",
                "December excluded: no prior month to carry stock from. Measured "
                "level is ~4,820 (PCR) / ~1,297 (TBR) = about 9 HOURS of production "
                "(Little's Law, I = lambda x W)."))

    # ---- demand ---------------------------------------------------------
    dd = CONFIG.paths.masters / "demand"
    for p in sorted(dd.glob("demand_2*.csv")) if dd.exists() else []:
        tag = p.stem.replace("demand_", "")
        made.append(write_master(
            out, f"Master_Demand_{tag}", pl.read_csv(p),
            f"DEMAND {tag} (daily, per GT)",
            "Tyres CURED per GT per day. Demand is FINISHED tyres -- a green tyre "
            "is an intermediate, not an order line.",
            "Derived from this month's own production, so planning this month with "
            "it is IN-SAMPLE. Use it to score against the plant, not as a forecast."))

    made = [m for m in made if m]
    for m in sorted(made):
        print(f"WROTE {m.name}  ({m.stat().st_size/1024:.0f} KB)")
    print(f"\n{len(made)} master files -> {out}")
    log.info("master_files.done", n=len(made))
    return 0


if __name__ == "__main__":
    d = Path(sys.argv[1]) if len(sys.argv) > 1 else CONFIG.paths.root / "output" / "master_files"
    sys.exit(main(d))
